"""The "Loss by direction" sheet — A->B, B->A and their average, per flagged cell.

The grid prints ONE number per cell, the bidirectional average, because that is
what FastReporter reports and what a reviewer hand-types.  The average alone
cannot be checked though: 0.30 dB reads the same whether both ends measured
0.30 or one end read 0.60 against a 0.00 gainer, and those are different
findings.  FastReporter's own export puts all three side by side.

Ground truth for the arithmetic, from the real FR exports (WSC<->SUI, 1152
fibers): FR's per-event "Avg." column equals (A->B + B->A) / 2 on 11,565 of
11,565 paired events -- zero exceptions.  This sheet prints that same
arithmetic, so these tests pin it exactly.

The scope is deliberately narrow and the tests hold it there: FLAGGED cells
only.  The engine keeps what it flags, so a per-fiber average taken from this
sheet would be an average of the failures -- which is why the header says so.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from conftest import (
    run_splicereport, FIXTURE_SPLICE_A_DIR, FIXTURE_SPLICE_B_DIR, SPLICEREPORT_DIR,
)

SHEET = "Loss by direction"


def _book(tmp_path, name="d.xlsx", **kw):
    out = tmp_path / name
    rc, m, err = run_splicereport(FIXTURE_SPLICE_A_DIR, FIXTURE_SPLICE_B_DIR,
                                  out, **kw)
    assert rc == 0 and m and m.get("ok"), f"run failed: {err[-1200:]}"
    return openpyxl.load_workbook(out), m


def _rows(ws):
    """[(fiber, splice, km, a, b, avg, label)] for the data rows."""
    out = []
    for r in ws.iter_rows(min_row=5, values_only=True):
        if r[0] is None:
            continue
        out.append(r[:7])
    return out


def test_sheet_exists_with_fr_column_shape(tmp_path):
    wb, _ = _book(tmp_path)
    assert SHEET in wb.sheetnames
    ws = wb[SHEET]
    hdr = [ws.cell(row=4, column=c).value for c in range(1, 8)]
    assert hdr == ["Fiber", "Splice", "Position (km)", "A->B (dB)",
                   "B->A (dB)", "Avg. (dB)", "What the report says"]


def test_average_is_exactly_the_mean_of_the_two_directions(tmp_path):
    """The headline contract, and FR's own rule: Avg. == (A->B + B->A) / 2."""
    wb, _ = _book(tmp_path)
    rows = _rows(wb[SHEET])
    assert rows, "fixture must produce flagged cells"
    checked = 0
    for fib, _sp, _km, a, b, avg in ((r[0], r[1], r[2], r[3], r[4], r[5]) for r in rows):
        if a is None or b is None:
            continue                      # single-direction cell, see below
        assert avg is not None, fib
        # All THREE values are rounded to 3 dp independently, so the displayed
        # numbers need not average exactly: (a,b) each carry up to 0.5 mdB of
        # rounding, and the printed average carries its own 0.5 mdB against the
        # raw mean.  1 mdB is the whole envelope.  FastReporter's export has
        # the same property -- recomputing its per-fiber average from its
        # DISPLAYED per-event values lands within half a millibel rather than
        # exactly, which is how we know FR averages raw and rounds once.
        assert abs((a + b) / 2.0 - avg) <= 0.001, (fib, a, b, avg)
        checked += 1
    assert checked, "no bidirectional rows to check"


def test_only_flagged_cells_are_listed(tmp_path):
    """The sheet explains the report's findings.  A cell the report did not
    flag must not appear -- including one the RETAIN_UNFLAGGED path kept."""
    wb, m = _book(tmp_path)
    rows = _rows(wb[SHEET])
    assert len(rows) == m["n_flagged"], (len(rows), m["n_flagged"])


def test_header_warns_against_taking_an_average_from_it(tmp_path):
    """A reader who assumed this was every splice would compute a per-fiber
    average from the failures alone.  The sheet has to say it is not that."""
    wb, _ = _book(tmp_path)
    ws = wb[SHEET]
    blurb = " ".join(str(ws.cell(row=r, column=1).value or "") for r in (1, 2))
    assert "not a census" in blurb
    assert re.search(r"per-fiber average must not be taken", blurb)
    assert "splices that passed are not listed" in blurb.lower()


def test_a_and_b_are_named_for_the_real_span_ends(tmp_path):
    """A->B / B->A mean nothing without knowing which end is which."""
    out = tmp_path / "named.xlsx"
    rc, m, err = run_splicereport(FIXTURE_SPLICE_A_DIR, FIXTURE_SPLICE_B_DIR,
                                  out, site_a="ELMHURST", site_b="MILLER")
    assert rc == 0 and m and m.get("ok"), err[-800:]
    ws = openpyxl.load_workbook(out)[SHEET]
    blurb = str(ws.cell(row=2, column=1).value)
    assert "ELMHURST" in blurb and "MILLER" in blurb


def test_single_direction_cells_leave_the_other_side_blank(tmp_path):
    """FR prints '---' for an end that could not see the event.  A blank is
    our equivalent; a 0.000 there would read as a measured gainer."""
    wb, _ = _book(tmp_path)
    for r in _rows(wb[SHEET]):
        a, b, avg = r[3], r[4], r[5]
        if a is None or b is None:
            assert not (a is None and b is None), "a cell with neither side"
        # Whatever is present must be a number, never a placeholder string.
        for v in (a, b, avg):
            assert v is None or isinstance(v, (int, float)), r


def test_sheet_order_keeps_the_audit_first(tmp_path):
    """test_stale_engine_gate pins sheet 0 as the acquisition audit; adding a
    sheet must not displace it, and the grid must stay where readers expect."""
    wb, _ = _book(tmp_path)
    assert wb.sheetnames[0] == "Acquisition Parameters"
    assert "Splice Report" in wb.sheetnames and "Legend" in wb.sheetnames
    assert wb.sheetnames.index(SHEET) > wb.sheetnames.index("Splice Report")


def test_positions_and_losses_are_numbers_not_text(tmp_path):
    """Someone will sort and filter this sheet.  Text would sort wrongly."""
    wb, _ = _book(tmp_path)
    ws = wb[SHEET]
    seen = 0
    for row in ws.iter_rows(min_row=5):
        if row[0].value is None:
            continue
        for c in (row[2], row[3], row[4], row[5]):
            assert c.value is None or isinstance(c.value, (int, float))
            if c.value is not None:
                assert c.number_format == '0.000'
        seen += 1
    assert seen


def test_the_engine_documents_the_fr_provenance():
    """The arithmetic is pinned to a measured fact, not a guess; keep the
    citation next to the code so a later editor sees what it must not break."""
    src = (Path(SPLICEREPORT_DIR) / "splicereportmatchexfo.py").read_text(encoding="utf-8")
    blk = src.split('ws_dir = wb.create_sheet("Loss by direction")')[0][-2600:]
    assert "11,565" in blk and "WSC" in blk, \
        "the FR-parity evidence for (A+B)/2 must stay documented here"


def test_retain_unflagged_ships_off_and_changes_nothing():
    """The flag that keeps passing splices is for the average-splice-loss work.
    It must default OFF: every shipped path wants the flagging behaviour, and
    a retained cell must never reach the grid or this sheet."""
    src = (Path(SPLICEREPORT_DIR) / "splicereportmatchexfo.py").read_text(encoding="utf-8")
    assert re.search(r"^RETAIN_UNFLAGGED\s*=\s*False\s*$", src, re.M)
    # The sheet writer gates on is_flagged, so a retained cell cannot render.
    blk = src.split('ws_dir = wb.create_sheet("Loss by direction")')[1][:1800]
    assert "if not _c.get('is_flagged'):" in blk and "continue" in blk
