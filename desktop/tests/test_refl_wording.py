"""Reflective events print REFL and the data — no place-name, no verdict.

Robert (2026-08-21): "for reflective events we can eliminate the verbiage
about tailbox or launch box.  we just say REFL and the data", "across uni and
splice report", "no, we don't change the flagging.  just the words we use".

The report prints DATA; the tech makes the call.  ``BAD`` was a verdict and
``TAILBOX`` / ``LAUNCH`` diagnosed a PLACE — and a place is precisely what we
have been wrong about before (a receive spool's bare end read as the cable's
tailbox, project_wscsui_connector).  The reflectance number is the fact; the
distance column is where the tech places it.

This module pins the printed forms so the old vocabulary cannot creep back,
and pins the three boundaries Robert drew explicitly:

  * the ``(A)`` / ``(B)`` / ``(B-fill)`` direction suffixes stay,
  * the loss tags (``LAUNCH_LOSS``, the ``.73 LAUNCH`` connector tag) stay,
  * ``bidi`` stays.

It ALSO pins the piece of the change that is easy to drop: renaming the two
reflectance tags without following them into the ``is_review`` prefix tuple
silently demotes both findings from REVIEW to WATCH.
"""
import os
import sys

import pytest

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
sys.path.insert(0, os.path.join(ROOT, 'splicereport'))
import splicereportmatchexfo as E          # noqa: E402

ENGINE_SRC = open(os.path.join(ROOT, 'splicereport', 'splicereportmatchexfo.py'),
                  encoding='utf-8').read()
FIX_CONT = os.path.join(HERE, 'fixtures', 'continuous')

SPAN = 50.0
LAUNCH = 1.0


# ── fixtures ────────────────────────────────────────────────────────────────

def _ev(km, loss, typ='0F9999', refl=0.0, end=False, tot=0):
    return {'dist_km': km, 'splice_loss': loss, 'is_end': end,
            'is_reflective': typ.startswith(('1F', '1E', '2F')),
            'reflection': refl, 'type': typ, 'time_of_travel': tot}


def _dir(launch_refl=-60.0, tailbox_refl=-57.0, span=SPAN):
    """One direction as pass 0 leaves it: port reflection, own launch
    connector, a mid-span splice, the far end's connector one reel back, EOF."""
    raw = [_ev(0.0, 0.0, '1F', -60.0, tot=0),
           _ev(LAUNCH, 0.05, '1F', launch_refl, tot=1000),
           _ev(25.0, 0.05, '0F', 0.0, tot=2000),
           _ev(span - LAUNCH, 0.05, '1F', tailbox_refl, tot=3000),
           _ev(span, 0.0, '1E', -60.0, end=True, tot=4000)]
    return {'_raw_events': raw,
            'events': E._normalize_untrimmed_events(list(raw))}


# ── the two ILA reflectance tags ────────────────────────────────────────────

def test_launch_reflectance_tag_is_bare_refl_plus_the_number():
    """The launch rule prints REFL and the reflectance.  Nothing else."""
    iss = E.detect_launch_issues({1: _dir(launch_refl=-20.0)},
                                 {1: _dir()})
    tags = iss[1]['a_tags']
    assert tags == ['REFL-20.0dB'], tags
    assert iss[1]['refl_rules']['A'] == ['launch'], iss[1]['refl_rules']


def test_tailbox_reflectance_tag_is_bare_refl_plus_the_number():
    """The tailbox rule prints the SAME word.  The report does not say where
    it thinks the reflection is — the distance column does that."""
    a = {i: _dir() for i in range(1, 21)}
    b = {i: _dir() for i in range(1, 21)}
    b[1] = _dir(tailbox_refl=-40.0)
    iss = E.detect_launch_issues(a, b)
    # The bad tailbox is in the B DIRECTION, and a tailbox reading is filed at
    # the end it is AT — the far end of a B shot is the A END.  Wording is
    # what this test is about; placement is test_ila_physical_end.py's.
    tags = iss[1]['a_tags']
    assert tags == ['REFL-40.0dB'], tags
    assert iss[1]['refl_rules']['A'] == ['tailbox'], iss[1]['refl_rules']
    assert iss[1]['b_tags'] == [], iss[1]['b_tags']


# `dirn` is the direction the fixture's bad fiber goes in; `end` is the tag
# list it files under.  They differ for the tailbox rule: launch fires at the
# end the shot was launched FROM, tailbox at the OTHER end, so the B
# direction's tailbox reading files at end A.
@pytest.mark.parametrize('kwargs,dirn,end', [
    ({'launch_refl': -20.0}, 'a', 'a_tags'),
    ({'tailbox_refl': -40.0}, 'b', 'a_tags'),
])
def test_no_reflectance_tag_names_a_place_or_passes_a_verdict(kwargs, dirn, end):
    a = {i: _dir() for i in range(1, 21)}
    b = {i: _dir() for i in range(1, 21)}
    (a if dirn == 'a' else b)[1] = _dir(**kwargs)
    tags = E.detect_launch_issues(a, b)[1][end]
    assert tags, 'fixture stopped firing the rule'
    for t in tags:
        assert 'TAILBOX' not in t, t
        assert 'LAUNCH' not in t, t
        assert 'BAD' not in t, t


def test_renaming_the_tags_did_not_demote_them_to_watch():
    """The trap: ``is_review`` matched on the OLD prefixes.  Rename the tags
    and forget the tuple and both findings drop REVIEW -> WATCH."""
    iss = E.detect_launch_issues({1: _dir(launch_refl=-20.0)}, {1: _dir()})
    assert iss[1]['severity'] == 'REVIEW', iss[1]


def test_refl_prefix_cannot_collide_with_another_launch_tag():
    """``is_review`` now keys on the bare prefix 'REFL'.  That is only safe
    while no other tag this function emits starts with it.

    Assert against the literals the ENGINE emits.  An earlier version checked
    this test's own token tuple, so it could never fail -- injecting a
    'REFLOW_NEEDED' tag left it green."""
    import re
    emitted = [ln.split('tags.append(', 1)[1] for ln in ENGINE_SRC.split('\n')
               if 'tags.append(' in ln]
    assert emitted, 'found no tags.append() sites -- the scrape broke'

    names = []
    for lit in emitted:
        m = re.search(r'''f?['"]([A-Za-z_][A-Za-z0-9_]*)''', lit)
        if m:
            names.append((m.group(1), lit))
    assert names, 'scraped no tag names -- the scrape broke'

    offenders = [l for n, l in names if n.startswith('REFL') and n != 'REFL']
    assert not offenders, (
        'a tag other than the bare reflectance label now starts with REFL, '
        'so is_review would silently capture it: %r' % (offenders,))


def test_ila_cell_text_survives_the_tag_splitter():
    """``build_ribbon_data`` prints ``tag.split('@')[0].split('+')[0]``.  A
    reflectance is always negative here, so the '+' split never bites — but
    the cell is what the tech reads, so pin the rendered text."""
    iss = E.detect_launch_issues({1: _dir(launch_refl=-20.0)}, {1: _dir()})
    _, lca, _ = E.build_ribbon_data({}, n_fibers=12, ribbon_size=12,
                                    n_splices=1, launch_issues=iss)
    assert lca[0]['text'] == '1 REFL-20.0dB', lca[0]


# ── in-grid reflective labels ───────────────────────────────────────────────
# Same fixture shapes as test_no_dirty_connector_label.py — proven to reach
# each label-emitting site.

GRID_SPAN = 40.0


def _gev(km, loss=0.0, typ='0F9999', refl=-60.0, end=False):
    return {'dist_km': km, 'splice_loss': loss, 'type': typ,
            'reflection': refl, 'is_end': end}


def test_bidirectional_reflective_cell_reads_refl():
    a = {'events': [_gev(0.0, 0.0, '1F9999', -35.0),
                    _gev(10.0, 0.30, '1F9999', -32.0),
                    _gev(30.0, 0.03),
                    _gev(GRID_SPAN, 0.0, end=True)],
         'wavelength': 1550.0}
    b = {'events': [_gev(0.0, 0.0, '1F9999', -35.0),
                    _gev(GRID_SPAN - 30.0, 0.03),
                    _gev(GRID_SPAN - 10.0, 0.30, '1F9999', -32.0),
                    _gev(GRID_SPAN, 0.0, end=True)],
         'wavelength': 1550.0}
    res = E.analyze_all({1: a}, {1: b}, [{'position_km': 10.0}],
                        E.REBURN_THRESHOLD)
    assert res[(1, 0)]['label'] == '1 REFL .300 (-32dB)', res[(1, 0)]['label']


def test_standalone_reflective_cell_reads_refl():
    a = {'events': [_gev(0.0, 0.0, '1F9999', -35.0),
                    _gev(20.0, 0.05),
                    _gev(25.0, 0.40, '1F9999', -30.0),
                    _gev(32.0, 0.03),
                    _gev(GRID_SPAN, 0.0, end=True)],
         'wavelength': 1550.0}
    res = E.scan_a_standalone_events({2: a}, [{'position_km': 20.0}],
                                     {}, GRID_SPAN)
    assert res[(2, 0)]['label'] == '2 REFL .400 (-30dB)', res[(2, 0)]['label']


def test_break_label_reads_refl_and_keeps_its_loss_wording():
    """A BREAK cell carries a reflectance too.  'uni' there qualifies the
    LOSS (the variable is ``uni_loss``), so it stays — Robert: 'keep the loss
    ones the same'.  Only the word for the reflectance changes."""
    a = {'events': [_gev(0.0, 0.0, '1F9999', -35.0),
                    _gev(20.0, 0.05),
                    _gev(25.0, 4.50, '1F9999', -30.0),   # the break
                    _gev(30.0, 0.0, end=True)],          # fiber dies short
         'wavelength': 1550.0}
    res = E.scan_a_standalone_events({3: a}, [{'position_km': 20.0}],
                                     {}, GRID_SPAN)
    assert res, 'fixture stopped producing the break'
    cell = list(res.values())[0]
    assert cell['is_break'] is True, cell
    lbl = cell['label']
    assert 'BREAK' in lbl, lbl
    assert 'uni REFL-30dB' in lbl, lbl
    assert 'reflection' not in lbl, lbl
    assert 'air gap' in lbl, lbl


def test_bidir_ghost_reflection_label_reads_refl():
    """Never exercised by any span on disk — pinned at fixture level."""
    a = {'events': [_gev(0.0, 0.0, '1F9999', -35.0),
                    _gev(20.0, 0.005, '1F9999', -71.0),
                    _gev(GRID_SPAN, 0.0, end=True)],
         'wavelength': 1550.0}
    b = {'events': [_gev(0.0, 0.0, '1F9999', -35.0),
                    _gev(GRID_SPAN - 20.0, 0.005, '1F9999', -70.0),
                    _gev(GRID_SPAN, 0.0, end=True)],
         'wavelength': 1550.0}
    res = E.scan_bidir_ghost_reflections({1: a}, {1: b},
                                         [{'position_km': 30.0}], {}, GRID_SPAN)
    assert res, 'fixture stopped producing a bidir ghost reflection'
    lbl = list(res.values())[0]['label']
    assert lbl == '1 REFL @ 20.00km (-71/-70dB bidir)', lbl


def test_merged_reflective_label_reads_refl():
    """Also never exercised by any span on disk."""
    a = {'events': [_gev(0.0, 0.0, '1F9999', -35.0),
                    _gev(20.0, 0.10, '1F9999', -56.0),
                    _gev(GRID_SPAN, 0.0, end=True)],
         'wavelength': 1550.0}
    b = {'events': [_gev(0.0, 0.0, '1F9999', -35.0),
                    _gev(GRID_SPAN, 0.0, end=True)],
         'wavelength': 1550.0}
    res = E.scan_merged_reflective_events({1: a}, {1: b},
                                          [{'position_km': 30.0}], {}, GRID_SPAN)
    assert res, 'fixture stopped producing a merged reflective event'
    lbl = list(res.values())[0]['label']
    assert lbl.startswith('1 REFL @ 20.00km (-56dB '), lbl
    assert 'own-frame)' in lbl, lbl


def test_splice_report_reflective_column_header_reads_refl():
    assert 'header = f"REFL @ {ref_km:.2f}km"' in ENGINE_SRC


# ── uni report ──────────────────────────────────────────────────────────────

@pytest.mark.skipif(not os.path.isdir(FIX_CONT), reason='fixture set missing')
def test_uni_report_says_refl_in_every_place_it_renders(tmp_path):
    """Grid column header, Flagged Events 'Column' + 'Kind' cells, and the
    hub grid's own column label — one vocabulary across all four."""
    import shutil
    from openpyxl import load_workbook
    src = tmp_path / 'in'
    src.mkdir()
    for f in sorted(os.listdir(FIX_CONT)):
        if f.endswith('.sor'):
            shutil.copy2(os.path.join(FIX_CONT, f), src)
    out = tmp_path / 'o.xlsx'
    res = E.uni_generate(str(src), str(out))
    assert res['reflective_columns'], 'fixture stopped producing a reflective column'

    # hub grid (app.py renders grid_columns[].label)
    hub = [c['label'] for c in res['grid_columns'] if c['kind'] == 'reflective']
    assert hub and all(l.startswith('REFL ') for l in hub), hub
    assert not any(l.startswith('Reflective') for l in hub), hub

    wb = load_workbook(out)
    texts = {}
    for name in wb.sheetnames:
        texts[name] = [[c.value for c in row]
                       for row in wb[name].iter_rows()]

    grid = [v for row in texts['Unidir Events'] for v in row
            if isinstance(v, str) and 'REFL' in v]
    assert grid, texts['Unidir Events'][:6]

    flat_flagged = [v for row in texts['Flagged Events'] for v in row
                    if isinstance(v, str)]
    assert 'REFL' in flat_flagged, 'Kind cell must read REFL'
    assert any(v.startswith('REFL ') for v in flat_flagged), 'Column cell'

    # nowhere in either sheet does the old column word survive
    for sheet in ('Unidir Events', 'Flagged Events'):
        for row in texts[sheet]:
            for v in row:
                if isinstance(v, str):
                    assert not v.startswith('Reflective '), (sheet, v)


# ── the boundaries Robert drew ──────────────────────────────────────────────

def test_direction_suffixes_are_untouched():
    """He was asked directly and said REFL only for now."""
    for form in ('(A)', '(B)', '(B-fill)'):
        assert form in ENGINE_SRC, form
    # Rendered through the cell writer itself (build_ribbon_data), which is
    # where the (A) / (B) / (B-fill) suffixes are printed.
    def _one(flag):
        return {'fiber': 7, 'bidir_loss': 0.40, 'a_loss': 0.40, 'b_loss': 0.40,
                'is_break': False, 'is_broke': False, 'is_bend': False,
                'is_ref': False, 'is_gainer': False,
                'is_bfill': flag == 'is_bfill',
                'is_a_only': flag == 'is_a_only',
                'is_b_only': flag == 'is_b_only',
                'is_flagged': True, 'label': ''}
    for flag, suffix in (('is_a_only', '(A)'), ('is_b_only', '(B)'),
                         ('is_bfill', '(B-fill)')):
        cells, _, _ = E.build_ribbon_data({(7, 0): _one(flag)}, n_fibers=12,
                                          ribbon_size=12, n_splices=1)
        text = cells[(0, 0)]['text']
        assert text == f'7 .400 {suffix}', (flag, text)


def test_loss_tags_are_untouched():
    """'keep the loss ones the same, just focus on REFL for now'."""
    assert "f'LAUNCH_LOSS{launch_loss_signed:+.2f}dB'" in ENGINE_SRC
    assert "'HIGH_LAUNCH_LOSS'" in ENGINE_SRC
    assert "' LAUNCH'" in ENGINE_SRC
    assert "' LAUNCH ' + side + ' side'" in ENGINE_SRC


def test_bidi_is_untouched():
    """'bidi stays the same' -- Robert drew this boundary explicitly.

    Key on the PRINTED template, not on the substring ' bidi': that also
    matches ' bidirectional', which appears ~45 times in comments, so the
    earlier form stayed green even when every printed label was renamed."""
    assert '{loss_str} bidi' in ENGINE_SRC, 'the printed bidi label changed'
    printed = [ln.strip() for ln in ENGINE_SRC.split('\n')
               if ' bidi' in ln and 'bidirectional' not in ln
               and ('f"' in ln or "f'" in ln)]
    assert len(printed) >= 6, (
        'expected >=6 printed bidi label templates, found %d: %r'
        % (len(printed), printed))


def test_the_old_reflective_vocabulary_is_gone_from_the_engine():
    for dead in ('BAD_LAUNCH_REFL', 'BAD_TAILBOX_REFL', 'uni reflection',
                 '(refl {', 'Ref @ {', 'f"Reflective ', "f'Reflective ",
                 "'Reflective'", 'Reflective event'):
        assert dead not in ENGINE_SRC, dead
