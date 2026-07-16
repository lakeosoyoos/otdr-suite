"""Regression tests for the Secret Sauce robust analysis window +
suspected-break reporting (sandbox/ss-window-robust) and the LONG-SPAN
WINDOW RESTORATION (sandbox/ss-window-longspan).

THE BUG: report_sor.py computed the folder's common analysis span as the
raw MINIMUM trace length over all files, so ONE broken fiber collapsed the
whole folder's window.  A-F West 145-288 (266 files, median EOF 2 037 m)
has port 198 physically broken ~1 km out (A-side EOF 1 005 m, B-side
1 036 m — the two sum to the span), which shrank the interior to 305 m,
hid the folder-wide similarity, misrouted the regime, and fed the
1 997-false-positive flood.

THE FIX (secretsauce/report_sor.py):
  * _robust_common_span — EOFs below 75 % of the folder median are
    suspected breaks (always reported) and are ALWAYS excluded from the
    common-span computation and pair metrics (they physically lack the
    glass being compared), provided ≥2 healthy files remain.  The first
    cut gated exclusion on the raw-min window collapsing below 2 km; that
    precondition was REMOVED on 2026-07-15 (Robert approved the baseline
    change: ELMMIL_1550's window was capped at 22.3 km of a 69.5 km span,
    SANDUR at 20.1 km of ~100.9 km).
  * SANITY GUARD — if more than 20 % of the folder's files sit below 75 %
    of the median (_INCONSISTENT_FOLDER_FRAC), the folder itself is
    inconsistent: exclude NOTHING, keep the raw-min window, and warn
    ("folder trace lengths are inconsistent (N of M below 75% of median)
    — window not restored; check folder contents").
  * _ab_break_notes — when BOTH directions of the same port are suspected
    breaks and their EOFs sum to the folder median (±10 %), each gets
    "A+B lengths are consistent with a break ~<EOF> m from the <prefix>
    end".
  * Reporting — additive `short_traces` + `window_warnings` manifest keys
    (present only when non-empty), a "Suspected broken / short fibers"
    PDF section (+ warning banner when the guard fired), and a "Suspected
    short fibers" XLSX sheet + Summary rows (all only when suspected
    breaks exist, so unaffected outputs stay byte-stable).

Calibration lock (2026-07-15, real folders): exclusion fires on A-F West
(2 of 266 short → healthy-min window 2 036.8 m), ELMMIL (1 of 1152:
ELMMIL0231 @ 22 288 m → window restored to 69 539 m) and SANDUR (2 of
864: SANDUR841/SANDUR229 → window restored to 100 548 m); the guard fires
on NONE of them (0.8 % / 0.09 % / 0.23 % short).  Homogeneous folders
(A-F East, SEANOR, PTL Panel A, TEST DUPE, the test fixtures) keep
byte-identical pair tables — verified by sha256 pair-table comparison
old-engine vs new-engine on 2026-07-15.

Namespace isolation rule: the Secret Sauce engine is only ever exercised
through subprocesses (run_secretsauce / sys.executable -c), never imported
into the test process.
"""
from __future__ import annotations

import json
import subprocess
import sys

from conftest import SECRETSAUCE_DIR, run_secretsauce, single_dir_fixture


def _run_engine_script(script, *args, timeout=240):
    p = subprocess.run([sys.executable, "-c", script, str(SECRETSAUCE_DIR),
                        *[str(a) for a in args]],
                       capture_output=True, text=True, timeout=timeout)
    assert p.returncode == 0, p.stderr[-2000:]
    return json.loads(p.stdout.strip().splitlines()[-1])


# ── _robust_common_span unit behavior ───────────────────────────────────────

_SPAN_UNIT_SCRIPT = r"""
import sys, json
sys.path.insert(0, sys.argv[1])
from report_sor import _robust_common_span

cases = {}

# Homogeneous folder: exact min preserved (byte-identity guarantee), no
# outliers, no exclusions.
cases['homog'] = _robust_common_span([2036.8, 2037.2, 2040.0, 2041.9])

# Homogeneous LONG folder (SEANOR-like): min is 99% of median — untouched.
cases['homog_long'] = _robust_common_span([108817.5] + [109954.0] * 9)

# A-F West shape: one A+B broken pair collapses the min below 2 km while
# the median says ~2 037 m -> both outliers excluded, healthy-min window.
west = [1005.0, 1036.4] + [2037.0] * 264
cases['west'] = _robust_common_span(west)

# ELMMIL shape: extreme outlier (32% of median) with a raw-min window
# >= 2 km.  LONG-SPAN RESTORATION: now ALWAYS excluded, window rebuilt
# from the healthy population (was: reported-but-kept under the removed
# 2 km-collapse gate).
cases['elmmil'] = _robust_common_span([22288.1] + [69554.0] * 9)

# SANDUR shape: two extreme outliers (20% / 59% of median) on a ~101 km
# span -> both excluded, window restored to the healthy min.
cases['sandur'] = _robust_common_span(
    [20144.2, 59219.2] + [100925.0] * 861 + [100547.7])

# Two-file folder: 1 of 2 short = 50% > the 20% guard -> no exclusion,
# raw-min window, inconsistency warning.
cases['twofile'] = _robust_common_span([1005.0, 2037.0])

# SANITY GUARD: 1 of 4 short (25% > 20%) -> exclude NOTHING, keep the
# raw-min window, warn.
cases['guard25'] = _robust_common_span([1000.0, 4000.0, 4001.0, 4002.0])

# Boundary: exactly 20% short is NOT "more than 20%" -> exclusion fires.
cases['exact20'] = _robust_common_span([1000.0] + [4000.0] * 4)

# Short-but-homogeneous panel (Deming-style): no outliers, min preserved —
# the short_panel / short-common-span regime routes keep working on it.
cases['short_panel'] = _robust_common_span([150.0] * 60)

# Outlier just ABOVE the 75% cut is span-length jitter, not a break.
cases['jitter'] = _robust_common_span([1530.0] + [2000.0] * 9)

print(json.dumps(cases))
"""

_GUARD_MSG_25 = ('folder trace lengths are inconsistent (1 of 4 below 75% '
                 'of median) — window not restored; check folder contents')


def test_robust_common_span_units():
    out = _run_engine_script(_SPAN_UNIT_SCRIPT)

    span, med, outliers, excluded, guard = out["homog"]
    assert span == 2036.8 and outliers == [] and excluded == []
    assert guard is None

    span, med, outliers, excluded, guard = out["homog_long"]
    assert span == 108817.5 and outliers == [] and excluded == []
    assert guard is None

    span, med, outliers, excluded, guard = out["west"]
    assert span == 2037.0, out["west"]
    assert med == 2037.0
    assert outliers == [0, 1] and excluded == [0, 1]
    assert guard is None

    span, med, outliers, excluded, guard = out["elmmil"]
    assert span == 69554.0, "long-span window must be RESTORED"
    assert outliers == [0] and excluded == [0], \
        "the short fiber is reported AND excluded"
    assert guard is None

    span, med, outliers, excluded, guard = out["sandur"]
    assert span == 100547.7, out["sandur"]
    assert outliers == [0, 1] and excluded == [0, 1]
    assert guard is None

    span, med, outliers, excluded, guard = out["twofile"]
    assert span == 1005.0 and excluded == []
    assert guard is not None and "inconsistent (1 of 2" in guard

    span, med, outliers, excluded, guard = out["guard25"]
    assert span == 1000.0, "guard keeps the raw-min window"
    assert outliers == [0], "the short file is still REPORTED"
    assert excluded == [], "guard: exclude nothing"
    assert guard == _GUARD_MSG_25, guard

    span, med, outliers, excluded, guard = out["exact20"]
    assert span == 4000.0 and outliers == [0] and excluded == [0]
    assert guard is None, "exactly 20% is not MORE than 20%"

    span, med, outliers, excluded, guard = out["short_panel"]
    assert span == 150.0 and outliers == [] and excluded == []
    assert guard is None

    span, med, outliers, excluded, guard = out["jitter"]
    assert outliers == [] and excluded == [] and span == 1530.0
    assert guard is None


# ── _ab_break_notes unit behavior ───────────────────────────────────────────

_AB_UNIT_SCRIPT = r"""
import sys, json
sys.path.insert(0, sys.argv[1])
from report_sor import _ab_break_notes

def entries(*pairs):
    return [{'file': f, 'eof_m': e} for f, e in pairs]

cases = {}

# The boss's West case: both directions of port 198, EOFs sum to ~median.
es = entries(('BCK1BCK60198', 1005.0), ('BCK6BCK10198', 1036.4))
_ab_break_notes(es, 2037.2)
cases['west'] = es

# Sum NOT consistent with the median span -> no note.
es = entries(('BCK1BCK60198', 1005.0), ('BCK6BCK10198', 1500.0))
_ab_break_notes(es, 2037.2)
cases['bad_sum'] = es

# Same prefix (same direction shot twice) -> not an A/B pair -> no note.
es = entries(('BCK1BCK60198', 1005.0), ('BCK1BCK60199', 1036.4))
_ab_break_notes(es, 2037.2)
cases['same_prefix'] = es

# Single-direction break (no partner) -> no note.
es = entries(('BCK1BCK60198', 1005.0),)
_ab_break_notes(es, 2037.2)
cases['single'] = es

# File without a trailing port number -> skipped, no crash.
es = entries(('NOPORT', 1005.0), ('BCK6BCK10198', 1036.4))
_ab_break_notes(es, 2037.2)
cases['no_port'] = es

print(json.dumps(cases))
"""


def test_ab_break_consistency_units():
    out = _run_engine_script(_AB_UNIT_SCRIPT)

    a, b = out["west"]
    assert a["break_note"] == ("A+B lengths are consistent with a break "
                               "~1005 m from the BCK1BCK6 end"), a
    assert b["break_note"] == ("A+B lengths are consistent with a break "
                               "~1036 m from the BCK6BCK1 end"), b

    for key in ("bad_sum", "same_prefix", "single", "no_port"):
        for e in out[key]:
            assert "break_note" not in e, (key, e)


# ── Pipeline e2e: _analyze_sor with a stubbed loader ────────────────────────
# Real fixtures are all healthy long fibers, so a broken strand is
# synthesized by stubbing report_sor.load_sor_file inside the engine
# subprocess — everything downstream of loading (robust window, exclusion,
# short_traces, pair metrics, regime, renderers) runs for real.

_STUB_PRELUDE = r"""
import sys, os, json, hashlib
import numpy as np
sys.path.insert(0, sys.argv[1])
import report_sor

def _stub_load(path):
    name = os.path.splitext(os.path.basename(path))[0]
    eof = float(open(path).read().strip())
    n = int(eof)                       # dz = 1 m
    seed = int(hashlib.md5(name.encode()).hexdigest()[:8], 16)  # deterministic
    rng = np.random.RandomState(seed)
    trace = (20.0 - 0.0002 * np.arange(n, dtype=np.float64)
             + rng.randn(n) * 0.05).astype(np.float32)
    return {
        'name': name, 'filepath': path,
        'trace': trace, 'pos': np.arange(n, dtype=np.float64) * 1.0,
        'length': eof, 'loss': None, 'max_splice_dB': None,
        'timestamp': None, 'wavelength': 1550, 'serial_number': None,
        'events': [],
    }

report_sor.load_sor_file = _stub_load

def make_folder(d, spec):
    os.makedirs(d, exist_ok=True)
    for name, eof in spec:
        with open(os.path.join(d, name + '.sor'), 'w') as fh:
            fh.write(str(eof))
    return d

HEALTHY = [('AAABBB%04d' % i, 2190.0 + 2 * i) for i in range(1, 9)]
BROKEN = [('AAABBB0198', 1000.4), ('BBBAAA0198', 1039.8)]
"""

_PIPELINE_SCRIPT = _STUB_PRELUDE + r"""
out = {}

folder = make_folder(os.path.join(sys.argv[2], 'broken'), HEALTHY + BROKEN)
a = report_sor._analyze_sor(folder)
out['min_L'] = a['min_L']
out['names'] = sorted(f['name'] for f in a['files'])
out['n_pairs'] = len(a['pairs'])
out['pair_names'] = sorted({n for p in a['pairs'] for n in (p['a'], p['b'])})
out['short_traces'] = a['short_traces']
out['flags'] = [a['n99'], a['n50']]
out['window_guard'] = a['window_guard']

folder2 = make_folder(os.path.join(sys.argv[2], 'clean'), HEALTHY)
a2 = report_sor._analyze_sor(folder2)
out['clean_min_L'] = a2['min_L']
out['clean_short_traces'] = a2['short_traces']
out['clean_n_pairs'] = len(a2['pairs'])
out['clean_window_guard'] = a2['window_guard']

# LONG-SPAN RESTORATION: the break sits at 3 000 m — ABOVE the 2 km floor
# that used to gate exclusion — on a ~9 km folder.  The removed gate would
# have kept the collapsed 3 000 m window; the strand must now be excluded
# and the window rebuilt from the healthy population.
LONG = [('CCCDDD%04d' % i, 9000.0 + 2 * i) for i in range(1, 9)]
folder3 = make_folder(os.path.join(sys.argv[2], 'longspan'),
                      LONG + [('CCCDDD0198', 3000.0)])
a3 = report_sor._analyze_sor(folder3)
out['long_min_L'] = a3['min_L']
out['long_names'] = sorted(f['name'] for f in a3['files'])
out['long_short_traces'] = a3['short_traces']
out['long_flags'] = [a3['n99'], a3['n50']]
out['long_window_guard'] = a3['window_guard']

# SANITY GUARD: 1 of 4 short (25% > 20%) — an inconsistent folder, not a
# broken strand.  Nothing excluded, raw-min window kept, warning emitted.
GUARD = [('EEEFFF%04d' % i, 8000.0 + 2 * i) for i in range(1, 4)]
folder4 = make_folder(os.path.join(sys.argv[2], 'guard'),
                      GUARD + [('EEEFFF0099', 2500.0)])
a4 = report_sor._analyze_sor(folder4)
out['guard_min_L'] = a4['min_L']
out['guard_names'] = sorted(f['name'] for f in a4['files'])
out['guard_n_pairs'] = len(a4['pairs'])
out['guard_short_traces'] = a4['short_traces']
out['guard_window_guard'] = a4['window_guard']

print(json.dumps(out))
"""


def test_pipeline_excludes_broken_strands_and_reports_them(tmp_path):
    out = _run_engine_script(_PIPELINE_SCRIPT, tmp_path)

    # Window rebuilt from the healthy population (min healthy EOF = 2192).
    assert out["min_L"] == 2192.0, out["min_L"]

    # Broken strands are gone from pair metrics…
    assert len(out["names"]) == 8 and "BBBAAA0198" not in out["names"]
    assert out["n_pairs"] == 28                      # C(8,2)
    assert "AAABBB0198" not in out["pair_names"]
    assert "BBBAAA0198" not in out["pair_names"]

    # …but NEVER silently: both surface as suspected breaks, with the
    # cross-direction A+B consistency note anchored at each launch end.
    # (Median over all 10 EOFs = (2196 + 2198) / 2 = 2197.)
    st = out["short_traces"]
    assert [e["file"] for e in st] == ["AAABBB0198", "BBBAAA0198"]
    for e in st:
        assert e["excluded"] is True
        assert e["median_eof_m"] == 2197.0
        assert "suspected break" in e["note"]
    assert st[0]["note"] == ("ends at 1000 m (folder median 2197 m) "
                             "— suspected break")
    assert st[0]["break_note"] == ("A+B lengths are consistent with a break "
                                   "~1000 m from the AAABBB end")
    assert st[1]["break_note"] == ("A+B lengths are consistent with a break "
                                   "~1040 m from the BBBAAA end")

    # Independent synthetic fibers must not be called duplicates.
    assert out["flags"] == [0, 0]
    assert out["window_guard"] is None

    # Homogeneous folder: raw-min behavior byte-identical, nothing reported.
    assert out["clean_min_L"] == 2192.0
    assert out["clean_short_traces"] == []
    assert out["clean_n_pairs"] == 28
    assert out["clean_window_guard"] is None


def test_pipeline_long_span_window_restoration(tmp_path):
    """A break ABOVE the old 2 km floor no longer caps the folder window:
    the strand is excluded and the window rebuilt (ELMMIL/SANDUR shape)."""
    out = _run_engine_script(_PIPELINE_SCRIPT, tmp_path)

    # Window restored to the healthy min (9002), not the 3 000 m break.
    assert out["long_min_L"] == 9002.0, out["long_min_L"]
    assert len(out["long_names"]) == 8
    assert "CCCDDD0198" not in out["long_names"]

    st = out["long_short_traces"]
    assert [e["file"] for e in st] == ["CCCDDD0198"]
    assert st[0]["excluded"] is True
    assert "suspected break" in st[0]["note"]

    assert out["long_flags"] == [0, 0]
    assert out["long_window_guard"] is None


def test_pipeline_sanity_guard_inconsistent_folder(tmp_path):
    """25% of files short → the folder is inconsistent: exclude NOTHING,
    keep the raw-min window, and warn."""
    out = _run_engine_script(_PIPELINE_SCRIPT, tmp_path)

    # Raw-min window kept; all 4 files still in the pair metrics.
    assert out["guard_min_L"] == 2500.0, out["guard_min_L"]
    assert len(out["guard_names"]) == 4
    assert "EEEFFF0099" in out["guard_names"]
    assert out["guard_n_pairs"] == 6                     # C(4,2)

    # The short file is still REPORTED — but not excluded.
    st = out["guard_short_traces"]
    assert [e["file"] for e in st] == ["EEEFFF0099"]
    assert st[0]["excluded"] is False

    assert out["guard_window_guard"] == _GUARD_MSG_25, out["guard_window_guard"]


# ── Renderer e2e: XLSX sheet appears only when suspected breaks exist ───────

_XLSX_SCRIPT = _STUB_PRELUDE + r"""
out = {}

folder = make_folder(os.path.join(sys.argv[2], 'broken'), HEALTHY + BROKEN)
xb = os.path.join(sys.argv[2], 'broken.xlsx')
report_sor.build_xlsx_sor(folder, 'T', xb)

folder2 = make_folder(os.path.join(sys.argv[2], 'clean'), HEALTHY)
xc = os.path.join(sys.argv[2], 'clean.xlsx')
report_sor.build_xlsx_sor(folder2, 'T', xc)

# Inconsistent folder (guard fires): Summary must carry the warning row,
# the short fiber shows Excluded=No.
GUARD = [('EEEFFF%04d' % i, 8000.0 + 2 * i) for i in range(1, 4)]
folder3 = make_folder(os.path.join(sys.argv[2], 'guard'),
                      GUARD + [('EEEFFF0099', 2500.0)])
xg = os.path.join(sys.argv[2], 'guard.xlsx')
report_sor.build_xlsx_sor(folder3, 'T', xg)

out['broken_xlsx'] = xb
out['clean_xlsx'] = xc
out['guard_xlsx'] = xg

# PDF section renderer (the HTML block the PDF embeds).
out['pdf_section'] = report_sor._short_trace_section_html([
    {'file': 'BCK1BCK60198', 'eof_m': 1005.0, 'median_eof_m': 2037.2,
     'excluded': True,
     'note': 'ends at 1005 m (folder median 2037 m) — suspected break',
     'break_note': ('A+B lengths are consistent with a break ~1005 m '
                    'from the BCK1BCK6 end')},
])
out['pdf_section_empty'] = report_sor._short_trace_section_html([])
out['pdf_section_guard'] = report_sor._short_trace_section_html(
    [{'file': 'EEEFFF0099', 'eof_m': 2500.0, 'median_eof_m': 8003.0,
      'excluded': False,
      'note': 'ends at 2500 m (folder median 8003 m) — suspected break'}],
    window_guard=('folder trace lengths are inconsistent (1 of 4 below '
                  '75% of median) — window not restored; check folder '
                  'contents'))
print(json.dumps(out))
"""


def test_xlsx_and_pdf_sections_only_when_breaks_exist(tmp_path):
    out = _run_engine_script(_XLSX_SCRIPT, tmp_path)

    from openpyxl import load_workbook   # test process: no engine import

    wb = load_workbook(out["broken_xlsx"])
    assert "Suspected short fibers" in wb.sheetnames
    # 'Confirmed duplicates' owns the first slot after Summary (boss request
    # 2026-07-15: duplicates lead the report); the breaks sheet comes next.
    assert wb.sheetnames[1] == "Confirmed duplicates"
    assert wb.sheetnames[2] == "Suspected short fibers"
    ws = wb["Suspected short fibers"]
    rows = list(ws.values)
    assert rows[0] == ("File", "Ends at (m)", "Folder median (m)",
                       "Excluded from pairs", "Finding")
    body = {r[0]: r for r in rows[1:]}
    assert set(body) == {"AAABBB0198", "BBBAAA0198"}
    assert body["AAABBB0198"][3] == "Yes"
    assert "A+B lengths are consistent with a break ~1000 m" in body["AAABBB0198"][4]
    # Summary carries the count row only for affected folders.
    summary = [r for r in wb["Summary"].values if r and r[0]]
    assert ("Suspected short fibers", 2) in [(r[0], r[1]) for r in summary]

    wb2 = load_workbook(out["clean_xlsx"])
    assert "Suspected short fibers" not in wb2.sheetnames
    assert all(r[0] != "Suspected short fibers"
               for r in wb2["Summary"].values if r and r[0])
    assert all(r[0] != "Window warning"
               for r in wb2["Summary"].values if r and r[0])

    # Guard folder: warning row on Summary, short fiber kept (Excluded=No).
    wb3 = load_workbook(out["guard_xlsx"])
    summary3 = {r[0]: r[1] for r in wb3["Summary"].values if r and r[0]}
    assert summary3.get("Window warning") == _GUARD_MSG_25, summary3
    ws3 = wb3["Suspected short fibers"]
    body3 = {r[0]: r for r in list(ws3.values)[1:]}
    assert body3["EEEFFF0099"][3] == "No"

    # PDF HTML block: boss-facing sentences present; empty input renders ''.
    sec = out["pdf_section"]
    assert "Suspected broken / short fibers" in sec
    assert "BCK1BCK60198" in sec
    assert "suspected break — excluded from pair comparison" in sec
    assert "A+B lengths are consistent with a break ~1005 m from the BCK1BCK6 end" in sec
    assert out["pdf_section_empty"] == ""

    # Guard banner renders above the table when the guard fired.
    gsec = out["pdf_section_guard"]
    assert "folder trace lengths are inconsistent (1 of 4" in gsec
    assert "<b>Warning:</b>" in gsec
    assert "EEEFFF0099" in gsec and "excluded from pair comparison" not in gsec


# ── Manifest contract ───────────────────────────────────────────────────────

def test_manifest_short_traces_absent_on_clean_folder(tmp_path):
    """Unaffected folders must not grow the keys at all (byte-stable
    manifests) — in pairs mode and in report mode."""
    d = single_dir_fixture(tmp_path)
    rc, m, stderr = run_secretsauce(d, tmp_path / "out", "pairs")
    assert rc == 0 and m and m.get("ok"), f"runner failed: {(stderr or '')[-800:]}"
    assert "short_traces" not in m, m.keys()
    assert "window_warnings" not in m, m.keys()

    rc, m2, stderr = run_secretsauce(d, tmp_path / "out2", "xlsx")
    assert rc == 0 and m2 and m2.get("ok"), f"runner failed: {(stderr or '')[-800:]}"
    assert "short_traces" not in m2, m2.keys()
    assert "window_warnings" not in m2, m2.keys()


def test_source_locks_window_and_manifest():
    """Pin the calibrated thresholds, the always-exclude rule + sanity
    guard, and the additive manifest wiring (calibrated against A-F West /
    A-F East / SEANOR / SANDUR / ELMMIL / PTL Panel A / TEST DUPE real
    spans on 2026-07-15)."""
    src = (SECRETSAUCE_DIR / "report_sor.py").read_text(encoding="utf-8")

    # Calibrated thresholds.
    assert "_BREAK_FRAC_OF_MEDIAN = 0.75" in src
    assert "_BREAK_AB_SUM_TOL = 0.10" in src
    assert "_INCONSISTENT_FOLDER_FRAC = 0.20" in src

    # LONG-SPAN RESTORATION: the 2 km-collapse precondition is GONE — the
    # only things standing between a suspected break and exclusion are the
    # inconsistent-folder guard and the >=2-healthy-files floor.  (The old
    # gate string must NOT come back; _SHORT_COMMON_SPAN_M keeps its
    # separate regime-routing role, locked in test_ss_regime_fix.)
    assert "raw_min < _SHORT_COMMON_SPAN_M" not in src
    assert "len(outlier_idx) > _INCONSISTENT_FOLDER_FRAC * n" in src
    assert "elif n_healthy >= 2:" in src
    assert "below 75% of median) — window not restored; " in src

    # The robust span feeds the SAME min_L every downstream consumer reads
    # (interior window, regime rules, Common span row).
    assert ("min_L, median_L, out_idx, excl_idx, window_guard = "
            "_robust_common_span(") in src

    # Report surfaces exist.
    assert "Suspected broken / short fibers" in src        # PDF banner
    assert "'Suspected short fibers'" in src               # XLSX sheet
    assert "A+B lengths are consistent with a break" in src
    assert "'Window warning'" in src                       # XLSX Summary row

    # Manifest wiring: additive, only-when-non-empty, in BOTH modes.
    runner = (SECRETSAUCE_DIR / "run_secretsauce.py").read_text(encoding="utf-8")
    assert runner.count("if short_traces_all:") == 2
    assert runner.count("payload['short_traces'] = short_traces_all") == 2
    assert runner.count("if window_warnings_all:") == 2
    assert runner.count("payload['window_warnings'] = window_warnings_all") == 2
