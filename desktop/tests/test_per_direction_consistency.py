"""A direction that disagrees with ITSELF is a finding; a healthy span is silent.

#86 removed `Per-trace detail` from the bidirectional sheet because its four
checks compared every A trace against every B trace — and a bidirectional span
is *defined* by A and B being shot from opposite ends, on different days, often
on a second unit.  "1152 of 2304 differ" was the span, not a defect.

The useful version was never built.  `audit_acquisition` receives `fibers_a`
and `fibers_b` separately and concatenates them immediately, so a direction
that was shot on two different instruments is invisible: its two halves land in
one pile with the other end's traces and nothing compares them.  That IS a
defect — losses from two units are not one calibration and cannot be compared
fibre-to-fibre inside one direction.

WHICH FIELDS, from a census of 43 direction folders / 29,520 traces on disk:

    OTDR unit (model + serial)    fires on  9 of 43 directions
    Wavelength                    tracks the unit — each has its own laser
    Test date (calendar day)      fires on 32 of 43 directions  ← NOT checked

Test date is excluded deliberately.  Sacramento↔Suisun — the FastReporter
ground-truth span — spans FIVE calendar days in its A direction alone, because
a 1152-fibre 64 km span takes a week to shoot.  Checking it per direction would
reproduce #86's noise one direction at a time.  With it excluded, 34 of the 43
directions on disk gain no rows at all.

Real bytes: HOWLAN309 (FTBx-730C-SM3-EA #1723374 @ 1539.8 nm) and HOWLAN631
(FTBx-730D-SM3 #1876268 @ 1554.7 nm) are both A-direction HOW→LAN acquisitions.
They are the real two-unit split, 432/432 across the full 864-fibre direction.
"""

import os
import sys

import openpyxl
import pytest

from conftest import (SPLICEREPORT_DIR, FIXTURE_SPLICE_A_DIR,
                      FIXTURE_SPLICE_B_DIR, run_splicereport)

sys.path.insert(0, str(SPLICEREPORT_DIR))

import acquisition_audit as aa            # noqa: E402
import sor_reader324802a as R             # noqa: E402

_FIX = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fixtures')
BANNER = "Per-direction acquisition consistency"
#  The FastReporter Test Settings panel is appended below this block on the
#  same sheet; its banner marks where the acquisition audit's rows stop.
TEST_SETTINGS_BANNER = "Test Settings"


def _rec(sub, name):
    return R.parse_sor_full(os.path.join(_FIX, sub, name), trim=False)


def _two_unit_direction():
    """One direction, real bytes, shot on two instruments."""
    return {1: _rec('endlaunch', 'HOWLAN309_1550.sor'),
            2: _rec('endlaunch', 'HOWLAN631_1550.sor')}


def _one_unit_direction():
    return {i: _rec('splice_B', 'MILELM%04d_1550.sor' % i) for i in range(1, 5)}


# ── the fixtures really are two units in one direction ──────────────────────

def test_the_fixtures_are_one_direction_on_two_instruments():
    a = _two_unit_direction()
    assert a[1]['otdr_serial'].strip() == '1723374'
    assert a[2]['otdr_serial'].strip() == '1876268'
    assert a[1]['otdr_model'].strip() != a[2]['otdr_model'].strip()
    # Both are HOW->LAN, i.e. the SAME direction of the same span.
    assert a[1]['filename'].startswith('HOWLAN')
    assert a[2]['filename'].startswith('HOWLAN')


# ── behaviour under test (fails on pristine main) ───────────────────────────

def test_a_two_unit_direction_is_reported():
    audit = aa.audit_acquisition(_two_unit_direction(), _one_unit_direction(),
                                 label_a='A-dir HOW', label_b='B-dir LAN')
    pd = audit['per_direction']
    assert len(pd) == 1, f'only the A direction should fire: {pd}'
    assert pd[0]['label'] == 'A-dir HOW'
    names = [r['name'] for r in pd[0]['rows']]
    assert names == ['OTDR unit', 'Wavelength']
    spec = pd[0]['rows'][0]['result']['spec']
    assert 'Mixed within this direction' in spec
    assert '#1723374' in spec and '#1876268' in spec
    # the model rides along with the serial, so one row names the instrument
    assert 'FTBx-730C-SM3-EA' in spec and 'FTBx-730D-SM3' in spec


def test_a_self_consistent_span_produces_nothing():
    """The whole point: silence unless one END disagrees with itself."""
    audit = aa.audit_acquisition(_one_unit_direction(), _one_unit_direction(),
                                 label_a='A', label_b='B')
    assert audit['per_direction'] == []


def test_the_concatenated_view_cannot_see_it():
    """Why a per-direction pass was needed at all.

    Put the two-unit direction opposite a THIRD unit.  The combined
    `file_fields` verdict reports a majority-vs-outliers split that says
    nothing about which end is broken — the finding is invisible in it.
    """
    audit = aa.audit_acquisition(_two_unit_direction(), _one_unit_direction(),
                                 label_a='A', label_b='B')
    combined = {f['name']: f['result'] for f in audit['file_fields']}
    assert not combined['OTDR serial']['all_match']      # it "differs"...
    # ...but the same verdict appears for a perfectly healthy bidirectional
    # span, where A and B are simply two ends on two units.
    healthy = aa.audit_acquisition(
        _one_unit_direction(),
        {1: _rec('splice_A', 'ELMMIL0001_1550.sor')},
        label_a='A', label_b='B')
    hc = {f['name']: f['result'] for f in healthy['file_fields']}
    assert not hc['OTDR serial']['all_match']
    assert healthy['per_direction'] == []               # ...and it stays quiet


def test_a_large_split_reports_the_split_not_a_file_roster():
    """A 432-of-864 split is a crew fact, not a to-do list of filenames."""
    big_a = {}
    for i in range(1, 41):
        r = dict(_rec('splice_B', 'MILELM0001_1550.sor'))
        r['filename'] = 'X%03d.sor' % i
        if i > 25:
            r['otdr_serial'] = '9999999'
        big_a[i] = r
    audit = aa.audit_acquisition(big_a, {}, label_a='A', label_b='B')
    unit = audit['per_direction'][0]['rows'][0]['result']
    assert len(unit['outliers']) == 0, 'roster should be suppressed above the cap'
    assert '× 25' in unit['spec'] and '× 15' in unit['spec'], unit['spec']
    assert 15 > aa.PER_DIRECTION_MAX_OUTLIER_FILES


def test_a_small_split_still_names_the_files():
    small_a = {}
    for i in range(1, 21):
        r = dict(_rec('splice_B', 'MILELM0001_1550.sor'))
        r['filename'] = 'X%03d.sor' % i
        if i > 18:
            r['otdr_serial'] = '9999999'
        small_a[i] = r
    audit = aa.audit_acquisition(small_a, {}, label_a='A', label_b='B')
    unit = audit['per_direction'][0]['rows'][0]['result']
    assert [fn for fn, _ in unit['outliers']] == ['X019.sor', 'X020.sor']


def test_test_date_is_not_a_per_direction_field():
    """Excluded on purpose — see the module docstring's census."""
    a = {}
    for i in range(1, 11):
        r = dict(_rec('splice_B', 'MILELM0001_1550.sor'))
        r['filename'] = 'X%03d.sor' % i
        r['date_time'] = 1_786_000_000 + (86_400 * (i % 3))   # three days
        a[i] = r
    audit = aa.audit_acquisition(a, {}, label_a='A', label_b='B')
    assert audit['per_direction'] == [], \
        'a multi-day shoot is not a per-direction defect'


# ── the rendered sheet ──────────────────────────────────────────────────────

def _sheet_rows(audit, **kw):
    wb = openpyxl.Workbook()
    aa.render_xlsx_sheet(wb, audit, **kw)
    ws = wb['Acquisition Parameters']
    return [[ws.cell(r, c).value for c in (1, 2)]
            for r in range(1, ws.max_row + 1)]


def test_the_block_renders_only_when_asked_and_only_when_it_fires():
    firing = aa.audit_acquisition(_two_unit_direction(), _one_unit_direction(),
                                  label_a='A-dir HOW', label_b='B-dir LAN')
    clean = aa.audit_acquisition(_one_unit_direction(), _one_unit_direction(),
                                 label_a='A', label_b='B')

    on = _sheet_rows(firing, per_trace_detail=False, per_direction_detail=True)
    assert any(r[0] == BANNER for r in on)
    assert any(r[0] == 'A-dir HOW — OTDR unit' for r in on), on

    off = _sheet_rows(firing, per_trace_detail=False)
    assert not any(r[0] == BANNER for r in off), 'default must be OFF'

    quiet = _sheet_rows(clean, per_trace_detail=False, per_direction_detail=True)
    assert not any(r[0] == BANNER for r in quiet)
    # ...and identical to the sheet with the feature switched off entirely.
    assert quiet == _sheet_rows(clean, per_trace_detail=False)


# ── end to end: a clean span gains ZERO rows ────────────────────────────────

def _acq_rows(tmp_path):
    out = tmp_path / 'report.xlsx'
    rc, m, stderr = run_splicereport(FIXTURE_SPLICE_A_DIR, FIXTURE_SPLICE_B_DIR,
                                     out, 'Elm', 'Mil')
    assert rc == 0 and m and m.get('ok'), f'runner failed: {stderr[-1200:]}'
    ws = openpyxl.load_workbook(out)['Acquisition Parameters']
    rows = [[ws.cell(r, c).value for c in (1, 2)]
            for r in range(1, ws.max_row + 1)]
    # Drop the trailing engine stamp ("Report engine") and ONLY the single
    # blank separator directly above it — not every blank row, because the
    # sheet's own grey spacer under the title is one of the 10 rows counted
    # below.  The stamp is a build-identity footer appended beneath the audit,
    # not an audit finding, so the counts here stay a statement about what the
    # AUDIT said — which is what this file pins.
    for i, r in enumerate(rows):
        if r[0] == 'Report engine':
            if i and all(v is None for v in rows[i - 1]):
                i -= 1
            return rows[:i]
    return rows


def test_clean_span_report_gains_no_rows(tmp_path):
    """splice_A / splice_B are each one unit, one wavelength — nothing to say.

    This is the invariance that separates the block from what #86 deleted, and
    it holds on pristine main as well (where the block does not exist).
    """
    rows = _acq_rows(tmp_path)
    assert not any(r[0] == BANNER for r in rows), rows
    assert not any(isinstance(r[0], str) and 'OTDR unit' in r[0] for r in rows)
    # Count the ACQUISITION-AUDIT portion only.  The sheet also carries the
    # FastReporter Test Settings panel, appended below this block behind its
    # own banner; that panel prints unconditionally (comparing A against B is
    # its whole job) and its rows say nothing about per-direction consistency.
    # Slicing at its banner keeps this guard measuring what it is named for —
    # a clean span adds no per-direction rows — instead of counting every row
    # on the sheet, which is what made it break the moment anything else was
    # appended.
    audit_rows = rows
    for i, r in enumerate(rows):
        if r[0] == TEST_SETTINGS_BANNER:
            audit_rows = rows[:i]
            break
    while audit_rows and not any(c is not None for c in audit_rows[-1]):
        audit_rows = audit_rows[:-1]           # blank spacer between sections
    assert len(audit_rows) == 10, audit_rows   # header + two wavelength blocks


# ── the unidirectional report must not move ─────────────────────────────────

def test_uni_call_site_does_not_turn_the_block_on():
    src = open(os.path.join(SPLICEREPORT_DIR, 'splicereportmatchexfo.py'),
               encoding='utf-8').read()   # explicit: Windows CI defaults to cp1252
    uni = src[src.index('audit = audit_acquisition(fibers, {})'):]
    uni = uni[:uni.index('except Exception')]
    assert 'per_trace_detail=True' in uni
    assert 'per_direction_detail' not in uni, \
        'uni already reports these facts per-direction; do not double it'


def test_render_default_leaves_uni_output_unchanged():
    """Uni renders with per_trace_detail=True and nothing else.

    Called the way uni calls it — positionally, no direction labels — so this
    pin holds on pristine main too.
    """
    audit = aa.audit_acquisition(_two_unit_direction(), {})
    uni_now = _sheet_rows(audit, per_trace_detail=True)
    assert any(r[0] == 'Per-trace detail' for r in uni_now)
    assert any(r[0] == 'OTDR serial' for r in uni_now)
    assert not any(r[0] == BANNER for r in uni_now)


@pytest.mark.parametrize('kw', [{}, {'per_trace_detail': True}])
def test_uni_sheet_is_byte_for_byte_what_it_was(kw):
    """No per-direction row can leak into the uni layout."""
    audit = aa.audit_acquisition(_two_unit_direction(), {})
    rows = _sheet_rows(audit, **kw)
    assert not any(isinstance(r[0], str) and r[0].startswith('A-dir ')
                   for r in rows)
    assert not any(isinstance(r[0], str) and 'Mixed within' in str(r[1])
                   for r in rows)
