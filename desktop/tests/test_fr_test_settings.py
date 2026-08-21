"""FastReporter's "Test Settings" panel, printed once per direction.

WHY THE TABLE EXISTS.  A bidirectional loss is the average of two shots.  That
average is only meaningful if both shots were taken with the same refractive
index, the same backscatter coefficient and the same detection thresholds — and
nothing in the report used to say whether they were.  The panel the boss reads
in FastReporter every day answers exactly that, so the report now prints it
twice, side by side, one table per direction.

WHERE THE NUMBERS COME FROM.  Seven of the eight rows are stored verbatim in
EXFO's proprietary block and are read back byte-for-byte:

    IOR                               Ior                          1.470000
    Backscatter                       Rbs                          -83.00 dB
    Helix factor                      HelixFactor                  0.00 %
    Splice loss detection threshold   SpliceLossThreshold          0.020 dB
    [ ] Splitter loss (SM Only)       SplitterDetection = 0        (unchecked)
                                      SplitterDetectionThreshold   2.500 dB
    Reflectance detection threshold   ReflectanceThreshold         -78.00 dB
    End-of-fiber detection threshold  EndOfFiberThreshold          5.000 dB

The eighth — Fiber core size — is NOT in the file.  A dump of the proprietary
block's 1,096 named fields contains no core-size or mode-field-diameter field,
and no numeric field anywhere holds 9 (or 9e-6).  The row prints the stored
glass designation instead of a micron figure it cannot prove.  See
test_core_size_is_not_invented.

THE MODE IS A TRAP, AND THE TRAP IS REAL.  Each direction is hundreds of traces
and the panel has one cell per row, so the cell shows the modal value.  On
PLACHE LS (1152 traces) the reflectance threshold is split 586 at -78.00 dB
against 566 at -72.00 dB, while its partner CHEPLA LS is a uniform -78.00 dB.
Compare modes alone and the two directions "match" — and the report would
quietly assert comparability for the 566 fibers where it does not hold.  So an
internally-mixed direction is a finding in its own right: see
test_internally_mixed_direction_is_flagged_even_though_modes_agree, which is
the regression guard for exactly that span.
"""

import os
import sys

import openpyxl
import pytest

from conftest import (SPLICEREPORT_DIR, FIXTURE_SPLICE_A_DIR,
                      FIXTURE_SPLICE_B_DIR, run_splicereport)

sys.path.insert(0, str(SPLICEREPORT_DIR))

import acquisition_audit as aa            # noqa: E402
import sor_reader324802a as R             # noqa: E402

_FIX = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fixtures')

GREEN = "E2EFDA"
AMBER = "FFF2CC"

#  FastReporter's panel, verbatim, in FR's own row order.  The report is judged
#  against this list: same labels, same order, same units, same digit counts.
FR_PANEL = [
    ("IOR",                              "1.470000"),
    ("Backscatter",                      "-83.00 dB"),
    ("Helix factor",                     "0.00 %"),
    ("Splice loss detection threshold",  "0.020 dB"),
    ("[ ] Splitter loss (SM Only)",      "2.500 dB"),
    ("Reflectance detection threshold",  "-78.00 dB"),
    ("End-of-fiber detection threshold", "5.000 dB"),
    ("Fiber core size",                  None),          # not stored — see below
]


# ─────────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────────
def _rec(sub, name):
    return R.parse_sor_full(os.path.join(_FIX, sub, name), trim=False)


def _direction(sub, names, rec_fields=None, **overrides):
    """Real fixture bytes, with optional per-field seeding.

    `**overrides` seed the Test Settings dict; `rec_fields` seed plain record
    keys (otdr_model / otdr_serial / exfo_wavelength_nm).  Both mutate the
    parsed record, never the extractor, so a seeded run exercises exactly the
    same code path a real mis-set unit would.
    """
    out = {}
    for i, nm in enumerate(names, start=1):
        r = _rec(sub, nm)
        if overrides:
            r['test_settings'] = dict(r['test_settings'])
            r['test_settings'].update(overrides)
        if rec_fields:
            r.update(rec_fields)
        out[i] = r
    return out


A_NAMES = ['ELMMIL0001_1550.sor', 'ELMMIL0002_1550.sor', 'ELMMIL0003_1550.sor']
B_NAMES = ['MILELM0001_1550.sor', 'MILELM0002_1550.sor', 'MILELM0003_1550.sor']


def _render(fibers_a, fibers_b, label_a='A-dir ELM', label_b='B-dir MIL'):
    audit = aa.audit_acquisition(fibers_a, fibers_b,
                                 label_a=label_a, label_b=label_b)
    wb = openpyxl.Workbook()
    aa.render_xlsx_sheet(wb, audit, per_trace_detail=False,
                         per_direction_detail=True, test_settings_block=True)
    return wb["Acquisition Parameters"], audit


def _fill(cell):
    f = cell.fill
    if not f or f.fill_type is None:
        return None
    rgb = f.start_color.rgb
    return rgb[-6:] if isinstance(rgb, str) else None


def _grid(ws):
    """iter_rows only — reading ws.cell(r, c) would CREATE cells and inflate
    max_row, which is exactly the bug this sheet's diff must not hide."""
    return {(c.row, c.column): c
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                    min_col=1, max_col=ws.max_column)
            for c in row}


def _content(ws):
    """Only cells that actually carry something.

    The additive check must compare CONTENT, not openpyxl's bounding box:
    widening the used range to column D materialises blank cells in every
    earlier row, and counting those as "changes" would drown the real diff.
    """
    return {k: c for k, c in _grid(ws).items()
            if c.value is not None or _fill(c) is not None}


def _block(ws):
    """The Test Settings block: {row_label: {a_label,a_value,a_fill,b_*}}.

    Located by the header row that carries "Test Settings" in BOTH panel label
    columns, so it cannot be confused with the grey banner above it.
    """
    g = _grid(ws)
    hdr = None
    for (r, c), cell in sorted(g.items()):
        if c == 1 and cell.value == "Test Settings" \
                and g.get((r, 3)) is not None \
                and g[(r, 3)].value == "Test Settings":
            hdr = r
            break
    assert hdr is not None, "Test Settings header row not found"
    rows = []
    for i in range(len(FR_PANEL)):
        r = hdr + 1 + i
        rows.append({
            "a_label": g[(r, 1)].value, "a_value": g[(r, 2)].value,
            "a_fill":  _fill(g[(r, 1)]), "a_vfill": _fill(g[(r, 2)]),
            "b_label": g[(r, 3)].value, "b_value": g[(r, 4)].value,
            "b_fill":  _fill(g[(r, 3)]), "b_vfill": _fill(g[(r, 4)]),
        })
    notes = [g[(r, 2)].value for (r, c) in sorted(g) if c == 2 and r > hdr + len(FR_PANEL)]
    return hdr, rows, [n for n in notes if n]


INFO_LABELS = ["OTDR model", "OTDR serial", "Wavelength"]


def _info_block(ws):
    """The instrument rows: [{label, value, lfill, vfill}, ...] for A then B."""
    g = _grid(ws)
    head = None
    for (r, c), cell in sorted(g.items()):
        if c == 1 and cell.value == "Instrument":
            head = r
            break
    assert head is not None, "Instrument heading not found"
    out = {"a": [], "b": []}
    for i in range(len(INFO_LABELS)):
        r = head + 1 + i
        out["a"].append({"label": g[(r, 1)].value, "value": g[(r, 2)].value,
                         "lfill": _fill(g[(r, 1)]), "vfill": _fill(g[(r, 2)])})
        out["b"].append({"label": g[(r, 3)].value, "value": g[(r, 4)].value,
                         "lfill": _fill(g[(r, 3)]), "vfill": _fill(g[(r, 4)])})
    return head, out


# ═════════════════════════════════════════════════════════════════════════
#  1. The parser reads the panel out of real bytes
# ═════════════════════════════════════════════════════════════════════════
def test_parser_reads_every_stored_setting_from_real_bytes():
    """All seven stored rows come back with the screenshot's exact values."""
    ts = _rec('splice_A', A_NAMES[0])['test_settings']
    assert ts == {
        'Ior': 1.47, 'Rbs': -83.0, 'HelixFactor': 0.0,
        'SpliceLossThreshold': 0.02, 'SplitterDetection': 0,
        'SplitterDetectionThreshold': 2.5, 'ReflectanceThreshold': -78.0,
        'EndOfFiberThreshold': 5.0, 'FiberCode': 0,
    }, ts


def test_ior_prefers_proprietary_precision_over_bellcore_quantisation():
    """FxdParams stores IOR quantised to 5 dp; the proprietary block keeps 6.

    FastReporter prints 6, so the proprietary value wins and the Bellcore
    group index is the fallback.  Both must parse.
    """
    path = os.path.join(_FIX, 'splice_A', A_NAMES[0])
    rec = R.parse_sor_full(path, trim=False)
    with open(path, 'rb') as f:
        data = f.read()
    blocks = R._parse_block_directory(data)
    fxd = R._parse_fxd_params(data, blocks)
    assert fxd['group_index'] == 1.47, fxd['group_index']       # Bellcore
    assert rec['ior'] == rec['test_settings']['Ior'] == 1.47    # proprietary


def test_short_field_names_do_not_match_the_tail_of_a_longer_field():
    """`Rbs` is a substring of `PeakReflectionToRbs`, which appears FIRST.

    A bare find() for "Rbs\\0" lands on that longer field's tail, where the
    bytes 12 back are not a descriptor — so the read returns None and the
    Backscatter row silently renders blank on files that plainly contain the
    value.  This is the bug the boundary-anchored reader fixes; it is a
    verification table, so a blank caused by a parser slip is indistinguishable
    from a genuinely absent setting and must not happen.
    """
    path = os.path.join(_FIX, 'splice_A', A_NAMES[0])
    with open(path, 'rb') as f:
        data = f.read()
    blocks = R._parse_block_directory(data)
    stream = R._decompress_proprietary(data, blocks)
    decoy = stream.find(b'PeakReflectionToRbs\x00')
    assert decoy >= 0, "fixture no longer contains the decoy field"
    # A bare find() lands on the decoy's TAIL, not on the real field.
    assert stream.find(b'Rbs\x00') == decoy + len(b'PeakReflectionTo')
    # …which is why the old unanchored helper comes back empty here.
    assert R._prop_f64(stream, 'Rbs') is None
    # The anchored reader finds the real field anyway.
    assert R._prop_scalar(stream, 'Rbs', 3, 8) == -83.0
    assert R._parse_test_settings(stream)['Rbs'] == -83.0


def test_parser_reports_absence_rather_than_a_default():
    """A file with no readable proprietary block yields {} — never stand-in
    numbers.  A fabricated value in a verification table is worse than a gap."""
    assert R._parse_test_settings(b'') == {}
    assert R._parse_test_settings(None) == {}
    assert R._parse_test_settings(b'no descriptors in here at all') == {}


# ═════════════════════════════════════════════════════════════════════════
#  2. The table renders FastReporter's panel, twice
# ═════════════════════════════════════════════════════════════════════════
def test_panel_renders_twice_with_fr_labels_order_and_formatting():
    ws, _ = _render(_direction('splice_A', A_NAMES),
                    _direction('splice_B', B_NAMES))
    _hdr, rows, _notes = _block(ws)
    assert len(rows) == len(FR_PANEL)
    for got, (label, value) in zip(rows, FR_PANEL):
        # Same labels, same order, in BOTH panels — "this exact table twice".
        assert got["a_label"] == label, (got["a_label"], label)
        assert got["b_label"] == label, (got["b_label"], label)
        if value is not None:
            # Same units and digit counts as FR: IOR 6 dp, dB 2 dp except the
            # splice-loss / splitter / end-of-fiber thresholds at 3, percent 2.
            assert got["a_value"] == value, (label, got["a_value"], value)
            assert got["b_value"] == value, (label, got["b_value"], value)


def test_both_panels_stay_row_aligned():
    """Side-by-side comparison only works if row N means the same parameter on
    both sides.  Nothing — not a mismatch, not an internally-mixed direction —
    may insert a row into one panel and not the other."""
    ws, _ = _render(
        _direction('splice_A', A_NAMES, ReflectanceThreshold=-72.0),
        _direction('splice_B', B_NAMES))
    _hdr, rows, _notes = _block(ws)
    for got, (label, _v) in zip(rows, FR_PANEL):
        assert got["a_label"] == got["b_label"] == label


def test_headline_confirms_a_matching_span():
    ws, audit = _render(_direction('splice_A', A_NAMES),
                        _direction('splice_B', B_NAMES))
    ts = audit["test_settings"]
    assert ts["clean"] is True
    assert ts["n_differ"] == 0
    assert ts["headline"].startswith("✓")
    _hdr, rows, _notes = _block(ws)
    for got in rows:
        assert got["a_fill"] == GREEN, got
        assert got["b_fill"] == GREEN, got


# ═════════════════════════════════════════════════════════════════════════
#  3. The values are read from the traces, not baked into the renderer
# ═════════════════════════════════════════════════════════════════════════
@pytest.mark.parametrize("field,seeded,row_label,expected", [
    ('Ior',                        1.468325, "IOR",                              "1.468325"),
    ('Rbs',                          -81.87, "Backscatter",                      "-81.87 dB"),
    ('HelixFactor',                     0.5, "Helix factor",                     "0.50 %"),
    ('SpliceLossThreshold',            0.01, "Splice loss detection threshold",  "0.010 dB"),
    ('SplitterDetectionThreshold',      1.5, "[ ] Splitter loss (SM Only)",      "1.500 dB"),
    ('ReflectanceThreshold',          -72.0, "Reflectance detection threshold",  "-72.00 dB"),
    ('EndOfFiberThreshold',             3.0, "End-of-fiber detection threshold", "3.000 dB"),
])
def test_each_cell_tracks_its_own_trace_field(field, seeded, row_label, expected):
    """Change the field in the trace, and only that cell changes.

    This is what stops the table from being decoration: if any row were a
    constant, or wired to the wrong field, its cell would not move when the
    underlying value does.  Every seeded value is one that genuinely occurs on
    disk (IOR 1.468325 / Rbs -81.87 on WSC_SUIsh; 0.010 dB and -72.00 dB on the
    tie-panel and PLACHE sets).
    """
    ws, _ = _render(_direction('splice_A', A_NAMES, **{field: seeded}),
                    _direction('splice_B', B_NAMES))
    _hdr, rows, _notes = _block(ws)
    by_label = {r["a_label"].replace("[x]", "[ ]"): r for r in rows}
    assert by_label[row_label]["a_value"] == expected
    # The untouched direction keeps the original value — so the cell really is
    # per-direction, not one number copied into both panels.
    assert by_label[row_label]["b_value"] != expected


def test_splitter_checkbox_reflects_the_stored_flag():
    """FR draws the checkbox from SplitterDetection; unchecked at 0, ticked
    at 1.  The state lives in the LABEL, where a difference between the two
    panels lands in a column the eye is already scanning."""
    ws, _ = _render(_direction('splice_A', A_NAMES, SplitterDetection=1),
                    _direction('splice_B', B_NAMES))
    _hdr, rows, _notes = _block(ws)
    row = rows[4]
    assert row["a_label"] == "[x] Splitter loss (SM Only)"
    assert row["b_label"] == "[ ] Splitter loss (SM Only)"


def test_core_size_is_not_invented():
    """The micron figure is not in the file, so the cell must not claim one.

    FastReporter shows "9 µm".  Nothing in the .sor supports that number: the
    proprietary field dump has no core-size or mode-field-diameter entry, and
    FiberCode is 0 on all 4,392 occurrences across a 1,106-trace census, so its
    mapping to a core size cannot be calibrated against any second value.  The
    row therefore reports the stored glass designation and says the rest is
    unavailable.  If someone later hard-codes 9, this test goes red.
    """
    ws, _ = _render(_direction('splice_A', A_NAMES),
                    _direction('splice_B', B_NAMES))
    _hdr, rows, notes = _block(ws)
    core = rows[7]
    assert core["a_label"] == "Fiber core size"
    assert "9" not in str(core["a_value"]).replace("G.652", "")
    assert "µm" not in str(core["a_value"])
    assert str(core["a_value"]).startswith("n/a")
    assert "G.652" in str(core["a_value"])          # the designation IS stored
    assert any("not stored" in str(n) for n in notes), notes


# ═════════════════════════════════════════════════════════════════════════
#  4. A mismatch is visible — the whole point of the table
# ═════════════════════════════════════════════════════════════════════════
def test_seeded_mismatch_is_flagged_on_both_panels():
    """A differs from B on the reflectance threshold.

    Everything a tech would use to notice must fire: the headline counts the
    difference, BOTH panels' cells go amber (you can be looking at either one),
    and a note names the parameter and quotes both values.

    MUTATION-CHECKED: deleting the `differs` term from `flag_a`/`flag_b` in
    acquisition_audit.compute_test_settings turns the fills green again and
    this test fails on the first fill assertion.
    """
    ws, audit = _render(
        _direction('splice_A', A_NAMES, ReflectanceThreshold=-72.0),
        _direction('splice_B', B_NAMES))
    ts = audit["test_settings"]
    assert ts["n_differ"] == 1, ts["headline"]
    assert ts["clean"] is False
    assert "DIFFER" in ts["headline"]

    _hdr, rows, notes = _block(ws)
    hit = rows[5]
    assert hit["a_label"] == "Reflectance detection threshold"
    assert hit["a_value"] == "-72.00 dB"
    assert hit["b_value"] == "-78.00 dB"
    # Both panels highlighted, label and value alike.
    assert hit["a_fill"] == AMBER, hit
    assert hit["b_fill"] == AMBER, hit
    assert hit["a_vfill"] == AMBER, hit
    assert hit["b_vfill"] == AMBER, hit
    # Every OTHER row stays green, so the amber means something.
    for i, got in enumerate(rows):
        if i == 5:
            continue
        assert got["a_fill"] == GREEN, (i, got)
        assert got["b_fill"] == GREEN, (i, got)
    assert any("Reflectance detection threshold" in str(n)
               and "-72.00 dB" in str(n) and "-78.00 dB" in str(n)
               for n in notes), notes


def test_every_row_can_carry_a_mismatch():
    """No row is comparison-dead: seeding a difference in any of the seven
    stored parameters raises the differ count."""
    for field in ('Ior', 'Rbs', 'HelixFactor', 'SpliceLossThreshold',
                  'SplitterDetection', 'SplitterDetectionThreshold',
                  'ReflectanceThreshold', 'EndOfFiberThreshold'):
        seeded = 1 if field == 'SplitterDetection' else 9.5
        _ws, audit = _render(
            _direction('splice_A', A_NAMES, **{field: seeded}),
            _direction('splice_B', B_NAMES))
        assert audit["test_settings"]["n_differ"] == 1, field


def test_internally_mixed_direction_is_flagged_even_though_modes_agree():
    """The PLACHE case, in miniature — and the one the mode alone gets wrong.

    A's own traces are split between two reflectance thresholds; B is uniform.
    A's MODE equals B's value, so a naive mode-vs-mode comparison reports a
    clean match.  It is not a clean match: half of A was shot with a different
    detection threshold and did not record the same events.

    MUTATION-CHECKED: dropping the `or ra["mixed"]` term from `flag_a` in
    compute_test_settings turns the cell green and this test fails.
    """
    mixed = _direction('splice_A', A_NAMES[:1], ReflectanceThreshold=-78.0)
    mixed[2] = _direction('splice_A', A_NAMES[1:2],
                          ReflectanceThreshold=-72.0)[1]
    ws, audit = _render(mixed, _direction('splice_B', B_NAMES))
    ts = audit["test_settings"]

    # The trap: the two directions' modal values are equal.
    assert ts["n_differ"] == 0
    # …and the report still refuses to call it a match.
    assert ts["clean"] is False
    assert "itself" in ts["headline"].lower() or "ITSELF" in ts["headline"]

    _hdr, rows, notes = _block(ws)
    hit = rows[5]
    assert hit["a_fill"] == AMBER, hit
    assert "mixed" in str(hit["a_value"]), hit["a_value"]
    assert hit["b_fill"] == GREEN, hit          # B really is uniform
    # The split is spelled out, with counts, so the tech knows how many fibers.
    assert any("-78.00 dB × 1" in str(n) and "-72.00 dB × 1" in str(n)
               for n in notes), notes


# ═════════════════════════════════════════════════════════════════════════
#  5. Additive: nothing that already existed moves
# ═════════════════════════════════════════════════════════════════════════
def test_one_sided_run_does_not_claim_a_clean_match():
    """With only one direction loaded there is nothing to compare against.

    The headline must say so.  It must NOT say "not stored" (the settings are
    stored — there is just no counterpart), and it must not show the green
    tick, which would read as "I checked A against B and they agree".
    """
    A = _direction('splice_A', A_NAMES)
    audit = aa.audit_acquisition(A, {}, label_a='A-dir ELM', label_b='B-dir MIL')
    ts = audit["test_settings"]
    assert ts["clean"] is False
    assert "not stored" not in ts["headline"]
    assert "one direction" in ts["headline"]


def test_block_is_appended_below_the_existing_audit():
    """Every pre-existing cell keeps its exact (row, column) and value.

    The block is appended, never inserted, so `first_data_row` — and the
    freeze_panes derived from it — are decided before the panel is written.
    """
    A, B = _direction('splice_A', A_NAMES), _direction('splice_B', B_NAMES)
    audit = aa.audit_acquisition(A, B, label_a='A-dir ELM', label_b='B-dir MIL')

    wb_off = openpyxl.Workbook()
    aa.render_xlsx_sheet(wb_off, audit, per_trace_detail=False,
                         per_direction_detail=True, test_settings_block=False)
    wb_on = openpyxl.Workbook()
    aa.render_xlsx_sheet(wb_on, audit, per_trace_detail=False,
                         per_direction_detail=True, test_settings_block=True)
    off, on = wb_off["Acquisition Parameters"], wb_on["Acquisition Parameters"]

    g_off, g_on = _content(off), _content(on)
    for key, cell in g_off.items():
        assert key in g_on, f"cell {key} disappeared"
        assert g_on[key].value == cell.value, f"cell {key} changed"
        assert _fill(g_on[key]) == _fill(cell), f"fill {key} changed"
    assert on.freeze_panes == off.freeze_panes
    # New cells only, and only below what was already there.
    added = set(g_on) - set(g_off)
    assert added
    assert min(r for r, _ in added) > max(r for r, _ in g_off)
    # Columns A and B keep the widths the audit table set for itself.
    for col in ('A', 'B'):
        assert on.column_dimensions[col].width == \
               off.column_dimensions[col].width


def test_unidirectional_report_does_not_get_the_block():
    """Uni's input IS one direction; a lone panel compares nothing.  Default
    off, and the uni call site must not pass it."""
    A = _direction('splice_A', A_NAMES)
    audit = aa.audit_acquisition(A, {}, label_a='A', label_b='B')
    wb = openpyxl.Workbook()
    aa.render_xlsx_sheet(wb, audit, per_trace_detail=True)   # no kwarg at all
    ws = wb["Acquisition Parameters"]
    g = _grid(ws)
    assert not any(c.value == "Test Settings" for c in g.values())

    # encoding= is mandatory: the Windows CI runner defaults to cp1252 and
    # this file is UTF-8 (see test_packaging_contract).
    src = (SPLICEREPORT_DIR / "splicereportmatchexfo.py").read_text(
        encoding="utf-8")
    uni = src[src.index('audit = audit_acquisition(fibers, {})'):]
    uni = uni[:uni.index('render_xlsx_sheet(wb, audit') + 200]
    assert 'test_settings_block' not in uni, \
        "uni report must not enable the two-direction panel"


# ═════════════════════════════════════════════════════════════════════════
#  6. End to end, through the real runner
# ═════════════════════════════════════════════════════════════════════════
def test_block_reaches_the_first_sheet_of_a_real_report(tmp_path):
    out = tmp_path / "splice.xlsx"
    run_splicereport(FIXTURE_SPLICE_A_DIR, FIXTURE_SPLICE_B_DIR, out,
                     site_a="ELM", site_b="MIL")
    wb = openpyxl.load_workbook(out)
    ws = wb[wb.sheetnames[0]]
    assert ws.title == "Acquisition Parameters"
    assert wb.active.title == ws.title
    _hdr, rows, _notes = _block(ws)
    for got, (label, value) in zip(rows, FR_PANEL):
        assert got["a_label"] == got["b_label"] == label
        if value is not None:
            assert got["a_value"] == value
            assert got["b_value"] == value


# ═════════════════════════════════════════════════════════════════════════
#  7. Instrument rows — printed beside the panel, never adjudicated
#
#  These are NOT FastReporter's rows.  FR's panel is what the boss asked to
#  see reproduced, so the eight rows stay intact and the banner keeps counting
#  eight; these sit below under their own heading and are read, not judged.
#
#  They are not compared because a bidirectional span is DEFINED by A and B
#  being shot from opposite ends — different day, often a second unit at its
#  own laser wavelength.  Comparing them would flag every healthy span.  What
#  IS worth reading is one direction shot on two instruments, so that case is
#  shown split rather than collapsed to a mode.
# ═════════════════════════════════════════════════════════════════════════
#  A real second instrument off the PLACHE set, chosen to differ from BOTH
#  fixtures (splice_A is 730D #1882155, splice_B is 730C #1723374) so a
#  "these differ" assertion cannot pass by coincidence.
_TWO_UNIT = {'otdr_model': 'FTBx-735D-SM3-EA', 'otdr_serial': '1982501',
             'exfo_wavelength_nm': 1559.4}


def test_instrument_rows_render_per_direction_from_the_traces():
    ws, _ = _render(_direction('splice_A', A_NAMES),
                    _direction('splice_B', B_NAMES))
    _head, info = _info_block(ws)
    for side in ('a', 'b'):
        assert [r["label"] for r in info[side]] == INFO_LABELS
    # Real values off the fixture bytes, not placeholders.
    rec = _rec('splice_A', A_NAMES[0])
    assert info['a'][0]["value"] == rec['otdr_model']
    assert info['a'][1]["value"] == rec['otdr_serial']
    assert info['a'][0]["value"] and info['a'][1]["value"]


def test_instrument_rows_carry_no_fill_even_when_the_directions_differ():
    """The absence of a fill IS the signal that nothing here was adjudicated.

    A and B on two different units at two different wavelengths — the normal
    shape of a bidirectional span — must produce no green and no amber.

    MUTATION-CHECKED: giving these cells the panel's fill (green or amber)
    fails this test.
    """
    ws, _ = _render(_direction('splice_A', A_NAMES, rec_fields=_TWO_UNIT),
                    _direction('splice_B', B_NAMES))
    _head, info = _info_block(ws)
    for side in ('a', 'b'):
        for r in info[side]:
            assert r["lfill"] is None, r
            assert r["vfill"] is None, r
    # …and they really are different, so this is not a vacuous pass.
    assert info['a'][0]["value"] != info['b'][0]["value"]
    assert info['a'][1]["value"] != info['b'][1]["value"]


def test_instrument_rows_never_reach_the_banner_or_the_mismatch_count():
    """The banner must keep saying 8, not 11.

    Model, serial and wavelength all differ between the directions here.  None
    of it may raise n_differ, lower n_comparable, add a mixed note, or clear
    the green tick — the panel's eight rows agree, and that is what the banner
    reports.

    MUTATION-CHECKED: appending the instrument rows to `verdicts` (so they are
    counted) fails this test on the banner text.
    """
    _ws, audit = _render(
        _direction('splice_A', A_NAMES, rec_fields=_TWO_UNIT),
        _direction('splice_B', B_NAMES))
    ts = audit["test_settings"]
    assert ts["n_comparable"] == 8, ts["n_comparable"]
    assert ts["n_differ"] == 0, ts["n_differ"]
    assert ts["mixed_notes"] == []
    assert ts["clean"] is True
    assert ts["headline"] == "✓ All 8 comparable parameter(s) match"
    assert len(ts["verdicts"]) == len(FR_PANEL) == 8


def test_a_direction_on_two_units_is_shown_split_not_collapsed():
    """The one case these rows exist for.

    KANLAN's A direction is genuinely #1723374 × 373 against #1876268 × 491.
    A bare mode would print one serial and hide that the direction is two
    calibrations — which is the only thing these rows are good for.
    """
    two = _direction('splice_A', A_NAMES[:1], rec_fields=_TWO_UNIT)
    two[2] = _direction('splice_A', A_NAMES[1:2])[1]        # the other unit
    ws, audit = _render(two, _direction('splice_B', B_NAMES))
    _head, info = _info_block(ws)
    serial = info['a'][1]["value"]
    assert '1982501' in serial and '1882155' in serial, serial
    assert '×' in serial and ';' in serial, serial
    # Same "value × count" presentation the mixed-threshold rows use — one
    # style for one fact, not two.
    assert '1982501 × 1' in serial and '1882155 × 1' in serial, serial
    # Still not flagged, and still not counted.
    assert info['a'][1]["vfill"] is None
    assert audit["test_settings"]["n_comparable"] == 8
    assert audit["test_settings"]["headline"].startswith("✓")


def test_instrument_block_is_visually_separate_from_the_fr_panel():
    """FR's panel must stay recognisably intact: the instrument rows sit below
    it, under their own heading, and never interleave with the eight."""
    ws, _ = _render(_direction('splice_A', A_NAMES),
                    _direction('splice_B', B_NAMES))
    panel_hdr, rows, _notes = _block(ws)
    info_head, _info = _info_block(ws)
    assert info_head > panel_hdr + len(FR_PANEL), "instrument block overlaps the panel"
    assert [r["a_label"] for r in rows] == [lbl for lbl, _v in FR_PANEL]
    g = _grid(ws)
    assert 'not compared' in str(g[(info_head, 2)].value)
