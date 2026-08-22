"""The two ILA columns are the two PHYSICAL CABLE ENDS, and each header names
the end its contents are actually at.

The columns used to be keyed on DIRECTION.  But the rules that populate them
fire at OPPOSITE ends of the span:

    launch  rule → events[0]          → the end the shot was launched FROM
    tailbox rule → last 1F before EOL → the OTHER end, up to a span away

Both landed in the measuring direction's column, so a WSC reading printed under
a header naming SUI — 64 km away.  Measured on origin/main across WSC<->SUI and
SANDUR: 9 of 19 ILA reflectance findings sat under a header naming the wrong
end.  Real cases: WSC<->SUI `AC44` held `486 …-48.8dB 491 …-46.5dB` in ONE cell
(486 at WSC, 491 at SUI); SANDUR F76 printed the SAME DUR connector in both
columns, so one fault read as two.

The ILA columns carry fiber + tag and NO distance column — `ILA:A`'s header
rows hold the `B->A:` / `A->B:` axis labels — so the header is the only thing
placing these findings for the tech.

Since #98 BOTH reflectance rules print the same bare `REFL-30.0dB`, so a
finding's rule CANNOT be read off its tag text — matching 'TAILBOX' here would
match nothing and quietly pass.  These tests ask the engine which rule fired
via the internal `refl_rules` dict, which is kept positionally parallel to each
END's tag list, and `_refl()` asserts that pairing on every call.

Engine tests run in a clean subprocess (3-engine sor_reader isolation).
"""
import subprocess
import sys
import textwrap

from conftest import REPO_ROOT

SPLICEREPORT_DIR = REPO_ROOT / "splicereport"

# Pre-dedented so it composes with a separately-dedented body.
_HELPERS = textwrap.dedent("""
    SPAN = 60.0            # cable end in each direction's own frame
    CLEAN = -60.0          # a healthy buried connector

    def rec(port_refl=CLEAN, tail_refl=CLEAN, span=SPAN, dur=60):
        '''One direction's record.  `port_refl` drives the LAUNCH rule (the
        end this direction was shot from); `tail_refl` drives the TAILBOX rule
        (the last 1F before EOL — the OTHER end of the cable).'''
        return {'events': [
            {'dist_km': 0.0, 'splice_loss': 0.0, 'is_end': False,
             'is_reflective': True, 'reflection': port_refl,
             'time_of_travel': 0, 'type': '1F9999LS'},
            {'dist_km': 1.0, 'splice_loss': 0.05, 'is_end': False,
             'is_reflective': True, 'reflection': CLEAN,
             'time_of_travel': 5000, 'type': '1F9999LS'},
            {'dist_km': 20.0, 'splice_loss': 0.05, 'is_end': False,
             'is_reflective': False, 'reflection': None, 'type': '0F9999LS'},
            {'dist_km': span - 1.0, 'splice_loss': 0.05, 'is_end': False,
             'is_reflective': True, 'reflection': tail_refl,
             'type': '1F9999LS'},
            {'dist_km': span, 'splice_loss': 0.0, 'is_end': True,
             'is_reflective': True, 'reflection': -14.0, 'type': '1E9999LS'},
        ], 'duration_sec': dur}

    def span_of(bad_fiber, **bad):
        '''12 fibers (one ribbon), all clean except `bad_fiber`, which takes
        the keyword overrides.  The clean majority is what makes the bad one an
        OUTLIER against the direction's population median.'''
        out = {}
        for f in range(1, 13):
            out[f] = rec(**bad) if f == bad_fiber else rec()
        return out

    def tags(issues, fnum):
        i = issues.get(fnum) or {}
        return list(i.get('a_tags') or []), list(i.get('b_tags') or [])

    def refl(issues, fnum, end, rule=None):
        '''The REFL tags at PHYSICAL end `end` ('A'/'B'), optionally only the
        ones the given RULE ('launch'/'tailbox') produced.

        Both rules print the same bare 'REFL' prefix, so the rule is not in
        the text.  refl_rules records it, positionally parallel to that end's
        tag list -- the length assert below is that contract, and it also
        catches rules keyed by DIRECTION while tags are keyed by END.'''
        i = issues.get(fnum) or {}
        tags  = list(i.get(end.lower() + '_tags') or [])
        rules = list(((i.get('refl_rules') or {}).get(end.upper())) or [])
        got = [t for t in tags if t.startswith('REFL')]
        assert len(got) == len(rules), (end, got, rules)
        if rule is None:
            return got
        return [t for t, rl in zip(got, rules) if rl == rule]
""")


def _run(body):
    header = ("import sys\n"
              f"sys.path.insert(0, {str(SPLICEREPORT_DIR)!r})\n"
              "import splicereportmatchexfo as E\n")
    p = subprocess.run([sys.executable, "-c", header + _HELPERS + textwrap.dedent(body)],
                       capture_output=True, text=True)
    assert p.returncode == 0, f"exit {p.returncode}\n{p.stdout}\n{p.stderr}"
    out = p.stdout.strip().splitlines()
    assert out and out[-1] == "OK", p.stdout


# ── 1. The far-end rule files under the far end ────────────────────────────

def test_a_direction_tailbox_finding_files_under_end_b():
    """A is shot FROM end A, so A's tailbox reading is at end B.  It must land
    in b_tags.  This is the WSC<->SUI F486 / F254 / F12 shape."""
    _run("""
        A = span_of(7, tail_refl=-30.0)      # A sees a bad connector at END B
        B = span_of(0)                       # B clean throughout
        issues = E.detect_launch_issues(A, B)
        a, b = tags(issues, 7)
        assert refl(issues, 7, 'B', 'tailbox') == ['REFL-30.0dB'], (a, b)
        assert refl(issues, 7, 'A') == [], f"far-end finding filed at end A: {a}"
        print('OK')
    """)


def test_b_direction_tailbox_finding_files_under_end_a():
    """Mirror: B is shot FROM end B, so B's tailbox reading is at end A."""
    _run("""
        A = span_of(0)
        B = span_of(7, tail_refl=-30.0)      # B sees a bad connector at END A
        issues = E.detect_launch_issues(A, B)
        a, b = tags(issues, 7)
        assert refl(issues, 7, 'A', 'tailbox') == ['REFL-30.0dB'], (a, b)
        assert refl(issues, 7, 'B') == [], f"far-end finding filed at end B: {b}"
        print('OK')
    """)


# ── 2. The near-end rule must NOT move (guards over-correction) ────────────

def test_launch_finding_stays_at_the_end_it_was_shot_from():
    """The launch rule already named the right end.  A's launch reading is at
    end A and must stay in a_tags; B's stays in b_tags."""
    _run("""
        A = span_of(7, port_refl=-30.0)
        B = span_of(9, port_refl=-31.0)
        issues = E.detect_launch_issues(A, B)
        a7, b7 = tags(issues, 7)
        a9, b9 = tags(issues, 9)
        assert refl(issues, 7, 'A', 'launch') == ['REFL-30.0dB'], (a7, b7)
        assert refl(issues, 7, 'B') == [], b7
        assert refl(issues, 9, 'B', 'launch') == ['REFL-31.0dB'], (a9, b9)
        assert refl(issues, 9, 'A') == [], a9
        print('OK')
    """)


# ── 3. One connector, seen from both ends, reads as ONE column ─────────────

def test_one_connector_seen_from_both_ends_lands_in_one_column():
    """SANDUR F76: the DUR connector is B's LAUNCH (-31.1) and A's TAILBOX
    (-32.8).  One physical fault, two views.  Both belong at end B — printed in
    two different columns it read as two separate faults at two sites."""
    _run("""
        A = span_of(7, tail_refl=-32.8)      # A's far-end view of end B
        B = span_of(7, port_refl=-31.1)      # B's own launch at end B
        issues = E.detect_launch_issues(A, B)
        a, b = tags(issues, 7)
        assert not a, f"nothing belongs at end A here: {a}"
        # Both views of the ONE connector, in end B's column, own near-end
        # view first and then the other direction's far view.
        assert refl(issues, 7, 'B') == ['REFL-31.1dB', 'REFL-32.8dB'], b
        assert (issues[7]['refl_rules']['B'] == ['launch', 'tailbox']), \\
            issues[7]['refl_rules']
        assert refl(issues, 7, 'A') == [], a
        print('OK')
    """)


# ── 4. The header over the receiving column names that end ────────────────

def test_header_names_the_end_the_finding_is_actually_at():
    """End to end, through the real workbook writer: a finding physically at
    end B must sit under the header naming SITE B.

    The column is located BY ITS HEADER TEXT, so this fails if the finding is
    routed to the other column AND if the two headers are swapped — the header
    and the contents are checked against each other, not against a literal.
    """
    _run("""
        import openpyxl, tempfile, os
        A = span_of(7, tail_refl=-30.0)      # bad connector at END B, seen from A
        B = span_of(0)
        issues = E.detect_launch_issues(A, B)
        cells, lca, lcb = E.build_ribbon_data({}, 12, 12, 1, launch_issues=issues)
        splices = [{'position_km': 20.0, 'position_km_refined': 20.0,
                    'position_km_display': 20.0, 'splice_display_num': 1}]
        out = os.path.join(tempfile.mkdtemp(), 'r.xlsx')
        E.write_xlsx(cells, splices, 12, 12, out, 'AAASITE', 'BBBSITE', SPAN,
                     launch_cells_a=lca, launch_cells_b=lcb)

        ws = openpyxl.load_workbook(out)['Splice Report']
        rows = list(ws.iter_rows())          # never ws.cell() — it inflates max_row
        hdr = rows[2]
        ila = {c.column: str(c.value) for c in hdr
               if c.value and 'ILA' in str(c.value)}
        assert len(ila) == 2, f"expected exactly two ILA headers: {ila}"
        col_a = [c for c, v in ila.items() if 'AAASITE' in v]
        col_b = [c for c, v in ila.items() if 'BBBSITE' in v]
        assert len(col_a) == 1 and len(col_b) == 1, ila
        col_a, col_b = col_a[0], col_b[0]

        found = {}
        for r in rows[3:]:
            for c in r:
                if c.value and 'REFL-30.0dB' in str(c.value):
                    found[c.column] = str(c.value)
        # Guards this test against the #98 tag rename going unnoticed: if the
        # printed text ever stops matching, this fires instead of the column
        # assertion below passing on an empty set.
        assert found, 'the finding vanished from the workbook'
        assert list(found) == [col_b], (
            'end-B finding is not under the header naming site B: '
            'found in %r, site-B column is %d, headers %r' % (found, col_b, ila))
        # …and the header does not call the column a direction.
        assert ila[col_b].startswith('B-end ILA:'), ila[col_b]
        assert ila[col_a].startswith('A-end ILA:'), ila[col_a]
        print('OK')
    """)


# ── 5. refl_rules stays paired with the list it describes ─────────────────

def test_refl_rules_stay_paired_with_each_ends_tag_list():
    """`refl_rules` is keyed by END and positionally parallel to that end's
    REFL tags.

    Since #98 both reflectance rules print the same bare `REFL` prefix, so
    every consumer that tells them apart zips the two lists by index (see
    test_tailbox_receive_reel.tags_for, which asserts the lengths match).
    Moving tags to the end they are AT while leaving the rules keyed by the
    measuring DIRECTION would mismatch that pairing silently — a launch
    finding would answer for a tailbox one.

    Fires BOTH rules in BOTH directions at once, with a non-REFL tag
    interleaved between them on end A so the filter is exercised too.
    """
    _run("""
        A = span_of(7, port_refl=-30.0, tail_refl=-32.0, dur=19)
        B = span_of(7, port_refl=-31.0, tail_refl=-33.0)
        issues = E.detect_launch_issues(A, B)
        a, b = tags(issues, 7)
        rr = issues[7]['refl_rules']

        # End A: A's own launch, A's duration tag, then B's far tailbox.
        assert a == ['REFL-30.0dB', 'DURATION_MISMATCH(19.0s vs 60.0s)',
                     'REFL-33.0dB'], a
        assert rr['A'] == ['launch', 'tailbox'], rr
        # End B: B's own launch, then A's far tailbox.
        assert b == ['REFL-31.0dB', 'REFL-32.0dB'], b
        assert rr['B'] == ['launch', 'tailbox'], rr

        # The contract every rule-vs-tag consumer relies on.
        for end, lst in (('A', a), ('B', b)):
            got = [t for t in lst if t.startswith('REFL')]
            assert len(got) == len(rr[end]), (end, got, rr[end])
        # And the pairing resolves to the right physical readings.
        assert refl(issues, 7, 'A', 'launch')  == ['REFL-30.0dB']
        assert refl(issues, 7, 'A', 'tailbox') == ['REFL-33.0dB']
        assert refl(issues, 7, 'B', 'launch')  == ['REFL-31.0dB']
        assert refl(issues, 7, 'B', 'tailbox') == ['REFL-32.0dB']
        print('OK')
    """)
