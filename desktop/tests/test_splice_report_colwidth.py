"""Regression: Splice Report columns are MINIMUM-fit — wide enough to fit their
content (including merged headers / merged loss cells that span the km+ft pair)
and no wider.  Guards against reverting to fixed-wide or pair-equalized columns.
"""
import subprocess
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
FX = REPO_ROOT / "desktop" / "tests" / "fixtures"


def test_splice_report_columns_minimum_fit(tmp_path):
    out = tmp_path / "report.xlsx"
    proc = subprocess.run(
        [sys.executable, str(REPO_ROOT / "splicereport" / "run_splicereport.py"),
         "--dir-a", str(FX / "splice_A"), "--dir-b", str(FX / "splice_B"),
         "--out", str(out), "--site-a", "HOWESPAN", "--site-b", "LANCASTER"],
        capture_output=True, text=True)
    assert proc.returncode == 0, proc.stderr[-600:]
    ws = openpyxl.load_workbook(out)["Splice Report"]

    # Widest content RELEVANT to each column = its own cells + any merged value
    # that spans it (the splice headers + merged loss cells span the km+ft pair).
    relevant = {c: 0 for c in range(1, ws.max_column + 1)}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v:
                relevant[c] = max(relevant[c],
                                  max(len(x) for x in str(v).splitlines()))
    for mr in ws.merged_cells.ranges:
        v = ws.cell(row=mr.min_row, column=mr.min_col).value
        if v:
            L = max(len(x) for x in str(v).splitlines())
            for c in range(mr.min_col, mr.max_col + 1):
                relevant[c] = max(relevant[c], L)

    for c in range(1, ws.max_column + 1):
        w = ws.column_dimensions[get_column_letter(c)].width
        assert w is not None, f"col {c} has no explicit width"
        assert 3.0 <= w <= 60.5, f"col {c} width {w} out of bounds"
        # MINIMUM-fit: never gratuitously wider than the widest content touching it.
        assert w <= relevant[c] * 1.5 + 4.0, (
            f"col {c} width {w} is wider than its content ({relevant[c]} chars)")

    # CONTENT NOT CLIPPED: each merged header / value fits across its span.
    for mr in ws.merged_cells.ranges:
        v = ws.cell(row=mr.min_row, column=mr.min_col).value
        if not v:
            continue
        L = max(len(x) for x in str(v).splitlines())
        span_w = sum(ws.column_dimensions[get_column_letter(c)].width
                     for c in range(mr.min_col, mr.max_col + 1))
        assert span_w + 1.0 >= L * 1.05, (
            f"merged '{str(v)[:24]}' ({L} chars) doesn't fit its {mr.coord} span ({span_w:.1f})")
