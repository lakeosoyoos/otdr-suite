"""Regression: the Splice Report must show WHICH ILA is the A-direction and
which is the B-direction (boss's request), using the real site names read from
the SOR GenParams rather than a literal "A"/"B".

app.py is importable in bare mode (Streamlit logs a harmless ScriptRunContext
warning), so we exercise the engine-free ILA helpers directly, plus assert the
engine writes the direction-labelled ILA headers into the Splice Report sheet.
"""
import subprocess
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))
import app  # noqa: E402  (bare-mode import; emits a harmless context warning)

FX = REPO_ROOT / "desktop" / "tests" / "fixtures"


def test_sor_locations_reads_genparams():
    f = sorted((FX / "splice_A").glob("*.sor"))[0]
    loc_a, loc_b = app._sor_locations(str(f))
    assert loc_a and loc_b, "should read both GenParams endpoints"
    assert loc_a == "ELMDALE"


def test_derive_ila_uses_filename_prefix_for_origin():
    # splice_A files are ELMMIL* (ELM origin); splice_B files are MILELM* (MIL
    # origin) — even though BOTH carry GenParams location_a=ELMDALE, the
    # filename prefix decides which end each direction was shot FROM.
    a_origin, _ = app._derive_ila(str(FX / "splice_A"))
    b_origin, _ = app._derive_ila(str(FX / "splice_B"))
    assert a_origin == "ELMDALE"
    assert b_origin.startswith("MIL")        # MILLER end, via the MILELM prefix
    assert a_origin != b_origin              # A and B directions are distinct ILAs


def test_report_labels_a_and_b_direction_ila(tmp_path):
    out = tmp_path / "report.xlsx"
    proc = subprocess.run(
        [sys.executable, str(REPO_ROOT / "splicereport" / "run_splicereport.py"),
         "--dir-a", str(FX / "splice_A"), "--dir-b", str(FX / "splice_B"),
         "--out", str(out), "--site-a", "ELMDALE", "--site-b", "MILLER"],
        capture_output=True, text=True)
    assert proc.returncode == 0, proc.stderr[-600:]
    ws = openpyxl.load_workbook(out)["Splice Report"]
    labels = [ws.cell(row=3, column=c).value
              for c in range(1, ws.max_column + 1)
              if ws.cell(row=3, column=c).value
              and "ILA" in str(ws.cell(row=3, column=c).value)]
    assert any(v == "A-dir ILA: ELMDALE" for v in labels), labels
    assert any(v == "B-dir ILA: MILLER" for v in labels), labels
