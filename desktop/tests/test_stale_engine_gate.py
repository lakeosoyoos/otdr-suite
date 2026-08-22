"""Two techs on different engines must not both be able to hand in a report.

The field complaint was "we have to restart the app even after we update".
Three things behind it, and this file locks the fixes for all three:

  * The update is applied by the LAUNCHER, at boot, into ~/.otdrSuite/engine.
    A process that is already running keeps the code it imported, so a restart
    is architecturally required — nothing can make new code appear in an old
    process.
  * launcher.main() returns EARLY when an instance is already serving
    ("Another instance is already serving — opening new tab"), which happens
    BEFORE _prepare_engine().  A tech who leaves the app open and
    double-clicks the icon therefore gets a new browser tab on the OLD engine
    and no update fetch at all.  Proven in test_launcher_early_return below.
  * The staleness check used to run once per session, so a machine that stays
    open for days never learned that a new engine had been published.

So: report generation is BLOCKED while the engine is behind (manifest 238 and
239 disagree about KANLAN F676 — .168 vs .167 — and nothing in either workbook
said which engine measured it), the check is TTL'd so a long session notices,
and every workbook now carries the engine that produced it.

The block FAILS OPEN by design.  Offline, a timed-out manifest, a garbled
version, a dev checkout — none of them block.  A tech in a truck with no
signal has to be able to work; blocking on a FAILED CHECK would be an outage
of our own making, so each fail-open case is tested explicitly.
"""
import ast
import importlib.util
import json
import os
import sys
import threading
import types
import urllib.request
from http.server import BaseHTTPRequestHandler, HTTPServer

from conftest import (REPO_ROOT, FIXTURE_SPLICE_A_DIR, FIXTURE_SPLICE_B_DIR,
                      run_streamlit)
import error_report as R

APP = REPO_ROOT / "app.py"
APP_SRC = APP.read_text(encoding="utf-8")


def _load_helper(name, **namespace):
    """Exec one top-level function out of app.py in a bare module (the
    test_update_nudge.py pattern) — no Streamlit, no network."""
    tree = ast.parse(APP_SRC)
    fn = next(n for n in tree.body if isinstance(n, ast.FunctionDef)
              and n.name == name)
    mod = types.ModuleType("stale_gate")
    mod.__dict__.update(namespace)
    exec(compile(ast.Module(body=[fn], type_ignores=[]), "app.py", "exec"),
         mod.__dict__)
    return getattr(mod, name)


def _fn_source(name):
    tree = ast.parse(APP_SRC)
    fn = next(n for n in tree.body if isinstance(n, ast.FunctionDef)
              and n.name == name)
    return ast.get_source_segment(APP_SRC, fn)


def _fn_code(name):
    """Executable code of one app.py function with the docstring and every
    comment stripped, so a source-lock greps what the function DOES rather
    than what its prose happens to mention.  (_latest_manifest_version's
    docstring says the launcher does the 'verify' — scanning raw source for
    that word flags the very comment that documents the boundary.)"""
    import copy
    tree = ast.parse(APP_SRC)
    fn = copy.deepcopy(next(n for n in tree.body
                            if isinstance(n, ast.FunctionDef)
                            and n.name == name))
    if (fn.body and isinstance(fn.body[0], ast.Expr)
            and isinstance(fn.body[0].value, ast.Constant)
            and isinstance(fn.body[0].value.value, str)):
        fn.body.pop(0)
    return ast.unparse(fn)          # unparse drops comments for us


def _stale_check(**ns):
    """_stale_check closes over _nudge_check, so inject the real one."""
    nudge = _load_helper("_nudge_check")
    return _load_helper("_stale_check", _nudge_check=nudge, **ns)


# ═════════════════════════════════════════════════════════════════════════
#  1. THE DIAGNOSIS — why a tech ends up on a stale engine
# ═════════════════════════════════════════════════════════════════════════
class _Health(BaseHTTPRequestHandler):
    def do_GET(self):
        if self.path == "/_stcore/health":
            self.send_response(200)
            self.end_headers()
            self.wfile.write(b"ok")
        else:
            self.send_response(404)
            self.end_headers()

    def log_message(self, *a):
        pass


def test_launcher_early_return_skips_the_signed_update_entirely(tmp_path,
                                                                monkeypatch):
    """THE root cause of "I updated but I'm still on the old version".

    launcher.main() health-checks port 8510 first and, if anything answers,
    opens a tab and returns — so _prepare_engine (the signed fetch/verify/swap)
    is never reached.  Double-clicking the icon while the app is open is
    therefore NOT a restart and NOT an update, which is exactly what left the
    fleet split across manifest 238 and 239.

    This test pins the behaviour so the block above it stays justified; if the
    launcher ever grows a while-running update path, this is the test that
    should be revisited."""
    spec = importlib.util.spec_from_file_location(
        "launcher_probe", str(REPO_ROOT / "desktop" / "launcher.py"))
    L = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(L)

    monkeypatch.setenv("HOME", str(tmp_path))
    monkeypatch.setenv("USERPROFILE", str(tmp_path))

    httpd = HTTPServer(("127.0.0.1", L.PORT), _Health)
    threading.Thread(target=httpd.serve_forever, daemon=True).start()
    calls = []
    monkeypatch.setattr(L, "_prepare_engine",
                        lambda: calls.append("prepare") or (L.bundled_dir(), "x"))
    monkeypatch.setattr(L.webbrowser, "open", lambda u: calls.append("browser"))
    real_stdout, real_stderr = sys.stdout, sys.stderr
    try:
        rc = L.main()
    finally:
        sys.stdout, sys.stderr = real_stdout, real_stderr
        httpd.shutdown()
        httpd.server_close()

    assert rc == 0
    assert "browser" in calls, "the already-serving path should still open a tab"
    assert "prepare" not in calls, (
        "launcher reached the signed-update path while an instance was serving "
        "— if this now passes, the early-return diagnosis has changed")


# ═════════════════════════════════════════════════════════════════════════
#  2. _stale_check — the TTL cache, with store/clock/fetch injected
# ═════════════════════════════════════════════════════════════════════════
def test_stale_check_reports_a_newer_published_engine():
    check = _stale_check()
    store = {}
    assert check(store, 1000.0, 300, lambda: 239, lambda: 238) == (239, 238)


def test_stale_check_is_silent_when_current_or_ahead():
    check = _stale_check()
    assert check({}, 1000.0, 300, lambda: 239, lambda: 239) is None
    assert check({}, 1000.0, 300, lambda: 238, lambda: 239) is None


def test_stale_check_does_not_refetch_inside_the_window():
    """A rerun-heavy page must not re-hit GitHub on every click."""
    check = _stale_check()
    hits = []
    store = {}

    def fetch():
        hits.append(1)
        return 239

    assert check(store, 1000.0, 300, fetch, lambda: 238) == (239, 238)
    assert check(store, 1100.0, 300, fetch, lambda: 238) == (239, 238)
    assert check(store, 1299.0, 300, fetch, lambda: 238) == (239, 238)
    assert hits == [1], f"re-fetched inside the TTL window: {len(hits)} fetches"


def test_stale_check_refetches_after_the_window():
    """THE 'live, not startup-only' guarantee: an always-on machine holds one
    session for days, so the answer has to be allowed to go stale and be
    re-asked — otherwise a publish at 09:00 is never noticed."""
    check = _stale_check()
    hits = []

    def fetch():
        hits.append(1)
        return 239 if len(hits) > 1 else 238

    store = {}
    assert check(store, 1000.0, 300, fetch, lambda: 238) is None
    assert check(store, 1299.0, 300, fetch, lambda: 238) is None, "inside window"
    assert check(store, 1301.0, 300, fetch, lambda: 238) == (239, 238), (
        "a session that outlives the TTL must notice a new publish")
    assert hits == [1, 1], f"expected exactly 2 fetches, got {len(hits)}"


def test_stale_check_caches_the_negative_answer_too():
    """Offline must not mean 'retry on every single rerun'."""
    check = _stale_check()
    hits = []

    def fetch():
        hits.append(1)
        raise OSError("network is unreachable")

    store = {}
    assert check(store, 1000.0, 300, fetch, lambda: 238) is None
    assert check(store, 1200.0, 300, fetch, lambda: 238) is None
    assert hits == [1], "an offline result must be cached like any other"


# ── fail-open ────────────────────────────────────────────────────────────
def test_stale_check_fails_open_when_the_fetch_raises():
    """Offline / DNS dead / TLS blocked → NOT stale → the tech may run."""
    check = _stale_check()

    def boom():
        raise OSError("no route to host")

    assert check({}, 1000.0, 300, boom, lambda: 238) is None


def test_stale_check_fails_open_when_the_fetch_returns_none():
    """A garbled or unparseable manifest degrades to None upstream."""
    check = _stale_check()
    assert check({}, 1000.0, 300, lambda: None, lambda: 238) is None


def test_stale_check_fails_open_when_the_version_reader_raises():
    """A broken version.json must not block, and must not even reach the
    network — an unknown running version can't be compared to anything."""
    check = _stale_check()
    hits = []

    def fetch():
        hits.append(1)
        return 999

    def boom():
        raise ValueError("garbled version.json")

    assert check({}, 1000.0, 300, fetch, boom) is None
    assert hits == [], "must short-circuit before the fetch"


def test_stale_check_fails_open_on_a_dev_checkout():
    """applied is None (a dev checkout can't be updated) → never blocked, and
    no network call — this is what keeps the whole suite off GitHub."""
    check = _stale_check()
    hits = []

    def fetch():
        hits.append(1)
        return 999

    assert check({}, 1000.0, 300, fetch, lambda: None) is None
    assert hits == []


# ═════════════════════════════════════════════════════════════════════════
#  3. Wiring — one version compare, one restart, every report control gated
# ═════════════════════════════════════════════════════════════════════════
def test_every_report_control_is_gated():
    """All three report-producing buttons must consult the gate AND honour it.
    A gate that renders a message but leaves the button live is the exact
    failure being fixed (a tech in a hurry clicks anyway)."""
    for label in ("Run analysis",                    # Secret Sauce
                  "Generate Splice Report",          # Splice Report (+ FR)
                  "Run unidirectional report"):      # Unidirectional
        i = APP_SRC.index(f"st.button('{label}'")
        window = APP_SRC[i - 400:i + 220]
        assert "_report_gate(" in window, f"{label!r} never calls the gate"
        assert "disabled=bool(_stale)" in window, (
            f"{label!r} renders a message but stays clickable")


def test_gate_reuses_the_one_version_compare_and_the_one_restart():
    """No second copy of either — that is how the paths drift apart."""
    assert APP_SRC.count("\ndef _nudge_check(") == 1
    assert APP_SRC.count("\ndef _stale_check(") == 1
    assert APP_SRC.count("\ndef _relaunch_and_exit(") == 1
    assert APP_SRC.count("\ndef _restart_command(") == 1
    gate = _fn_code("_report_gate")
    assert "_update_state()" in gate
    assert "_relaunch_and_exit()" in gate, "the block must offer a way forward"
    for dup in ("Popen", "os._exit", "Start-Process", "/bin/sh", "urlopen"):
        assert dup not in gate, f"{dup} must not be copied into the gate"


def test_gate_never_applies_an_update_itself():
    """The trust boundary: the signed fetch/verify/swap stays in the frozen
    launcher.  app.py may READ a version to decide whether to block; it must
    never fetch, verify or apply code."""
    for name in ("_report_gate", "_stale_check", "_update_state",
                 "_latest_manifest_version"):
        code = _fn_code(name)
        for forbidden in ("Ed25519", "verify", "sha256", "hashlib",
                          "rename", "write_bytes", "shutil"):
            assert forbidden not in code, f"{name} must not {forbidden}"


def test_gate_fails_open_if_the_whole_check_explodes():
    """Belt and braces: even an unexpected exception inside _update_state
    must return 'not stale', never block."""
    gate = _fn_code("_report_gate")
    assert "except Exception:" in gate
    body = gate[gate.index("try:"):]
    assert body.index("return None") < body.index("st.error("), (
        "the failure path must return before anything is blocked/rendered")


def test_engine_files_list_is_unchanged():
    """app.py and splicereport/acquisition_audit.py already ship in
    ENGINE_FILES, so this change auto-updates with NO manifest file-set change
    (the launcher rejects a manifest whose set differs)."""
    launcher = (REPO_ROOT / "desktop" / "launcher.py").read_text(encoding="utf-8")
    block = launcher.split("ENGINE_FILES = [", 1)[1].split("]", 1)[0]
    assert '"app.py",' in block
    assert '"splicereport/acquisition_audit.py",' in block
    assert len([l for l in block.splitlines() if l.strip().startswith('"')]) == 21


# ═════════════════════════════════════════════════════════════════════════
#  4. AppTest — what the tech actually sees
# ═════════════════════════════════════════════════════════════════════════
def _fake_manifest(version):
    body = json.dumps({"version": version, "commit": "abc1234"}).encode()

    class _Resp:
        def read(self):
            return body

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    return lambda req, timeout=None, **kw: _Resp()


def _arm(monkeypatch, tmp_path, applied_label, urlopen):
    monkeypatch.setenv("HOME", str(tmp_path))
    monkeypatch.setenv("USERPROFILE", str(tmp_path))
    drive, tail = os.path.splitdrive(str(tmp_path))
    monkeypatch.setenv("HOMEDRIVE", drive)
    monkeypatch.setenv("HOMEPATH", tail)
    monkeypatch.delenv("SS_ERROR_WEBHOOK", raising=False)
    monkeypatch.setattr(R, "version_labels", lambda *a, **k: applied_label)
    monkeypatch.setattr(urllib.request, "urlopen", urlopen)


APPLIED_238 = ("build 238 (2026-08-20)", "update 238 applied 2026-08-20 09:00 PDT")


def _page(name, **state):
    """Boot the hub, seed the page's folder slots, then switch to it — the
    run()-first order the rest of the suite uses (the nav radio does not exist
    as a widget until the first render)."""
    at = run_streamlit().run()
    for k, v in state.items():
        at.session_state[k] = v
    at.sidebar.radio[0].set_value(name).run()
    return at


def _splice_page():
    return _page("Splice Report",
                 view_dir_a_input=str(FIXTURE_SPLICE_A_DIR),
                 view_dir_b_input=str(FIXTURE_SPLICE_B_DIR))


def _find_button(at, label):
    for b in at.button:
        if b.label == label:
            return b
    raise AssertionError(f"{label!r} not rendered; saw {[b.label for b in at.button]}")


def test_stale_engine_disables_generate_splice_report(monkeypatch, tmp_path):
    """238 applied, 239 published → the button is there but NOT clickable."""
    _arm(monkeypatch, tmp_path, APPLIED_238, _fake_manifest(239))
    at = _splice_page()
    assert not at.exception, f"page raised: {list(at.exception)}"
    assert _find_button(at, "Generate Splice Report").disabled is True


def test_stale_engine_disables_secret_sauce(monkeypatch, tmp_path):
    _arm(monkeypatch, tmp_path, APPLIED_238, _fake_manifest(239))
    at = _page("Secret Sauce", ss_folder_input=str(FIXTURE_SPLICE_A_DIR))
    assert not at.exception, f"page raised: {list(at.exception)}"
    assert _find_button(at, "Run analysis").disabled is True


def test_stale_engine_disables_unidirectional(monkeypatch, tmp_path):
    _arm(monkeypatch, tmp_path, APPLIED_238, _fake_manifest(239))
    at = _page("Unidirectional", uni_folder_input=str(FIXTURE_SPLICE_A_DIR))
    assert not at.exception, f"page raised: {list(at.exception)}"
    assert _find_button(at, "Run unidirectional report").disabled is True


def test_the_tech_is_told_why_and_given_a_way_forward(monkeypatch, tmp_path):
    """A block with no explanation and no button is a support call."""
    _arm(monkeypatch, tmp_path, APPLIED_238, _fake_manifest(239))
    at = _splice_page()
    assert not at.exception, f"page raised: {list(at.exception)}"
    text = " ".join(e.value for e in at.error)
    assert "Report generation is paused" in text, text
    assert "engine 238" in text and "engine 239" in text, text
    assert "Nothing is lost" in text, text
    assert any("Restart the app to apply" in c.value for c in at.caption), (
        "a dev/unfrozen run must still say how to clear the block")


# ── fail-open, end to end ────────────────────────────────────────────────
def test_offline_does_not_block_the_report(monkeypatch, tmp_path):
    """THE truck-with-no-signal case: the update server is unreachable, so the
    tech must still be able to generate."""
    def boom(*a, **k):
        raise OSError("no route to host")

    _arm(monkeypatch, tmp_path, APPLIED_238, boom)
    at = _splice_page()
    assert not at.exception, f"page raised: {list(at.exception)}"
    assert _find_button(at, "Generate Splice Report").disabled is False
    assert not at.error, [e.value for e in at.error]


def test_up_to_date_does_not_block_the_report(monkeypatch, tmp_path):
    _arm(monkeypatch, tmp_path,
         ("build 239 (2026-08-21)", "update 239 applied 2026-08-21 09:00 PDT"),
         _fake_manifest(239))
    at = _splice_page()
    assert not at.exception, f"page raised: {list(at.exception)}"
    assert _find_button(at, "Generate Splice Report").disabled is False


def test_dev_checkout_does_not_block_and_never_calls_out(monkeypatch, tmp_path):
    """An unknown running version is 'could not tell', not 'stale'."""
    hits = []

    def spy(*a, **k):
        hits.append(1)
        raise OSError("blocked")

    monkeypatch.delenv("OTDR_SUITE_SOURCE", raising=False)
    _arm(monkeypatch, tmp_path, ("dev", "dev"), spy)
    at = _splice_page()
    assert not at.exception, f"page raised: {list(at.exception)}"
    assert _find_button(at, "Generate Splice Report").disabled is False
    assert hits == [], "a dev run must not hit the update server"


def test_a_garbled_manifest_does_not_block(monkeypatch, tmp_path):
    """Not JSON / no version key → unknown → fail open."""
    class _Resp:
        def read(self):
            return b"<!DOCTYPE html>captive portal"

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    _arm(monkeypatch, tmp_path, APPLIED_238, lambda *a, **k: _Resp())
    at = _splice_page()
    assert not at.exception, f"page raised: {list(at.exception)}"
    assert _find_button(at, "Generate Splice Report").disabled is False


# ═════════════════════════════════════════════════════════════════════════
#  5. The report stamp
# ═════════════════════════════════════════════════════════════════════════
def _audit_stub():
    return {"n_files": 2, "earliest_iso": "2026-01-01", "latest_iso": "2026-01-02",
            "file_fields": [], "per_wavelength": [], "per_direction": []}


def _render(**env):
    """Render the Acquisition Parameters sheet in isolation and return it."""
    import openpyxl
    sys.path.insert(0, str(REPO_ROOT / "splicereport"))
    import acquisition_audit as A
    wb = openpyxl.Workbook()
    wb.active.title = "Splice Report"
    A.render_xlsx_sheet(wb, _audit_stub())
    return wb, A


def _last_row(ws):
    """Last row carrying a value — via iter_rows(), never ws.cell() (which
    would CREATE cells and inflate max_row)."""
    last = None
    for rowtuple in ws.iter_rows():
        for c in rowtuple:
            if c.value not in (None, ""):
                last = c.row
    return last


def test_stamp_is_the_last_row_of_the_acquisition_sheet():
    wb, _ = _render()
    ws = wb["Acquisition Parameters"]
    row = _last_row(ws)
    vals = {c.column: c.value for rt in ws.iter_rows(min_row=row, max_row=row)
            for c in rt}
    assert vals[1] == "Report engine", vals
    assert str(vals[2]).startswith("OTDR Suite"), vals


def test_stamp_does_not_disturb_the_sheet_it_appends_to():
    """Additive only: sheet 0, still active, freeze pane still on the header."""
    wb, _ = _render()
    assert wb.sheetnames[0] == "Acquisition Parameters"
    assert wb.active.title == "Acquisition Parameters"
    ws = wb["Acquisition Parameters"]
    assert ws.freeze_panes == "A4", ws.freeze_panes
    assert ws.column_dimensions["A"].width == 38
    assert ws.column_dimensions["B"].width == 80
    hdr = {c.column: c.value for rt in ws.iter_rows(min_row=3, max_row=3)
           for c in rt}
    assert hdr[1] == "Parameter" and hdr[2] == "Result", hdr


def test_stamp_names_an_applied_update(monkeypatch, tmp_path):
    """'update N applied' — the label that disambiguates a swapped-in engine
    from the one frozen into the exe."""
    _, A = _render()
    monkeypatch.setattr(R, "version_labels",
                        lambda *a, **k: ("build 239 (2026-08-21)",
                                         "update 239 applied 2026-08-21 09:10 PDT"))
    text = A.engine_stamp_text()
    assert "update 239 applied" in text, text
    assert "app build 239" in text, text


def test_stamp_distinguishes_bundled_from_an_update(monkeypatch):
    _, A = _render()
    monkeypatch.setattr(R, "version_labels",
                        lambda *a, **k: ("build 239 (2026-08-21)", "bundled"))
    text = A.engine_stamp_text()
    assert "engine: bundled" in text, text
    assert "applied" not in text, text


def test_stamp_degrades_instead_of_taking_the_sheet_down(monkeypatch):
    """version_labels blowing up must not lose the tech their whole report."""
    _, A = _render()
    def boom(*a, **k):
        raise RuntimeError("no version.json")

    monkeypatch.setattr(R, "version_labels", boom)
    assert A.engine_stamp_text() == "OTDR Suite · unknown"


def test_stamp_matches_the_sidebar_vocabulary():
    """The boss compares the workbook stamp against what the tech's sidebar
    shows; two different wordings for one fact is a translation step."""
    assert "'OTDR Suite · app {_appv} · engine: {_engv}'" in APP_SRC.replace(
        'f\'OTDR Suite · app {_appv} · engine: {_engv}\'',
        "'OTDR Suite · app {_appv} · engine: {_engv}'")


def test_stamp_is_resolved_in_the_engine_not_handed_down_by_the_hub():
    """The number in the workbook came from the ENGINE process, so the stamp
    must be read there too.  The hub must never compute or forward it — after
    a cache swap the hub is running the app.py it imported at boot while the
    subprocess loads today's engine off disk, and the two can genuinely
    differ."""
    assert "engine_stamp_text" not in APP_SRC, (
        "the hub must not compute the engine stamp")
    for name in ("splicereport_cmd", "uni_cmd"):
        src = _fn_source(name)
        for leak in ("version", "engine_stamp", "build"):
            assert leak not in src, f"{name} must not forward a version"


def test_stamp_reflects_the_subprocess_environment(tmp_path):
    """End to end through the REAL engine subprocess, with the engine identity
    set ONLY in the child's environment.

    This is the parent-vs-child proof for task 5.  The pytest process has no
    OTDR_SUITE_SOURCE, so it would stamp 'dev'; the child is handed
    'bundled .exe' and the workbook must come back saying 'engine: bundled'.
    If the stamp were ever computed in the hub and forwarded, this reads 'dev'
    and the test fails.

    (HOME is deliberately NOT redirected here: moving it changes where Python
    resolves user site-packages, and the engine child then cannot import
    numpy.  The meta.json 'update N applied' branch is covered in-process by
    test_stamp_names_an_applied_update.)"""
    import subprocess
    import openpyxl

    assert os.environ.get("OTDR_SUITE_SOURCE") in (None, "", "dev"), (
        "this test only means something while the PARENT is not a frozen build")

    out = tmp_path / "span.xlsx"
    env = dict(os.environ, OTDR_SUITE_SOURCE="bundled .exe",
               PYTHONUNBUFFERED="1")
    p = subprocess.run(
        [sys.executable, str(REPO_ROOT / "splicereport" / "run_splicereport.py"),
         "--dir-a", str(FIXTURE_SPLICE_A_DIR), "--dir-b", str(FIXTURE_SPLICE_B_DIR),
         "--out", str(out), "--site-a", "A", "--site-b", "B"],
        capture_output=True, text=True, env=env)
    assert p.returncode == 0, p.stderr[-2000:]
    assert out.exists(), p.stderr[-2000:]

    ws = openpyxl.load_workbook(out)["Acquisition Parameters"]
    row = _last_row(ws)
    vals = {c.column: c.value for rt in ws.iter_rows(min_row=row, max_row=row)
            for c in rt}
    assert vals[1] == "Report engine", vals
    assert "engine: bundled" in vals[2], vals[2]
    assert vals[2] != "OTDR Suite · dev", (
        "the stamp took the HUB's identity, not the engine's")
