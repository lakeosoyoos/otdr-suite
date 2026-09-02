"""Phase 2 of the AWS / IIG MT.1085 work: the acquisition checked against the
customer's own figures, and the grading wavelength made explicit.

Everything the acquisition audit did before asked "do these traces agree with
EACH OTHER".  A contract block asks a different question -- "do they agree
with what the contract says the cable IS" -- and a whole span shot at the
wrong group index is perfectly self-consistent, so the old check could never
see it.

Three properties matter and each has a quiet failure mode:

  1. REPORTED, NEVER CORRECTED.  A contract IOR must not re-scale distances:
     FastReporter reads the file's own IOR and matching FR is the north star.
     So the contract must reach the audit sheet and NOTHING else.
  2. It rides its own channel.  --overrides setattr's engine constants; the
     contract is a dict of expected values, and the runner must whitelist it
     so a profile typo cannot put arbitrary content in a customer's report.
  3. Absence changes nothing.  A run with no customer profile (Default /
     Lumen / Zayo) must render the audit sheet exactly as before.

Plus the grading wavelength: a folder holding both wavelengths of a fiber
collides on fiber number and the loader keeps the FIRST by name, so such a
run graded at 1550 by alphabetical accident.  GRADE_WAVELENGTH_NM makes that
a decision, and the report states which it was.
"""
from __future__ import annotations

import json
import os
import re
import subprocess
import sys
import textwrap
from pathlib import Path

from conftest import (
    run_splicereport, FIXTURE_SPLICE_A_DIR, FIXTURE_SPLICE_B_DIR, REPO_ROOT,
    SPLICEREPORT_DIR,
)

import app as hub

IIG = "AWS / IIG MT.1085"
IIG_CONTRACT = {"name": IIG, "ior": 1.467, "backscatter_db": -81.4,
                "wavelengths_nm": [1550.0, 1625.0], "graded_nm": 1550.0}


def _acq_text(xlsx_path) -> str:
    """Every populated cell of the Acquisition Parameters sheet, row-joined."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    assert ws.title == "Acquisition Parameters", ws.title
    return "\n".join(
        " | ".join(str(c.value) for c in row if c.value is not None)
        for row in ws.iter_rows())


def _legend_text(xlsx_path) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    return "\n".join(
        " | ".join(str(c.value) for c in row if c.value is not None)
        for row in wb["Legend"].iter_rows())


def _run_raw(out, extra):
    """The runner as the hub invokes it, plus `extra` argv."""
    cmd = [sys.executable, str(SPLICEREPORT_DIR / "run_splicereport.py"),
           "--dir-a", str(FIXTURE_SPLICE_A_DIR), "--dir-b", str(FIXTURE_SPLICE_B_DIR),
           "--out", str(out), "--site-a", "A", "--site-b", "B", *extra]
    return subprocess.run(cmd, capture_output=True, text=True)


# ── 1. The conformance computation itself ────────────────────────────────
def _audit():
    # acquisition_audit imports nothing engine-side at module level, so it is
    # safe to import in-process (unlike the engine, which drags in its own
    # sor_reader copy).
    if str(SPLICEREPORT_DIR) not in sys.path:
        sys.path.insert(0, str(SPLICEREPORT_DIR))
    import acquisition_audit as AA
    return AA


def _rec(ior=1.47, rbs=-83.0, nm=1548.0):
    return {"filename": "f", "ior": ior, "backscatter_db": rbs, "wavelength": nm}


def test_ior_mismatch_is_stated_in_metres():
    """The only form in which an IOR delta means anything to a reader.  An OTDR
    set to 1.4700 on 1.467 glass reports every distance SHORT -- true distance
    is d * 1.4700 / 1.467 -- so events sit ~204 m per 100 km farther out."""
    AA = _audit()
    con = AA.compute_contract_conformance([_rec()], [_rec()], IIG_CONTRACT)
    row = next(r for r in con["rows"] if r["name"] == "Group index (IOR)")
    assert row["conforms"] is False
    assert "+2,045 ppm" in row["note"], row["note"]
    assert re.search(r"≈20[45] m per 100 km farther out", row["note"]), row["note"]
    assert con["clean"] is False
    assert "not corrected" in con["headline"].lower()


def test_ior_within_display_precision_conforms():
    """EXFO shows IOR to 5 decimals; a last-digit difference is not a finding."""
    AA = _audit()
    con = AA.compute_contract_conformance([_rec(ior=1.46703)], [], {"ior": 1.467})
    assert con["rows"][0]["conforms"] is True
    assert con["rows"][0]["note"] == ""


def test_missing_wavelength_is_informational_not_amber():
    """One folder per wavelength is the normal shape of a run (grading at
    1550 means the tech loads the 1550 shots), so '1625 not in this folder'
    must NOT open the sheet amber.  Only a wavelength the contract never asked
    for is a finding."""
    AA = _audit()
    con = AA.compute_contract_conformance(
        [_rec(ior=1.467, rbs=-81.4)], [], {"wavelengths_nm": [1550, 1625]})
    row = con["rows"][0]
    assert row["name"] == "Wavelengths acquired"
    assert row["conforms"] is None, row
    assert "Not a finding" in row["note"]
    assert con["clean"] is True

    con = AA.compute_contract_conformance(
        [_rec(nm=1548.0), _rec(nm=1310.2)], [], {"wavelengths_nm": [1550, 1625]})
    row = con["rows"][0]
    assert row["conforms"] is False
    assert "1310.2 nm" in row["note"]


def test_matched_lot_lasers_count_as_their_nominal_wavelength():
    """A '1550' unit reports ~1545.8 and a '1625' ~1625.5; the window has to
    absorb that or every healthy span would 'lack' its own wavelength."""
    AA = _audit()
    con = AA.compute_contract_conformance(
        [_rec(nm=1545.8), _rec(nm=1625.5)], [], {"wavelengths_nm": [1550, 1625]})
    assert con["rows"][0]["conforms"] is True


def test_no_contract_means_no_block():
    AA = _audit()
    assert AA.compute_contract_conformance([_rec()], [_rec()], None) is None
    assert AA.compute_contract_conformance([_rec()], [_rec()], {}) is None
    # A contract with only a name has nothing to check -> no rows -> None.
    assert AA.compute_contract_conformance([_rec()], [], {"name": "X"}) is None


def test_graded_wavelength_row_explains_the_drops():
    """The row that turns the loader's silent keep-first into a stated fact."""
    AA = _audit()
    drops = [{"fiber": 1, "kept_nm": 1548.0, "dropped_nm": 1625.5,
              "kept_file": "a", "dropped_file": "b", "direction": "A"}] * 24
    dups = [{"fiber": 7, "kept_nm": 1548.0, "dropped_nm": 1548.0,
             "kept_file": "c", "dropped_file": "d", "direction": "A"}]
    con = AA.compute_contract_conformance(
        [_rec()], [], {"graded_nm": 1550}, load_drops=drops + dups)
    row = next(r for r in con["rows"] if r["name"] == "Graded wavelength")
    # By filename order, and it happened to be the contract's wavelength.
    assert "by filename order" in row["actual"]
    assert row["conforms"] is True
    assert "24 trace(s) at another wavelength were NOT graded" in row["note"]
    assert "1 further file(s) were dropped as duplicate" in row["note"]

    # Explicitly selected: says so.
    con = AA.compute_contract_conformance(
        [_rec()], [], {"graded_nm": 1550}, load_drops=drops, graded_nm=1550.0)
    row = next(r for r in con["rows"] if r["name"] == "Graded wavelength")
    assert row["actual"] == "1550 nm (selected)"
    assert row["conforms"] is True

    # Selected the OTHER one: the contract row says so, in amber.
    con = AA.compute_contract_conformance(
        [_rec()], [], {"graded_nm": 1550}, graded_nm=1625.0)
    row = next(r for r in con["rows"] if r["name"] == "Graded wavelength")
    assert row["conforms"] is False


# ── 2. The loader's wavelength preference ────────────────────────────────
def test_grading_wavelength_beats_filename_order(tmp_path):
    """'_1550' sorts before '_1625', so keep-first always graded 1550.  With
    GRADE_WAVELENGTH_NM the requested wavelength wins regardless of name, an
    unmatched request falls back to keep-first, and every collision is
    recorded so the report can state it."""
    a = tmp_path / "A"; b = tmp_path / "B"
    a.mkdir(); b.mkdir()
    for n in ("X0001_1550.sor", "X0001_1625.sor"):
        (a / n).write_bytes(b"")
    (b / "Y0001_1550.sor").write_bytes(b"")
    src = textwrap.dedent(f"""
        import sys, os
        sys.path.insert(0, {str(SPLICEREPORT_DIR)!r})
        import splicereportmatchexfo as E
        def fake(path, trim=False):
            n = os.path.basename(path)
            return {{'filename': n, 'events': [],
                     'wavelength': 1625.5 if '1625' in n else 1548.0}}
        E.parse_sor_full = fake
        def run(want):
            E.GRADE_WAVELENGTH_NM = want
            E.LOAD_DROPS.clear()
            fa, fb = E.load_all({str(a)!r}, {str(b)!r})
            d = E.LOAD_DROPS[0]
            return fa[1]['filename'], d['dropped_file'], len(E.LOAD_DROPS)
        print(run(0.0)); print(run(1625.0)); print(run(1550.0)); print(run(1310.0))
    """)
    p = subprocess.run([sys.executable, "-c", src], capture_output=True, text=True)
    assert p.returncode == 0, p.stderr[-1500:]
    lines = [l for l in p.stdout.splitlines() if l.startswith("(")]
    assert lines[0] == "('X0001_1550.sor', 'X0001_1625.sor', 1)", lines   # default: first by name
    assert lines[1] == "('X0001_1625.sor', 'X0001_1550.sor', 1)", lines   # 1625 wins by request
    assert lines[2] == "('X0001_1550.sor', 'X0001_1625.sor', 1)", lines
    assert lines[3] == "('X0001_1550.sor', 'X0001_1625.sor', 1)", lines   # no 1310: keep-first


def test_grading_wavelength_ships_off():
    """0.0 = today's behaviour.  A non-zero default would silently change
    which trace every mixed folder grades."""
    src = (SPLICEREPORT_DIR / "splicereportmatchexfo.py").read_text(encoding="utf-8")
    assert re.search(r"^GRADE_WAVELENGTH_NM\s*=\s*0\.0\s*$", src, re.M)


# ── 3. Through the runner: its own channel, whitelisted, never a measurement ──
def test_contract_reaches_the_audit_sheet_without_overrides(tmp_path):
    """--contract with NO --overrides.  The overrides block imports math
    locally, so a contract block that borrowed it raised NameError on exactly
    the hub's most common invocation (no active overrides)."""
    out = tmp_path / "c.xlsx"
    p = _run_raw(out, ["--contract", json.dumps(IIG_CONTRACT)])
    assert p.returncode == 0, p.stderr[-1500:]
    text = _acq_text(out)
    assert "Contract" in text and IIG in text
    assert "Group index (IOR)" in text and "ppm" in text
    assert "Backscatter coefficient" in text
    assert "Graded wavelength" in text


def test_contract_changes_no_measurement(tmp_path):
    """Property 1.  The Splice Report grid must be byte-identical with and
    without the contract -- the contract is printed, it is not applied."""
    import openpyxl

    def grid(path):
        ws = openpyxl.load_workbook(path)["Splice Report"]
        return {(c.row, c.column): str(c.value)
                for row in ws.iter_rows() for c in row if c.value is not None}

    base = tmp_path / "base.xlsx"; con = tmp_path / "con.xlsx"
    assert _run_raw(base, []).returncode == 0
    # A contract IOR far from the file's, so a re-scale would be obvious.
    assert _run_raw(con, ["--contract", json.dumps({"ior": 1.30})]).returncode == 0
    assert grid(base) == grid(con)


def test_runner_whitelists_the_contract(tmp_path):
    """A malformed or hostile contract can only cost the contract block, never
    the run.  Non-numeric figures are dropped, unknown keys ignored, and a
    contract with nothing checkable left renders no block at all."""
    out = tmp_path / "w.xlsx"
    p = _run_raw(out, ["--contract",
                       json.dumps({"ior": "abc", "backscatter_db": float("nan")
                                   if False else "nan", "evil": 1,
                                   "wavelengths_nm": "1550", "name": "X" * 500})])
    assert p.returncode == 0, p.stderr[-1500:]
    assert "Group index (IOR)" not in _acq_text(out)
    # Non-JSON and non-dict JSON degrade the same way.
    for bad in ("{not json", "[1,2]", "5"):
        out2 = tmp_path / "b.xlsx"
        p = _run_raw(out2, ["--contract", bad])
        assert p.returncode == 0, (bad, p.stderr[-800:])
        assert "Contract" not in _acq_text(out2).split("Test Settings")[0]


def test_no_contract_leaves_the_audit_sheet_as_before(tmp_path):
    """Property 3: Default / Lumen / Zayo runs carry no contract and their
    audit sheet must not gain a block."""
    out = tmp_path / "n.xlsx"
    assert _run_raw(out, []).returncode == 0
    head = _acq_text(out).split("Test Settings")[0]
    assert "Contract" not in head and "Group index (IOR)" not in head


def test_legend_states_the_graded_wavelength(tmp_path):
    out = tmp_path / "g.xlsx"
    assert _run_raw(out, []).returncode == 0
    assert re.search(r"Graded wavelength \| no preference", _legend_text(out))
    out2 = tmp_path / "g2.xlsx"
    assert _run_raw(out2, ["--overrides", json.dumps({"GRADE_WAVELENGTH_NM": 1550})]).returncode == 0
    assert "1550 nm (selected)" in _legend_text(out2)


# ── 4. The hub: the profile carries it, the command forwards it ──────────
def test_iig_profile_carries_the_contract_and_grades_at_1550():
    con = hub._contract_from_profile(IIG)
    assert con["ior"] == 1.467
    assert con["backscatter_db"] == -81.4
    assert con["wavelengths_nm"] == [1550.0, 1625.0]
    assert con["graded_nm"] == 1550.0
    assert hub._engine_extras_from_profile(IIG) == {"GRADE_WAVELENGTH_NM": 1550.0}


def test_existing_profiles_carry_neither():
    for prof in ("Default (engine baseline)", "Lumen", "Zayo"):
        assert hub._contract_from_profile(prof) is None, prof
        assert hub._engine_extras_from_profile(prof) == {}, prof


def test_engine_extras_are_whitelisted():
    hub.CUSTOMER_PROFILES["__t__"] = {
        "apply": set(), "thresholds": {},
        "engine": {"REBURN_THRESHOLD": 0.01, "GRADE_WAVELENGTH_NM": "1625",
                   "GRADE_WAVELENGTH_NM_TYPO": 1}}
    try:
        assert hub._engine_extras_from_profile("__t__") == {"GRADE_WAVELENGTH_NM": 1625.0}
    finally:
        hub.CUSTOMER_PROFILES.pop("__t__", None)


def test_cmd_forwards_contract_on_its_own_channel():
    cmd = hub.splicereport_cmd("/a", "/b", "/o.xlsx", "X", "Y",
                               overrides={"REBURN_THRESHOLD": 0.2},
                               contract={"ior": 1.467})
    assert json.loads(cmd[cmd.index("--contract") + 1]) == {"ior": 1.467}
    assert json.loads(cmd[cmd.index("--overrides") + 1]) == {"REBURN_THRESHOLD": 0.2}
    assert "--contract" not in hub.splicereport_cmd("/a", "/b", "/o.xlsx", "X", "Y")


def test_run_button_passes_the_active_profiles_contract():
    """Source-level pin, in the repo's own style: the pending command is built
    with the ACTIVE profile's contract and engine extras."""
    src = (REPO_ROOT / "app.py").read_text(encoding="utf-8")
    site = src.split("_pending_cmd'] = splicereport_cmd(", 1)
    assert len(site) == 2
    before = site[0][-900:]
    assert "overrides.update(_engine_extras_from_profile(_prof_name))" in before
    assert "contract=_contract_from_profile(_prof_name)" in site[1][:300]
