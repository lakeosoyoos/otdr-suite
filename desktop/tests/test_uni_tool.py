"""Unidirectional one-shot (uni_* engine section) — regression tests.

Ground truth: LAMBEY 432 (Lumen Border project) vs the Zach-approved
LAM_BEY_unidir ZK.xlsx.  The three detection deltas vs the old standalone
tool are each locked here with the LAMBEY numbers that motivated them:

  * break floor — 37 fibers break at 0.99 km; the old `eof <= 1.0` floor
    silently dropped every one (Zach hand-typed them into the approved
    sheet as 'Break 1').
  * pre-break damage — 51 events at ~0.57 km sit AHEAD of the 0.99/1.14
    breaks; the old blanket end-region exclusion (each fiber's own EOF
    minus 0.5 km) blanked all of them (Zach's hand-typed 'Bend/Damage 1').
  * launch dead zone — LAMBEY has no launch reels (6/432 fibers with a
    launch reflection); the old unconditional 3.0 km exclusion hid every
    near-launch plant event.

NOT a rule (measured refutation, LAMBEY): demoting sparse splice
candidates by occupancy — the real splice @1.52 km has LOWER occupancy
(10%) than the tech-demoted @4.05 km candidate (20%).  The @4.05 rename
was field knowledge (HH8 handhole), not a trace signal, so it stays the
tech's call in the workbook.
"""
import json
import os
import subprocess
import sys

import pytest

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
SPLICE_DIR = os.path.join(ROOT, 'splicereport')
sys.path.insert(0, SPLICE_DIR)

import splicereportmatchexfo as E  # noqa: E402

SPAN = 11.4


def _ev(km, loss=0.0, typ='0F9999LS', end=False, refl=False):
    return {'dist_km': km, 'splice_loss': loss, 'type': typ,
            'is_end': end, 'is_reflective': refl, 'reflection': 0.0}


def _fiber(events, loc_a='LA Media Rd MH', loc_b='Beyer'):
    return {'events': events, 'gen_loc_a': loc_a, 'gen_loc_b': loc_b,
            'gen_cable_id': 'BC'}


def _full_span_fiber(extra=()):
    evs = [_ev(0.0, typ='1F9999LS', refl=True)]
    evs += list(extra)
    evs.append(_ev(SPAN, typ='1E9999LS', end=True, refl=True))
    return _fiber(evs)


def _broken_fiber(eof_km, extra=()):
    evs = [_ev(0.0, typ='1F9999LS', refl=True)]
    evs += list(extra)
    evs.append(_ev(eof_km, typ='1E9999LS', end=True, refl=True))
    return _fiber(evs)


# ── Break floor ─────────────────────────────────────────────────────────

def test_break_floor_keeps_sub_1km_breaks():
    fibers = {1: _broken_fiber(0.99), 2: _full_span_fiber()}
    breaks = E.uni_find_breaks(fibers, [], SPAN)
    assert [b['fiber'] for b in breaks] == [1]      # old `eof <= 1.0` dropped it


def test_break_floor_still_excludes_front_panel():
    fibers = {1: _broken_fiber(0.2)}                # inside UNI_BREAK_MIN_KM
    assert E.uni_find_breaks(fibers, [], SPAN) == []


def test_break_clustering_splits_099_and_113():
    fibers = {f: _broken_fiber(0.99) for f in range(1, 5)}
    fibers.update({f: _broken_fiber(1.14) for f in range(5, 9)})
    breaks = E.uni_find_breaks(fibers, [], SPAN)
    cols = E.uni_cluster_breaks(breaks)
    assert len(cols) == 2                            # 150 m apart: two events
    assert [len(c['members']) for c in cols] == [4, 4]


# ── Pre-break damage + launch dead zone ─────────────────────────────────

def test_prebreak_damage_is_reported_without_launch_box():
    fibers = {1: _broken_fiber(0.99, extra=[_ev(0.57, loss=0.15)])}
    evs = E.uni_find_off_splice_events(fibers, [], launch_box_present=False,
                                       span_km=SPAN)
    assert [(e['fiber'], e['position_km']) for e in evs] == [(1, 0.57)]


def test_prebreak_guard_still_masks_events_at_the_break():
    # 0.05 km ahead of the EOF is inside UNI_PREBREAK_GUARD_KM — that is the
    # break's own signature, not separate damage.
    fibers = {1: _broken_fiber(0.99, extra=[_ev(0.94, loss=0.30)])}
    evs = E.uni_find_off_splice_events(fibers, [], launch_box_present=False,
                                       span_km=SPAN)
    assert evs == []


def test_full_span_fiber_keeps_end_region_exclusion():
    fibers = {1: _full_span_fiber(extra=[_ev(SPAN - 0.3, loss=0.20)])}
    evs = E.uni_find_off_splice_events(fibers, [], launch_box_present=False,
                                       span_km=SPAN)
    assert evs == []                                 # within 0.5 km of real EOL


def test_launch_box_present_keeps_3km_exclusion():
    fibers = {1: _broken_fiber(0.99, extra=[_ev(0.57, loss=0.15)])}
    evs = E.uni_find_off_splice_events(fibers, [], launch_box_present=True,
                                       span_km=SPAN)
    assert evs == []                                 # launch reel covers 0.57


def test_launch_box_detection_fraction():
    with_box = {f: _full_span_fiber(extra=[_ev(1.0, typ='1F9999LS', refl=True)])
                for f in range(1, 4)}
    with_box.update({f: _full_span_fiber() for f in range(4, 6)})
    present, frac = E.uni_detect_launch_box(with_box)   # 3/5 = 60%
    assert present and frac == pytest.approx(0.6)
    no_box = {f: _full_span_fiber() for f in range(1, 6)}
    present, frac = E.uni_detect_launch_box(no_box)
    assert not present and frac == 0.0


# ── Closure discovery + phantom validation (ported behavior preserved) ──

def _population(n, splice_km, loss_fn):
    return {f: _full_span_fiber(extra=[_ev(splice_km + (f % 5) * 0.005,
                                           loss=loss_fn(f))])
            for f in range(1, n + 1)}


def test_phantom_closure_rejected_no_gainers_high_median():
    fibers = _population(30, 5.0, lambda f: 0.151)   # bend-like: all lose big
    cands = E.uni_discover_splices(fibers)
    assert len(cands) == 1
    assert E.uni_refine_and_validate(fibers, cands) == []


def test_real_closure_survives_validation():
    # Splice-like: small median, some gainers (MFD mismatch both ways)
    fibers = _population(30, 5.0, lambda f: -0.05 if f % 6 == 0 else 0.04)
    cands = E.uni_discover_splices(fibers)
    valid = E.uni_refine_and_validate(fibers, cands)
    assert len(valid) == 1
    assert valid[0]['position_km_refined'] == pytest.approx(5.0, abs=0.05)


def test_discovery_needs_min_population():
    fibers = _population(E.UNI_MIN_POP_SPLICE - 1, 5.0, lambda f: 0.05)
    assert E.uni_discover_splices(fibers) == []


# ── Labels + grid ───────────────────────────────────────────────────────

def test_cell_label_formats():
    assert E.uni_format_cell_label([(23, 0.18)]) == 'F23 .180'
    assert E.uni_format_cell_label([(23, 0.18), (47, 0.22)]) == 'F23,F47 .220'
    assert E.uni_format_cell_label([(12, None), (19, None)]) == 'F12,F19 broke'
    assert E.uni_format_cell_label([(23, -0.105)]) == 'F23 -.105'


def test_ribbon_label_tube_naming():
    assert E.uni_ribbon_label(0, 12, 432) == 'Fiber 1-12 (1) (A1)'
    assert E.uni_ribbon_label(5, 12, 432) == 'Fiber 61-72 (6) (C2)'


def test_direction_signature_from_genparams():
    assert E.uni_direction_signature(_fiber([])) == 'LA Media Rd MH->Beyer'
    assert E.uni_direction_signature(
        {'gen_loc_a': '', 'gen_loc_b': '', 'gen_cable_id': 'BC42'}) == 'BC42'


def test_short_code():
    assert E.uni_short_code('LA Media Rd MH') == 'LAM'
    assert E.uni_short_code('Beyer Caliente Airway Border') == 'BEY'


# ── Writer: ZK five-sheet contract + Handholes annotation row ───────────

def _flagged_splice_population(n=24):
    """A VALID closure (gainers present → survives phantom validation) where
    every fiber still carries |loss| >= 0.1 so every cell flags."""
    return {f: _full_span_fiber(extra=[_ev(5.0 + (f % 5) * 0.005,
                                           loss=(-0.12 if f % 6 == 0 else 0.105))])
            for f in range(1, n + 1)}


def test_writer_zk_sheet_set_and_handholes_row(tmp_path):
    import openpyxl
    fibers = _flagged_splice_population()
    cands = E.uni_discover_splices(fibers)
    valid = E.uni_refine_and_validate(fibers, cands)
    cols = E.uni_build_columns(valid, [], [])
    grid = E.uni_build_ribbon_grid(fibers, cols, 12)
    out = str(tmp_path / 'uni.xlsx')
    res = E.uni_write_xlsx(grid, cols, 24, 12, SPAN, out,
                           site_a='LAM', site_b='BEY', fibers=None)
    wb = openpyxl.load_workbook(out)
    # fibers=None → no audit sheet; Reburn inserts at index 1.  (The real-run
    # five-sheet ZK order is locked by test_writer_full_sheet_order_with_fibers.)
    assert wb.sheetnames == ['Unidir Events', 'Reburn Percentage', 'Legend',
                             'Flagged Events']
    ws = wb['Unidir Events']
    assert ws.cell(row=1, column=1).value == 'LAM→BEY: ft'
    assert ws.cell(row=2, column=1).value == 'LAM→BEY: km'
    assert ws.cell(row=3, column=1).value == 'Handholes:'   # tech-fill row
    assert ws.cell(row=4, column=1).value == 'Ribbon'
    assert ws.cell(row=4, column=2).value == 'Splice 1'
    assert ws.freeze_panes == 'B5'
    assert res['flagged_rows'] == 24                        # every fiber ≥ 0.1


def test_writer_full_sheet_order_with_fibers(tmp_path):
    import openpyxl
    fibers = _flagged_splice_population()
    for r in fibers.values():                # audit fields the auditor reads
        r.update({'filename': f'LAMBEY_{id(r)}.sor', 'otdr_model': 'FTBx-735D',
                  'otdr_serial': '123', 'date_time': 1_780_000_000,
                  'wavelength': '1550'})
    cands = E.uni_discover_splices(fibers)
    valid = E.uni_refine_and_validate(fibers, cands)
    cols = E.uni_build_columns(valid, [], [])
    grid = E.uni_build_ribbon_grid(fibers, cols, 12)
    out = str(tmp_path / 'uni_full.xlsx')
    E.uni_write_xlsx(grid, cols, 24, 12, SPAN, out,
                     site_a='LAM', site_b='BEY', fibers=fibers)
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ['Acquisition Parameters', 'Reburn Percentage',
                             'Unidir Events', 'Legend', 'Flagged Events']


# ── Runner contract ─────────────────────────────────────────────────────

def test_runner_uni_needs_no_dir_b(tmp_path):
    """--uni with an empty input folder emits a graceful ok:False manifest —
    and must not demand --dir-b."""
    empty = tmp_path / 'empty'
    empty.mkdir()
    proc = subprocess.run(
        [sys.executable, os.path.join(SPLICE_DIR, 'run_splicereport.py'),
         '--uni', '--dir-a', str(empty), '--out', str(tmp_path / 'o.xlsx')],
        capture_output=True, text=True, timeout=120)
    manifest = json.loads(proc.stdout.strip().splitlines()[-1])
    assert manifest['ok'] is False
    assert 'no SOR/JSON files' in manifest['error'] or 'folder' in manifest['error']


# ── Damage-zone completion (trace-measured, LAMBEY BD1 ground truth) ────

def _zone_population(n_broken=25, anchor=0.57, eof_broken=1.1):
    fibers = {}
    for f in range(1, n_broken + 1):
        fibers[f] = _broken_fiber(eof_broken, extra=[_ev(anchor + (f % 3) * 0.004,
                                                         loss=0.05)])
    return fibers


def test_certified_zone_completes_live_fibers(monkeypatch):
    fibers = _zone_population()
    fibers[99] = _full_span_fiber()                  # live, NO stored event
    fibers[98] = _full_span_fiber()                  # live, undamaged
    def fake_fixed(r, km, **kw):
        if r is fibers[99] and abs(km - 0.57) < 0.1:
            return 0.05                              # real step on the glass
        return 0.004                                 # noise floor
    monkeypatch.setattr(E, 'measure_grey_loss_from_sor', fake_fixed)
    monkeypatch.setattr(E, 'measure_grey_loss_from_sor_event',
                        lambda r, e, **kw: e.get('splice_loss'))
    cols = E.uni_prebreak_damage(fibers, SPAN, launch_box_present=False,
                                 break_centers=[1.0])
    assert len(cols) == 1 and cols[0]['damage_zone_certified']
    members = cols[0]['prebreak_members']
    assert 99 in members                             # live fiber recovered
    assert 98 not in members                         # noise stays out


def test_uncertified_zone_stays_broken_only(monkeypatch):
    fibers = _zone_population()
    fibers[99] = _full_span_fiber()
    monkeypatch.setattr(E, 'measure_grey_loss_from_sor',
                        lambda r, km, **kw: 0.05)
    monkeypatch.setattr(E, 'measure_grey_loss_from_sor_event',
                        lambda r, e, **kw: e.get('splice_loss'))
    cols = E.uni_prebreak_damage(fibers, SPAN, launch_box_present=False,
                                 break_centers=[5.0])   # no break near anchor
    assert len(cols) == 1 and not cols[0]['damage_zone_certified']
    assert 99 not in cols[0]['prebreak_members']


def test_membership_floors_stored_vs_sweep(monkeypatch):
    fibers = _zone_population()
    fibers[97] = _full_span_fiber(extra=[_ev(0.57, loss=0.01)])  # stored, tiny
    fibers[98] = _full_span_fiber()                              # no stored
    steps = {97: 0.022, 98: 0.022}
    monkeypatch.setattr(E, 'measure_grey_loss_from_sor',
                        lambda r, km, **kw: next(
                            (v for f, v in steps.items() if r is fibers[f]), 0.05))
    monkeypatch.setattr(E, 'measure_grey_loss_from_sor_event',
                        lambda r, e, **kw: e.get('splice_loss'))
    cols = E.uni_prebreak_damage(fibers, SPAN, launch_box_present=False,
                                 break_centers=[1.0])
    members = cols[0]['prebreak_members']
    assert 97 in members       # .022 WITH a stored event (table+trace agree)
    assert 98 not in members   # .022 sweep-only sits under the .03 floor


def test_zone_sweep_respects_eof_margin(monkeypatch):
    fibers = _zone_population()
    fibers[50] = _broken_fiber(0.99)                 # dies just past the zone
    calls = []
    def fake_fixed(r, km, **kw):
        if r is fibers[50]:
            calls.append(km)
        return 0.004
    monkeypatch.setattr(E, 'measure_grey_loss_from_sor', fake_fixed)
    monkeypatch.setattr(E, 'measure_grey_loss_from_sor_event',
                        lambda r, e, **kw: e.get('splice_loss'))
    E.uni_prebreak_damage(fibers, SPAN, launch_box_present=False,
                          break_centers=[0.99, 1.1])
    assert calls, "the 0.99-breaker must still be swept where safe"
    assert max(calls) <= 0.99 - E.UNI_ZONE_EOF_MARGIN_KM + 1e-6


# ── Landmarks: labels + demotion ────────────────────────────────────────

def test_landmarks_demote_and_label():
    columns = [
        {'kind': 'splice', 'position_km_refined': 4.05,
         'position_km_display': 4.05, 'fiber_count': 60},
        {'kind': 'splice', 'position_km_refined': 7.91,
         'position_km_display': 7.91, 'fiber_count': 300},
        {'kind': 'bend_damage', 'position_km_refined': 10.41,
         'position_km_display': 10.41, 'members': []},
    ]
    landmarks = [
        {'km': 4.05, 'label': 'HH8', 'closure': False},
        {'km': 7.78, 'label': 'HH5', 'closure': False},   # 130 m from splice
        {'km': 7.91, 'label': 'HH4', 'closure': True},
        {'km': 10.41, 'label': 'HH2', 'closure': False},
    ]
    demoted = E.uni_apply_landmarks(columns, landmarks)
    assert columns[0]['kind'] == 'bend_damage'            # HH8 demotes
    assert columns[0]['landmark'] == 'HH8'
    assert columns[1]['kind'] == 'splice'                 # nearest = HH4 closure
    assert columns[1]['landmark'] == 'HH4'
    assert columns[2]['landmark'] == 'HH2'                # label only
    assert demoted == [4.05]


def test_landmark_demote_radius_tighter_than_label():
    # A non-closure landmark 130 m out labels the column but must NOT demote.
    columns = [{'kind': 'splice', 'position_km_refined': 7.91,
                'position_km_display': 7.91, 'fiber_count': 300}]
    demoted = E.uni_apply_landmarks(
        columns, [{'km': 7.78, 'label': 'HH5', 'closure': False}])
    assert columns[0]['kind'] == 'splice' and demoted == []
    assert columns[0]['landmark'] == 'HH5'


def test_writer_handholes_row_renders_labels(tmp_path):
    import openpyxl
    fibers = _flagged_splice_population()
    cands = E.uni_discover_splices(fibers)
    valid = E.uni_refine_and_validate(fibers, cands)
    cols = E.uni_build_columns(valid, [], [])
    E.uni_apply_landmarks(cols, [{'km': cols[0]['position_km_refined'],
                                  'label': 'HH7', 'closure': True}])
    grid = E.uni_build_ribbon_grid(fibers, cols, 12)
    out = str(tmp_path / 'uni_lm.xlsx')
    E.uni_write_xlsx(grid, cols, 24, 12, SPAN, out, site_a='LAM', site_b='BEY')
    ws = openpyxl.load_workbook(out)['Unidir Events']
    assert ws.cell(row=3, column=1).value == 'Handholes:'
    assert ws.cell(row=3, column=2).value == 'HH7'


def test_landmarks_text_parser():
    """app.py's landmarks box parser — AST-loaded so importing the Streamlit
    page is not required."""
    import ast as _ast, types
    # encoding= is load-bearing: Windows CI reads cp1252 by default and
    # app.py carries UTF-8 (emoji in page copy).
    src = open(os.path.join(ROOT, 'app.py'), encoding='utf-8').read()
    tree = _ast.parse(src)
    fn = next(n for n in tree.body if isinstance(n, _ast.FunctionDef)
              and n.name == '_parse_landmarks_text')
    mod = types.ModuleType('lm'); exec(compile(_ast.Module(
        body=[fn], type_ignores=[]), 'app.py', 'exec'), mod.__dict__)
    lms, bad = mod._parse_landmarks_text(
        '0.57, Replaced section\n4.05, HH8\n7.91, HH4, splice\n'
        '# comment\nnot-a-km, X\n')
    assert lms == [
        {'km': 0.57, 'label': 'Replaced section', 'closure': False},
        {'km': 4.05, 'label': 'HH8', 'closure': False},
        {'km': 7.91, 'label': 'HH4', 'closure': True}]
    assert bad == ['not-a-km, X']
