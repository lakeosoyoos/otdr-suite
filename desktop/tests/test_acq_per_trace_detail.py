"""The Acquisition Parameters sheet: no per-trace detail on the BIDIRECTIONAL
splice report, per-wavelength Pulse width / Averaging untouched.

Background
----------
`Per-trace detail` held four file-level checks — Test date, OTDR model, OTDR
serial, Wavelength — each comparing every A-direction trace against every
B-direction trace.  A bidirectional span is *defined* by A and B being shot
from opposite ends, usually on a different day and (on WSC/SUI-class spans)
with a second unit at a different wavelength, so three of the four reported
exactly half the run as "differing" and then listed ~250 rows of filenames.
On WSC↔SUI the block ran rows 13–261 of a 261-row sheet.

What must survive is everything ABOVE it: the per-wavelength Pulse width and
Averaging rollups.  On WSC↔SUI the Averaging row reads
    ⚠ Majority: 60 s (1151 of 1152) — 1 differ
          SUIWSC1100.sor = 19 s
which is how we match the field team's sheet entry "1100 19 Sec (reshoot)".

The block is NOT deleted — it stays ON for the UNIDIRECTIONAL report, whose
input is one direction, where a model/serial/date disagreement is a genuine
finding (validated on LAMBEY: 216 fibers on an FTBx-735D, 216 on an
FTBx-730D, inside a single direction).
"""
import subprocess
import sys
import textwrap

import openpyxl

from conftest import (SPLICEREPORT_DIR, FIXTURE_SPLICE_A_DIR,
                      FIXTURE_SPLICE_B_DIR, run_splicereport)

FILE_LEVEL_LABELS = {"Test date (calendar day)", "OTDR model",
                     "OTDR serial", "Wavelength"}


def _acq_sheet(tmp_path):
    out = tmp_path / "report.xlsx"
    rc, m, stderr = run_splicereport(FIXTURE_SPLICE_A_DIR, FIXTURE_SPLICE_B_DIR,
                                     out, "Elm", "Mil")
    assert rc == 0 and m and m.get("ok"), f"runner failed: {stderr[-1200:]}"
    wb = openpyxl.load_workbook(out)
    assert "Acquisition Parameters" in wb.sheetnames
    return wb["Acquisition Parameters"]


def _col(ws, n):
    return [ws.cell(r, n).value for r in range(1, ws.max_row + 1)]


def test_bidir_report_has_no_per_trace_detail(tmp_path):
    ws = _acq_sheet(tmp_path)
    labels = [v for v in _col(ws, 1) if v]
    assert "Per-trace detail" not in labels, \
        f"the per-trace detail banner is back on the bidirectional sheet: {labels}"
    # The LABEL alone is no longer the signal.  Below the FastReporter Test
    # Settings panel the sheet prints an "Instrument" block whose rows are
    # named "OTDR model" / "OTDR serial" / "Wavelength" per direction — plain
    # values, deliberately NOT compared, because A and B are shot from
    # opposite ends and comparing them would flag every healthy span, which is
    # precisely the noise #86 removed.
    #
    # What must never come back is those fields carrying an A-vs-B VERDICT.
    # So the check is on the verdict, not the label: a "✓ All match" / "⚠
    # Majority:" spec beside one of these labels means the file-level block is
    # back.  This tests the defect rather than a proxy for it.
    verdicted = {ws.cell(r, 1).value for r in range(1, ws.max_row + 1)
                 if isinstance(ws.cell(r, 2).value, str)
                 and ws.cell(r, 2).value.lstrip().startswith(("✓", "⚠"))}
    leaked = FILE_LEVEL_LABELS & verdicted
    assert not leaked, f"file-level A-vs-B checks still emitted: {sorted(leaked)}"
    # No filename outlier list from those checks either.
    body = [v for v in _col(ws, 2) if isinstance(v, str)]
    assert not [v for v in body if v.strip().startswith("… and ")], \
        "a truncated '… and N more' outlier list survived"


def test_per_wavelength_pulse_and_averaging_survive(tmp_path):
    ws = _acq_sheet(tmp_path)
    labels = [v for v in _col(ws, 1) if v]
    assert any(isinstance(v, str) and v.startswith("— ") and "nm" in v
               for v in labels), f"per-wavelength banner missing: {labels}"
    assert "Pulse width" in labels, labels
    assert "Averaging" in labels, labels
    # Both rows carry a verdict, not a blank.
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value in ("Pulse width", "Averaging"):
            assert ws.cell(r, 2).value, f"row {r} verdict is empty"


# ── Renderer contract, exercised through the real audit payload ──────────
def _run(body):
    header = ("import sys\n"
              f"sys.path.insert(0, {str(SPLICEREPORT_DIR)!r})\n"
              "import openpyxl\n"
              "import acquisition_audit as aa\n"
              + textwrap.dedent("""
                def _rec(fn, wl, dur, model, serial, day):
                    return {'filename': fn, '_source': 'sor',
                            'exfo_wavelength_nm': wl, 'duration_sec': dur,
                            'otdr_model': model, 'otdr_serial': serial,
                            'date_time': day,
                            'exfo_calibration': {'NominalPulseWidth': 500e-9}}

                # A WSC/SUI-shaped span: A and B on different units, different
                # exact wavelengths, different days — plus ONE 19 s reshoot.
                A = {i: _rec('WSCSUI%04d.sor' % i, 1542.8, 60.0,
                             'MAX-730C-SM2-EA', '1310137', 1_786_000_000)
                     for i in range(1, 5)}
                B = {i: _rec('SUIWSC%04d.sor' % i, 1556.1, 60.0,
                             'FTBx-735D-SM3-EA', '1978245', 1_786_200_000)
                     for i in range(1, 5)}
                B[4] = _rec('SUIWSC1100.sor', 1542.8, 19.0,
                            'FTBx-735D-SM3-EA', '1978245', 1_786_200_000)
                audit = aa.audit_acquisition(A, B)

                def _sheet(**kw):
                    wb = openpyxl.Workbook()
                    aa.render_xlsx_sheet(wb, audit, **kw)
                    ws = wb['Acquisition Parameters']
                    return ws, [[ws.cell(r, c).value for c in (1, 2)]
                                for r in range(1, ws.max_row + 1)]
                """))
    p = subprocess.run([sys.executable, "-c", header + textwrap.dedent(body)],
                       capture_output=True, text=True)
    assert p.returncode == 0, f"exit {p.returncode}\n{p.stdout}\n{p.stderr}"
    assert p.stdout.strip().splitlines()[-1] == "OK", p.stdout


def test_reshoot_outlier_still_fires_with_detail_off():
    """The 19 s reshoot is the finding the sheet exists for — suppressing the
    per-trace block must not touch it."""
    _run("""
        ws, rows = _sheet(per_trace_detail=False)
        flat = [c for row in rows for c in row if isinstance(c, str)]
        assert not any(c == 'Per-trace detail' for c in flat), flat
        assert not any(c in ('OTDR model', 'OTDR serial', 'Wavelength',
                             'Test date (calendar day)') for c in flat), flat
        # Averaging majority verdict + the named outlier line both present.
        avg = [row for row in rows if row[0] == 'Averaging']
        assert avg, rows
        assert any('60 s' in (row[1] or '') and 'differ' in (row[1] or '')
                   for row in avg), avg
        assert any('SUIWSC1100.sor' in c and '19 s' in c for c in flat), flat
        assert any(row[0] == 'Pulse width' for row in rows), rows
        print('OK')
    """)


def test_unidirectional_report_keeps_per_trace_detail():
    """Uni feeds one direction, where a model/serial/date split IS a finding.
    Default stays ON so that report is unchanged."""
    _run("""
        for kw in ({}, {'per_trace_detail': True}):
            ws, rows = _sheet(**kw)
            flat = [c for row in rows for c in row if isinstance(c, str)]
            assert 'Per-trace detail' in flat, (kw, flat)
            assert 'OTDR serial' in flat, (kw, flat)
        print('OK')
    """)
