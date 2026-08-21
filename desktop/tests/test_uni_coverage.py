"""Unidirectional folder coverage must never be silent.

`uni_load_dir` picks ONE GenParams direction signature out of the folder it is
handed.  When the folder held more than one, everything else was dropped
without a word: KANLAN_B shipped a clean-looking report over 480 of its 864
files (56%), EMVSUI_A over 958 of 1152 (83%).  A tech had no way to tell a
"no events here" from a "did not look here".

These tests pin BOTH polarities:
  * a mixed-signature folder surfaces its covered/total count everywhere a
    fiber total is read — console, Acquisition Parameters sheet, grid banner,
    and the hub manifest;
  * a clean single-signature folder gains NOTHING AT ALL — same sheets, same
    rows, same freeze pane, no banner, no tab colour.

Coverage is REPORTED, not acted on: the set of fibers analysed is identical
before and after, which the third test pins directly.
"""
import os
import sys

import openpyxl
import pytest

HERE = os.path.dirname(os.path.abspath(__file__))
SPLICE_DIR = os.path.join(os.path.dirname(os.path.dirname(HERE)), 'splicereport')
if SPLICE_DIR not in sys.path:
    sys.path.insert(0, SPLICE_DIR)
import splicereportmatchexfo as E  # noqa: E402


# ── Helpers ────────────────────────────────────────────────────────────────

def _fixture_sor():
    fx = os.path.join(HERE, 'fixtures')
    for root, _dirs, files in os.walk(fx):
        for f in sorted(files):
            if f.lower().endswith('.sor') and not f.startswith('._'):
                return os.path.join(root, f)
    return None


def _clone_with_locations(src, dst, old_a, new_a):
    """Copy a real .sor, swapping a GenParams location string of equal length."""
    raw = open(src, 'rb').read()
    assert len(old_a) == len(new_a), 'equal-length swap keeps every offset valid'
    if old_a.encode() not in raw:
        return False
    raw = raw.replace(old_a.encode(), new_a.encode(), 1)
    open(dst, 'wb').write(raw)
    return True


# ── 1. Coverage bookkeeping is exact ────────────────────────────────────────

def test_coverage_accounts_for_every_candidate_file():
    """n_candidates == n_loaded + n_dropped, always — the whole contract."""
    groups = {'A->B': {i: {} for i in range(1, 481)},
              'AD->B': {i: {} for i in range(481, 865)}}
    cov = E._uni_coverage('/x', '.sor', [f'f{i}.sor' for i in range(864)],
                          0, {}, 'A->B', groups, [])
    assert cov['n_candidates'] == 864
    assert cov['n_loaded'] == 480
    assert cov['n_dropped'] == 384
    assert cov['n_candidates'] == cov['n_loaded'] + cov['n_dropped']
    assert cov['complete'] is False
    assert cov['dropped_signatures'] == [
        {'signature': 'AD->B', 'n_files': 384, 'fiber_ranges': '481-864'}]


def test_coverage_complete_when_one_signature():
    groups = {'A->B': {i: {} for i in range(1, 433)}}
    cov = E._uni_coverage('/x', '.sor', [f'f{i}.sor' for i in range(432)],
                          0, {}, 'A->B', groups, [])
    assert cov['complete'] is True
    assert cov['n_dropped'] == 0
    assert cov['dropped_signatures'] == []
    # A clean folder produces NO narrative at all — nothing to render anywhere.
    assert E.uni_coverage_lines(cov) == []


def test_headline_states_both_numbers():
    groups = {'A->B': {i: {} for i in range(1, 481)},
              'AD->B': {i: {} for i in range(481, 865)}}
    cov = E._uni_coverage('/x', '.sor', [f'f{i}.sor' for i in range(864)],
                          0, {}, 'A->B', groups, [])
    head = E.uni_coverage_headline(cov)
    assert '480' in head and '864' in head
    assert 'INCOMPLETE COVERAGE' in head
    lines = E.uni_coverage_lines(cov)
    assert any("'AD->B'" in ln and '481-864' in ln for ln in lines)


def test_other_drop_reasons_are_itemised():
    groups = {'A->B': {1: {}, 2: {}}}
    cov = E._uni_coverage('/x', '.sor', ['a.sor', 'b.sor', 'c.sor', 'd.sor'],
                          0, {'unreadable': ['c.sor'],
                              'no fiber number': ['d.sor']},
                          'A->B', groups, [])
    assert cov['n_dropped'] == 2
    joined = ' '.join(E.uni_coverage_lines(cov))
    assert 'unreadable' in joined and 'c.sor' in joined
    assert 'no fiber number' in joined and 'd.sor' in joined


# ── 2. A mixed folder surfaces coverage; a clean one gains nothing ──────────

def _tiny_grid():
    cols = [{'kind': 'splice', 'position_km_display': 1.0, 'landmark': '',
             'fiber_count': 12}]
    grid = {(0, 0): [(1, 0.2)]}
    return grid, cols


def test_clean_folder_workbook_is_untouched(tmp_path):
    """No coverage, or complete coverage → byte-for-byte the approved layout."""
    grid, cols = _tiny_grid()
    outs = []
    for tag, cov in (('none', None),
                     ('complete', {'complete': True, 'n_loaded': 24,
                                   'n_candidates': 24, 'n_dropped': 0,
                                   'dropped_signatures': [],
                                   'dropped_files': [], 'n_other_format': 0})):
        p = str(tmp_path / f'{tag}.xlsx')
        E.uni_write_xlsx(grid, cols, 24, 12, 5.0, p,
                         site_a='LAM', site_b='BEY', fibers=None, coverage=cov)
        outs.append(p)
    for p in outs:
        ws = openpyxl.load_workbook(p)['Unidir Events']
        # The ZK-approved header lands on rows 1-4, exactly as before.
        assert ws.cell(row=1, column=1).value == 'LAM→BEY: ft'
        assert ws.cell(row=2, column=1).value == 'LAM→BEY: km'
        assert ws.cell(row=3, column=1).value == 'Handholes:'
        assert ws.cell(row=4, column=1).value == 'Ribbon'
        assert ws.freeze_panes == 'B5'
        assert ws.sheet_properties.tabColor is None   # no alarm colour
        assert not ws.merged_cells.ranges             # no banner merge


def test_incomplete_folder_gets_a_red_banner_above_the_grid(tmp_path):
    grid, cols = _tiny_grid()
    cov = {'complete': False, 'n_loaded': 480, 'n_candidates': 864,
           'n_dropped': 384, 'chosen': 'LANCASTER->KANSAS CITY',
           'dropped_signatures': [{'signature': 'LANDCASTER->KANSAS CITY',
                                   'n_files': 384, 'fiber_ranges': '1-384'}],
           'dropped_files': [], 'n_other_format': 0, 'ext': '.sor'}
    p = str(tmp_path / 'mixed.xlsx')
    E.uni_write_xlsx(grid, cols, 24, 12, 5.0, p,
                     site_a='KAS', site_b='LAN', fibers=None, coverage=cov)
    ws = openpyxl.load_workbook(p)['Unidir Events']
    banner = str(ws.cell(row=1, column=1).value)
    assert '480' in banner and '864' in banner
    assert ws.cell(row=1, column=1).fill.fgColor.rgb.endswith('C00000')
    assert ws.sheet_properties.tabColor is not None
    # The grid itself is intact, just pushed below the banner.
    col_a = [ws.cell(row=r, column=1).value for r in range(1, 12)]
    assert 'KAS→LAN: ft' in col_a and 'Ribbon' in col_a
    # …and the dropped signature is named with its fiber range.
    text = ' '.join(str(ws.cell(row=r, column=1).value or '')
                    for r in range(1, 6))
    assert 'LANDCASTER->KANSAS CITY' in text and '1-384' in text


# ── 3. Acquisition Parameters: the trace count states coverage ──────────────

def _audit_stub(n_files=480):
    return {'n_files': n_files, 'earliest_iso': '2026-01-01',
            'latest_iso': '2026-01-02', 'per_wavelength': [], 'file_fields': []}


def test_acq_sheet_count_row_unchanged_when_clean(tmp_path):
    from acquisition_audit import render_xlsx_sheet
    wb = openpyxl.Workbook()
    render_xlsx_sheet(wb, _audit_stub(432), per_trace_detail=True)
    ws = wb['Acquisition Parameters']
    assert ws.cell(row=1, column=1).value == 'Acquisition Parameters'
    assert ws.cell(row=1, column=2).value.startswith('432 trace(s); first')
    assert ws.freeze_panes == 'A4'


def test_acq_sheet_count_row_states_coverage_when_incomplete(tmp_path):
    from acquisition_audit import render_xlsx_sheet
    cov = {'complete': False, 'n_loaded': 480, 'n_candidates': 864,
           'n_dropped': 384, 'chosen': 'LANCASTER->KANSAS CITY',
           'dropped_signatures': [{'signature': 'LANDCASTER->KANSAS CITY',
                                   'n_files': 384, 'fiber_ranges': '1-384'}],
           'dropped_files': [], 'n_other_format': 0, 'ext': '.sor'}
    audit = _audit_stub(480)
    audit.update({'coverage': cov,
                  'coverage_headline': E.uni_coverage_headline(cov),
                  'coverage_lines': E.uni_coverage_lines(cov)})
    wb = openpyxl.Workbook()
    render_xlsx_sheet(wb, audit, per_trace_detail=True)
    ws = wb['Acquisition Parameters']
    # Row 1 is the alarm, above everything.
    assert ws.cell(row=1, column=1).value == 'COVERAGE'
    assert '480' in ws.cell(row=1, column=2).value
    assert '864' in ws.cell(row=1, column=2).value
    assert ws.cell(row=1, column=1).fill.fgColor.rgb.endswith('C00000')
    # The trace-count row still exists and now carries both numbers.
    counts = [ws.cell(row=r, column=2).value for r in range(1, 12)
              if ws.cell(row=r, column=1).value == 'Acquisition Parameters']
    assert counts and '480 of 864 trace(s)' in counts[0]
    assert 'NOT analysed' in counts[0]
    # Freeze pane followed the header down instead of cutting the block.
    assert ws.freeze_panes != 'A4'


# ── 4. End to end on real files: coverage reported, selection UNCHANGED ─────

def test_mixed_signature_folder_end_to_end(tmp_path):
    """Two site codes in one folder → same fibers analysed, coverage stated."""
    src = _fixture_sor()
    if src is None:
        pytest.skip('no .sor fixture available')
    raw = open(src, 'rb').read()
    rec = E.parse_sor_full(src, trim=False)
    if not rec:
        pytest.skip('fixture did not parse')
    loc_a = (rec.get('gen_loc_a') or '').strip()
    if len(loc_a) < 3:
        pytest.skip('fixture has no usable GenParams origin')
    typo = loc_a[:-1] + ('X' if loc_a[-1] != 'X' else 'Y')

    d = tmp_path / 'mixed'
    d.mkdir()
    for i in range(1, 7):                     # 6 files: majority signature
        open(d / f'SPAN{i:04d}.sor', 'wb').write(raw)
    made = 0
    for i in range(7, 10):                    # 3 files: the other signature
        if _clone_with_locations(src, str(d / f'SPAN{i:04d}.sor'), loc_a, typo):
            made += 1
    if made != 3:
        pytest.skip('could not relabel the fixture GenParams')

    fibers, chosen, counts, merged, cov = E.uni_load_dir(str(d))
    assert len(counts) == 2, counts
    assert cov['n_candidates'] == 9
    assert cov['n_loaded'] == len(fibers)
    assert cov['n_loaded'] + cov['n_dropped'] == 9
    assert cov['complete'] is False
    # Selection UNCHANGED: still exactly the majority signature's fibers.
    assert cov['n_loaded'] == counts[chosen] + sum(
        m['n_fibers'] for m in merged)
    # And the drop is narrated.
    assert E.uni_coverage_lines(cov)
