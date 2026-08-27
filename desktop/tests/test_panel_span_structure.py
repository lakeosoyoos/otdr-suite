"""A panel-to-panel span must publish its SHAPE, without reporting anything twice.

Zach's 2026-08-25 field issue.  #114 stopped the BLDG1<->BLDG3 crash; the
report then opened showing only its two ILA end columns, because a tie
between two panels has no midspan splices and nothing else can make a
column.

Those ILA columns are NOT empty and were never the problem — on Defuniak
they already flag 123 of 144 fibers for connector loss, by the calibrated
launch-connector rule (min(A,B) >= 0.62, max(A,B) >= 0.65, average).  This
pass does not touch them.

What is missing is the span's structure.  FastReporter prints it as events
interleaved with sections (verified in FR full mode, 12 Defuniak fibers):

    Event 2  Reflective  1.0049 km   Loss -0.333  Refl. -52.3
    Section              0.0311 km   Loss 0.000   Att. 0.000 dB/km
    Event 3  Reflective  1.0360 km   Loss 0.616   Refl. -53.3

The middle line is the one no ILA column can say: the 31 m of cable between
the panels loses NOTHING.  Everything worth reporting on this span is in its
two connectors and the glass between them is fine.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from conftest import (run_splicereport, FIXTURE_DIR,
                      FIXTURE_SPLICE_A_DIR, FIXTURE_SPLICE_B_DIR)

PANEL_DIR = FIXTURE_DIR / "panelspan"


def _run(tmp_path, name="panel.xlsx"):
    out = tmp_path / name
    rc, m, stderr = run_splicereport(PANEL_DIR, PANEL_DIR, out, "DNN1", "DNN2")
    assert rc == 0 and m and m.get("ok"), f"runner failed: {stderr[-1200:]}"
    return m, stderr, out


def test_panel_span_publishes_connector_section_connector(tmp_path):
    """THE regression: before this, a no-closure span produced n_columns 0."""
    m, _, _ = _run(tmp_path)
    assert m["n_splices"] == 0, "this span genuinely has no closures"
    kinds = [c["kind"] for c in m["columns"]]
    assert kinds, "panel span still publishes no structure"
    assert kinds == ["connector", "section", "connector"], kinds


def test_the_cable_between_the_panels_reads_zero(tmp_path):
    """FR reports 0.000 dB / 0.000 dB/km across the 31 m tie.  That reading —
    the glass is fine, it is all in the connectors — is the whole point of
    the section column."""
    m, _, _ = _run(tmp_path)
    sect = [c for c in m["columns"] if c["kind"] == "section"]
    assert len(sect) == 1, m["columns"]
    cells = [c for c in m["cells"] if c["splice"] == sect[0]["index"]]
    assert cells, "the section column carries no cells"
    for c in cells:
        assert abs(c["loss"]) < 0.0005, c
        assert not c["is_flagged"], "sections are descriptive, never findings"


def test_connectors_are_typed_never_numbered(tmp_path):
    """FR names events by TYPE and numbers them only ordinally.  The header
    must read 'Connector @ …', never 'Splice N'."""
    m, _, out = _run(tmp_path)
    assert m["columns"], "no columns — nothing to check numbering on"
    assert all(c["num"] is None for c in m["columns"]), m["columns"]
    ws = openpyxl.load_workbook(out)["Splice Report"]
    heads = [str(ws.cell(3, c).value) for c in range(1, ws.max_column + 1)
             if ws.cell(3, c).value]
    assert any(h.startswith("Connector @") for h in heads), heads
    assert any(h.startswith("Section ") for h in heads), heads
    assert not any(h.startswith("Splice ") for h in heads), heads


def test_no_fiber_is_reported_at_both_ila_and_a_connector_column(tmp_path):
    """The constraint.  A connector column keeps what analyze_all could
    measure — loss ILA's 0.62/0.65 gates sit above and the tech never sees —
    but never for a fiber the ILA column already names at that end."""
    m, _, out = _run(tmp_path)
    ws = openpyxl.load_workbook(out)["Splice Report"]
    ncol = ws.max_column
    conn = [c["index"] for c in m["columns"] if c["kind"] == "connector"]
    # Fail on the premise rather than IndexError-ing: with no connector
    # columns there is nothing to double-report and the test would be
    # asserting about an empty set.
    assert conn, "no connector columns — the constraint is untestable"
    for col_x, si in ((2, conn[0]), (ncol, conn[-1])):
        ila = set()
        for r in range(4, ws.max_row + 1):
            for mt in re.finditer(r'(?:^|\s)(\d+)[,\d]*\s',
                                  str(ws.cell(r, col_x).value or '')):
                ila.add(int(mt.group(1)))
        col = {c["fiber"] for c in m["cells"] if c["splice"] == si}
        assert not (ila & col), (
            f"fibers reported twice at one end: {sorted(ila & col)[:8]}")


def test_a_span_with_closures_is_untouched(tmp_path):
    """Scope guard: the pass runs ONLY when discovery found no closures."""
    out = tmp_path / "normal.xlsx"
    rc, m, stderr = run_splicereport(FIXTURE_SPLICE_A_DIR, FIXTURE_SPLICE_B_DIR, out)
    assert rc == 0 and m and m.get("ok"), stderr[-800:]
    assert m["n_splices"] >= 1, "fixture span should have closures"
    assert not any(c["kind"] in ("connector", "section") for c in m["columns"])
    assert "span structure" not in stderr


def test_nothing_past_the_cable_end_becomes_a_connector(tmp_path):
    """The receive reel tables its own events BEYOND the cable end, and they
    are not plant.

        LSC1<->LSC5   0.0000 1F   0.0316 1E is_end   1.0365 1F
        FTH01         0.0000 1F   0.0624 1E is_end   0.0775 1F   1.0956 1F

    Treating those trailing reflections as connectors put a column on the
    reel and then fitted a "section" ACROSS the reel — which is how FTH came
    to report a NEGATIVE attenuation, a physically impossible number in a
    field report.  Every column this pass publishes must sit at or before
    the cable end, and no section may report a negative loss.
    """
    m, _, _ = _run(tmp_path)
    # Every fibre's cable end on this fixture is ~0.031 km; the receive reel
    # runs a further kilometre past it.  A column out there is a column on
    # the reel.
    for c in m["columns"]:
        assert c["km"] <= 0.05, f"column past the cable end: {c}"
    for cell in m["cells"]:
        if cell["loss"] is not None:
            assert cell["loss"] >= -0.0005, f"negative section loss: {cell}"
