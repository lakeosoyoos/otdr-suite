"""AWS / IIG MT.1085 customer profile — thresholds, connector knobs, receipt.

Sources for every number asserted here: RFP-FOT-2025-001 (Intermountain
Infrastructure Group, issued 09 Jul 2026) and the Zero DB Statement of Work
(DocuSigned 06 Aug 2026), reconciled by Northcentral Telcom on 24 Aug 2026.

Three things have to hold, and each has a way of quietly breaking:

  1. The profile's contract values reach the ENGINE GLOBALS.  A profile is
     just data; if a key stops being mapped in _OTDR_KEY_TO_ENGINE_GLOBAL the
     row still renders and grades at the baseline instead.
  2. The profile reaches the CONNECTOR panel.  Customer profiles used to
     cover only OTDR_ROWS, so IIG's one-sided connector rule — the change
     with real teeth (154 flags to 7 on Frenchtown) — would have rendered as
     a chosen profile and changed nothing.
  3. Turning that gate off must SURVIVE the trip.  run_splicereport rejects
     non-finite and (for some keys) non-positive overrides, so a 0.0 that
     gets skipped there would silently restore 0.65 with no error anywhere.

Plus: adding IIG must not move Default / Lumen / Zayo.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from conftest import (
    run_splicereport, FIXTURE_SPLICE_A_DIR, FIXTURE_SPLICE_B_DIR, REPO_ROOT,
)

import app as hub

IIG = "AWS / IIG MT.1085"


# ── 1. The contract thresholds reach the engine globals ──────────────────
def test_iig_contract_thresholds_reach_engine_globals():
    """The three rows the contract actually specifies, at their governing
    values, mapped onto the globals the engine reads at run time."""
    assert IIG in hub.CUSTOMER_PROFILES, "the IIG profile must be selectable"
    ov = hub._overrides_from_settings(hub._otdr_settings_from_profile(IIG))

    # Bidir splice loss <= 0.20 dB (RFP and SOW agree).
    assert ov["REBURN_THRESHOLD"] == 0.200
    # Bidir connector loss <= 0.50 dB — the executed SOW governs, not the
    # RFP's 0.30.  On Span 29 that is 3 failures against 152.
    assert ov["BIDIR_CONNECTOR_LOSS"] == 0.500
    # Connector reflectance <= -55 dB, SIGNED (a less negative reading fails).
    assert ov["LAUNCH_BAD_REFL_DB"] == -55.0


def test_iig_bidir_splice_is_looser_than_the_engine_baseline():
    """0.20 dB is LOOSER than the 0.160 we ship, so IIG flags FEWER splice
    cells than Default.  That is the contract, but it is surprising enough
    that it should fail loudly if someone 'corrects' it downward."""
    src = (Path(hub.SPLICEREPORT_DIR) / "splicereportmatchexfo.py").read_text(
        encoding="utf-8")
    base = float(re.search(r"^REBURN_THRESHOLD\s*=\s*([\d.]+)", src, re.M).group(1))
    iig = hub._otdr_settings_from_profile(IIG)["bidir_splice_loss"]["fail"]
    assert iig > base, f"IIG {iig} must be looser than the baseline {base}"


def test_iig_keeps_unidir_splice_loss_on_at_the_default():
    """The contract sets no single-direction threshold.  Leaving the row ON at
    the engine default grades those cells as today; unticking it would HIDE
    events (the disable sentinel), which is not what 'unspecified' means."""
    row = hub._otdr_settings_from_profile(IIG)["unidir_splice_loss"]
    assert row["apply"] is True
    assert row["fail"] == 0.250
    ov = hub._overrides_from_settings(hub._otdr_settings_from_profile(IIG))
    assert ov["SINGLE_DIR_THRESHOLD"] == 0.250
    assert ov["SINGLE_DIR_THRESHOLD"] != hub._OTDR_DISABLE_SENTINEL


def test_iig_does_not_invent_gates_the_engine_cannot_grade():
    """Average splice loss, fiber attenuation, link ORL, OLTS, PMD and CD are
    all in the contract and none of them is a per-cell OTDR gate.  None may
    appear in the override payload — a knob that reaches nothing is worse
    than no knob."""
    ov = hub._overrides_from_settings(hub._otdr_settings_from_profile(IIG))
    for key in ("fiber_section_atten", "span_loss", "span_orl", "splitter_loss",
                "avg_splice_loss", "pmd", "cd", "olts"):
        assert key not in ov
    # Every emitted key must be a real engine global, not a hopeful name.
    eng = (Path(hub.SPLICEREPORT_DIR) / "splicereportmatchexfo.py").read_text(
        encoding="utf-8")
    for g in ov:
        assert re.search(r"^%s\s*=" % re.escape(g), eng, re.M), \
            f"{g} is not a module-level engine global"


# ── 2. The profile reaches the CONNECTOR panel ───────────────────────────
def test_iig_turns_the_one_sided_connector_gate_off():
    """The change with teeth.  One-sided panel offsets on this job are a
    backscatter / mode-field mismatch against the 200 um span fiber (the
    near end reads a GAINER, which neither contamination nor a bad connector
    can produce), so a one-sided reading is not a failure here."""
    conn = hub._conn_settings_from_profile(IIG)
    assert conn["LAUNCH_CONN_UNI_MIN_DB"] == 0.0


def test_iig_keeps_the_bidirectional_connector_gate_running():
    """Turning the one-sided gate off must not blind the connector check
    entirely — a connector BOTH directions see as bad still has to flag.
    (Mark's counts are 26->2, 154->7, 18->2: survivors, not zero.)"""
    conn = hub._conn_settings_from_profile(IIG)
    assert conn["LAUNCH_CONN_LOSS_MIN_DB"] == 0.62
    assert conn["LAUNCH_CONN_LOSS_MIN_DB"] > 0


def test_engine_treats_zero_as_off_not_as_flag_everything():
    """A `>=` gate at 0 would flag EVERY connector.  The engine must guard it
    with an explicit `> 0`, and must not skip the whole connector block just
    because one of the three gates is zero."""
    eng = (Path(hub.SPLICEREPORT_DIR) / "splicereportmatchexfo.py").read_text(
        encoding="utf-8")
    assert re.search(
        r"_uni_fires\s*=\s*\(_both and LAUNCH_CONN_UNI_MIN_DB > 0", eng), \
        "the one-sided gate must short-circuit on > 0 before comparing"
    assert "(LAUNCH_CONN_LOSS_MIN_DB or 0) > 0" in eng and \
           "(LAUNCH_CONN_UNI_MIN_DB or 0) > 0" in eng, \
        "zeroing one connector gate must not take the others down with it"


def test_zero_survives_the_runner_override_guard():
    """run_splicereport skips non-finite overrides, and forces a few keys to
    stay positive.  LAUNCH_CONN_UNI_MIN_DB must NOT be one of those, or the
    0.0 is skipped and the gate silently returns to 0.65."""
    src = (Path(hub.SPLICEREPORT_DIR) / "run_splicereport.py").read_text(
        encoding="utf-8")
    block = src.split("_positive_float_globals", 1)[1].split("}", 1)[0]
    assert "LAUNCH_CONN_UNI_MIN_DB" not in block, \
        "a positive-only guard would silently drop the OFF value"


def test_conn_overrides_ride_the_same_channel_to_the_engine():
    """The connector knobs reach the engine through --overrides alongside the
    threshold table, and only globals the panel renders are forwarded."""
    src = (REPO_ROOT / "app.py").read_text(encoding="utf-8")
    assert "if g in _CONN_DEFAULTS" in src, \
        "connector overrides must be filtered to rendered knobs"
    conn = hub._conn_settings_from_profile(IIG)
    assert set(conn) == set(hub._CONN_DEFAULTS), \
        "a profile may retune knobs, never invent or drop them"


def test_profile_switch_reloads_the_connector_knobs():
    """Picking a customer must reload BOTH panels, and re-mount the connector
    component (its key encodes the profile) — otherwise the iframe keeps
    showing, and re-committing, the previous customer's knobs."""
    src = (REPO_ROOT / "app.py").read_text(encoding="utf-8")
    switch = src.split("if _picked != _cur:", 1)[1][:900]
    assert "_conn_settings_from_profile(_picked)" in switch
    assert 'key=f"conn_settings_component::' in src


def test_custom_profile_keeps_the_techs_own_connector_edits():
    """'Custom' is the sentinel that preserves manual edits — it must not
    reset the connector knobs any more than it resets the threshold table."""
    src = (REPO_ROOT / "app.py").read_text(encoding="utf-8")
    switch = src.split("if _picked != _cur:", 1)[1][:900]
    conn_line = switch.index("_conn_settings_from_profile(_picked)")
    guard_line = switch.index("if 'Custom' not in _picked:")
    assert guard_line < conn_line, \
        "the connector reload must sit INSIDE the non-Custom guard"


# ── 3. Adding IIG must not move the profiles that shipped before it ──────
def test_existing_profiles_keep_engine_default_connector_knobs():
    """Only a profile that declares a "conn" block may differ from the
    engine defaults.  Default / Lumen / Zayo declare none, so their connector
    behavior is byte-identical to what they shipped with."""
    for prof in ("Default (engine baseline)", "Lumen", "Zayo"):
        assert hub._conn_settings_from_profile(prof) == hub._CONN_DEFAULTS, \
            f"{prof} connector knobs moved"


def test_unknown_conn_global_in_a_profile_is_ignored():
    """A typo in a profile's conn block must not push an unwired constant at
    the engine, nor invent a knob the panel never renders."""
    hub.CUSTOMER_PROFILES["__test__"] = {
        "apply": set(), "thresholds": {},
        "conn": {"NOT_A_REAL_GLOBAL": 1.0, "LAUNCH_CONN_UNI_MIN_DB": 0.0},
    }
    try:
        out = hub._conn_settings_from_profile("__test__")
        assert "NOT_A_REAL_GLOBAL" not in out
        assert out["LAUNCH_CONN_UNI_MIN_DB"] == 0.0
    finally:
        hub.CUSTOMER_PROFILES.pop("__test__", None)


# ── 4. The report states the thresholds it applied ───────────────────────
def test_report_prints_the_thresholds_it_applied(tmp_path):
    """The IIG contract review's explicit ask: a report that does not state
    its thresholds cannot be checked against the RFP or the SOW, which
    disagree by 152 failures against 3.  Run with the IIG overrides and read
    the numbers back out of the Legend sheet."""
    import openpyxl

    settings = hub._otdr_settings_from_profile(IIG)
    ov = hub._overrides_from_settings(settings)
    ov.update(hub._conn_settings_from_profile(IIG))

    out = tmp_path / "iig.xlsx"
    rc, m, err = run_splicereport(FIXTURE_SPLICE_A_DIR, FIXTURE_SPLICE_B_DIR,
                                  out, overrides=ov)
    assert rc == 0 and m and m.get("ok"), f"IIG run failed: {err[-1500:]}"

    wb = openpyxl.load_workbook(out)
    assert "Legend" in wb.sheetnames
    text = "\n".join(
        " ".join(str(c.value) for c in row if c.value is not None)
        for row in wb["Legend"].iter_rows())

    assert "THRESHOLDS APPLIED" in text
    # The contract values this run actually graded on.
    assert "0.2 dB" in text, "bidir splice loss 0.20 not stated"
    assert "0.5 dB" in text, "bidir connector loss 0.50 not stated"
    assert "-55 dB" in text, "connector reflectance -55 not stated"
    # The one-sided gate must be reported as off IN WORDS, not as a bare 0.
    assert re.search(r"Connector loss . 1 direction.*OFF", text), \
        "the disabled one-sided gate must say it was not graded"


def test_disabled_row_never_prints_the_raw_sentinel(tmp_path):
    """Unticking a row sends 1e9, which must render as OFF rather than
    '1000000000 dB' in front of a customer."""
    import openpyxl

    s = hub._otdr_settings_from_profile(IIG)
    s["midspan_reflectance"]["apply"] = False
    ov = hub._overrides_from_settings(s)
    assert ov["MIDSPAN_REFL_WARN_DB"] == hub._OTDR_DISABLE_SENTINEL

    out = tmp_path / "off.xlsx"
    rc, m, err = run_splicereport(FIXTURE_SPLICE_A_DIR, FIXTURE_SPLICE_B_DIR,
                                  out, overrides=ov)
    assert rc == 0 and m and m.get("ok"), f"run failed: {err[-1200:]}"
    wb = openpyxl.load_workbook(out)
    text = "\n".join(
        " ".join(str(c.value) for c in row if c.value is not None)
        for row in wb["Legend"].iter_rows())
    assert "1e+09" not in text and "1000000000" not in text
    assert "OFF" in text


# ── 5. The real gesture: pick the customer from the dropdown ─────────────
def test_picking_iig_in_the_dropdown_moves_both_panels():
    """End to end through the actual widget, because everything above tests
    helpers.  Picking the customer must move the threshold table AND the
    connector knob — and switching to another customer must take IIG's
    connector setting back off, or the next span silently inherits it."""
    from conftest import run_streamlit

    def _get(at, key, default=None):
        # AppTest's session_state proxy has no .get().
        try:
            return at.session_state[key]
        except Exception:
            return default

    at = run_streamlit().run()
    at.session_state["view_dir_a_input"] = str(FIXTURE_SPLICE_A_DIR)
    at.session_state["view_dir_b_input"] = str(FIXTURE_SPLICE_B_DIR)
    at.sidebar.radio[0].set_value("Splice Report").run()
    assert not at.exception, list(at.exception)

    assert IIG in at.selectbox[0].options, "IIG must be pickable"
    assert (_get(at, "conn_settings") or {})["LAUNCH_CONN_UNI_MIN_DB"] == 0.65

    at.selectbox[0].set_value(IIG).run()
    assert not at.exception, list(at.exception)
    assert _get(at, "otdr_profile") == IIG
    s = _get(at, "otdr_settings") or {}
    assert s["bidir_splice_loss"]["fail"] == 0.200
    assert s["bidir_connector_loss"]["fail"] == 0.500
    assert s["reflectance"]["fail"] == -55.0
    assert (_get(at, "conn_settings") or {})["LAUNCH_CONN_UNI_MIN_DB"] == 0.0

    # Leaving IIG must not leave its connector rule behind.
    at.selectbox[0].set_value("Lumen").run()
    assert not at.exception, list(at.exception)
    assert (_get(at, "conn_settings") or {})["LAUNCH_CONN_UNI_MIN_DB"] == 0.65, \
        "IIG's one-sided-gate-off leaked into the next customer"
