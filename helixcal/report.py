"""report — render a CalibrationResult to an .xlsx in the suite's house style.

Mirrors the suite's existing Excel idioms:
  * QA / summary sheet  — the Parameter | Result table from
    ``splicereport/acquisition_audit.py::render_xlsx_sheet`` (Calibri, light
    grey header, green = pass / amber = warning, thin BFBFBF box borders,
    freeze header).
  * data grids          — the Calibri-12 / header-fill-1F4E79 grid style from
    ``splicereport/splicereportmatchexfo.py::write_xlsx`` for the per-anchor
    residual and per-fiber spread tables.

Style primitives are COPIED here (not imported) so we match house style
without depending on / re-running those purpose-built report functions.
Distances are shown in BOTH meters and feet, per house convention.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .anchors import m_to_ft


FONT_NAME = "Calibri"
FSIZE = 12

# QA sheet palette (acquisition_audit idiom)
FILL_HEADER = PatternFill("solid", fgColor="D9D9D9")
FILL_GREEN = PatternFill("solid", fgColor="E2EFDA")   # pass
FILL_AMBER = PatternFill("solid", fgColor="FFF2CC")   # warning
FILL_RED = PatternFill("solid", fgColor="FFC7CE")     # hard fail / outlier
FILL_GREY = PatternFill("solid", fgColor="F2F2F2")    # section banner

# Grid palette (write_xlsx idiom)
GRID_HDR_FILL = PatternFill("solid", fgColor="1F4E79")

_thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

_fnt = Font(name=FONT_NAME, size=FSIZE)
_fnt_bold = Font(name=FONT_NAME, size=FSIZE, bold=True)
_fnt_header = Font(name=FONT_NAME, size=FSIZE + 1, bold=True)
_fnt_small = Font(name=FONT_NAME, size=FSIZE - 1, italic=True, color="555555")
_grid_hdr_font = Font(name=FONT_NAME, size=FSIZE, bold=True, color="FFFFFF")


def _fmt(v, nd=4):
    if v is None:
        return "—"
    return f"{v:.{nd}f}"


def _summary_sheet(wb, result):
    ws = wb.create_sheet("Helix Calibration", 0)
    wb.active = wb.sheetnames.index(ws.title)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 86

    row = 1
    ws.cell(row=row, column=1, value="Helix / EFL Calibration").font = _fnt_header
    wl = result.wavelength
    ws.cell(row=row, column=2, value=(
        f"{result.n_traces} trace(s), {result.n_anchors} anchor pair(s)"
        + (f"; {wl:g} nm" if wl else ""))).font = _fnt_small
    row += 1
    ws.cell(row=row, column=1, value="").fill = FILL_GREY
    ws.cell(row=row, column=2, value="").fill = FILL_GREY
    row += 1

    ws.cell(row=row, column=1, value="Parameter").font = _fnt_bold
    ws.cell(row=row, column=2, value="Result").font = _fnt_bold
    for c in (1, 2):
        ws.cell(row=row, column=c).fill = FILL_HEADER
        ws.cell(row=row, column=c).border = BOX
    row += 1

    def emit(name, value, fill=None):
        nonlocal row
        a = ws.cell(row=row, column=1, value=name)
        b = ws.cell(row=row, column=2, value=value)
        a.font = _fnt
        b.font = _fnt
        a.border = BOX
        b.border = BOX
        a.alignment = Alignment(vertical="top")
        b.alignment = Alignment(vertical="top", wrap_text=True)
        if fill is not None:
            a.fill = fill
            b.fill = fill
        row += 1

    # Band verdict drives the pass/warn color of the headline rows.
    band_pass = result.band_verdict.startswith("PASS")
    headline_fill = FILL_GREEN if band_pass else FILL_AMBER

    emit("Conversion factor m (cable-sheath / OTDR-fiber)",
         _fmt(result.m, 5), headline_fill)
    emit("Excess Fiber Length  EFL% = (1/m − 1)·100",
         (_fmt(result.efl_pct, 4) + " %") if result.efl_pct is not None else "—",
         headline_fill)
    b_ft = m_to_ft(result.b_m) if result.b_m is not None else None
    emit("Fixed offset b (launch / patch-cord / dead-zone)",
         (f"{_fmt(result.b_m, 3)} m  ({_fmt(b_ft, 2)} ft)"
          if result.b_m is not None else "—"),
         headline_fill)
    emit("R² (goodness of fit)", _fmt(result.r2, 6),
         FILL_GREEN if (result.r2 is not None and result.r2 >= 0.99)
         else FILL_AMBER)

    # AEN142 band
    emit("AEN142 band verdict", result.band_verdict,
         FILL_GREEN if band_pass else FILL_AMBER)
    # How the cable type was obtained (manual setting / GenParams auto-detect /
    # default fallback) — and the expected band it implies.
    src_label = {
        "manual": "manual setting",
        "genparams": "auto-detected from SOR GenParams",
        "default": "default fallback (GenParams empty on this span)",
    }.get(result.cable_type_source, result.cable_type_source)
    emit("Cable type", f"{result.cable_type}  [{src_label}]", FILL_GREY)
    if result.cable_type_note:
        emit("  how it was picked", result.cable_type_note, FILL_GREY)
    ce = result.cable_entry
    if ce is not None:
        emit("  construction / source", f"{ce.construction} — {ce.source}",
             FILL_GREY)
        emit("  expected helix band",
             f"m {ce.m_low:.3f}–{ce.m_high:.3f}  "
             f"(EFL {ce.efl_low:.2f}–{ce.efl_high:.2f} %)", FILL_GREY)

    # IOR guardrail
    emit("Stored IOR cohort median", _fmt(result.cohort_ior, 5), FILL_GREY)
    if result.expected_ior is not None:
        emit("Expected fiber-spec IOR", _fmt(result.expected_ior, 5), FILL_GREY)
    emit("IOR guardrail label", result.ior_label,
         FILL_GREEN if result.ior_verified else FILL_AMBER)
    if result.ior_flags:
        for f in result.ior_flags[:60]:
            c = ws.cell(row=row, column=2, value=f"      {f}")
            ws.cell(row=row, column=1, value="").border = BOX
            c.font = _fnt_small
            c.fill = FILL_AMBER
            c.border = BOX
            c.alignment = Alignment(vertical="top", wrap_text=True)
            row += 1

    # Cross-fiber consistency
    emit("Per-fiber m mean", _fmt(result.fiber_m_mean, 5), FILL_GREY)
    emit("Per-fiber m spread (std dev)",
         _fmt(result.fiber_m_std, 5),
         FILL_GREEN if (result.fiber_m_std is not None
                        and result.fiber_m_std <= 0.003) else FILL_AMBER)
    emit("Per-fiber m range (max − min)", _fmt(result.fiber_m_range, 5),
         FILL_GREY)
    if result.outlier_fibers:
        emit("Outlier fibers (disagree with cohort)",
             ", ".join(result.outlier_fibers), FILL_RED)
    else:
        emit("Outlier fibers (disagree with cohort)", "none", FILL_GREEN)

    # Warnings banner
    if result.warnings:
        row += 1
        a = ws.cell(row=row, column=1, value="Warnings")
        a.font = _fnt_bold
        a.fill = FILL_AMBER
        ws.cell(row=row, column=2, value="").fill = FILL_AMBER
        row += 1
        for w in result.warnings:
            emit("•", w, FILL_AMBER)

    ws.freeze_panes = "A4"


def _grid_header(ws, headers, row=1):
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = _grid_hdr_font
        c.fill = GRID_HDR_FILL
        c.border = BOX
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)


def _autofit(ws, headers, n_cols):
    from openpyxl.utils import get_column_letter
    for ci in range(1, n_cols + 1):
        width = max(
            [len(str(headers[ci - 1]))]
            + [len(str(ws.cell(row=r, column=ci).value or ""))
               for r in range(2, ws.max_row + 1)]
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(
            max(width * 1.15 + 2.0, 8), 60)


def _anchor_sheet(wb, result):
    ws = wb.create_sheet("Per-Anchor Residuals")
    headers = ["Fiber", "Closure", "Type", "Direction",
               "OTDR fiber dist (m)", "OTDR fiber dist (ft)",
               "Known sheath (m)", "Known sheath (ft)",
               "Residual (m)", "Residual (ft)"]
    _grid_header(ws, headers)
    r = 2
    for a in result.anchor_fits:
        x_ft = m_to_ft(a.x_otdr_m)
        y_ft = m_to_ft(a.y_known_m)
        res_ft = m_to_ft(a.residual_m) if a.residual_m is not None else None
        vals = [a.fiber_key, a.closure_name, a.anchor_type, a.direction_used,
                round(a.x_otdr_m, 3), round(x_ft, 2),
                round(a.y_known_m, 3), round(y_ft, 2),
                round(a.residual_m, 3) if a.residual_m is not None else None,
                round(res_ft, 2) if res_ft is not None else None]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = _fnt
            c.border = BOX
        # Flag large residuals amber (> 1 m).
        if a.residual_m is not None and abs(a.residual_m) > 1.0:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=r, column=ci).fill = FILL_AMBER
        r += 1
    _autofit(ws, headers, len(headers))
    ws.freeze_panes = "A2"


def _fiber_sheet(wb, result):
    ws = wb.create_sheet("Per-Fiber Spread")
    headers = ["Fiber", "# anchors", "m", "EFL %", "b (m)", "R²",
               "Stored IOR", "Outlier?", "Note"]
    _grid_header(ws, headers)
    r = 2
    for f in result.fiber_fits:
        vals = [f.fiber_key, f.n_anchors,
                round(f.m, 5) if f.m is not None else None,
                round(f.efl_pct, 4) if f.efl_pct is not None else None,
                round(f.b, 3) if f.b is not None else None,
                round(f.r2, 6) if f.r2 is not None else None,
                round(f.stored_ior, 5) if f.stored_ior is not None else None,
                "YES" if f.outlier else "",
                f.outlier_note or f.ior_note]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = _fnt
            c.border = BOX
        if f.outlier:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=r, column=ci).fill = FILL_RED
        r += 1
    _autofit(ws, headers, len(headers))
    ws.freeze_panes = "A2"


def build_workbook(result):
    """Build and return an openpyxl Workbook for ``result`` (does not save)."""
    wb = Workbook()
    # The default workbook ships with one empty 'Sheet'; our sheets are added
    # explicitly (summary at index 0), so drop the stock one.
    _summary_sheet(wb, result)
    _anchor_sheet(wb, result)
    _fiber_sheet(wb, result)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def write_report(result, output_path):
    """Render ``result`` to ``output_path`` (.xlsx). Returns the path."""
    wb = build_workbook(result)
    wb.save(output_path)
    return output_path
