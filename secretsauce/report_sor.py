"""
report_sor.py — SOR-file variant of the clean report.

Takes a folder of .sor files, runs the same classification logic (single
wavelength), and produces the clean HTML + PDF output with likelihood column.
"""
import os, re, sys, glob, base64, subprocess, argparse
from datetime import datetime
from itertools import combinations
from io import BytesIO
import numpy as np
from scipy.stats import norm

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from sor_reader324802a import parse_sor_full

from report import (  # reuse helpers — all neutral
    _BASE_CSS, _embed_logo, _find_chrome, _outlier_probability,
    html_to_pdf_bytes, _fmt_time_gap, _detrend, _shape_color,
    _COLOR_HIGH, _COLOR_MID, _COLOR_LOW,
    _event_match_quality, _events_agree,
)

_IOR = 1.4682
_LAUNCH_SKIP_M = 500
_END_BUFFER_M  = 200


def load_sor_file(path):
    r = parse_sor_full(path, trim=False)
    if r is None:
        raise ValueError(f'unparseable: {path}')
    trace = r['trace']
    sp = r.get('exfo_sampling_period')
    if not sp or sp <= 0:
        raise ValueError(f'bad sampling period: {path}')
    dz_m = 2.998e8 * sp / (2.0 * _IOR)
    pos = np.arange(len(trace)) * dz_m
    length_m = r.get('exfo_spans_length') or (pos[-1] if len(pos) else 0.0)
    events = r.get('events') or []
    # Pulse width expressed in SAMPLES.  The speckle high-pass has to cut at
    # this scale, and it varies 15x across the acquisitions on disk, so the
    # filter width cannot be a fixed sample count (see _SPECKLE_HP_WIDTH).
    _cal = r.get('exfo_calibration') or {}
    _cpw = _cal.get('CalibratedPulseWidth') or _cal.get('NominalPulseWidth')
    pulse_samples = (float(_cpw) / sp) if (_cpw and sp and sp > 0) else None
    # Max splice loss from event table (firmware-reported, interior events only)
    splice_vals = [e.get('splice_loss') for e in events
                   if e.get('splice_loss') is not None
                   and not e.get('is_end')
                   and (e.get('dist_km') or 0) > 0.01]
    max_splice = max((abs(v) for v in splice_vals), default=None) if splice_vals else None
    # Pull OTDR serial number from GenParams/SupParams so we can flag pairs
    # acquired by different OTDRs in the confirmed-duplicate detail table.
    from sor_reader324802a import parse_gen_params
    gp = parse_gen_params(path) or {}
    serial = (gp.get('serial_number') or '').strip() or None
    return {
        'name':     os.path.splitext(os.path.basename(path))[0],
        'filepath': path,
        'trace':    trace.astype(np.float32),
        'pos':      pos,
        'length':   float(length_m),
        'loss':     r.get('exfo_spans_loss'),
        'max_splice_dB': max_splice,
        'timestamp': r.get('date_time'),
        'wavelength': r.get('exfo_wavelength_nm') or r.get('wavelength'),
        'serial_number': serial,
        'events':   events,
        'pulse_samples': pulse_samples,
    }


def _pair_score(a, b, interior_start, interior_end):
    pa, pb = a['pos'], b['pos']
    ta, tb = a['trace'], b['trace']
    n = min(len(ta), len(tb))
    mask = (pa[:n] > interior_start) & (pa[:n] < interior_end)
    if mask.sum() < 50:
        return None
    return float(np.std(ta[:n][mask] - tb[:n][mask]))


def _compute_pair_metrics_batch(files, interior_start, interior_end, min_samples=50,
                                  tie_panel_mode=False):
    """Vectorized pair-metric computation. For N files this scales as O(N²·S)
    via two matmuls instead of O(N²) Python loops, so 864-file runs go from
    hours to seconds.

    Returns (sigma_matrix, r_matrix, valid_file_indices) where the matrices
    are indexed by position within `valid_file_indices` (NOT the original
    `files` list). σ is computed on raw traces; r on detrended traces.
    """
    interior = []
    valid_idx = []
    for i, f in enumerate(files):
        ta, pa = f['trace'], f['pos']
        n = len(ta)
        mask = (pa[:n] > interior_start) & (pa[:n] < interior_end)
        if mask.sum() < min_samples:
            continue
        interior.append((ta[mask].astype(np.float32),
                         pa[mask].astype(np.float32)))
        valid_idx.append(i)
    if len(interior) < 2:
        return None

    N = min(len(d[0]) for d in interior)
    K = len(interior)
    M_raw = np.empty((K, N), dtype=np.float32)
    M_det = np.empty((K, N), dtype=np.float32)
    for k, (ts, ps) in enumerate(interior):
        ts = ts[:N]; ps = ps[:N]
        M_raw[k] = ts
        # Detrend per-row: subtract best-fit linear (slope·pos + intercept).
        # Closed-form: slope = cov(p, t) / var(p), intercept = mean(t) - slope·mean(p).
        pm = ps.mean(); tm = ts.mean()
        denom = ((ps - pm) ** 2).sum()
        slope = float(((ps - pm) * (ts - tm)).sum() / denom) if denom > 0 else 0.0
        intercept = float(tm - slope * pm)
        M_det[k] = ts - (slope * ps + intercept)

    # σ(M[i] - M[j]) for all pairs via the variance-decomposition identity
    # on MEAN-CENTERED rows:
    #     var(A - B) = var(A') + var(B') - 2·E[A'·B']   with A' = A - E[A]
    # Centering first is load-bearing, not cosmetic: on raw ~46 dB trace
    # levels the uncentered identity (m2a + m2b - 2C - Δmean²) subtracts
    # ~2000-magnitude terms to extract a variance of ~1e-4, and the float32
    # trace quantization (~5.5e-6 dB/sample at 46 dB) alone puts ~2.6e-4 of
    # error into the cross term — catastrophic cancellation.  On a pristine
    # span (true pair σ ~0.01) that error DOMINATES: σ collapsed to 0.0000
    # for high-injection-offset pairs and the σ-outlier tier confirmed 67
    # numerical artifacts as duplicates (Lumen Border LAM/BEY, 2026-07-23).
    # Centered values span ~±1 dB, so the same identity is exact to ~1e-8.
    M64 = M_raw.astype(np.float64)
    M0 = M64 - M64.mean(axis=1, keepdims=True)
    v = (M0 ** 2).mean(axis=1)
    C0 = (M0 @ M0.T) / float(N)
    var_ij = v[:, None] + v[None, :] - 2.0 * C0
    sigma_matrix = np.sqrt(np.maximum(var_ij, 0.0))

    # Pearson r on detrended traces, after FINGERPRINT EXTRACTION:
    # subtract the per-position MEDIAN trace across files so the launch
    # reflection, attenuation slope, and shared connector signatures that
    # every fiber sees through the same launch box get cancelled. What
    # remains is each fiber's unique Rayleigh-scatterer fingerprint +
    # shot noise — the actual basis for "same fiber" calls.
    #
    # Why median (not mean): in datasets where duplicates make up a large
    # fraction of the files (e.g. TEST DUPE has 12 of 18 fibers in
    # duplicate pairs), the mean is biased toward the duplicate signal and
    # subtracting it weakens the same-fiber agreement. The median is
    # robust to that — it represents the typical "non-duplicate" trace
    # even when ~half the dataset is duplicates of the other half.
    #
    # Without this step, tie panels (short fibers with no splice events)
    # show inflated r because the shared launch+connector features
    # dominate the trace. With it, two truly-different short fibers
    # uncorrelate to near zero.
    M_det64 = M_det.astype(np.float64)
    if tie_panel_mode:
        # Subtract the median trace across all files: removes the shared
        # launch + connector signal so the per-fiber Rayleigh fingerprint
        # is what r actually measures. Median (not mean) is robust to the
        # presence of real duplicates in the dataset.
        group_ref = np.median(M_det64, axis=0, keepdims=True)
        M_fingerprint = M_det64 - group_ref
    else:
        # Production mode: skip fingerprint extraction. Real same-fiber
        # duplicates with naturally-low r (0.85-0.94) on long fibers
        # shouldn't be demoted by an aggressive shared-signal subtraction.
        M_fingerprint = M_det64
    # Re-center each row's residual fingerprint (should already be near zero).
    Mc = M_fingerprint - M_fingerprint.mean(axis=1, keepdims=True)
    std = np.sqrt((Mc ** 2).mean(axis=1))
    std_outer = np.outer(std, std)
    np.maximum(std_outer, 1e-12, out=std_outer)
    r_matrix = (Mc @ Mc.T) / (float(N) * std_outer)
    np.clip(r_matrix, -1.0, 1.0, out=r_matrix)
    return sigma_matrix, r_matrix, valid_idx


def _pair_shape_r(a, b, interior_start, interior_end):
    """Detrended Pearson r in the interior window. r ≈ 1 → same fiber."""
    pa = a['pos']
    ta, tb = a['trace'], b['trace']
    n = min(len(ta), len(tb))
    mask = (pa[:n] > interior_start) & (pa[:n] < interior_end)
    if mask.sum() < 50:
        return None
    pp = pa[:n][mask].astype(np.float64)
    da = _detrend(ta[:n][mask].astype(np.float64), pp)
    db = _detrend(tb[:n][mask].astype(np.float64), pp)
    sa, sb = np.std(da), np.std(db)
    if sa == 0 or sb == 0:
        return None
    return float(np.dot(da - da.mean(), db - db.mean()) / (sa * sb * len(da)))


def _distribution_chart(scores, p_dup, stats, shape_rs=None):
    """2x2 grid of panels (4-mode) or stacked 2 (2-mode):
        top-left:    level-of-disagreement distribution (histogram + cluster fit)
        top-right:   similarity score distribution (histogram + same-fiber tiers)
        bottom-left: per-pair likelihood vs level of disagreement
        bottom-right: per-pair likelihood vs similarity score
    When `shape_rs` is None, reverts to a 2-panel column (top-left + bottom-left)."""
    if shape_rs is not None:
        # 13x6 keeps the chart compact enough that section 1 banner + the 2x2
        # grid fit on the same landscape page as the title/cards header.
        fig, axes = plt.subplots(2, 2, figsize=(13, 6))
        ax1, axR  = axes[0, 0], axes[0, 1]
        ax2, axRS = axes[1, 0], axes[1, 1]
    else:
        fig, axes = plt.subplots(2, 1, figsize=(13, 5.5))
        ax1, ax2 = axes
        axR = axRS = None
    legend_kw = dict(loc='upper center', bbox_to_anchor=(0.5, -0.30),
                     ncol=2, fontsize=7.5, frameon=False)

    log_s = np.log10(np.maximum(scores, 1e-9))
    counts, bin_edges, _ = ax1.hist(log_s, bins=50, color='#4A90D9',
                                    alpha=0.75, edgecolor='white')
    bin_width = bin_edges[1] - bin_edges[0]
    # Scale the Gaussian PDF to raw-count units so it overlays the histogram.
    x = np.linspace(log_s.min() - 0.2, log_s.max() + 0.2, 400)
    ax1.plot(x, norm.pdf(x, stats['center_log'], stats['spread_log']) * len(log_s) * bin_width,
             color='#b97000', linewidth=2, label='cluster fit')
    ax1.axvline(stats['center_log'], linestyle='--', color='#b97000', alpha=0.7)
    for z_line in (-3, -5, -10):
        ax1.axvline(stats['center_log'] + z_line * stats['spread_log'],
                    linestyle=':', color='#888', alpha=0.5)
    ax1.set_xticklabels([])
    ax1.set_xlabel('level of disagreement (log scale)')
    ax1.set_ylabel('Number of pairs')
    ax1.set_title('Pair level-of-disagreement distribution with cluster fit', fontweight='bold')
    ax1.legend(**legend_kw)
    ax1.grid(alpha=0.3)

    if axR is not None:
        rs = np.asarray([r if r is not None else np.nan for r in shape_rs],
                        dtype=np.float64)
        rs_valid = rs[~np.isnan(rs)]
        # Always show out to similarity = 1.0 with the 0.95/0.99 thresholds
        # visible, so the reference lines anchor the reader's eye.
        lo = min(0.4, float(rs_valid.min()) - 0.02) if rs_valid.size else 0.4
        hi = 1.005
        if rs_valid.size:
            bins = np.linspace(lo, hi, 60)
            axR.hist(rs_valid, bins=bins, color='#4A90D9', alpha=0.75,
                     edgecolor='white')
            # Tier markers: green ≥ 0.99, orange 0.95–0.99, grey < 0.95.
            axR.axvspan(0.99, hi, color=_COLOR_HIGH, alpha=0.10)
            axR.axvspan(0.95, 0.99, color=_COLOR_MID, alpha=0.10)
            axR.axvline(0.99, linestyle='--', color=_COLOR_HIGH, linewidth=1.3,
                        label='≥ 0.99 (same fiber)')
            axR.axvline(0.95, linestyle=':', color=_COLOR_MID, linewidth=1.2,
                        label='= 0.95 (borderline floor)')
        axR.set_xlim(lo, hi)
        axR.set_xlabel('similarity score per pair')
        axR.set_ylabel('Number of pairs')
        ttl = ('Similarity score distribution — duplicates concentrate near 1.0'
               if rs_valid.size else 'Similarity score unavailable')
        axR.set_title(ttl, fontweight='bold')
        axR.legend(**legend_kw)
        axR.grid(axis='y', alpha=0.3)

    # Tier masks: high ≥ 0.9, mid 0.5–0.9, low ≤ 0.5. Colors match the tables.
    p = np.asarray(p_dup)
    m_hi = p > 0.9
    m_md = (p > 0.5) & (~m_hi)
    m_lo = ~(m_hi | m_md)
    if m_lo.any():
        ax2.scatter(log_s[m_lo], p[m_lo], s=45, alpha=0.6, color=_COLOR_LOW,
                    edgecolor='white', linewidth=0.5,
                    label=f'Non-duplicate (n={int(m_lo.sum())})')
    if m_md.any():
        ax2.scatter(log_s[m_md], p[m_md], s=120, alpha=0.95,
                    color=_COLOR_MID, edgecolor='black', linewidth=1, zorder=4,
                    label=f'Borderline 50–90% (n={int(m_md.sum())})')
    if m_hi.any():
        ax2.scatter(log_s[m_hi], p[m_hi], s=140, alpha=0.95,
                    color=_COLOR_HIGH, edgecolor='black', linewidth=1, zorder=5,
                    label=f'Duplicate ≥90% (n={int(m_hi.sum())})')
    ax2.axhline(0.9, color=_COLOR_HIGH, linestyle=':', alpha=0.4, linewidth=1)
    ax2.axhline(0.5, color=_COLOR_MID, linestyle='--', alpha=0.5, linewidth=1)
    ax2.set_xticklabels([])
    ax2.set_xlabel('level of disagreement (log scale)')
    ax2.set_ylabel('duplicate likelihood')
    ax2.set_title('Per-pair likelihood vs level of disagreement', fontweight='bold')
    ax2.legend(**legend_kw)
    ax2.grid(alpha=0.3)

    if axRS is not None:
        # Per-pair likelihood vs similarity score (Pearson r). Same tier-color
        # masks as the disagreement scatter, so high/mid/low pairs render
        # consistently between panels.
        rs_full = np.asarray([r if r is not None else np.nan for r in shape_rs],
                             dtype=np.float64)
        valid = ~np.isnan(rs_full)
        m_hi_v = m_hi & valid
        m_md_v = m_md & valid
        m_lo_v = m_lo & valid
        if m_lo_v.any():
            axRS.scatter(rs_full[m_lo_v], p[m_lo_v], s=45, alpha=0.6,
                         color=_COLOR_LOW, edgecolor='white', linewidth=0.5,
                         label=f'Non-duplicate (n={int(m_lo_v.sum())})')
        if m_md_v.any():
            axRS.scatter(rs_full[m_md_v], p[m_md_v], s=120, alpha=0.95,
                         color=_COLOR_MID, edgecolor='black', linewidth=1, zorder=4,
                         label=f'Borderline 50–90% (n={int(m_md_v.sum())})')
        if m_hi_v.any():
            axRS.scatter(rs_full[m_hi_v], p[m_hi_v], s=140, alpha=0.95,
                         color=_COLOR_HIGH, edgecolor='black', linewidth=1, zorder=5,
                         label=f'Duplicate ≥90% (n={int(m_hi_v.sum())})')
        axRS.axhline(0.9, color=_COLOR_HIGH, linestyle=':', alpha=0.4, linewidth=1)
        axRS.axhline(0.5, color=_COLOR_MID, linestyle='--', alpha=0.5, linewidth=1)
        axRS.axvline(0.99, color=_COLOR_HIGH, linestyle=':', alpha=0.4, linewidth=1)
        axRS.axvline(0.95, color=_COLOR_MID, linestyle='--', alpha=0.5, linewidth=1)
        # Lock x-axis so the 0.95 / 0.99 reference lines always show.
        rs_valid_pts = rs_full[valid]
        rs_lo = min(0.4, float(rs_valid_pts.min()) - 0.02) if rs_valid_pts.size else 0.4
        axRS.set_xlim(rs_lo, 1.005)
        axRS.set_xlabel('similarity score per pair')
        axRS.set_ylabel('duplicate likelihood')
        axRS.set_title('Per-pair likelihood vs similarity score', fontweight='bold')
        axRS.legend(**legend_kw)
        axRS.grid(alpha=0.3)

    plt.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close()
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('ascii')


# ── Raw-identity short-circuit ────────────────────────────────────────────
# A pair whose RAW interior metrics are essentially identical is a copy of
# the same acquisition data — no regime routing may hide it.  Applied to the
# PRE-fingerprint metrics (σ on raw traces, r on detrended-only traces), so
# tie_panel median-subtraction can't cancel the shared signal first: a
# byte-identical file pair in a tie-panel folder previously came back
# "no duplicates" because both residuals against the group median were
# equal (or exactly zero when the copies ARE the median) and σ-outlier is
# bypassed in that regime.
#
# Thresholds (calibrated 2026-07-14 on real spans):
#   byte-identical pair (measured through this float32 pipeline)
#                              →  raw σ = 3.8e-6 dB, raw r = 1.000000
#   closest NON-copy pair seen →  SANDUR 107↔146: σ 0.00495 dB, r 0.9982
#                                 (864-file span; NOT md5-identical, 5 m
#                                  length delta — near-identical re-shoots
#                                  that current main does NOT flag >0.5)
#   A-F West 145-288 tie panel →  min raw σ 0.0290 dB, max raw r 0.810
#   SEANOR / ELMMIL            →  min raw σ 0.0162 / 0.0079 dB
# The σ floor 0.001 dB sits ~260× above a true copy and ~5× below the
# closest real-world non-copy, so only a literal copy / re-export of the
# same acquisition data can trip it.  raw r ≥ 0.98 is a co-gate (copies
# read exactly 1.0).  The short-circuit only ever RAISES p_dup (to 1.0)
# — it can't demote, so no existing detection is weakened.
_RAW_IDENT_R = 0.98
_RAW_IDENT_SIGMA_DB = 0.001

# ── Distance-decay (shared-glass) tie-panel routing ───────────────────────
# Tie panels whose files share physical glass (jumper feed + ribbon) show
# raw-r that DECAYS with port distance: neighbors correlate strongly, far
# ports don't.  Real duplicate files don't care about port distance, so
# decay is a folder-level signature of shared path — route those folders to
# tie_panel (fingerprint extraction) even when bulk_r / frac_high_r stay
# low because the correlation tops out among neighbors only.
# Calibration (2026-07-14):
#   A-F West 145-288 (the FP flood): near r 0.580 vs far r 0.096 → decay 0.48
#   SEANOR 432-file production span: near r 0.574 vs far r 0.472 → decay 0.10
#   SANDUR (if it were consulted):   near r 0.928 vs far r 0.895 → decay 0.03
# 0.30 splits those cleanly.
#
# ── AUDITED 2026-08-29: RETIREMENT CONSIDERED AND REJECTED ────────────────
# After the same-instrument fix (`serials`), the drop measured on every
# folder on disk is:
#   A-F West 0.0344 | ELMMIL 0.0353 | SANDUR 0.0655 | EMVSUI Long 0.0745
#   SEANOR 0.1024 | NIL->MEC 0.1466 | SUIEMV Long 0.2065 | MEC->NIL 0.2286
# ZERO of the 13 folders is routed by this rule, and deleting it changes no
# folder's regime.  Its own calibration case is also gone: the A-F West
# figure above was taken on a 305 m collapsed window, and the later
# window-restoration fix took that folder to 2037 m where it measures 0.0344
# and routes tie_panel on bulk_r 0.9485 / frac_high_r 0.4829 instead.  (It
# was never solely dependent on decay anyway — its raw min_L of 1005 m also
# trips _SHORT_COMMON_SPAN_M, and both constants landed in commit 777c5fb.)
#
# RETIRING IT ANYWAY WAS REJECTED, on measurement.  Across the corpus the
# rule's effect is: zero duplicates suppressed, some false positives
# suppressed.  The one folder that still crosses the trigger is SUIEMV Long
# restricted to serial 989584 (547 files, decay 0.3255); deleting the route
# there takes the report from 0 pairs to 10, and all 10 are false positives
# on a folder with no duplicates — 8 of which already print in the
# as-delivered run.  EMVSUI Long's worst single-instrument half measures
# 0.0826, a 3.6x margin against the trigger, so the folder carrying the four
# known duplicates is not close to being re-broken.
#
# KNOWN, UNFIXED: what the near/far drop measures is confounded with
# ELAPSED ACQUISITION TIME, not just port distance — near pairs are shot
# minutes apart and far pairs days apart.  Time-matched on SUIEMV/989584
# (both buckets dt < 1 h) the drop is -0.008.  Ordinary long cables do carry
# a real port gradient too (0.13-0.25 after time-matching), so the premise
# "decay implies shared launch glass" is not safe in either direction.  The
# rule is left in place because it currently costs nothing and removing it
# has no measured benefit; the near/far medians are now printed on EVERY run
# so the first folder it does misroute is visible in the log.  The rule is only consulted for folders the
# existing rules would have called 'production' (additive routing), so
# all_dups / short_panel / tie_panel folders can never be re-routed by it.
_DECAY_NEAR_GAP = 3        # |port Δ| ≤ 3 → "neighbor" pair
_DECAY_FAR_GAP = 30        # |port Δ| ≥ 30 → "far" pair
_DECAY_MIN_PAIRS = 10      # need this many near AND far pairs to judge
_DECAY_MIN_DROP = 0.30     # near_r − far_r ≥ this → shared-path structure
# Very short common spans are launch+connector dominated — the interior
# window is too short for the Rayleigh fingerprint to separate fibers, so
# σ-outlier cascades on shared structure (A-F West: 1 005 m common span,
# 305 m interior → 1 997 false positives in production mode).
_SHORT_COMMON_SPAN_M = 2000.0
_ALLDUPS_MIN_SPAN_M = 15000.0   # all_dups needs >= this much common window

# all_dups SELF-REFUTATION check (2026-07-31).  The all_dups claim is
# "every file in this folder IS the same physical fiber".  If that were
# true MOST pairs would be near-identical by definition, so frac_high_r
# (the fraction of pairs at raw r >= 0.95) would be near 1.  A folder that
# claims all_dups while frac_high_r reads 0.00 is refuting itself: it has
# NO near-identical pairs at all, and the only thing that put it in the
# regime is a bulk median sitting a hair inside the gate.
#   BKF->DEL 80 km span, boss's 2026-07-30 run (all 48 pairs adjudicated
#   FALSE POSITIVE):  A dir bulk σ 0.0979 (< 0.10 by 2 mdB), bulk r 0.8014
#   (>= 0.7), span 80 km — routed all_dups with frac_high_r = 0.00, which
#   skips the σ-outlier tier, skips the twin gate, and widens the r-ramp
#   to 0.85-0.95 so an ordinary r ~0.91 walks to "Likely duplicate".
#   Under production routing the SAME folder reports ZERO pairs (measured).
# 0.50 is deliberately generous: a genuine all-duplicates folder sits at
# ~1.0.  Measured across the whole 2026-07-31 ripple corpus (55 folder
# runs, 21 spans), 12 folders routed all_dups and EVERY ONE of them read
# frac_high_r <= 0.03 — bkfdel A/B (0.00), tulbar B/AB (0.02/0.01) and the
# eight span3 Indio-Mecca runs (0.00-0.03).  Not one folder on disk is
# genuinely all-duplicates, so in practice this check turns the regime OFF
# for the corpus we have; that is the finding, not a side effect.  A folder
# that really is one fiber shot 400 times would clear 0.50 by a mile.
# The signal was already computed and PRINTED on the Regime: line — it just
# was never consulted.
_ALLDUPS_MIN_HIGHR_FRAC = 0.5
# NOT applied to the tie_panel route.  tie_panel is the CONSERVATIVE
# destination (fingerprint extraction + the 0.999-0.9999 ramp + σ-outlier
# bypassed), so demanding a high frac_high_r there would push folders the
# other way — into production, where σ-outlier is live.  Measured
# counter-example: the BKF+DEL combined 864-file folder routes tie_panel on
# bulk r 0.7256 with frac_high_r 0.00 and reports ZERO pairs; a
# frac_high_r sanity on that route would have sent it to production
# instead.  Leave tie_panel routing untouched.

# Uniqueness (twin) gate — ALL regimes.  A true duplicate is
# UNIQUELY close to its twin: its pair σ sits far below its σ to every
# other file.  Ribbon-family members (same tube position in adjacent
# ribbons: Δ24/Δ48 ladders like Lumen Border 146-194-218-242) share helix
# micro-structure on pristine cables and land at σ ~0.009 / r ~0.99
# against SEVERAL partners at once — no unique twin, so they are cable
# geometry, not duplication.  A flagged pair must have pair-σ ≤ this
# fraction of the smaller member's next-best σ, or it caps at borderline.
#
# 2026-07-31: the gate used to run in the production regime ONLY, on the
# reasoning that the other regimes "have their own machinery".  They don't
# — all_dups / tie_panel / short_panel bypass σ-outlier entirely, so a
# regime misroute took the twin gate off the board at exactly the moment
# it was needed.  On the BKF↔DEL all_dups misroute every one of the 47
# flagged pairs had a twin ratio of 1.00-1.71 (its σ was no closer to its
# "twin" than to the rest of the folder) and the gate would have capped
# all 47.  It now runs in every regime; the semantics are unchanged, and
# it reads the RAW σ matrix in all regimes (tie_panel_mode only changes r).
_UNIQ_TWIN_RATIO = 0.5
                                # (ELMMIL long shots: 69.5 km — comfortably in;
                                # Span 7 short shots: 5 km — routed tie_panel)

# ── Rayleigh-speckle confirmation gate ────────────────────────────────────
# The class-closer for "similar-looking but different fiber".  Below the
# splice/attenuation structure that σ and detrended-r measure sits the
# Rayleigh backscatter speckle: the frozen-in, sub-pulse-width interference
# pattern of each individual fiber's scattering centres.  It is
# DETERMINISTIC per fiber — the same fiber re-shot on the same instrument
# reproduces it — and independent between fibers, even fibers in the same
# ribbon of the same cable that share every macroscopic feature.
#
# Method: subtract a moving average of _SPECKLE_HP_WIDTH samples (the
# low-pass that carries splice/attenuation structure), keep the residual,
# and Pearson-correlate the two residuals inside each analysis window.
# Windows are FRACTIONS of the interior, taken from the launch-side (high
# SNR) 60% — the far end runs into the noise floor where every pair
# decorrelates.  Score = MAX over windows (the most permissive combiner:
# one window agreeing is enough to confirm).
#
# NO FIXED CONFIRM THRESHOLD — that was the first cut and it is WRONG.
# A same-fiber pair's speckle correlation falls off with how much
# acquisition noise separates the two shots: s²/(s²+σ²/2) for band
# amplitude s and pair disagreement σ.  Measured band amplitudes are ~7.4
# mdB (BKF, dz 2.55 m) and ~3.6 mdB (DEL), so a genuine re-shoot at σ =
# 20 mdB reads only ~0.35, at σ = 40 mdB only ~0.08.  Any fixed threshold
# high enough to reject the BKF false positives (r_hp <= 0.078) would also
# reject real duplicates at those σ — measured directly by injecting white
# noise into a real file:
#     BKF  σ 0.0149 -> same-fiber control 0.320   (fixed 0.25 keeps it)
#     BKF  σ 0.0396 -> same-fiber control 0.077   (fixed 0.25 KILLS it)
#     CHEPLA σ 0.0098 -> control 0.091            (fixed 0.25 KILLS it)
#     NILMEC σ 0.0125 -> control 0.217            (fixed 0.25 KILLS it)
# So the gate is scored against the pair's OWN same-fiber floor instead:
#
#   r_floor = s²/(s²+σ²/2)   (_speckle_same_fiber_floor — the LOWEST value
#             the same-fiber hypothesis can produce at this pair's σ; any
#             low-frequency part of the difference only raises the truth)
#   null_q  = the 99th percentile of r_hp over a sample of KNOWN-different
#             pairs in this same folder (what chance looks like here)
#
#   VETO iff   r_floor >= null_q                (the statistic can tell
#                                                same from different at
#                                                this pair's σ, on this
#                                                folder — otherwise abstain)
#        and   r_hp <= r_floor / _SPECKLE_FLOOR_MARGIN
#                                               (the pair is far below
#                                                even the worst same-fiber
#                                                case — not one fiber)
#
# Worked (2026-07-31).  Across the whole 62-run ripple corpus the gate
# vetoed exactly THREE pairs — it is a narrow, high-confidence instrument,
# NOT the thing that fixes BKF.  All three were then corroborated by an
# INDEPENDENT check the gate never sees: the same two fiber numbers shot
# from the OTHER end.  Same glass reads the same from both ends.
#   PLACHE0665↔0666  A-dir σ 0.0057 r 0.998 (looks like a copy)
#                    r_hp 0.0008  floor 0.206  null_q 0.137  -> VETO
#                    B-dir CHEPLA0665↔0666: σ 0.0993 r 0.336 = different
#                    fibers.  Veto correct.
#   WNHNIL413↔414    A-dir σ 0.0070 r 0.992, events 11/12 @ 1.1 mdB —
#                    the classic true-duplicate signature, and the only
#                    gate that catches it.  r_hp 0.0586 floor 0.445
#                    null_q 0.110 -> VETO.  B-dir NILWNH413↔414: σ 0.0537
#                    r 0.747 = different fibers.  Veto correct.
#   NILMEC418↔424    σ 0.0125 r_hp 0.0395 floor 0.209 null_q 0.201 ->
#                    VETO, but the competence margin is THIN (1.04x).
#                    B-dir MECNIL418↔424 reads r_hp -0.047 against floor
#                    0.308 — both ends independently say "no shared
#                    speckle".  Veto correct, margin noted.
# Kept (gate declined to act):
#   TUCROM453↔454    σ 0.0040 r_hp 0.184 floor 0.250 — within 1.4x of the
#                    floor, not far enough below: kept.
#   CHEPLA0167↔0168  σ 0.0098 floor 0.091 < null_q 0.154: ABSTAIN, kept.
#   BKF's 47 + DELBKF138↔162 (σ 0.021-0.12): floor at/below null_q, so the
#                    gate abstains; the router / twin / events gates are
#                    what clear those.
#   byte-identical copy: σ 0 -> floor 1.0, r_hp 1.0 -> never vetoed.
#   BKF synthetic re-shoots (+10/+15/+20 mdB white noise on a real file):
#                    r_hp 0.594/0.443/0.349 vs floor 0.44/0.32/0.25 —
#                    all comfortably above floor/3, all kept.
#
# DEMOTE-ONLY and FAIL-SAFE, matching the splicereport re-measure gates:
# the gate can only cap a pair some other tier already pushed over the
# print threshold, it never promotes, and anything UNMEASURABLE
# (mismatched sample spacing, too few samples, a flat/saturated window,
# too few files to build a null) confirms by default.
# HIGH-PASS WIDTH IS PER-FOLDER (2026-08-29).  21 samples was tuned on
# 500 ns / 25 ns acquisitions, where it happens to equal one pulse width.
# The pulse expressed in SAMPLES varies 15x over the acquisitions on disk
# (275ns/25ns = 11, 500ns/25ns = 20, 2500ns/50ns = 50, 10ns/3.125ns = 3.2,
# 5ns/0.78ns = 6.4), so a fixed sample count is a different filter on every
# span.  Run WIDER than the pulse and splice steps survive the moving
# average and land in the residual as a bipolar spike of the same sign in
# every fiber, which is common-mode and inflates the folder null.
#
# MEASURED on Mecca<->Niland (275 ns / 25 ns, pulse = 11 samples, so the
# shipped 21 is 1.9 pulse widths): 10.2% of the null sample's residual
# energy sits beyond 3 MAD, concentrated exactly on the splice set
# (5.45-5.71 km, 10.46-10.83 km, 21.4 km — events the table places in
# 150-454 of the 576 files).  Folder null p99 reads 0.201.  Narrowed to the
# pulse it reads 0.072, and NILMEC498<->504 goes from r_hp +0.068 (gate
# ABSTAINS, floor 0.124 < null 0.201, pair prints at p_dup 1.0) to -0.132
# against floor 0.095 (VETO).  The B direction agrees independently:
# MECNIL498<->504 +0.147 -> -0.030, null 0.092 -> 0.067, keep -> VETO.
# Robert (field) states that pair is not a duplicate; both ends now say so.
#
# THIS IS AN EMPIRICAL RULE, NOT A DERIVED ONE.  "Match the filter to the
# pulse" is NOT sufficient on its own: at matched width the null reads 0.072
# (NILMEC), 0.086 (EMVSUI L), 0.363 (SEANOR), 0.493 (SANDUR), 0.955
# (A-F West), 0.976 (EMVSUI Short) — a 13x spread, so w/pulse is not the
# controlling variable.  The cap at 21 is what keeps the long-pulse folders
# (SEANOR/SANDUR, pulse 50 samples) on the width they were calibrated with.
# The rule therefore only ever NARROWS, and only on acquisitions whose pulse
# is shorter than the calibrated 21.
#
# CONTROL SET, stated honestly: across the 13 folders on disk the speckle
# gate evaluates FIVE pairs in total, all of them in the two Mecca<->Niland
# directions.  Every other folder is a no-op because the gate never runs
# there, not because the width was shown to be safe there.  Full-engine A/B
# runs: NIL->MEC 1 of 165,600 pairs changed, MEC->NIL 1 of 165,600, and ZERO
# of 1.39 million across A-F West, A-F East, LAMBEY, EMVSUI Long, EMVSUI
# Short (at w=5), SUIEMV Long and SUIEMV Short.  EMVSUI's four confirmed
# duplicates keep their verdicts (79/80 0.467, 511/512 0.909, 563/564 0.615,
# 296/308 0.478 — unchanged, the folder resolves to 21).  The veto is not
# knife-edged: 498/504 vetoes at w = 7, 9, 11, 13 and 15.
_SPECKLE_HP_WIDTH = 21          # moving-average width in SAMPLES (odd); CAP
_SPECKLE_HP_WIDTH_MIN = 5       # never narrow below this, however short the pulse
_SPECKLE_WINDOWS = ((0.02, 0.20), (0.20, 0.40), (0.40, 0.60))
_SPECKLE_MIN_SAMPLES = 500      # per window, after high-pass edge trim
_SPECKLE_DZ_TOL = 1e-6          # relative sample-spacing match required
_SPECKLE_FLOOR_MARGIN = 3.0     # r_hp must be this far below r_floor to veto
_SPECKLE_NULL_FILES = 60        # evenly-spaced folder sample for the null
_SPECKLE_NULL_PCT = 99.0        # percentile of that null a veto must clear
_SPECKLE_NULL_MIN_PAIRS = 100   # fewer than this -> no null -> no vetoes
# Multiple of the folder null a pair must clear for its fingerprint to
# REFUTE the twin gate's sigma-ratio proxy (see the twin-gate block).  The
# null is already the 99th percentile of known-different pairs, so this is
# a deliberately high bar.  Measured on EMVSUI0 Long Shots: over 4,005
# known-different pairs sampled from 90 fibers across the cable the
# statistic reads p50 0.024, p99 0.084, p99.9 0.103 and MAXIMUM 0.111,
# while the four confirmed same-fiber pairs read 0.467, 0.478, 0.615 and
# 0.909.  3x p99 (0.258 there) sits in the empty band between them.
_SPECKLE_CONFIRM_NULL_MULT = 3.0
# A pair must be at least this sigma-likely before a sigma-bypassed regime
# will even look at its fingerprint.  0.99 is deliberately extreme: on the
# whole corpus it selects TWO pairs, both on MILTOP, and nothing at all on
# the folders whose cascades the bypass exists to prevent.
_SIGMA_RESCUE_MIN = 0.99

# ── Robust common span + suspected-break reporting ────────────────────────
# The common analysis span used to be the raw MINIMUM EOF over all files,
# so ONE broken fiber collapsed the whole folder's window.  A-F West
# 145-288 (266 files, median EOF 2 037 m) has port 198 physically broken
# ~1 km out — A-side EOF 1 005 m, B-side 1 036 m, and the two EOFs sum to
# the span — and that single strand shrank the interior to 305 m, hiding
# the folder-wide similarity and misrouting the regime (the 1 997-false-
# positive flood).  Two rules here:
#   1. Files whose EOF is far below the folder median are SUSPECTED BREAKS
#      — a finding in its own right, always reported (manifest key
#      `short_traces` + a "Suspected broken / short fibers" report section).
#   2. Suspected breaks are ALWAYS excluded from the common-span
#      computation and pair metrics — they physically lack the glass being
#      compared (their post-break samples are noise floor, not
#      backscatter) — provided ≥2 healthy files remain and the folder
#      passes the consistency guard below.
#
# LONG-SPAN WINDOW RESTORATION (2026-07-15): the first cut of rule 2 only
# excluded when the raw-min window had collapsed below the 2 km
# launch+connector floor (_SHORT_COMMON_SPAN_M), which preserved ELMMIL's
# and SANDUR's historical pair tables but left ELMMIL_1550 analyzing
# 22 288 m of a 69 554 m span (ELMMIL0231's break) and SANDUR 20 144 m of
# ~100 925 m (SANDUR841).  Robert approved changing those baselines
# (better detection is worth the numbers moving), so the 2 km-collapse
# precondition is gone: the window is always rebuilt from the healthy
# population.  _SHORT_COMMON_SPAN_M keeps its separate role in regime
# routing (short-common-span → tie_panel).
#
# SANITY GUARD (_INCONSISTENT_FOLDER_FRAC): if MORE than 20 % of the
# folder's files sit below 75 % of the median, that isn't "a few broken
# strands" — it's an inconsistent folder (mixed spans, wrong files) where
# the median itself is untrustworthy.  Exclude NOTHING, keep the raw-min
# window, and warn ("folder trace lengths are inconsistent …").
#
# Calibration (2026-07-15, per-file EOFs on real folders — the guard must
# fire on none of these):
#   A-F West 145-288: 2 of 266 (0.8 %) below cut → excluded, window
#                     restored to the healthy min 2 036.8 m (unchanged
#                     from the first cut — its raw min was < 2 km).
#   A-F East 1-144:   0 outliers (min 2 036.8 m ≈ median) → byte-identical.
#   SEANOR (432):     0 outliers (min 108 818 m = 99 % of median)
#                     → byte-identical.
#   ELMMIL (1152):    1 of 1152 (0.09 %): ELMMIL0231 ends 22 288 m (32 %
#                     of the 69 554 m median) → NOW excluded, window
#                     restored 22 288 → 69 549 m (baseline change,
#                     approved).
#   SANDUR (864):     2 of 864 (0.23 %): SANDUR841 @ 20 144 m (20 %) and
#                     SANDUR229 @ 59 219 m (59 %) → NOW excluded, window
#                     restored 20 144 → ~100 856 m (baseline change,
#                     approved).
_BREAK_FRAC_OF_MEDIAN = 0.75   # EOF below this × median ⇒ suspected break
_BREAK_AB_SUM_TOL = 0.10       # |EOF_A + EOF_B − median| ≤ this × median
_INCONSISTENT_FOLDER_FRAC = 0.20   # > this fraction short ⇒ guard, no exclusion

# Port extraction mirrors run_secretsauce._extract_fiber_num (fiber number
# before a wavelength suffix first, else trailing digits) so the two agree
# on which digits are the port:  ELMMIL0001_1550 → ('ELMMIL', 1),
# BCK1BCK60145 → ('BCK1BCK6', 145).
_PORT_WL_RE = re.compile(r'(\d{3,4})_\d{3,4}\b')
_PORT_TAIL_RE = re.compile(r'(\d{3,4})$')
# A trailing run of NON-digits hides the port from _PORT_TAIL_RE, and the
# whole filename then becomes its own one-member prefix group.  Measured on
# disk: 'DNW5DNW10271withstartstop' and friends, 347 files across 9 folders,
# 331 of them in folders where EVERY file collapses this way.  Split the
# suffix off, parse the head, then put the suffix BACK on the prefix so a
# folder mixing 'X0001' and 'X0001withstartstop' still gets two groups.
_PORT_SUFFIX_RE = re.compile(r'(\d)([^\d]+)$')
# Ports under 100 are 1-2 digits ('RCHDNW-A-66').  Only reachable once both
# rules above have failed, which guarantees the trailing digit run is 1 or 2
# long — a longer run already matched _PORT_TAIL_RE.  The separator
# lookbehind stops a route code ('...BCK6') or an ordinary word ending in a
# digit ('panel1') from donating its last digit as a port.
_PORT_SHORT_TAIL_RE = re.compile(r'(?<=[^0-9A-Za-z])(\d{1,2})$')


def _port_split(name):
    """Split a filename stem into (prefix, port).  The prefix doubles as the
    direction/route group so A→B and B→A shots never mix in gap stats.

    STRICTLY ADDITIVE (2026-08-29): the two original patterns are tried
    first and unchanged, so any name that parses today parses identically.
    The two fallbacks below only ever turn `port None` into a parsed port.
    Swept over 41,268 distinct stems / 57,628 files on disk: 330 stems
    change, every one of them None -> parsed, and ZERO where the old and new
    parsers both return a port and disagree.
    """
    m = _PORT_WL_RE.search(name) or _PORT_TAIL_RE.search(name)
    if m:
        return name[:m.start(1)], int(m.group(1))
    suf = _PORT_SUFFIX_RE.search(name)
    if suf:
        pref, port = _port_split(name[:suf.end(1)])
        if port is not None:
            return pref + suf.group(2), port
    short = _PORT_SHORT_TAIL_RE.search(name)
    if short:
        return name[:short.start(1)], int(short.group(1))
    return name, None


def _merge_zero_pad_prefixes(prefixes):
    """Fold '<P>0' into '<P>' when BOTH forms occur in the same folder.

    The greedy 4-digit tail eats a padding zero when the zero-padding is
    inconsistent: 'PTL5PTL1sh0232' -> ('PTL5PTL1sh', 232) but
    'PTL5PTL1sh00309' -> ('PTL5PTL1sh0', 309), orphaning that one file into
    its own group.  Data-driven and folder-scoped — it only fires when both
    the padded and unpadded form are present in the same call, so a prefix
    that legitimately ends in '0' with no shorter sibling is never touched.
    Measured: fires on 1 folder / 1 file across all 230 folders on disk.
    """
    uniq = set(prefixes)
    canon = {}
    for p in uniq:
        q = p
        while q.endswith('0') and q[:-1] in uniq:
            q = q[:-1]
        canon[p] = q
    return [canon[p] for p in prefixes]


def _neighbor_decay(names, r_matrix, serials=None,
                    near_gap=_DECAY_NEAR_GAP, far_gap=_DECAY_FAR_GAP,
                    min_pairs=_DECAY_MIN_PAIRS):
    """Neighbor-vs-far raw-r structure for the distance-decay regime test.

    `names` are filename stems aligned with `r_matrix` rows.  Pairs are
    compared only within the same prefix group (same route/direction) and
    only when both files carry a trailing port number.  Returns
    (near_r_median, far_r_median, n_near, n_far) or None when either bucket
    has fewer than `min_pairs` pairs (small folders can't trip this rule).

    SAME INSTRUMENT ONLY (`serials`, 2026-08-29).  The rule reads a drop in
    r between near and far port pairs as evidence of shared launch glass.
    That inference only holds when the two buckets differ in port distance
    and NOTHING ELSE.  On a folder shot by more than one OTDR the far
    bucket fills up with cross-instrument pairs, which decorrelate for
    reasons that have nothing to do with port distance, and the rule fires
    on a folder that has no tie panel in it.

    Measured on EMVSUI0 Long Shots (1152 fibers, 78.5 km, serials 1723356
    and 1876271 interleaved across the port range over 5 days):

        near (gap <= 3,  same OTDR)  r 0.746
        far  (gap >= 30, same OTDR)  r 0.677   <- port distance costs 0.069
        far  (gap >= 30, DIFF OTDR)  r 0.205   <- instrument costs 0.472

    The pooled far median landed at 0.37, the folder tripped the 0.30 drop,
    routed tie_panel, and tie_panel bypasses sigma-outlier — so the single
    tightest pair of all 662,976 (sigma 0.0095 dB against a bulk of 0.1238,
    z = -9.1) scored 0.0 and the report flagged nothing at all.  Restricted
    to one instrument the drop is 0.069 and the rule correctly stays quiet.

    Fails OPEN per pair when either serial is missing, matching the
    different-OTDR verdict gate: an unknown instrument is not evidence.
    """
    K = len(names)
    if K < 2 or r_matrix.shape[0] != K:
        return None
    prefixes, ports = [], []
    for n in names:
        pref, port = _port_split(n)
        prefixes.append(pref)
        ports.append(port)
    prefixes = _merge_zero_pad_prefixes(prefixes)
    port_arr = np.array([p if p is not None else -1 for p in ports],
                        dtype=np.int64)
    has_port = port_arr >= 0
    codes = {p: i for i, p in enumerate(sorted(set(prefixes)))}
    pref_arr = np.array([codes[p] for p in prefixes], dtype=np.int64)

    same_pref = pref_arr[:, None] == pref_arr[None, :]
    both_ports = has_port[:, None] & has_port[None, :]
    gap = np.abs(port_arr[:, None] - port_arr[None, :])
    upper = np.triu(np.ones((K, K), dtype=bool), k=1)
    # Same-instrument mask.  Missing serial fails open (pair stays eligible),
    # so a folder with no serials at all behaves exactly as it did before.
    if serials is None:
        same_ser = np.ones((K, K), dtype=bool)
    else:
        codes_s = {v: i for i, v in enumerate(sorted({x for x in serials if x}))}
        ser_arr = np.array([codes_s.get(x, -1) if x else -1 for x in serials],
                           dtype=np.int64)
        known = ser_arr >= 0
        same_ser = ((~known[:, None]) | (~known[None, :])
                    | (ser_arr[:, None] == ser_arr[None, :]))
    eligible = upper & same_pref & both_ports & same_ser
    near_mask = eligible & (gap >= 1) & (gap <= near_gap)
    far_mask = eligible & (gap >= far_gap)
    n_near = int(near_mask.sum())
    n_far = int(far_mask.sum())
    if n_near < min_pairs or n_far < min_pairs:
        return None
    near_r = float(np.median(r_matrix[near_mask]))
    far_r = float(np.median(r_matrix[far_mask]))
    return near_r, far_r, n_near, n_far


def _speckle_hp_width(files):
    """Moving-average width in samples for THIS folder's acquisition.

    Returns _SPECKLE_HP_WIDTH (the calibrated cap) unless every file agrees
    on a pulse shorter than that, in which case it narrows to the pulse.

    ACQUISITION-UNIFORMITY GUARD.  The width is taken from the MEDIAN over
    all files and only applied when the folder is uniform, because Secret
    Sauce runs one report over whatever folder is uploaded.  A mixed folder
    is reachable and would otherwise have its filter width decided by which
    file sorted first: 275 ns and 500 ns EXFO acquisitions share a
    bit-identical sample spacing (dz = 2.552445171 m), so _SPECKLE_DZ_TOL
    considers their traces comparable and the existing grid guard does not
    catch the mismatch.  When the folder is not uniform we decline to
    narrow and keep the calibrated width, which is the fail-safe direction:
    a wider filter only ever inflates the null, and a higher null makes the
    gate abstain rather than veto.
    """
    vals = [f.get('pulse_samples') for f in files]
    vals = [float(v) for v in vals if v and np.isfinite(v) and v > 0]
    if len(vals) < len(files) or not vals:
        return _SPECKLE_HP_WIDTH        # a file could not report its pulse
    lo, hi = min(vals), max(vals)
    if hi - lo > 0.01 * hi:
        return _SPECKLE_HP_WIDTH        # mixed acquisition: do not narrow
    w = int(np.floor(float(np.median(vals))))
    if w % 2 == 0:
        w += 1                          # the kernel must be odd
    return max(_SPECKLE_HP_WIDTH_MIN, min(_SPECKLE_HP_WIDTH, w))


def _speckle_windows(f, interior_start, interior_end, hp_width=None):
    """Unit-normalized speckle-band residual of one trace per analysis
    window (see the _SPECKLE_* calibration block).

    The residual is the trace minus an `hp_width`-sample moving
    average (see _speckle_hp_width; defaults to the calibrated cap) — the low-pass carries splice steps and attenuation slope, the
    residual carries each fiber's own frozen-in Rayleigh interference
    pattern.  Window bounds are computed from the sample spacing (not from
    a boolean position mask) so two files on the same grid always get
    byte-identical index ranges.

    Returns {'dz': sample spacing,
             'win': [ (i0, i1, unit_residual, residual_std_dB) | None ]}
    or None when the file can't be measured at all (fail-safe: unmeasurable
    never vetoes).  residual_std_dB is the fiber's own speckle amplitude in
    that window — the scale that sets how much acquisition noise the
    statistic can survive (see _speckle_same_fiber_floor).
    """
    if f is None:
        return None
    trace, pos = f.get('trace'), f.get('pos')
    if trace is None or pos is None or len(trace) < 4 or len(pos) < 2:
        return None
    dz = float(pos[1] - pos[0])
    if not np.isfinite(dz) or dz <= 0:
        return None
    w = _SPECKLE_HP_WIDTH if hp_width is None else int(hp_width)
    kern = np.ones(w) / w
    n = len(trace)
    span = interior_end - interior_start
    if span <= 0:
        return None
    out = []
    for f0, f1 in _SPECKLE_WINDOWS:
        i0 = int(np.floor((interior_start + span * f0) / dz)) + 1
        i1 = int(np.ceil((interior_start + span * f1) / dz))
        i0 = max(i0, 0)
        i1 = min(i1, n)
        # Need the window plus the convolution edge trim on both sides.
        if i1 - i0 < _SPECKLE_MIN_SAMPLES + 2 * w:
            out.append(None)
            continue
        x = trace[i0:i1].astype(np.float64)
        h = (x - np.convolve(x, kern, mode='same'))[w:-w]
        h = h - h.mean()
        nrm = float(np.sqrt(np.dot(h, h)))
        if not np.isfinite(nrm) or nrm <= 0:
            # Flat / saturated / all-NaN window — nothing to fingerprint.
            out.append(None)
            continue
        out.append((i0, i1, h / nrm, float(nrm / np.sqrt(len(h)))))
    if all(v is None for v in out):
        return None
    return {'dz': dz, 'win': out}


def _speckle_comparable(ra, rb):
    """Windows the two files can actually be compared in, or None."""
    if ra is None or rb is None:
        return None
    dz_ref = max(ra['dz'], rb['dz'])
    if abs(ra['dz'] - rb['dz']) > _SPECKLE_DZ_TOL * dz_ref:
        return None                      # different acquisition grid
    out = []
    for wa, wb in zip(ra['win'], rb['win']):
        if wa is None or wb is None:
            continue
        if wa[0] != wb[0] or wa[1] != wb[1] or len(wa[2]) != len(wb[2]):
            continue
        out.append((wa, wb))
    return out or None


def _speckle_pair_r(ra, rb):
    """MAX speckle-band Pearson r across the analysis windows, or None when
    the pair is UNMEASURABLE (different sample spacing, no window long
    enough on both sides, flat residuals).  MAX is the most permissive
    combiner — one agreeing window is enough to confirm a pair."""
    cw = _speckle_comparable(ra, rb)
    if cw is None:
        return None
    return max(float(np.dot(wa[2], wb[2])) for wa, wb in cw)


def _speckle_same_fiber_floor(ra, rb, sigma_pair):
    """LOWEST speckle r the same-fiber hypothesis can produce for a pair
    that disagrees by `sigma_pair` dB — i.e. the number this pair would
    still have to beat if it really were one fiber shot twice.

    Two shots of one fiber share the speckle exactly; everything that makes
    them differ is acquisition noise.  Put ALL of that difference into the
    speckle band (the worst case — real re-shoot differences are launch
    level and thermal drift, which the high-pass removes) and split it
    evenly between the two shots.  With band amplitude s and per-shot noise
    σ/√2 the correlation is s² / (s² + σ²/2).  Verified against
    white-noise-injected controls on real files: predicted 0.331 vs
    measured 0.320 (BKF, σ 0.0149); predicted 0.056 vs measured 0.056
    (DEL, σ 0.0210); predicted 0.065 vs measured 0.077 (BKF, σ 0.0396).

    Any low-frequency component in the real difference only pushes the true
    value ABOVE this, so it is a genuine lower bound.  Uses the SMALLER of
    the two files' band amplitudes (conservative) and the best window.
    Returns None when the pair is unmeasurable.
    """
    cw = _speckle_comparable(ra, rb)
    if cw is None:
        return None
    sig2 = max(float(sigma_pair), 0.0) ** 2 / 2.0
    best = None
    for wa, wb in cw:
        s2 = min(wa[3], wb[3]) ** 2
        if s2 <= 0:
            continue
        v = s2 / (s2 + sig2)
        if best is None or v > best:
            best = v
    return best


def _robust_common_span(lengths):
    """Robust common analysis span over per-file EOFs (meters).

    Returns (span_m, median_m, outlier_idx, excluded_idx, guard_note):
      span_m       — the common span the interior window is built from
      median_m     — folder median EOF
      outlier_idx  — indices whose EOF < _BREAK_FRAC_OF_MEDIAN × median
                     (suspected breaks — ALWAYS reported, never silent)
      excluded_idx — indices to drop from pair metrics.  Equal to
                     outlier_idx whenever suspected breaks exist, ≥2
                     healthy files remain, and the consistency guard
                     does not fire; else empty.
      guard_note   — None normally; the warning string when the
                     inconsistent-folder guard fired (no exclusion, span
                     stays the raw min).
    Homogeneous folders (no extreme outliers) return exactly min(lengths),
    so their windows — and pair tables — are byte-identical to the
    historical raw-min behavior.  See the calibration block above
    _BREAK_FRAC_OF_MEDIAN for the long-span window restoration + guard.
    """
    arr = np.asarray(list(lengths), dtype=np.float64)
    raw_min = float(arr.min())
    med = float(np.median(arr))
    if med <= 0:
        return raw_min, med, [], [], None
    cut = _BREAK_FRAC_OF_MEDIAN * med
    outlier_idx = [int(i) for i in np.flatnonzero(arr < cut)]
    span, excluded_idx, guard_note = raw_min, [], None
    n = len(arr)
    n_healthy = n - len(outlier_idx)
    if outlier_idx:
        if len(outlier_idx) > _INCONSISTENT_FOLDER_FRAC * n:
            guard_note = (
                f'folder trace lengths are inconsistent ({len(outlier_idx)} '
                f'of {n} below 75% of median) — window not restored; '
                f'check folder contents')
        elif n_healthy >= 2:
            healthy_min = float(arr[arr >= cut].min())
            if healthy_min > raw_min:
                span = healthy_min
                excluded_idx = outlier_idx
    return span, med, outlier_idx, excluded_idx, guard_note


def _ab_break_notes(entries, median_m):
    """Cross-direction A+B consistency for suspected breaks (in place).

    `entries` are short-trace dicts carrying 'file' and 'eof_m'.  When the
    folder holds BOTH directions of the same port (same trailing port
    number, different prefix — the same direction/prefix grouping
    _neighbor_decay uses) and the two EOFs sum to the folder median span
    within ±_BREAK_AB_SUM_TOL, the two shots are the two sides of ONE
    physical break: each entry gains a 'break_note' anchored at its own
    launch end, e.g. "A+B lengths are consistent with a break ~1005 m from
    the BCK1BCK6 end"."""
    if not median_m or median_m <= 0:
        return
    by_port = {}
    for e in entries:
        pref, port = _port_split(e['file'])
        if port is None:
            continue
        by_port.setdefault(port, []).append((pref, e))
    for port, lst in by_port.items():
        if len(lst) != 2:
            continue
        (pref_a, ea), (pref_b, eb) = lst
        if pref_a == pref_b:
            continue
        total = ea['eof_m'] + eb['eof_m']
        if abs(total - median_m) <= _BREAK_AB_SUM_TOL * median_m:
            for pref, e in lst:
                e['break_note'] = (
                    f'A+B lengths are consistent with a break '
                    f'~{e["eof_m"]:.0f} m from the {pref} end')


def _short_trace_section_html(short_traces, window_guard=None):
    """PDF section for suspected broken / short fibers.  Returns '' when
    there are none, so unaffected reports stay byte-stable.  When the
    inconsistent-folder guard fired, a warning banner renders above the
    table (the guard can only fire when suspected breaks exist)."""
    if not short_traces:
        return ''
    guard_html = ''
    if window_guard:
        guard_html = (f'<div class="verdict-box verdict-dispute">'
                      f'<b>Warning:</b> {window_guard}</div>')
    rows = ''
    for e in short_traces:
        finding = 'suspected break'
        if e.get('excluded'):
            finding += ' — excluded from pair comparison'
        if e.get('break_note'):
            finding += f'. {e["break_note"]}'
        rows += (f'<tr><td class="pair-cell">{e["file"]}</td>'
                 f'<td class="center">{e["eof_m"]:.0f}</td>'
                 f'<td class="center">{e["median_eof_m"]:.0f}</td>'
                 f'<td>{finding}</td></tr>')
    return f'''
<div class="section-block">
<div class="dir-banner">Suspected broken / short fibers</div>{guard_html}
<table class="vote-table">
<tr><th style="text-align:left">File</th><th>Ends at (m)</th>
    <th>Folder median (m)</th><th style="text-align:left">Finding</th></tr>
{rows}
</table>
</div>
'''


def _analyze_sor(folder):
    """Shared SOR analysis: load files, compute pair metrics, apply
    physical-reality filters, pick best partners. Returns a dict the
    PDF and XLSX renderers can both consume.
    """
    paths = sorted(glob.glob(os.path.join(folder, '*.sor')))
    files = []
    for p in paths:
        try:
            files.append(load_sor_file(p))
        except Exception as e:
            print(f'  skip {os.path.basename(p)}: {e}')
    if len(files) < 2:
        raise RuntimeError(f'Not enough usable .sor files in {folder}')
    print(f'Loaded {len(files)} .sor files from {folder}')

    # Robust common span: rebuilt from the healthy population whenever
    # suspected breaks exist — see the calibration block above
    # _BREAK_FRAC_OF_MEDIAN.  Suspected breaks (EOF far below the folder
    # median) are always surfaced via `short_traces` AND excluded from pair
    # metrics (they physically lack the glass being compared), unless the
    # inconsistent-folder guard fired (window_guard below).
    sized = [f for f in files if f['length'] > 0]
    min_L, median_L, out_idx, excl_idx, window_guard = _robust_common_span(
        [f['length'] for f in sized])
    excluded_names = {sized[i]['name'] for i in excl_idx}
    short_traces = []
    for i in out_idx:
        f = sized[i]
        short_traces.append({
            'file': f['name'],
            'eof_m': round(float(f['length']), 1),
            'median_eof_m': round(median_L, 1),
            'excluded': f['name'] in excluded_names,
            'note': (f'ends at {f["length"]:.0f} m (folder median '
                     f'{median_L:.0f} m) — suspected break'),
        })
    short_traces.sort(key=lambda e: (e['eof_m'], e['file']))
    _ab_break_notes(short_traces, median_L)
    for e in short_traces:
        state = 'excluded from pair metrics' if e['excluded'] else 'kept'
        line = f'  Suspected break: {e["file"]} {e["note"]} [{state}]'
        if e.get('break_note'):
            line += f' — {e["break_note"]}'
        print(line)
    if window_guard:
        print(f'  WARNING: {window_guard}')
    if excluded_names:
        files = [f for f in files if f['name'] not in excluded_names]
        print(f'Common span restored to {min_L:.0f} m '
              f'({len(excluded_names)} suspected-broken trace(s) excluded, '
              f'{len(files)} files remain)')
    interior_start = _LAUNCH_SKIP_M
    interior_end = min_L - _END_BUFFER_M
    if interior_end - interior_start < 100:
        interior_start = max(2.0, min_L * 0.05)
        interior_end = max(interior_start + 2.0, min_L * 0.95)
    print(f'Interior window: {interior_start:.0f}–{interior_end:.0f} m  '
          f'(common span {min_L:.0f} m)')

    print(f'Computing pair metrics for {len(files)} files '
          f'({len(files) * (len(files) - 1) // 2} pairs)...')
    # Three-regime classifier (replaces the old file-count-floor heuristic):
    #
    #   PRODUCTION — typical case. Bulk pair-r low (~0.3), σ-outlier detector
    #                works because non-duplicate pairs define a clear bulk.
    #   TIE-PANEL  — many fibers sharing a launch+connector signal. Bulk r
    #                high (~0.95) AND bulk σ moderate (~0.15 dB) — the
    #                shared signal pulls r up but the fibers are physically
    #                different so σ doesn't collapse. Needs fingerprint
    #                extraction + tightened r-ramp + r-confirmation gate.
    #   ALL-DUPS   — every file is the same physical fiber. Bulk r high
    #                (~0.95) AND bulk σ at shot-noise floor (~0.06 dB).
    #                σ-outlier detector breaks (no non-duplicate bulk), so
    #                bypass it and use a widened r-ramp.
    #
    # First pass: compute pair metrics WITHOUT fingerprint extraction so
    # the classifier can see the raw σ/r distributions.
    batch_raw = _compute_pair_metrics_batch(files, interior_start, interior_end,
                                            tie_panel_mode=False)
    if batch_raw is None:
        raise RuntimeError('No comparable pairs after interior masking')
    sigma_raw, r_raw, valid_idx_raw = batch_raw
    iu_raw = np.triu_indices(sigma_raw.shape[0], k=1)
    bulk_sigma = float(np.median(sigma_raw[iu_raw])) if len(iu_raw[0]) else 0.0
    bulk_r = float(np.median(r_raw[iu_raw])) if len(iu_raw[0]) else 0.0
    # Fraction of pairs with elevated raw r. Catches tie panels whose MEDIAN
    # r is low (most ports mutually uncorrelated) but a large minority of
    # pairs share cable structure at r ~ 1.0. Example: 2 km tie panels
    # (CLQTILA) where median r ~ 0.18 yet ~48% of pairs sit at r >= 0.95
    # because they run the same route. Median alone misses these and they
    # cascade into 20k+ false positives in production mode — which also
    # OOM-kills the renderer (a 20k-row confirmed-duplicate table).
    frac_high_r = float((r_raw[iu_raw] >= 0.95).mean()) if len(iu_raw[0]) else 0.0
    # Four-regime classifier:
    #   all_dups    — every file IS the same fiber. High r, low σ.
    #   short_panel — many short fibers (< 200 m interior) in a panel where
    #                 the interior trace is too featureless for σ-outlier
    #                 to discriminate. Without this gate σ-outlier cascades
    #                 into thousands of false positives (BETA Raywood/
    #                 Sorrento etc.). Bulk r stays LOW on short panels
    #                 because the shared launch+connector signal doesn't
    #                 dominate a featureless interior, so the tie_panel
    #                 trigger never fires for these.
    #   tie_panel   — many fibers with shared structure. Triggered by EITHER
    #                 high median r (>=0.7, classic short-launch tie panels
    #                 like Deming) OR a high FRACTION of elevated-r pairs
    #                 (>=30% at r>=0.95, long tie panels like CLQTILA whose
    #                 median is low). Either way fingerprint extraction +
    #                 the tight ramp sort true re-shoots from shared cable
    #                 structure, so over-classifying here is self-correcting.
    #   production  — typical case.
    # Order matters: all_dups checked first so a hypothetical all-duplicates
    # short-fiber dataset doesn't get misrouted to short_panel.
    #
    # Two ADDITIVE tie_panel routes run after the existing rules (they can
    # only re-route folders that would otherwise land in 'production'):
    #   neighbor-decay — raw r falls off with port distance (shared glass:
    #                    jumper feed + ribbon). Copies don't care about port
    #                    distance, so decay ⇒ shared path, not duplication.
    #                    A-F West: 1 997 σ-outlier false positives at ≥99%
    #                    in production mode; its bulk_r stayed 0.05 because
    #                    shared-glass r tops out ~0.8 among NEIGHBORS only,
    #                    which the bulk_r / frac_high_r triggers can't see.
    #   short common span — < 2 km of shared window is launch+connector
    #                    dominated; too little Rayleigh fingerprint for the
    #                    σ-outlier bulk to mean anything.
    regime_reason = None
    # all_dups additionally requires a LONG-ENOUGH common window for bulk_r
    # to mean anything.  Short-shot folders of UNIQUE fibers (Span 7
    # Tularosa-Orogrande: 864 fibers, ~5 km common span) correlate broadband
    # over the short shared window — bulk r lands INSIDE the all_dups
    # 0.85-0.95 ramp and the ordinary bulk walks over 50% (62,014 false
    # pairs, issue #9), with the self-refuting signature frac_high_r = 0.00
    # and ZERO pairs at >=99%.  Same physics as the < 2 km tie_panel rule,
    # applied to the gate that claims folders FIRST.  A short folder with
    # high bulk r falls through to the bulk_r >= 0.7 tie_panel route
    # (fingerprint extraction + the 0.999 ramp) — true re-shoots still land
    # at r >= 0.999 there, and byte-copies are caught regime-independently
    # by the raw-identity short-circuit.
    #
    # all_dups ALSO has to survive its own self-refutation check: a folder
    # where every file is the same fiber has most pairs near-identical, so
    # frac_high_r must be high.  frac_high_r = 0.00 with an all_dups claim
    # is self-refuting (BKF↔DEL 80 km — see _ALLDUPS_MIN_HIGHR_FRAC), and
    # such folders route to production, where the σ-outlier bulk, the twin
    # gate, and the 0.95-0.99 ramp all apply.
    # sigma_ratio (the standalone Secret Sauce's noise-relative all_dups
    # gate, bulk_sigma / (sqrt(2) * noise_floor) <= 3.0) was evaluated for
    # porting here on 2026-08-29 and CLOSED AS SUPERSEDED.  Measured: the
    # 3.0 threshold is unreachable — real same-fiber duplicate pairs read
    # 5.53 / 6.24 / 10.59 / 14.37 and the loosest folder on disk (LAMBEY)
    # reads 15.3, so it would not fire even on a genuine all-duplicates
    # folder.  Its noise_floor is also quantization-limited rather than a
    # noise measurement: the 2nd-difference MAD lands on integer multiples
    # of the 0.000999 dB Bellcore storage quantum (SEANOR pegged at exactly
    # 1.000, EMVSUI Long 2.004, EMVSUI Short 9.008), so it takes about six
    # values corpus-wide.  And DURANC, the folder it was written for, is
    # already blocked twice here by _robust_common_span (5 broken traces
    # excluded, min_L 6985 -> 89902 m, bulk_r 0.9873 -> 0.5073) and by
    # _ALLDUPS_MIN_SPAN_M.  NOT by _ALLDUPS_MIN_HIGHR_FRAC — DURANC's
    # frac_high_r is 0.7758, which clears 0.5.  The three guards are
    # complementary, not interchangeable.
    alldups_refuted = (bulk_r >= 0.7 and bulk_sigma < 0.10
                       and min_L >= _ALLDUPS_MIN_SPAN_M
                       and frac_high_r < _ALLDUPS_MIN_HIGHR_FRAC)
    if (bulk_r >= 0.7 and bulk_sigma < 0.10
            and min_L >= _ALLDUPS_MIN_SPAN_M
            and frac_high_r >= _ALLDUPS_MIN_HIGHR_FRAC):
        regime = 'all_dups'
    elif alldups_refuted:
        # Self-refuted all_dups claim.  Route PRODUCTION, not tie_panel:
        # bulk_r >= 0.7 would otherwise hand the folder straight to the
        # tie_panel route below, which bypasses σ-outlier — and σ-outlier
        # against a real non-duplicate bulk is exactly the detector a
        # 432-unique-fiber long span needs.
        regime = 'production'
        regime_reason = (f'all_dups refuted: frac high-r {frac_high_r:.2f} '
                         f'< {_ALLDUPS_MIN_HIGHR_FRAC:.2f}')
    elif min_L < 200 and len(files) >= 50:
        regime = 'short_panel'
    elif bulk_r >= 0.7 or frac_high_r >= 0.30:
        regime = 'tie_panel'
    else:
        regime = 'production'
    # Measured on EVERY run, routed on only in the 'production' branch below.
    # The rule fires on no folder on disk (see the _DECAY_* audit note), so
    # this line is how the first folder it does misroute becomes visible.
    names_raw = [files[i]['name'] for i in valid_idx_raw]
    serials_raw = [files[i].get('serial_number') for i in valid_idx_raw]
    decay = _neighbor_decay(names_raw, r_raw, serials_raw)
    if regime == 'production':
        # Additive tie_panel re-routes: only ever applied to folders that
        # landed on 'production' (including via the all_dups refutation).
        _extra = None
        if decay is not None and (decay[0] - decay[1]) >= _DECAY_MIN_DROP:
            regime = 'tie_panel'
            _extra = (f'neighbor-decay: near r {decay[0]:.2f} '
                      f'vs far r {decay[1]:.2f}')
        elif min_L < _SHORT_COMMON_SPAN_M:
            regime = 'tie_panel'
            _extra = f'short common span: {min_L:.0f} m'
        if _extra:
            regime_reason = (f'{regime_reason}; {_extra}' if regime_reason
                             else _extra)
    _reason_sfx = f', {regime_reason}' if regime_reason else ''
    print(f'Regime: {regime} (bulk σ={bulk_sigma:.4f} dB, '
          f'bulk r={bulk_r:.4f}, frac high-r={frac_high_r:.2f}{_reason_sfx})')
    # Diagnostic, always logged, never routed on outside the branch above.
    if decay is not None:
        print(f'Port-distance decay: near r {decay[0]:.4f} vs far r '
              f'{decay[1]:.4f} (drop {decay[0] - decay[1]:.4f}, trigger '
              f'{_DECAY_MIN_DROP:.2f}; {decay[2]} near / {decay[3]} far pairs, '
              f'same instrument)')
    else:
        print('Port-distance decay: not measurable on this folder')
    tie_panel_mode = (regime == 'tie_panel')
    if regime == 'tie_panel':
        # Re-compute with fingerprint extraction (median-trace subtraction)
        # so the r-tier sees per-fiber residuals instead of shared signal.
        batch = _compute_pair_metrics_batch(files, interior_start, interior_end,
                                              tie_panel_mode=True)
    else:
        batch = batch_raw
    sigma_matrix, r_matrix, valid_idx = batch
    # Raw-identity short-circuit inputs: σ is computed on raw traces in BOTH
    # batch passes (tie_panel_mode only changes r), so sigma_matrix is already
    # the raw σ.  Raw r comes from the first (pre-fingerprint) pass.  valid_idx
    # selection is deterministic on (files, window) so the two passes align;
    # guard anyway — a mismatch disables the short-circuit rather than
    # mis-indexing a matrix.
    r_raw_aligned = r_raw if list(valid_idx) == list(valid_idx_raw) else None
    pairs = []
    K = len(valid_idx)
    for ki in range(K):
        i = valid_idx[ki]
        name_i = files[i]['name']
        len_i = files[i].get('length')
        for kj in range(ki + 1, K):
            j = valid_idx[kj]
            len_j = files[j].get('length')
            len_delta = (abs(len_i - len_j) if (len_i and len_j) else None)
            sigma_ij = float(sigma_matrix[ki, kj])
            raw_r_ij = (float(r_raw_aligned[ki, kj])
                        if r_raw_aligned is not None else None)
            pairs.append({
                'a': name_i,
                'b': files[j]['name'],
                'score': sigma_ij,
                'shape_r': float(r_matrix[ki, kj]),
                'shape_r_raw': raw_r_ij,
                'raw_identical': bool(raw_r_ij is not None
                                      and raw_r_ij >= _RAW_IDENT_R
                                      and sigma_ij <= _RAW_IDENT_SIGMA_DB),
                'length_delta_m': len_delta,
            })
    if not pairs:
        raise RuntimeError('No comparable pairs after interior masking')
    print(f'Pair metrics ready: {len(pairs)} pairs')

    scores = np.array([p['score'] for p in pairs], dtype=np.float64)
    p_dup_sigma, stats = _outlier_probability(scores)

    # Pearson-shape contribution. Each regime uses its own r-ramp:
    #   production: (0.95 → 0.99)     standard
    #   tie-panel:  (0.999 → 0.9999)  tightened — fingerprint extraction
    #               on tie panels leaves residual r up to ~0.998 between
    #               physically-different fibers (shared 2-km-scale bend
    #               structure the median can't fully capture). True same-
    #               fiber re-shoots in a tie panel land at r ≥ 0.9999.
    #   all-dups:   (0.85 → 0.95)     widened — every pair is genuinely
    #               a same-fiber re-shoot, so even pairs with r as low as
    #               0.85 (short-fiber shot-noise spread) are real duplicates.
    if regime == 'tie_panel':
        R_LO, R_HI = 0.999, 0.9999
    elif regime == 'all_dups':
        R_LO, R_HI = 0.85, 0.95
    elif regime == 'short_panel':
        # Standard production ramp — true same-fiber re-shoots in a short
        # panel still produce r ≥ 0.95. With σ-outlier disabled below,
        # the r-tier is the entire detector for this regime.
        R_LO, R_HI = 0.95, 0.99
    else:
        R_LO, R_HI = 0.95, 0.99
    _R_SPAN = R_HI - R_LO
    def _r_to_p(r):
        if r is None:
            return 0.0
        if r >= R_HI:
            return 1.0
        if r <= R_LO:
            return 0.0
        return float((r - R_LO) / _R_SPAN)

    p_dup_r = np.array([_r_to_p(p.get('shape_r')) for p in pairs],
                       dtype=np.float64)

    # σ-outlier handling: ONLY production mode trusts it. Every other regime
    # bypasses σ-outlier and lets the regime-specific r-ramp drive the verdict.
    #   production  — standard max(σ-outlier, r-tier) combiner.
    #   tie_panel   — bypass σ. The fingerprint-extracted tight r-ramp
    #                 (0.999-0.9999) is the detector. σ-outlier would cascade
    #                 on shared cable structure: on a 2 km tie panel (CLQTILA)
    #                 ~48% of pairs share enough route structure that σ looks
    #                 like an outlier AND post-fingerprint r still sits above
    #                 0.9, so the old r≥0.9 confirmation gate let 20k false
    #                 positives through. True re-shoots survive (post-FP r→1.0).
    #   all_dups    — no non-duplicate bulk to define an "outlier".
    #   short_panel — short featureless fibers give a narrow σ bulk that
    #                 cascades.
    # ── Shared Rayleigh-speckle context, built at most once per run ───────
    # Both the twin-gate refutation below and the speckle confirmation gate
    # further down read the same folder null and the same per-file windows.
    _by_name = {f['name']: f for f in files}
    # One filter width for the whole folder, from its own acquisition.
    _hp_w = _speckle_hp_width(files)
    if _hp_w != _SPECKLE_HP_WIDTH:
        print(f'Speckle high-pass: {_hp_w} samples '
              f'(pulse-matched; calibrated cap is {_SPECKLE_HP_WIDTH})')
    _spk = {'built': False, 'cache': {}, 'null_q': None}

    def _spk_null():
        """Folder null: what the statistic reads between files KNOWN to be
        different fibers here.  Evenly-spaced sample (no RNG — the run has
        to be reproducible)."""
        if not _spk['built']:
            _spk['built'] = True
            step = max(1, len(files) // _SPECKLE_NULL_FILES)
            null_res = [_speckle_windows(f, interior_start, interior_end,
                                         hp_width=_hp_w)
                        for f in files[::step][:_SPECKLE_NULL_FILES]]
            null_res = [r for r in null_res if r is not None]
            null_vals = [v for a_i in range(len(null_res))
                         for b_i in range(a_i + 1, len(null_res))
                         for v in (_speckle_pair_r(null_res[a_i], null_res[b_i]),)
                         if v is not None]
            if len(null_vals) >= _SPECKLE_NULL_MIN_PAIRS:
                _spk['null_q'] = float(np.percentile(null_vals,
                                                     _SPECKLE_NULL_PCT))
        return _spk['null_q']

    def _spk_win(name):
        if name not in _spk['cache']:
            _spk['cache'][name] = _speckle_windows(_by_name.get(name),
                                                   interior_start, interior_end,
                                                   hp_width=_hp_w)
        return _spk['cache'][name]

    if regime in ('tie_panel', 'all_dups', 'short_panel'):
        p_dup_sigma_eff = np.zeros_like(p_dup_sigma)
    else:
        p_dup_sigma_eff = p_dup_sigma
    # Combined likelihood = max of (possibly confirmed) σ-outlier and r tiers.
    p_dup_raw = np.maximum(p_dup_sigma_eff, p_dup_r)

    # ── Fingerprint rescue from a sigma-bypassed regime ───────────────────
    # tie_panel / all_dups / short_panel throw the sigma-outlier result away
    # wholesale, because on a genuine shared-glass folder it cascades.  That
    # is right for the bulk and wrong for the tail: a pair that is an EXTREME
    # sigma outlier and ALSO carries the fiber's own Rayleigh fingerprint is
    # not shared structure, and zeroing it is the exact failure PR #122
    # repaired on EMVSUI by a different route.
    #
    # MEASURED on MILTOP (Miller->Topeka, 1146 files after break exclusion).
    # It clears the tie_panel trigger by 0.0214 - bulk_r 0.7214 against 0.70 -
    # and the bypass then discards:
    #     MILTOPls0329/0330  sigma 0.00985  p_sigma 0.9991  r 0.9965 -> 0.0000
    #     MILTOPls0830/0831  sigma 0.00941  p_sigma 0.9995  r 0.9964 -> 0.0000
    # Fingerprinted against a 1,770-pair known-different null on that folder
    # (p50 0.0310, p99 0.1066, MAX 0.2470): 329/330 reads 0.8243, which is
    # 3.3x the maximum any different-fiber pair there reaches, with identical
    # EOF.  830/831 reads 0.1225 against a same-fiber floor of 0.3055 and is
    # NOT rescued - the bar is doing real work, not waving both through.
    #
    # WHY THIS CANNOT RE-OPEN THE CASCADES THE REGIME EXISTS TO STOP: the
    # candidate set is empty on every other LONG sigma-bypassed folder on
    # disk.  Measured p_dup_sigma > _SIGMA_RESCUE_MIN: A-F West 0 (the
    # 1,997-FP panel), A-F East 0, BKF<->DEL 0 (the 47-FP set), LAMBEY 0
    # (the 67-FP set), TULORO 0 (the 62k flood), ELMMIL short 0.  MILTOP's 2
    # are the only candidates among them.  A LONG folder whose sigma bulk is
    # genuinely cascading has no extreme outliers to rescue, by construction.
    #
    # THE WORD "LONG" IS LOAD-BEARING.  Short panels are NOT empty: the 20
    # on disk carry 132 to 6,081 candidates each (2,953 on BETA LFY East
    # 144f DW Tray A-F, which is the historic flood number).  Nothing is
    # rescued there today only because the confirm bar is
    # _SPECKLE_CONFIRM_NULL_MULT x the folder's own null p99, and on every
    # span class below 78 km that product EXCEEDS 1.0 - a value a Pearson r
    # cannot take.  Measured bars: EMVSUI Long 78.5 km 0.257 (usable), BETA
    # 62 m 1.718, LSC 31 m 1.800, Reubensville 31 m 2.064, Dinwiddie 2.07 km
    # 2.883, EMVSUI Short 3.99 km 2.927, ELMMIL sh 4.99 km 2.929.
    #
    # So on short panels this rescue is safe by ARITHMETIC, not by the
    # emptiness argument above.  Anyone repairing that bar must re-measure
    # the short-panel candidate sets BEFORE lowering it, or this path
    # inherits the flood.  (Measured 2026-08-31.)
    #
    # Rescued pairs re-enter at their sigma likelihood and then face EVERY
    # downstream gate - length, events, twin, serial and the speckle veto -
    # exactly as a production-regime pair does.
    n_sig_rescued = 0
    if regime in ('tie_panel', 'all_dups', 'short_panel'):
        _cands = [i for i in range(len(pairs))
                  if p_dup_sigma[i] > _SIGMA_RESCUE_MIN]
        if _cands:
            _nq = _spk_null()
            if _nq is not None:
                _bar = _nq * _SPECKLE_CONFIRM_NULL_MULT
                for i in _cands:
                    pr = pairs[i]
                    ra, rb = _spk_win(pr['a']), _spk_win(pr['b'])
                    r_hp = _speckle_pair_r(ra, rb)
                    r_floor = _speckle_same_fiber_floor(ra, rb, pr['score'])
                    if r_hp is None or r_floor is None:
                        continue
                    # Both: clearly above what different fibers do here, AND
                    # at least what the same-fiber hypothesis predicts at this
                    # pair's own sigma.
                    if r_hp < _bar or r_hp < r_floor:
                        continue
                    p_dup_raw[i] = max(p_dup_raw[i], float(p_dup_sigma[i]))
                    pr['sigma_rescued'] = True
                    pr['speckle_r'] = round(float(r_hp), 4)
                    n_sig_rescued += 1
            print(f'Sigma rescue: {len(_cands)} extreme outlier(s) in a '
                  f'{regime} folder, {n_sig_rescued} confirmed by fingerprint')

    # Physical-reality filter: same fiber must produce the same end-of-fiber
    # length to within launch-connector + IOR + sample-resolution variation.
    # Tolerance scales with fiber length but is bounded:
    #   - floor 0.5 m  (launch-mating + OTDR sample resolution dominate at small spans)
    #   - 0.01 % of length above 5 km
    #   - cap 2 m      (avoid being too permissive on 100 km+ spans)
    # When a pair's length delta exceeds tol, cap likelihood at 0.5 (borderline) —
    # different physical fibers can't be the same fiber regardless of how similar
    # their splice profiles look. Pairs with no length info pass through.
    LEN_CAP = 0.5
    def _len_tol_m(length_m):
        # Tolerance accommodates launch-cable-swap systematic offsets (~5 m
        # observed in real re-shoots) but still catches physically-different-
        # fiber routing differences (typically tens to hundreds of meters
        # when paths diverge at closures). The event filter does the
        # fine-grained discrimination — length is just a coarse pre-filter.
        if length_m is None or length_m <= 0:
            return 10.0
        return max(10.0, length_m * 5e-4)
    length_deltas = np.array([(p.get('length_delta_m') or 0.0) for p in pairs], dtype=np.float64)
    has_lengths = np.array([p.get('length_delta_m') is not None for p in pairs])
    # Use the LONGER of the two fibers in the pair to set tolerance.
    name_to_length = {f['name']: (f.get('length') or 0) for f in files}
    pair_max_len = np.array([
        max(name_to_length.get(p['a'], 0), name_to_length.get(p['b'], 0))
        for p in pairs
    ], dtype=np.float64)
    tols = np.array([_len_tol_m(L) for L in pair_max_len], dtype=np.float64)
    length_violation = has_lengths & (length_deltas > tols)

    # Event-table consistency gate: same physical fiber → splice events match
    # in count, position, and loss. Different fibers can share σ/r and even
    # length (paths diverge then reconverge) but their event tables disagree.
    # Only evaluate pairs that survived the σ/r screen, since pairs already
    # at p_dup_raw < 0.1 won't be flagged regardless.
    file_events = {f['name']: f.get('events') for f in files}
    events_violation = np.zeros(len(pairs), dtype=bool)
    EVENT_CHECK_THRESHOLD = 0.10
    for i, p in enumerate(pairs):
        if p_dup_raw[i] < EVENT_CHECK_THRESHOLD:
            continue
        (n_match, n_max, n_min, mean_dloss, max_dloss,
         median_dloss, n_max_sig) = _event_match_quality(
            file_events.get(p['a']), file_events.get(p['b']))
        p['events_n_match'] = int(n_match)
        p['events_n_max']   = int(n_max)
        p['events_n_min']   = int(n_min)
        p['events_mean_dloss_db'] = float(mean_dloss)
        p['events_max_dloss_db']  = float(max_dloss)
        p['events_median_dloss_db'] = float(median_dloss)
        p['events_n_max_significant'] = int(n_max_sig)
        if not _events_agree(n_match, n_max, n_min, mean_dloss,
                             median_dloss_db=median_dloss,
                             n_max_significant=n_max_sig):
            events_violation[i] = True
            # Distinguish "the tables disagree" from "the table is present
            # but too thin to check" in the internals (both cap
            # identically).  See _events_agree: BKFDEL028/040 are the only
            # 2 of 432 files on that span with <= 2 interior events, and
            # every one of the 47 false positives contained one of them.
            if n_min < 3:
                p['events_unverifiable'] = True

    # Uniqueness (twin) gate — ALL regimes (2026-07-31; was production-only,
    # which took it off the board on exactly the misroutes it guards
    # against).  For each pair that would flag, ask whether the two files
    # are each other's UNIQUE twin: pair σ must be ≤ _UNIQ_TWIN_RATIO x the
    # smaller of the two members' next-best σ against anyone else.
    # Family/ladder members (several equally-close partners) fail; a
    # genuine re-shoot or copy passes even on a pristine featureless span
    # (its twin is still several x closer than the field).  σ is the RAW
    # pair σ in every regime, so the comparison means the same thing
    # everywhere.  Verdict-level guard only — flag or don't, no review
    # tier; and the raw-identity short-circuit still overrides it, so a
    # literal copy can never be capped by this.
    uniq_violation = np.zeros(len(pairs), dtype=bool)
    Ksz = sigma_matrix.shape[0]
    sig_self_inf = sigma_matrix + np.diag(np.full(Ksz, np.inf))
    sig_sorted = np.sort(sig_self_inf, axis=1)
    best1, best2 = sig_sorted[:, 0], sig_sorted[:, 1]
    arg_sorted = np.argsort(sig_self_inf, axis=1)
    arg1, arg2 = arg_sorted[:, 0], arg_sorted[:, 1]
    uniq_rival = {}                       # pair index -> (row, rival row)
    pidx = 0
    for ki in range(Ksz):
        for kj in range(ki + 1, Ksz):
            if p_dup_raw[pidx] > 0.5:
                s = float(sigma_matrix[ki, kj])
                take_i = s <= best1[ki]
                take_j = s <= best1[kj]
                nb_i = float(best2[ki] if take_i else best1[ki])
                nb_j = float(best2[kj] if take_j else best1[kj])
                if s > _UNIQ_TWIN_RATIO * min(nb_i, nb_j):
                    uniq_violation[pidx] = True
                    pairs[pidx]['uniq_next_best_db'] = round(min(nb_i, nb_j), 4)
                    # Remember the file that raised the objection so the
                    # fingerprint can be asked about that specific rival.
                    if nb_i <= nb_j:
                        owner = ki
                        rival = int(arg2[ki] if take_i else arg1[ki])
                    else:
                        owner = kj
                        rival = int(arg2[kj] if take_j else arg1[kj])
                    uniq_rival[pidx] = (owner, rival)
            pidx += 1

    # Different-OTDR gate: duplication (a copied file, or the same fiber
    # re-shot and presented as another) is a SINGLE-instrument phenomenon.
    # A pair acquired by two different physical OTDRs is two independent
    # acquisitions — it can only be "the same data" at raw-identity grade,
    # and the raw-identity short-circuit below fires regardless of this
    # cap.  (Lumen Border LAMBEY170/241: serials 1876272 vs 1978245, 84
    # min apart, 3.4 dB injection delta, r 0.992 on a pristine span —
    # and the SAME two fibers from the OPPOSITE end read σ 0.05, five
    # times the flag level.  Different boxes -> different fibers.)
    # Fails open when either serial is missing.
    name_to_serial = {f['name']: f.get('serial_number') for f in files}
    serial_violation = np.zeros(len(pairs), dtype=bool)
    for i, p in enumerate(pairs):
        if p_dup_raw[i] <= 0.5:
            continue
        sa, sb_ = name_to_serial.get(p['a']), name_to_serial.get(p['b'])
        if sa and sb_ and sa != sb_ and not p.get('raw_identical'):
            serial_violation[i] = True
            p['serial_mismatch'] = f'{sa} != {sb_}'

    # ── Twin-gate refutation by fingerprint ───────────────────────────────
    # The twin gate asks "is this pair's partner UNIQUE?" and answers it
    # with a sigma ratio, which is a proxy.  The Rayleigh speckle answers
    # the same question by direct measurement, so where the two disagree
    # the measurement decides.  A pair is restored only when BOTH hold:
    # the pair itself fingerprints far above what different fibers produce
    # in this folder, AND the specific rival that raised the objection
    # fingerprints at the null, i.e. is demonstrably NOT a second twin.
    #
    # Measured on EMVSUI0 Long Shots (folder null p99 = 0.086):
    #   563/564 sigma 0.0182  fingerprint 0.6150   <- confirmed duplicate
    #     rival 564/566 sigma 0.0331  fingerprint 0.0170  <- different fiber
    #   296/308 sigma 0.0246  fingerprint 0.4780   <- confirmed duplicate
    #     rival 308/336 sigma 0.0350  fingerprint 0.0739  <- different fiber
    # Both were capped at 0.5 because their twin was "only" 1.8x and 1.4x
    # closer than a rival that shares no fingerprint with them at all.
    #
    # This CANNOT re-open a ribbon-ladder false positive (the case the twin
    # gate was built for): a ladder pair fails the first condition — it has
    # no shared fingerprint either — so it never reaches the rival test.
    # Applies only to pairs whose ONLY objection is the twin gate; length,
    # events and serial violations are untouched.
    n_uniq_refuted = 0
    twin_only = [i for i in range(len(pairs))
                 if uniq_violation[i] and not (length_violation[i]
                                               or events_violation[i]
                                               or serial_violation[i])]
    if twin_only:
        nq = _spk_null()
        if nq is not None:
            bar = nq * _SPECKLE_CONFIRM_NULL_MULT
            for i in twin_only:
                p = pairs[i]
                r_pair = _speckle_pair_r(_spk_win(p['a']), _spk_win(p['b']))
                if r_pair is None or r_pair < bar:
                    continue
                owner, rival = uniq_rival.get(i, (None, None))
                if owner is None:
                    continue
                r_rival = _speckle_pair_r(
                    _spk_win(files[valid_idx[owner]]['name']),
                    _spk_win(files[valid_idx[rival]]['name']))
                if r_rival is None or r_rival >= nq:
                    continue
                uniq_violation[i] = False
                p['uniq_refuted_by_speckle'] = True
                p['uniq_rival_speckle_r'] = round(r_rival, 4)
                n_uniq_refuted += 1
    if n_uniq_refuted:
        print(f'Twin gate: {n_uniq_refuted} of {len(twin_only)} sigma-ratio '
              f'objection(s) refuted by the fingerprint')

    physical_violation = (length_violation | events_violation
                          | uniq_violation | serial_violation)
    p_dup = np.where(physical_violation, np.minimum(p_dup_raw, LEN_CAP), p_dup_raw)

    # Rayleigh-speckle confirmation gate (see the _SPECKLE_* calibration
    # block).  LAST gate before the raw-identity short-circuit: a pair still
    # standing above the print threshold is demoted when its sub-pulse-width
    # backscatter fingerprint sits far below anything the same-fiber
    # hypothesis could produce at that pair's own σ — and only when the
    # statistic has been shown to separate same from different on THIS
    # folder at THAT σ.  Otherwise the gate abstains.  Only the survivors
    # are measured, so the cost is O(files in candidate pairs) + the one
    # folder-null sample, not O(pairs).
    speckle_violation = np.zeros(len(pairs), dtype=bool)
    cand = [i for i in range(len(pairs)) if p_dup[i] > 0.5]
    n_unmeas = n_abstain = 0
    null_q = None
    if cand:
        # Same windows and same folder null the twin-gate refutation used;
        # built once, on first demand, either here or up there.
        null_q = _spk_null()
        for i in cand:
            p = pairs[i]
            ra, rb = _spk_win(p['a']), _spk_win(p['b'])
            r_hp = _speckle_pair_r(ra, rb)
            r_floor = _speckle_same_fiber_floor(ra, rb, p['score'])
            if r_hp is None or r_floor is None or null_q is None:
                p['speckle_unmeasurable'] = True     # fail-safe: no veto
                n_unmeas += 1
                continue
            p['speckle_r'] = round(r_hp, 4)
            p['speckle_floor'] = round(r_floor, 4)
            if r_floor < null_q:
                # At this pair's σ the statistic cannot separate a
                # same-fiber re-shoot from two random fibers here.  Abstain.
                p['speckle_abstain'] = True
                n_abstain += 1
                continue
            if r_hp <= r_floor / _SPECKLE_FLOOR_MARGIN:
                speckle_violation[i] = True
                p['speckle_capped'] = True
        p_dup = np.where(speckle_violation, np.minimum(p_dup, LEN_CAP), p_dup)
    # Always logged (even at 0 candidates) so the gate is auditable from any
    # run log — silence would be indistinguishable from the gate not running.
    print(f'Speckle gate: {len(cand)} candidate pair(s), '
          f'{int(speckle_violation.sum())} demoted, '
          f'{n_abstain} inconclusive (kept), '
          f'{n_unmeas} unmeasurable (kept)'
          + (f', folder null p{_SPECKLE_NULL_PCT:.0f}={null_q:.3f}'
             if null_q is not None else ''))

    # Raw-identity short-circuit: a pair whose RAW interior trace is the
    # same data (σ ≤ 0.001 dB, r ≥ 0.98 — see the calibration block above)
    # is a CONFIRMED copy regardless of regime routing.  Applied last so no
    # regime bypass (tie_panel fingerprint subtraction, all_dups σ bypass),
    # stored-event-table disagreement, or gate above it (twin, speckle) can
    # hide a literal file copy: the trace itself is the identity proof.
    # Raises to 1.0 only — never lowers.  (Byte-identical copies also pass
    # the speckle gate trivially: identical traces give r_hp = 1.0.)
    raw_ident_mask = np.array([bool(p.get('raw_identical')) for p in pairs],
                              dtype=bool)
    if raw_ident_mask.any():
        p_dup = np.where(raw_ident_mask, 1.0, p_dup)

    for i, p in enumerate(pairs):
        p['p_dup_sigma']   = float(p_dup_sigma[i])
        p['p_dup_r']       = float(p_dup_r[i])
        p['p_dup_raw']     = float(p_dup_raw[i])
        p['p_dup']         = float(p_dup[i])
        p['length_capped'] = bool(length_violation[i])
        p['events_capped'] = bool(events_violation[i])
        p['z']             = float(stats['z'][i])

    order = np.argsort(scores)
    n99 = int((p_dup > 0.99).sum())
    n50 = int((p_dup > 0.5).sum())
    n10 = int((p_dup > 0.1).sum())
    print(f'Likelihood >99%: {n99}   >50%: {n50}   >10%: {n10}')

    # For each file, pick the partner that gives the HIGHEST duplicate
    # likelihood (tie-broken by smallest disagreement). This ensures the
    # per-file table is symmetric: if pair (A,B) is the most-likely
    # duplicate for both A and B, both rows point at each other. Earlier
    # logic picked by smallest σ alone, which could leave a confirmed-
    # duplicate flag on one row while the partner's row pointed elsewhere.
    best_partner = {}
    for idx, f in enumerate(files):
        best = None
        for p in pairs:
            if f['name'] not in (p['a'], p['b']):
                continue
            if best is None:
                best = p
            elif (p['p_dup'] > best['p_dup']
                  or (p['p_dup'] == best['p_dup'] and p['score'] < best['score'])):
                best = p
        best_partner[f['name']] = best

    return {
        'files': files,
        'pairs': pairs,
        'scores': scores,
        'stats': stats,
        'p_dup': p_dup,
        'best_partner': best_partner,
        'n99': n99, 'n50': n50, 'n10': n10,
        'interior_start': interior_start, 'interior_end': interior_end,
        'min_L': min_L,
        'short_traces': short_traces,
        'window_guard': window_guard,
        'order_by_score': order,
        'regime': regime,
        'regime_reason': regime_reason,
        'bulk_sigma': bulk_sigma,
        'bulk_r': bulk_r,
        'frac_high_r': frac_high_r,
    }


def build_report_sor(folder, title, out_pdf, meta=None):
    analysis = _analyze_sor(folder)
    if meta is not None:
        # Additive side-channel for the runner's manifest (`short_traces`);
        # optional so every existing caller is untouched.
        meta['short_traces'] = analysis.get('short_traces') or []
        if analysis.get('window_guard'):
            meta['window_guard'] = analysis['window_guard']
        # The counts the REPORT prints, so the caller can stop recomputing
        # its own (see run_sor_bytes).
        meta['n_files'] = len(analysis['files'])
        meta['n_pairs'] = len(analysis['pairs'])
    files = analysis['files']
    pairs = analysis['pairs']
    scores = analysis['scores']
    stats = analysis['stats']
    p_dup = analysis['p_dup']
    best_partner = analysis['best_partner']
    n99, n50, n10 = analysis['n99'], analysis['n50'], analysis['n10']
    order = analysis['order_by_score']

    verdict_block = (f'<div class="verdict-box verdict-confirm">'
                     f'<b>{n50} duplicate pair(s) identified</b> at ≥50% likelihood; '
                     f'{n99} at ≥99% likelihood across {len(pairs)} pairs.</div>'
                     if n50 else
                     '<div class="verdict-box verdict-dispute">'
                     '<b>No duplicate pairs identified</b> at ≥50% likelihood.</div>')

    shape_rs = [p.get('shape_r') for p in pairs]
    dist_chart = _distribution_chart(scores, p_dup, stats, shape_rs=shape_rs)

    # '' when the folder has no suspected breaks — unaffected reports stay
    # byte-stable (no empty section, no renumbering).
    short_block = _short_trace_section_html(
        analysis.get('short_traces'),
        window_guard=analysis.get('window_guard'))

    file_by_name = {f['name']: f for f in files}

    def _gap_str(name_a, name_b):
        _fa, _fb = file_by_name.get(name_a), file_by_name.get(name_b)
        _ta = _fa.get('timestamp') if _fa else None
        _tb = _fb.get('timestamp') if _fb else None
        return _fmt_time_gap(abs(_ta - _tb)) if _ta and _tb else '—'

    file_rows = ''
    for f in sorted(files, key=lambda x: x['name']):
        bp = best_partner.get(f['name'])
        if bp is None:
            continue
        partner = bp['b'] if bp['a'] == f['name'] else bp['a']
        pd_val = bp['p_dup']
        pd_color = '#2d8f48' if pd_val > 0.9 else ('#b97000' if pd_val > 0.1 else '#888')
        verdict_cell = (f'<span class="dup">DUPLICATE of {partner}</span>'
                        if pd_val > 0.5 else
                        f'<span class="na">unique (closest: {partner})</span>')
        loss_cell = f'{f["loss"]:.3f}' if f['loss'] is not None else '—'
        r_val = bp.get('shape_r')
        r_cell = ('<td class="center na">—</td>' if r_val is None else
                  f'<td class="center" style="color:{_shape_color(r_val)};font-weight:600">{r_val:.4f}</td>')
        file_rows += (f'<tr><td class="pair-cell">{f["name"]}</td>'
                      f'<td class="center">{f["length"]/1000:.3f}</td>'
                      f'<td class="center">{_gap_str(f["name"], partner)}</td>'
                      f'<td class="center">{loss_cell}</td>'
                      f'<td class="center">{bp["score"]:.4f}</td>'
                      f'<td class="center" style="color:{pd_color};font-weight:600">{pd_val*100:.2f}%</td>'
                      f'{r_cell}'
                      f'<td class="center">{verdict_cell}</td></tr>')

    top_rows = ''
    for rank, k in enumerate(order[:30], 1):
        p = pairs[k]
        pd_val = p['p_dup']
        pd_color = '#2d8f48' if pd_val > 0.9 else ('#b97000' if pd_val > 0.1 else '#888')
        r_val = p.get('shape_r')
        r_cell = ('<td class="center na">—</td>' if r_val is None else
                  f'<td class="center" style="color:{_shape_color(r_val)};font-weight:600">{r_val:.4f}</td>')
        top_rows += (f'<tr><td class="center">{rank}</td>'
                     f'<td class="pair-cell">{p["a"]} ↔ {p["b"]}</td>'
                     f'<td class="center">{_gap_str(p["a"], p["b"])}</td>'
                     f'<td class="center">{p["score"]:.4f}</td>'
                     f'<td class="center" style="color:{pd_color};font-weight:600">{pd_val*100:.2f}%</td>'
                     f'{r_cell}</tr>')

    # Top 30 by similarity (highest first). Skip pairs where similarity is None.
    sim_pairs = [(i, p) for i, p in enumerate(pairs) if p.get('shape_r') is not None]
    sim_order = sorted(sim_pairs, key=lambda x: -x[1]['shape_r'])[:30]
    sim_rows = ''
    for rank, (k, p) in enumerate(sim_order, 1):
        pd_val = p['p_dup']
        pd_color = '#2d8f48' if pd_val > 0.9 else ('#b97000' if pd_val > 0.1 else '#888')
        r_val = p['shape_r']
        sim_rows += (f'<tr><td class="center">{rank}</td>'
                     f'<td class="pair-cell">{p["a"]} ↔ {p["b"]}</td>'
                     f'<td class="center">{_gap_str(p["a"], p["b"])}</td>'
                     f'<td class="center" style="color:{_shape_color(r_val)};font-weight:600">{r_val:.4f}</td>'
                     f'<td class="center">{p["score"]:.4f}</td>'
                     f'<td class="center" style="color:{pd_color};font-weight:600">{pd_val*100:.2f}%</td></tr>')

    # Confirmed-duplicate detail table (p_dup > 0.5)
    dup_pairs_sorted = sorted([p for p in pairs if p['p_dup'] > 0.5],
                              key=lambda q: -q['p_dup'])
    # PDF cap (Zach 2026-07-21): an all_dups folder produced 62,014 pairs
    # >=50% — the unbounded table blew Chrome's print budget and crashed the
    # run.  The PDF renders the top PDF_DUP_ROWS_CAP by likelihood with an
    # overflow note; the Excel report always carries the complete list.
    from report import _capped_rows, PDF_DUP_ROWS_CAP
    dup_pairs_render, dup_overflow = _capped_rows(dup_pairs_sorted,
                                                  PDF_DUP_ROWS_CAP)
    dup_detail_rows = ''
    for p in dup_pairs_render:
        fa = file_by_name.get(p['a']); fb = file_by_name.get(p['b'])
        if fa is None or fb is None:
            continue
        ta, tb = fa.get('timestamp'), fb.get('timestamp')
        gap_str = _fmt_time_gap(abs(ta - tb)) if ta and tb else '—'
        a_sl, b_sl = fa.get('loss'), fb.get('loss')
        # Max splice Δ at MATCHED events (For-Romeo style): for each splice
        # closure that exists in both fibers, |Δloss|, then max across closures.
        # Falls back to '—' when no events were matched.
        max_dloss = p.get('events_max_dloss_db')
        n_match_pair = p.get('events_n_match', 0)
        ms_cell = (f'<td class="center">{max_dloss*1000:.0f}</td>'
                   if max_dloss is not None and n_match_pair >= 1
                   else '<td class="center na">—</td>')
        sl_cell = (f'<td class="center">{abs(a_sl - b_sl)*1000:.0f}</td>'
                   if a_sl is not None and b_sl is not None
                   else '<td class="center na">—</td>')
        # Same OTDR serial → both shots came from the same instrument.
        sn_a, sn_b = fa.get('serial_number'), fb.get('serial_number')
        if sn_a and sn_b:
            same_sn = (sn_a == sn_b)
            sn_cell = (f'<td class="center" style="color:#2d8f48;font-weight:700">Yes</td>'
                       if same_sn else
                       f'<td class="center" style="color:#c0392b;font-weight:700">No</td>')
        else:
            sn_cell = '<td class="center na">—</td>'
        pd_val = p['p_dup']
        pd_color = '#2d8f48' if pd_val > 0.9 else '#b97000'
        r_val = p.get('shape_r')
        r_cell = ('<td class="center na">—</td>' if r_val is None else
                  f'<td class="center" style="color:{_shape_color(r_val)};font-weight:600">{r_val:.4f}</td>')
        dup_detail_rows += (f'<tr><td class="pair-cell">{p["a"]} ↔ {p["b"]}</td>'
                            f'<td class="center">{gap_str}</td>'
                            f'{ms_cell}{sl_cell}{r_cell}{sn_cell}'
                            f'<td class="center" style="color:{pd_color};font-weight:600">{pd_val*100:.2f}%</td></tr>')
    # Boss request (2026-07-15): duplicates lead the report — this block is
    # section 1 on page one, with an explicit "none" line when the folder is
    # clean so the verdict is visible at a glance.
    if dup_detail_rows:
        wl_hdr = f'{int(files[0].get("wavelength") or 0)} nm' if files else ''
        dup_detail_block = f'''
<div class="section-block">
<div class="dir-banner">1. Confirmed duplicate pairs (≥50% likelihood) — detail ({wl_hdr})</div>
<table class="vote-table">
<tr><th style="text-align:left">Pair</th><th>Time gap</th>
  <th>max splice Δ (mdB)</th><th>span loss Δ (mdB)</th>
  <th>similarity</th><th>Same OTDR</th><th>Duplicate likelihood</th></tr>
{dup_detail_rows}
</table>
{('<div style="padding:8px 4px;color:#b97000;font-weight:600">… and '
  f'{dup_overflow:,} more pairs at ≥50% likelihood — the complete list is '
  'in the Excel report.</div>') if dup_overflow else ''}
</div>
'''
    else:
        dup_detail_block = (
            '<div class="section-block">'
            '<div class="dir-banner">1. Confirmed duplicate pairs (\u226550% likelihood)</div>'
            '<div style="padding:10px 4px;color:#2d8f48;font-weight:600">'
            'None \u2014 no pairs at \u226550% duplicate likelihood.</div></div>')

    generated = datetime.now().strftime('%Y-%m-%d %H:%M')
    html = f'''<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<title>{title}</title>
<style>{_BASE_CSS}</style></head><body>
{_embed_logo()}
<h1>{title}</h1>
<div class="subtitle">{len(files)} files &bull; {len(pairs)} pairs &bull; generated {generated}</div>

{verdict_block}

{dup_detail_block}
{short_block}
<div class="section-block">
<div class="dir-banner">2. Distribution</div>
<img src="data:image/png;base64,{dist_chart}" class="chart-img" />
</div>

<div class="cards">
  <div class="card"><div class="card-label">Files</div><div class="card-value">{len(files)}</div></div>
  <div class="card"><div class="card-label">Pairs</div><div class="card-value">{len(pairs)}</div></div>
  <div class="card"><div class="card-label">Likelihood &gt; 99%</div>
    <div class="card-value good">{n99}</div></div>
  <div class="card"><div class="card-label">Likelihood &gt; 50%</div>
    <div class="card-value">{n50}</div></div>
  <div class="card"><div class="card-label">Likelihood &gt; 10%</div>
    <div class="card-value">{n10}</div></div>
</div>

<div class="section-block">
<div class="dir-banner">3. Per-file verdict</div>
<table class="vote-table">
<tr><th style="text-align:left">File</th>
    <th>Length (km)</th><th>Time gap (closest)</th><th>Span loss (dB)</th>
    <th>lowest disagreement</th><th>Duplicate likelihood</th>
    <th>similarity</th><th>Verdict</th></tr>
{file_rows}
</table>
</div>

<div class="section-block">
<div class="dir-banner">4. Top 30 pairs — lowest level of disagreement</div>
<table class="vote-table">
<tr><th>Rank</th><th style="text-align:left">Pair</th><th>Time gap</th>
    <th>level of disagreement</th><th>Duplicate likelihood</th><th>similarity</th></tr>
{top_rows}
</table>
</div>

<div class="section-block">
<div class="dir-banner">5. Top 30 pairs — highest similarity</div>
<table class="vote-table">
<tr><th>Rank</th><th style="text-align:left">Pair</th><th>Time gap</th>
    <th>similarity</th><th>level of disagreement</th><th>Duplicate likelihood</th></tr>
{sim_rows}
</table>
</div>
</body></html>'''

    pdf_bytes = html_to_pdf_bytes(html, base_url=folder)
    with open(out_pdf, 'wb') as fh:
        fh.write(pdf_bytes)
    print(f'PDF:  {out_pdf}')
    return out_pdf


def run_sor_bytes(folder, title, meta=None):
    """Run SOR mode and return (pdf_bytes, n_files, n_pairs).  Pass a dict
    as `meta` to receive additive analysis facts (currently
    `short_traces`) without changing the return contract."""
    import tempfile
    _meta = meta if meta is not None else {}
    with tempfile.TemporaryDirectory() as td:
        tmp_pdf = os.path.join(td, 'report.pdf')
        build_report_sor(folder, title, tmp_pdf, meta=_meta)
        with open(tmp_pdf, 'rb') as fh:
            pdf_bytes = fh.read()
    # Report what was ANALYSED, not what was globbed.  These used to
    # recount the staged folder after rendering, which disagreed with the
    # report's own header on any span where a trace is excluded:
    #
    #   ELMDALE TO MILER   glob 1152 / 662,976   analysed 1151 / 661,825
    #                      (ELMMIL0231_1550 ends at 22,288 m against a
    #                       69,567 m median — a real break)
    #   DURANC 1-144       glob  144 /  10,296   analysed  141 /   9,870
    #
    # The glob number reached the download-button label and the green
    # "N SOR files processed" line while the workbook's own Summary sheet
    # printed the smaller one.  Same folder, two numbers, no explanation.
    #
    # The glob was also case-sensitive on a path that is not: _inventory
    # matches on a lowercased name, so a file saved as .SOR was inventoried
    # and staged but missed here.  On POSIX that made the count too LOW and
    # the trace was silently dropped from the analysis; on Windows
    # ntpath.normcase lowercases, so the same folder behaved differently in
    # the field than on the dev machine.  Reading the analysis removes the
    # second parser entirely rather than teaching it the same rules.
    n_files = _meta.get('n_files', 0)
    n_pairs = _meta.get('n_pairs', 0)
    return pdf_bytes, n_files, n_pairs


def build_xlsx_sor(folder, title, out_xlsx, meta=None):
    """SOR-mode Excel renderer. Same analysis as build_report_sor, but
    output is an .xlsx workbook with one sheet per table (no rendered
    charts — Excel users typically filter / sort the raw numbers).

    Sheets:
      Summary                — header counts and verdict
      Suspected short fibers — only when suspected breaks exist
      Per-file verdict
      Confirmed duplicates   — pairs at ≥50% likelihood, with detail columns
      Top 30 — lowest disagreement
      Top 30 — highest similarity
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XlsxImage

    analysis = _analyze_sor(folder)
    if meta is not None:
        meta['short_traces'] = analysis.get('short_traces') or []
        if analysis.get('window_guard'):
            meta['window_guard'] = analysis['window_guard']
        meta['n_files'] = len(analysis['files'])
        meta['n_pairs'] = len(analysis['pairs'])
    files = analysis['files']
    pairs = analysis['pairs']
    best_partner = analysis['best_partner']
    n99, n50, n10 = analysis['n99'], analysis['n50'], analysis['n10']
    order = analysis['order_by_score']

    wb = Workbook()

    # Unified font: Calibri 12 everywhere. Bold variant for headers and
    # labels keeps the same size/family for visual consistency.
    BASE = Font(name='Calibri', size=12)
    BASE_BOLD = Font(name='Calibri', size=12, bold=True)
    TITLE_FONT = Font(name='Calibri', size=14, bold=True)
    HDR_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='2C3E50')

    # ---------- Summary ----------
    ws = wb.active
    ws.title = 'Summary'

    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f'Generated {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = BASE

    rows = [
        ('Files', len(files)),
        ('Pairs', len(pairs)),
        ('Regime', analysis.get('regime', 'production')),
    ]
    # Only present when one of the additive tie_panel routes fired, so
    # reports from unaffected folders keep their exact row layout.
    if analysis.get('regime_reason'):
        rows.append(('Regime reason', analysis['regime_reason']))
    rows += [
        ('Bulk pair-σ (dB)', f'{analysis.get("bulk_sigma", 0.0):.4f}'),
        ('Bulk pair-r',      f'{analysis.get("bulk_r", 0.0):.4f}'),
        ('Frac pairs r≥0.95', f'{analysis.get("frac_high_r", 0.0):.2f}'),
        ('Likelihood ≥ 99%', n99),
        ('Likelihood ≥ 50%', n50),
        ('Likelihood ≥ 10%', n10),
        ('Common span (m)', f'{analysis["min_L"]:.1f}'),
        ('Interior window (m)',
         f'{analysis["interior_start"]:.0f}–{analysis["interior_end"]:.0f}'),
    ]
    # Only when suspected breaks exist — unaffected Summary layouts stay
    # byte-stable (same pattern as the Regime-reason row above).
    short_traces = analysis.get('short_traces') or []
    if short_traces:
        rows.append(('Suspected short fibers', len(short_traces)))
    if analysis.get('window_guard'):
        rows.append(('Window warning', analysis['window_guard']))
    for i, (k, v) in enumerate(rows, start=4):
        c1 = ws.cell(row=i, column=1, value=k); c1.font = BASE_BOLD
        c2 = ws.cell(row=i, column=2, value=v); c2.font = BASE
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 24

    def _write_table(ws, headers, rows_data, col_widths=None):
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = hdr_fill
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal='center')
        for r, row in enumerate(rows_data, start=2):
            for c, v in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = BASE
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = (f'A1:{get_column_letter(len(headers))}'
                              f'{1 + len(rows_data)}')
        if col_widths:
            for c, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(c)].width = w

    # ---------- Suspected short fibers (only when present) ----------
    if short_traces:
        ws = wb.create_sheet('Suspected short fibers', 1)   # after Summary
        headers = ['File', 'Ends at (m)', 'Folder median (m)',
                   'Excluded from pairs', 'Finding']
        rows_data = []
        for e in short_traces:
            finding = 'suspected break'
            if e.get('break_note'):
                finding += f' — {e["break_note"]}'
            rows_data.append([e['file'], e['eof_m'], e['median_eof_m'],
                              'Yes' if e.get('excluded') else 'No', finding])
        _write_table(ws, headers, rows_data,
                     col_widths=[18, 13, 17, 18, 72])

    # ---------- Per-file verdict ----------
    ws = wb.create_sheet('Per-file verdict')
    headers = ['File', 'Length (km)', 'Span loss (dB)',
               'Lowest disagreement', 'Duplicate likelihood (%)',
               'Similarity', 'Best partner', 'Verdict']
    rows_data = []
    for f in sorted(files, key=lambda x: x['name']):
        bp = best_partner.get(f['name'])
        if bp is None:
            rows_data.append([f['name'], None, None, None, None, None, None, '—'])
            continue
        partner = bp['b'] if bp['a'] == f['name'] else bp['a']
        verdict = (f'DUPLICATE of {partner}' if bp['p_dup'] > 0.5
                   else f'unique (closest: {partner})')
        rows_data.append([
            f['name'],
            (f['length'] / 1000.0) if f.get('length') else None,
            f.get('loss'),
            bp['score'],
            bp['p_dup'] * 100.0,
            bp.get('shape_r'),
            partner,
            verdict,
        ])
    _write_table(ws, headers, rows_data,
                 col_widths=[18, 12, 14, 18, 22, 12, 20, 32])

    # ---------- Confirmed duplicates (≥50% likelihood) ----------
    ws = wb.create_sheet('Confirmed duplicates')
    headers = ['Pair A', 'Pair B', 'Time gap (s)',
               'Max splice Δ at matched events (mdB)',
               'Span loss Δ (mdB)', 'Similarity', 'Same OTDR',
               'Duplicate likelihood (%)']
    file_by_name = {f['name']: f for f in files}
    dup_sorted = sorted([p for p in pairs if p['p_dup'] > 0.5],
                        key=lambda q: -q['p_dup'])
    rows_data = []
    for p in dup_sorted:
        fa = file_by_name.get(p['a'])
        fb = file_by_name.get(p['b'])
        ta, tb = (fa.get('timestamp') if fa else None,
                  fb.get('timestamp') if fb else None)
        gap = abs(ta - tb) if ta and tb else None
        a_sl = fa.get('loss') if fa else None
        b_sl = fb.get('loss') if fb else None
        sl_d = abs(a_sl - b_sl) * 1000 if a_sl is not None and b_sl is not None else None
        max_d = p.get('events_max_dloss_db')
        ms_d = max_d * 1000 if (max_d is not None and p.get('events_n_match', 0) >= 1) else None
        sn_a = fa.get('serial_number') if fa else None
        sn_b = fb.get('serial_number') if fb else None
        if sn_a and sn_b:
            same_sn = 'Yes' if sn_a == sn_b else 'No'
        else:
            same_sn = '—'
        rows_data.append([
            p['a'], p['b'], gap, ms_d, sl_d,
            p.get('shape_r'), same_sn, p['p_dup'] * 100.0,
        ])
    _write_table(ws, headers, rows_data,
                 col_widths=[18, 18, 13, 32, 18, 12, 11, 22])

    # ---------- Top 30 — lowest disagreement ----------
    def _gap_s(name_a, name_b):
        _fa, _fb = file_by_name.get(name_a), file_by_name.get(name_b)
        _ta = _fa.get('timestamp') if _fa else None
        _tb = _fb.get('timestamp') if _fb else None
        return abs(_ta - _tb) if _ta and _tb else None

    ws = wb.create_sheet('Top 30 lowest disagreement')
    headers = ['Rank', 'Pair A', 'Pair B', 'Time gap (s)',
               'Level of disagreement',
               'Duplicate likelihood (%)', 'Similarity']
    rows_data = []
    for rank, k in enumerate(order[:30], 1):
        p = pairs[k]
        rows_data.append([
            rank, p['a'], p['b'], _gap_s(p['a'], p['b']), p['score'],
            p['p_dup'] * 100.0, p.get('shape_r'),
        ])
    _write_table(ws, headers, rows_data,
                 col_widths=[6, 18, 18, 13, 22, 22, 12])

    # ---------- Top 30 — highest similarity ----------
    ws = wb.create_sheet('Top 30 highest similarity')
    headers = ['Rank', 'Pair A', 'Pair B', 'Time gap (s)', 'Similarity',
               'Level of disagreement', 'Duplicate likelihood (%)']
    sim_sorted = sorted([(i, p) for i, p in enumerate(pairs)
                         if p.get('shape_r') is not None],
                        key=lambda x: -x[1]['shape_r'])[:30]
    rows_data = []
    for rank, (_, p) in enumerate(sim_sorted, 1):
        rows_data.append([
            rank, p['a'], p['b'], _gap_s(p['a'], p['b']), p['shape_r'],
            p['score'], p['p_dup'] * 100.0,
        ])
    _write_table(ws, headers, rows_data,
                 col_widths=[6, 18, 18, 13, 12, 22, 22])

    # ---------- Charts ----------
    # Generate the same 2x2 distribution chart used in the PDF and embed
    # it on its own sheet so Excel users have the visual context too.
    try:
        shape_rs = [p.get('shape_r') for p in pairs]
        chart_b64 = _distribution_chart(
            analysis['scores'], analysis['p_dup'], analysis['stats'],
            shape_rs=shape_rs)
        png_bytes = base64.b64decode(chart_b64)
        img_buf = BytesIO(png_bytes)
        img = XlsxImage(img_buf)
        # Matplotlib rendered at figsize (13, 6) at 150 dpi → ~1950×900 px
        # native. Keep aspect ratio while scaling to a sensible Excel width.
        orig_w, orig_h = img.width, img.height
        target_w = 1400  # matches the PDF body's max content width
        img.width = target_w
        img.height = int(target_w * orig_h / orig_w) if orig_w else target_w // 2
        ws = wb.create_sheet('Charts')
        ws['A1'] = 'Distribution charts'
        ws['A1'].font = TITLE_FONT
        ws.add_image(img, 'A3')
    except Exception as exc:
        # Charts are nice-to-have — never fail the whole report on a render error.
        print(f'  warn: skipped Charts sheet ({exc})')

    # Boss request: duplicates up front — first sheet after Summary.
    if 'Confirmed duplicates' in wb.sheetnames:
        wb.move_sheet('Confirmed duplicates',
                      offset=1 - wb.sheetnames.index('Confirmed duplicates'))
    wb.save(out_xlsx)
    print(f'XLSX: {out_xlsx}')
    return out_xlsx


def run_sor_xlsx_bytes(folder, title, meta=None):
    """Run SOR mode and return (xlsx_bytes, n_files, n_pairs).  Pass a dict
    as `meta` to receive additive analysis facts (currently
    `short_traces`) without changing the return contract."""
    import tempfile
    _meta = meta if meta is not None else {}
    with tempfile.TemporaryDirectory() as td:
        tmp = os.path.join(td, 'report.xlsx')
        build_xlsx_sor(folder, title, tmp, meta=_meta)
        with open(tmp, 'rb') as fh:
            xlsx_bytes = fh.read()
    # Report what was ANALYSED, not what was globbed.  These used to
    # recount the staged folder after rendering, which disagreed with the
    # report's own header on any span where a trace is excluded:
    #
    #   ELMDALE TO MILER   glob 1152 / 662,976   analysed 1151 / 661,825
    #                      (ELMMIL0231_1550 ends at 22,288 m against a
    #                       69,567 m median — a real break)
    #   DURANC 1-144       glob  144 /  10,296   analysed  141 /   9,870
    #
    # The glob number reached the download-button label and the green
    # "N SOR files processed" line while the workbook's own Summary sheet
    # printed the smaller one.  Same folder, two numbers, no explanation.
    #
    # The glob was also case-sensitive on a path that is not: _inventory
    # matches on a lowercased name, so a file saved as .SOR was inventoried
    # and staged but missed here.  On POSIX that made the count too LOW and
    # the trace was silently dropped from the analysis; on Windows
    # ntpath.normcase lowercases, so the same folder behaved differently in
    # the field than on the dev machine.  Reading the analysis removes the
    # second parser entirely rather than teaching it the same rules.
    n_files = _meta.get('n_files', 0)
    n_pairs = _meta.get('n_pairs', 0)
    return xlsx_bytes, n_files, n_pairs


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--sor-dir', required=True)
    parser.add_argument('--title', required=True)
    parser.add_argument('--out-pdf', help='Path for PDF output')
    parser.add_argument('--out-xlsx', help='Path for XLSX output')
    args = parser.parse_args()
    if args.out_pdf:
        build_report_sor(args.sor_dir, args.title, args.out_pdf)
    if args.out_xlsx:
        build_xlsx_sor(args.sor_dir, args.title, args.out_xlsx)
    if not args.out_pdf and not args.out_xlsx:
        parser.error('Specify at least one of --out-pdf or --out-xlsx')


if __name__ == '__main__':
    main()
