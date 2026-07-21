"""
report.py
---------
Generates a duplicate-classification report for a folder of OTDR JSON files.
"""
import os, sys, json, glob, base64, struct, subprocess
from datetime import datetime
from itertools import combinations
from io import BytesIO
import numpy as np
from scipy.stats import norm

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt


HERE = os.path.dirname(os.path.abspath(__file__))
JSON_FOLDER = os.path.join(HERE, 'data')
WL_ORDER = [1310, 1550, 1625]
WL_COLOR = {1310: '#1F77B4', 1550: '#2CA02C', 1625: '#D62728'}

_INTERIOR_MIN_M = 1100
_INTERIOR_MAX_M = 60000
_SCORE_GATE     = 0.025


def _decode(pts_b64, n):
    raw = base64.b64decode(pts_b64)
    s = np.frombuffer(raw, dtype='<u2')[:n].astype(np.float64)
    return 64.0 - s / 1024.0


def _fmt_time_gap(sec):
    """Render an integer seconds count as a compact human string."""
    if sec is None:
        return '—'
    sec = int(sec)
    if sec < 60:
        return f'{sec}s'
    if sec < 3600:
        return f'{sec//60}m {sec%60:02d}s'
    if sec < 86400:
        h, r = divmod(sec, 3600)
        return f'{h}h {r//60:02d}m'
    d, r = divmod(sec, 86400)
    return f'{d}d {r//3600:02d}h'


def _parse_iso_ts(s):
    """Return (raw_string, epoch_seconds_or_None). Handles ISO-8601 w/ 'Z'."""
    if not s:
        return '', None
    try:
        from datetime import datetime as _dt
        s2 = s.replace('Z', '+00:00')
        return s, _dt.fromisoformat(s2).timestamp()
    except Exception:
        return s, None


def load_file(path):
    # utf-8-sig: OTDR JSON is UTF-8 (possibly BOM-prefixed); without an explicit
    # encoding Windows defaults to cp1252 and crashes on non-ASCII bytes. (Mac
    # defaults UTF-8, so this only ever bites techs on Windows.)
    with open(path, encoding="utf-8-sig") as f:
        d = json.load(f)
    name = os.path.basename(path).split('_')[0].strip()
    per_wl = {}
    for meas in (d.get('Measurement') or {}).get('OtdrMeasurements') or []:
        # An acquisition block can be incomplete (e.g. a JSON export that
        # dropped Wavelength or its DataPoints sub-keys).  Skip just that block
        # with a visible warning instead of KeyError-ing the whole batch; the
        # required keys are exactly the ones we hard-subscript just below.
        dp = meas.get('DataPoints') or {}
        missing = ([k for k in ('Wavelength',) if meas.get(k) is None]
                   + [k for k in ('NumberOfPoints', 'Resolution',
                                  'FirstPointPosition', 'Points')
                      if dp.get(k) is None])
        if missing:
            print(f'  warn: {os.path.basename(path)}: skipped acquisition block '
                  f'missing {", ".join(missing)}', file=sys.stderr)
            continue
        wl = int(meas['Wavelength'])
        n = int(dp['NumberOfPoints'])
        res = float(dp['Resolution'])
        fp = float(dp['FirstPointPosition'].replace(',', ''))
        trace = _decode(dp['Points'], n)
        pos = np.arange(n) * res + fp
        results = meas.get('Results') or {}
        def _num(k):
            v = results.get(k)
            try:
                return float(v) if v is not None else None
            except (TypeError, ValueError):
                return None
        # KeyEvents → list of {dist_km, splice_loss, is_end} (matches the SOR
        # event shape so the cross-mode event filter can use one helper).
        events = []
        length_for_end = _num('Length') or 0.0
        ke_list = (meas.get('KeyEvents') or {}).get('KeyEventList') or []
        for ke in ke_list:
            dist = ke.get('Distance')
            try:
                dist_m = float(str(dist).replace(',', '')) if dist is not None else None
            except (TypeError, ValueError):
                dist_m = None
            if dist_m is None:
                continue
            try:
                splice = ke.get('SpliceLoss')
                splice = float(str(splice).replace(',', '')) if splice is not None else None
            except (TypeError, ValueError):
                splice = None
            # Position-vs-length detection of end-of-fiber events (works across
            # Bellcore code variants where 'EventType' string can be coded
            # differently). Position past the fiber length is also "end-or-past".
            is_end = (length_for_end > 0 and dist_m >= length_for_end * 0.999)
            events.append({
                'dist_km': dist_m / 1000.0,
                'splice_loss': splice,
                'is_end': is_end,
            })
        per_wl[wl] = {
            'trace': trace, 'pos': pos,
            'max_splice_dB': _num('MaximumSpliceLoss'),
            'span_loss_dB':  _num('AveragedLoss'),
            'length_m':      _num('Length'),
            'events':        events,
        }
    if not per_wl:
        raise ValueError(f'{os.path.basename(path)}: no usable acquisition '
                         f'blocks (all missing required keys)')
    dt_raw, dt_epoch = _parse_iso_ts(d.get('TestDateTime', ''))
    return {'name': name, 'filepath': path,
            'test_dt': dt_raw, 'test_epoch': dt_epoch, 'wl': per_wl}


def _load_json_files(paths):
    """Load every JSON path, skipping (with a visible warning) any file that is
    malformed or carries no usable acquisition block — so one bad file in a
    batch of a dozen doesn't abort the whole run.  Raises only if NOTHING
    loads, naming why."""
    files, skipped = [], []
    for p in paths:
        try:
            files.append(load_file(p))
        except (ValueError, KeyError, json.JSONDecodeError) as exc:
            skipped.append((os.path.basename(p), exc))
            print(f'  warn: skipped {os.path.basename(p)}: {exc}', file=sys.stderr)
    if not files:
        detail = '; '.join(f'{n}: {e}' for n, e in skipped) or 'no files'
        raise RuntimeError(f'No usable JSON files loaded ({detail})')
    return files


def load_trc_file(path):
    """Parse a .trc file into the same per-file dict shape as load_file (JSON)."""
    from trc_parser import parse_trc_file
    r = parse_trc_file(path)
    name = os.path.basename(path).split('_')[0].split('.')[0].strip()
    per_wl = {}
    IOR = 1.4682
    for wlblock in r.get('wavelengths') or []:
        wl_nm = int(wlblock['wavelength_nm'])
        sp = wlblock.get('sampling_period_s')
        if not sp:
            continue
        dz = 2.998e8 * sp / (2.0 * IOR)
        samples = wlblock['samples']
        trace = 64.0 - samples.astype(np.float64) / 1024.0
        pos = np.arange(len(trace)) * dz
        # Max interior splice loss from event table (skip end-of-fiber events)
        raw_events = wlblock.get('events') or []
        spl_vals = [abs(e.get('loss_db'))
                    for e in raw_events
                    if e.get('loss_db') is not None
                    and (e.get('position_m') or 0) > 0.01
                    and (str(e.get('type', '')).lower() != 'end')]
        max_splice = max(spl_vals) if spl_vals else None
        # Normalize TRC events to {dist_km, splice_loss, is_end} so the
        # cross-mode event filter can use a single helper. TRC type codes are
        # numeric, not 'end' strings — detect end-of-fiber by position vs the
        # fiber's reported length so dead-zone reflectors past the end are
        # also dropped.
        events = []
        L = wlblock.get('length_m') or 0.0
        for e in raw_events:
            pos_m = e.get('position_m')
            if pos_m is None:
                continue
            is_end = (L > 0 and pos_m >= L * 0.999)
            events.append({
                'dist_km': pos_m / 1000.0,
                'splice_loss': e.get('loss_db'),
                'is_end': is_end,
            })
        per_wl[wl_nm] = {
            'trace': trace, 'pos': pos,
            'max_splice_dB': max_splice,
            'span_loss_dB': wlblock.get('span_loss_db'),
            'length_m':     wlblock.get('length_m'),
            'events':       events,
        }
    ts = r.get('timestamp')
    if ts:
        from datetime import datetime as _dt
        dt_raw = _dt.fromtimestamp(ts).isoformat()
    else:
        dt_raw = ''
    return {'name': name, 'filepath': path,
            'test_dt': dt_raw, 'test_epoch': float(ts) if ts else None,
            'wl': per_wl}


def _load_trc_files(paths):
    """Load every .trc path, skipping (with a visible warning) any file that is
    malformed or carries no usable acquisition block — so one bad .trc in a
    batch doesn't abort the whole run.  Mirrors _load_json_files; raises only
    if NOTHING loads, naming why."""
    files, skipped = [], []
    for p in paths:
        try:
            files.append(load_trc_file(p))
        except (ValueError, KeyError, struct.error, OSError) as exc:
            skipped.append((os.path.basename(p), exc))
            print(f'  warn: skipped {os.path.basename(p)}: {exc}', file=sys.stderr)
    if not files:
        detail = '; '.join(f'{n}: {e}' for n, e in skipped) or 'no files'
        raise RuntimeError(f'No usable TRC files loaded ({detail})')
    return files


def _event_match_quality(a_events, b_events, pos_tol_m=100.0):
    """Greedy match interior splice/event detections by closest position.
    Skips end-of-fiber and very-near-launch (< 10 m) events.

    Returns (n_matched, n_max_events, n_min_events, mean_dloss_db,
    max_dloss_db). max_dloss_db is the For-Romeo-style 'max splice Δ at
    matched events': for each splice closure that appears in both fibers,
    compute |Δloss|, then take the max across matched closures. When
    n_min_events < 3 the agreement metric isn't meaningful — caller should
    treat as 'agree' by default.
    """
    def _interior(events):
        out = []
        for e in events or []:
            if e.get('is_end'):
                continue
            d = e.get('dist_km') or 0
            if d < 0.01:
                continue
            sl = e.get('splice_loss')
            # Skip events with no/NaN loss — they pollute mean_dloss.
            if sl is None or (isinstance(sl, float) and np.isnan(sl)):
                continue
            out.append((d * 1000.0, sl))
        return out

    a = _interior(a_events)
    b = _interior(b_events)
    if not a or not b:
        return 0, 0, 0, 0.0, 0.0
    used_b = [False] * len(b)
    matched_dloss = []
    for pa, la in a:
        best_j = -1
        best_d = pos_tol_m + 1.0
        for j, (pb, _) in enumerate(b):
            if used_b[j]:
                continue
            d = abs(pa - pb)
            if d < best_d:
                best_d = d
                best_j = j
        if best_j >= 0 and best_d <= pos_tol_m:
            matched_dloss.append(abs(la - b[best_j][1]))
            used_b[best_j] = True
    n_match = len(matched_dloss)
    n_max = max(len(a), len(b))
    n_min = min(len(a), len(b))
    mean_dloss_db = float(np.mean(matched_dloss)) if matched_dloss else 0.0
    max_dloss_db = float(max(matched_dloss)) if matched_dloss else 0.0
    return n_match, n_max, n_min, mean_dloss_db, max_dloss_db


def _events_agree(n_match, n_max, n_min, mean_dloss_db,
                  min_count=3, frac_thresh=0.85, loss_thresh_db=0.010):
    """Return True iff the pair's events look like the same physical fiber.

    Calibrated against measured-truth datasets:
      - True same-fiber re-shoots: 100% match rate, mean |Δloss| ~1 mdB,
        equal event counts.
      - Different fibers in the same cable (DURSAN-style): 25-90% match
        rate, mean |Δloss| 10-40 mdB, asymmetric event counts.

    Default thresholds:
      - at least 3 matched events
      - ≥ 85% of the LONGER event list matched (penalizes asymmetric counts;
        a real duplicate detects the same splices in both shots)
      - mean loss difference ≤ 10 mdB (true dups are <2 mdB; this is
        generously above noise but catches splice-aligned non-duplicates)
    """
    if n_min < min_count or n_max == 0:
        return True  # too few events to evaluate — don't penalize
    return (n_match >= min_count
            and n_match / n_max >= frac_thresh
            and mean_dloss_db <= loss_thresh_db)


def _outlier_probability(values):
    """P(duplicate) per pair via robust-bulk fit + Bonferroni tail in log space."""
    v = np.asarray(values, dtype=np.float64)
    N = len(v)
    log_v = np.log10(np.maximum(v, 1e-9))
    med = float(np.median(log_v))
    mad = float(np.median(np.abs(log_v - med)))
    spread = max(mad * 1.4826, 1e-6)
    z = (log_v - med) / spread
    p_tail = norm.cdf(z)
    expected_fp = N * p_tail
    p_dup = np.clip(1.0 - expected_fp, 0.0, 1.0)
    return p_dup, {'center_log': med, 'spread_log': spread, 'N': N,
                   'z': z, 'p_tail': p_tail, 'expected_fp': expected_fp}


def _score(a, b, wl):
    ta, tb = a['wl'][wl]['trace'], b['wl'][wl]['trace']
    pa = a['wl'][wl]['pos']
    n = min(len(ta), len(tb))
    # Use length-aware interior window so short coils aren't discarded
    length_m = a['wl'][wl].get('length_m') or b['wl'][wl].get('length_m')
    mask = _interior_mask(pa[:n], length_m=length_m)
    if mask.sum() < 50:
        return None
    return float(np.std(ta[:n][mask] - tb[:n][mask]))


def _detrend(trace, pos):
    """Subtract best-fit linear (offset + slope) so two traces with different
    launch power / attenuation gain can still be shape-compared."""
    A = np.vstack([pos, np.ones_like(pos)]).T
    m, c = np.linalg.lstsq(A, trace, rcond=None)[0]
    return trace - (m * pos + c)


def _interior_mask(pos, length_m=None):
    """Pick an interior window that works for both km-scale fibers and short
    coils. For short fibers (< 800 m), use a 1 m launch buffer + 5% end
    buffer — anything tighter on coils discards the very fiber region we
    want to compare. For long fibers, use the production 1100–60000 m window."""
    if length_m is not None and length_m > 0 and length_m < 800:
        lo = max(1.0, length_m * 0.03)
        hi = max(lo + 1.0, length_m - max(0.5, length_m * 0.03))
    else:
        lo, hi = _INTERIOR_MIN_M, _INTERIOR_MAX_M
    return (pos > lo) & (pos < hi)


def _shape_r(a, b, wl):
    """Detrended Pearson correlation between two traces at one wavelength.
    Returns r in [-1, 1] or None if insufficient samples. r ≈ 1 → same fiber."""
    if wl not in a['wl'] or wl not in b['wl']:
        return None
    ta, tb = a['wl'][wl]['trace'], b['wl'][wl]['trace']
    pa = a['wl'][wl]['pos']
    L = min(a['wl'][wl].get('length_m') or 0, b['wl'][wl].get('length_m') or 0) or None
    n = min(len(ta), len(tb))
    mask = _interior_mask(pa[:n], length_m=L)
    if mask.sum() < 50:
        return None
    da = _detrend(ta[:n][mask].astype(np.float64), pa[:n][mask])
    db = _detrend(tb[:n][mask].astype(np.float64), pa[:n][mask])
    sa, sb = np.std(da), np.std(db)
    if sa == 0 or sb == 0:
        return None
    return float(np.dot(da - da.mean(), db - db.mean()) / (sa * sb * len(da)))


def _compute_pair_metrics_batch_multiwl(files, wl_list, min_samples=50,
                                          tie_panel_mode=False):
    """Vectorized per-wavelength σ and r matrices. Mirrors the SOR-mode
    `_compute_pair_metrics_batch` but iterates wavelengths so JSON / TRC
    multi-λ inputs get the same matmul speed-up and the same optional
    fingerprint-extraction step.

    Returns a dict keyed by wavelength:
        {wl: {'sigma_matrix': (K,K), 'r_matrix': (K,K), 'valid_idx': [int,...]}}

    `valid_idx` is the list of indices into `files` that had enough interior
    samples at that wavelength to participate (K = len(valid_idx) for the wl).

    When tie_panel_mode=True, the per-wavelength r matrix is computed AFTER
    subtracting the per-position median trace across the K participating
    files — strips the shared launch+connector signature so r reflects each
    fiber's unique scatter fingerprint.
    """
    results = {}
    for wl in wl_list:
        interior = []
        valid_idx = []
        for i, f in enumerate(files):
            wlblock = (f.get('wl') or {}).get(wl)
            if wlblock is None:
                continue
            ta = wlblock.get('trace')
            pa = wlblock.get('pos')
            if ta is None or pa is None:
                continue
            n = min(len(ta), len(pa))
            length_m = wlblock.get('length_m')
            mask = _interior_mask(pa[:n], length_m=length_m)
            if mask.sum() < min_samples:
                continue
            interior.append((ta[:n][mask].astype(np.float32),
                             pa[:n][mask].astype(np.float32)))
            valid_idx.append(i)
        if len(interior) < 2:
            continue
        N = min(len(d[0]) for d in interior)
        K = len(interior)
        M_raw = np.empty((K, N), dtype=np.float32)
        M_det = np.empty((K, N), dtype=np.float32)
        for k, (ts, ps) in enumerate(interior):
            ts = ts[:N]; ps = ps[:N]
            M_raw[k] = ts
            pm = ps.mean(); tm = ts.mean()
            denom = ((ps - pm) ** 2).sum()
            slope = float(((ps - pm) * (ts - tm)).sum() / denom) if denom > 0 else 0.0
            intercept = float(tm - slope * pm)
            M_det[k] = ts - (slope * ps + intercept)

        # σ matrix via variance-decomposition identity (no K×N intermediate).
        m1 = M_raw.mean(axis=1)
        m2 = (M_raw.astype(np.float64) ** 2).mean(axis=1)
        C = (M_raw.astype(np.float64) @ M_raw.astype(np.float64).T) / float(N)
        var_ij = (m2[:, None] + m2[None, :] - 2.0 * C
                  - (m1[:, None] - m1[None, :]) ** 2)
        sigma_matrix = np.sqrt(np.maximum(var_ij, 0.0))

        # r matrix on detrended traces, with optional fingerprint extraction.
        M_det64 = M_det.astype(np.float64)
        if tie_panel_mode:
            group_ref = np.median(M_det64, axis=0, keepdims=True)
            M_fingerprint = M_det64 - group_ref
        else:
            M_fingerprint = M_det64
        Mc = M_fingerprint - M_fingerprint.mean(axis=1, keepdims=True)
        std = np.sqrt((Mc ** 2).mean(axis=1))
        std_outer = np.outer(std, std)
        np.maximum(std_outer, 1e-12, out=std_outer)
        r_matrix = (Mc @ Mc.T) / (float(N) * std_outer)
        np.clip(r_matrix, -1.0, 1.0, out=r_matrix)

        results[wl] = {
            'sigma_matrix': sigma_matrix,
            'r_matrix': r_matrix,
            'valid_idx': valid_idx,
        }
    return results


def _shape_tier(r):
    """Bin a Pearson r into a same-fiber tier."""
    if r is None:
        return None
    if r >= 0.99:
        return 'high'
    if r >= 0.95:
        return 'mid'
    return 'low'


def _shape_color(r):
    t = _shape_tier(r)
    return _COLOR_HIGH if t == 'high' else (_COLOR_MID if t == 'mid' else _COLOR_LOW)


def _find_chrome():
    """Locate a Chromium-based browser for headless PDF rendering, across
    macOS / Linux / Windows. On Windows, Microsoft Edge is always installed,
    so PDF works on any tech's machine even without Chrome."""
    candidates = [
        # macOS
        '/Applications/Google Chrome.app/Contents/MacOS/Google Chrome',
        '/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge',
        '/Applications/Chromium.app/Contents/MacOS/Chromium',
        # Linux
        '/usr/bin/google-chrome', '/usr/bin/google-chrome-stable',
        '/usr/bin/chromium-browser', '/usr/bin/chromium',
        '/usr/bin/microsoft-edge',
    ]
    # Windows — Chrome then Edge (Edge ships with every Windows install)
    for env in ('PROGRAMFILES', 'PROGRAMFILES(X86)', 'LOCALAPPDATA'):
        root = os.environ.get(env)
        if not root:
            continue
        candidates += [
            os.path.join(root, 'Google', 'Chrome', 'Application', 'chrome.exe'),
            os.path.join(root, 'Microsoft', 'Edge', 'Application', 'msedge.exe'),
        ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    # last resort: anything named chrome/edge on PATH
    import shutil as _sh
    for name in ('google-chrome', 'chrome', 'chromium', 'msedge',
                 'microsoft-edge'):
        found = _sh.which(name)
        if found:
            return found
    return None


def _embed_logo():
    logo_path = os.path.join(HERE, 'zerodblogo.png')
    if not os.path.exists(logo_path):
        return ''
    with open(logo_path, 'rb') as f:
        b64 = base64.b64encode(f.read()).decode('ascii')
    return (f'<div style="text-align:center; margin-bottom:16px;">'
            f'<img src="data:image/png;base64,{b64}" style="height:60px; margin-left:-30px;" />'
            f'</div>')


_BASE_CSS = """
@page { size: landscape; margin: 10mm 10mm 18mm 10mm;
  @bottom-center { content: "Page " counter(page) " of " counter(pages); font-size: 8px; } }
* { box-sizing:border-box; margin:0; padding:0; }
body { font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
        color:#2c2c2a; padding:16px; font-size:11px; max-width:1400px; margin:0 auto; }
h1 { font-size:20px; font-weight:500; margin-bottom:2px; }
h2 { font-size:14px; font-weight:500; margin:24px 0 8px; page-break-after:avoid; }
.subtitle { font-size:11px; color:#888; margin-bottom:16px; }
.chart-img { width:100%; border-radius:8px; border:1px solid #ddd; margin-bottom:16px; }
.cards { display:flex; gap:10px; margin-bottom:16px; }
.card { flex:1; background:#fff; border:1px solid rgba(0,0,0,.08); border-radius:10px; padding:12px 14px; }
.card-label { font-size:9px; color:#999; text-transform:uppercase; letter-spacing:.04em; }
.card-value { font-size:22px; font-weight:600; }
.card-value.good { color:#2d8f48; }
.card-sub { font-size:9px; color:#999; margin-top:2px; }
.vote-table { width:100%; border-collapse:collapse; font-size:9.5px;
               font-family:'SF Mono','Courier New',monospace; margin-bottom:16px;
               page-break-inside:avoid; break-inside:avoid; }
.vote-table tr { page-break-inside:avoid; break-inside:avoid; }
.vote-table th { background:#f4f3f0; padding:5px 6px; text-align:center;
                  font-weight:600; border:0.5px solid #ddd; font-size:8px; color:#555; }
.vote-table td { padding:4px 6px; border:0.5px solid #ddd; }
.pair-cell { text-align:left !important; font-weight:600; }
.center { text-align:center; }
.dup { color:#2d8f48; font-weight:700; }
.na  { color:#888;    font-weight:500; }
.dir-banner { background:#2C3E50; color:white; padding:10px 16px; border-radius:8px;
               font-size:14px; font-weight:600; margin:28px 0 12px;
               page-break-after:avoid; break-after:avoid; }
h1, h2 { page-break-after:avoid; break-after:avoid; }
.chart-img { page-break-inside:avoid; break-inside:avoid; }
/* Wrap each section banner with its content so they travel together
   across page breaks. For sections whose content is a long table that
   exceeds one page, the renderer falls back to breaking at row
   boundaries (vote-table tr also has page-break-inside:avoid). */
.section-block { page-break-inside:avoid; break-inside:avoid; }
.verdict-box { padding:14px 18px; border-radius:10px; font-size:13px; font-weight:600;
               margin:16px 0; }
.verdict-confirm { background:#e8f5ec; color:#1f6b35; border:1px solid #bce0c6; }
.verdict-dispute { background:#fbeedf; color:#8a5200; border:1px solid #f0d2a3; }
"""


_COLOR_HIGH = '#2d8f48'   # p_dup > 0.9  — solid duplicate (green)
_COLOR_MID  = '#b97000'   # 0.5 < p_dup ≤ 0.9  — borderline (orange)
_COLOR_LOW  = '#888'      # p_dup ≤ 0.5  — non-duplicate (grey)


def _tier(p):
    """Return 'high' / 'mid' / 'low' based on p_dup (or is_dup = high)."""
    if p.get('is_dup') or p.get('p_dup', 0) > 0.9:
        return 'high'
    if p.get('p_dup', 0) > 0.5:
        return 'mid'
    return 'low'


def _is_highlighted(p):
    return _tier(p) != 'low'


def _tier_split(all_pairs_list, key_fn):
    """Split pair values into (high, mid, low) lists using key_fn(pair)->value|None."""
    hi, md, lo = [], [], []
    for p in all_pairs_list:
        v = key_fn(p)
        if v is None:
            continue
        t = _tier(p)
        (hi if t == 'high' else md if t == 'mid' else lo).append(v)
    return hi, md, lo


def chart_distribution(all_pairs_list):
    rng = np.random.default_rng(42)
    fig, axes = plt.subplots(4, 1, figsize=(13, 9), sharex=False)
    panels = [(1310, axes[0]), (1550, axes[1]), (1625, axes[2])]
    for wl, ax in panels:
        hi, md, lo = _tier_split(all_pairs_list, lambda p: p['score'].get(wl))
        dup_v = hi + md  # combined "highlighted" for separation-band math
        if dup_v and lo:
            dup_max = max(dup_v); non_min = min(lo)
            if non_min > dup_max:
                ax.axvspan(dup_max, non_min, color=_COLOR_HIGH, alpha=0.15,
                           label=f'separation band ({non_min/dup_max:.2f}×)')
            ax.set_title(f'{wl} nm — duplicates separate {non_min/dup_max:.1f}× below non-duplicates',
                         fontweight='bold', loc='left')
        else:
            ax.set_title(f'{wl} nm — level-of-disagreement distribution', fontweight='bold', loc='left')
        if lo:
            ax.scatter(lo, rng.uniform(0.25, 0.55, len(lo)),
                       color=_COLOR_LOW, alpha=0.55, s=55, edgecolor='white', linewidth=0.4,
                       label=f'Non-duplicate (n={len(lo)})')
        if md:
            ax.scatter(md, rng.uniform(0.55, 0.70, len(md)),
                       color=_COLOR_MID, alpha=0.95, s=140, edgecolor='black', linewidth=1,
                       zorder=4, label=f'Borderline 50–90% (n={len(md)})')
        if hi:
            ax.scatter(hi, rng.uniform(0.70, 0.85, len(hi)),
                       color=_COLOR_HIGH, alpha=0.95, s=170, edgecolor='black', linewidth=1,
                       zorder=5, label=f'Duplicate ≥90% (n={len(hi)})')
        ax.axvline(_SCORE_GATE, color=_COLOR_MID, linestyle='--', linewidth=1.3,
                   label='decision threshold')
        ax.set_xscale('log')
        all_v = hi + md + lo
        if all_v:
            ax.set_xlim(min(all_v) * 0.7, max(all_v) * 1.3)
        ax.set_ylim(0, 1)
        ax.set_yticks([])
        ax.set_xticklabels([])
        ax.set_xlabel(f'level of disagreement @ {wl} nm (log scale)', fontsize=10)
        ax.grid(axis='x', alpha=0.3, which='both')
        ax.legend(loc='upper right', fontsize=8, ncol=2)

    ax = axes[3]
    hi, md, lo = _tier_split(all_pairs_list, lambda p: sum(p['score'].values()))
    dup_sum = hi + md
    if dup_sum and lo:
        d_max = max(dup_sum); n_min = min(lo)
        if n_min > d_max:
            ax.axvspan(d_max, n_min, color=_COLOR_HIGH, alpha=0.15,
                       label=f'separation band ({n_min/d_max:.2f}×)')
    if lo:
        ax.scatter(lo, rng.uniform(0.25, 0.55, len(lo)),
                   color=_COLOR_LOW, alpha=0.55, s=55, edgecolor='white', linewidth=0.4,
                   label=f'Non-duplicate (n={len(lo)})')
    if md:
        ax.scatter(md, rng.uniform(0.55, 0.70, len(md)),
                   color=_COLOR_MID, alpha=0.95, s=140, edgecolor='black', linewidth=1,
                   zorder=4, label=f'Borderline 50–90% (n={len(md)})')
    if hi:
        ax.scatter(hi, rng.uniform(0.70, 0.85, len(hi)),
                   color=_COLOR_HIGH, alpha=0.95, s=170, edgecolor='black', linewidth=1,
                   zorder=5, label=f'Duplicate ≥90% (n={len(hi)})')
    ax.set_xscale('log')
    all_sum = hi + md + lo
    if all_sum:
        ax.set_xlim(min(all_sum) * 0.7, max(all_sum) * 1.3)
    ax.set_ylim(0, 1)
    ax.set_yticks([])
    ax.set_xticklabels([])
    ax.set_xlabel('combined level of disagreement across 3 wavelengths (log scale)', fontsize=10)
    ax.grid(axis='x', alpha=0.3, which='both')
    ax.set_title('Combined 3λ level-of-disagreement distribution', fontweight='bold', loc='left')
    ax.legend(loc='upper right', fontsize=8, ncol=2)

    fig.suptitle(f'Level-of-disagreement distribution across {len(all_pairs_list)} pairs',
                 fontsize=13, fontweight='bold', y=1.00)
    plt.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close()
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('ascii')


def chart_histogram(all_pairs_list):
    hi, md, lo = _tier_split(all_pairs_list, lambda p: sum(p['score'].values()))
    dup_sum = hi + md
    all_sum = hi + md + lo
    fig, ax = plt.subplots(figsize=(13, 4.5))
    if not all_sum:
        return None
    bins = np.linspace(0, max(all_sum) * 1.05, 60)
    if lo:
        ax.hist(lo, bins=bins, color=_COLOR_LOW, alpha=0.75,
                label=f'Non-duplicate (n={len(lo)})')
    counts, _ = np.histogram(lo or [0], bins=bins)
    y_mark = max(counts) * 0.75 if len(counts) else 1
    # Draw vertical lines and markers per tier so the color matches the table
    for d in md:
        ax.axvline(d, color=_COLOR_MID, linewidth=2, alpha=0.9)
    for d in hi:
        ax.axvline(d, color=_COLOR_HIGH, linewidth=2, alpha=0.9)
    if md:
        ax.scatter(md, [y_mark]*len(md), color=_COLOR_MID, s=160, zorder=4,
                   edgecolor='black', linewidth=1.0,
                   label=f'Borderline 50–90% (n={len(md)})')
    if hi:
        ax.scatter(hi, [y_mark]*len(hi), color=_COLOR_HIGH, s=200, zorder=5,
                   edgecolor='black', linewidth=1.2,
                   label=f'Duplicate ≥90% (n={len(hi)})')
    if dup_sum and lo and min(lo) > max(dup_sum):
        ax.axvspan(max(dup_sum), min(lo), color=_COLOR_HIGH, alpha=0.12,
                   label='separation band')
    ax.axvline(_SCORE_GATE * 3, color=_COLOR_MID, linestyle='--', linewidth=1.3,
               label='decision threshold')
    ax.set_xticklabels([])
    ax.set_xlabel('combined level of disagreement across 3 wavelengths')
    ax.set_ylabel('Number of non-duplicate pairs')
    ax.set_title('Histogram — combined 3λ level of disagreement', fontweight='bold')
    ax.legend(loc='upper right', fontsize=9)
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close()
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('ascii')


def _finalize_pairs_multiwl(files, all_pairs_list, regime='production'):
    """Run the σ-outlier + r-tier + physical-reality math on a freshly built
    multi-λ pair list. Mutates each pair dict in place (sets `p_dup`,
    `p_dup_r`, `z`, `length_capped`, `events_capped`, `length_delta_m`,
    `events_*`, `events_max_dloss_per_wl`) and returns a dict with:
        'p_dup_arr'      — np.array of final P(dup) values, pair-aligned
        'p_dup_r_arr'    — np.array of r-tier P(dup) values
        'prob_stats'     — diagnostic dict from _outlier_probability
        'best_partner'   — {file_name: {'partner', 'sum_score', 'p_dup', 'pair'}}
        'pair_lookup'    — {(sorted(a,b)): pair_dict}

    `regime` is one of:
        'production' — standard σ-outlier + r-tier (0.95-0.99 ramp).
        'tie_panel'  — fingerprint-extracted r + tight ramp (0.999-0.9999)
                       + r-confirmation gate on σ-outlier.
        'all_dups'   — σ-outlier bypassed (broken on homogeneous data);
                       widened r-ramp (0.85-0.95) catches every pair.

    Used by both build_report (PDF/HTML) and build_xlsx_multiwl (Excel) so
    the two renderers can never drift.
    """
    pair_lookup = {tuple(sorted([p['a'], p['b']])): p for p in all_pairs_list}

    combined = np.array([p['sum_score'] for p in all_pairs_list], dtype=np.float64)
    p_dup_sigma_arr, prob_stats = _outlier_probability(combined)

    # Pearson-shape contribution. Per regime:
    #   production: (0.95 → 0.99)     standard
    #   tie-panel:  (0.999 → 0.9999)  tightened (residual structure after FP)
    #   all-dups:   (0.85 → 0.95)     widened — every pair IS a duplicate
    if regime == 'tie_panel':
        _R_LO, _R_HI = 0.999, 0.9999
    elif regime == 'all_dups':
        _R_LO, _R_HI = 0.85, 0.95
    elif regime == 'short_panel':
        # Standard production ramp — true same-fiber re-shoots on a short
        # panel still produce r ≥ 0.95. r-tier is the entire detector here.
        _R_LO, _R_HI = 0.95, 0.99
    else:
        _R_LO, _R_HI = 0.95, 0.99
    _R_SPAN = _R_HI - _R_LO
    def _r_to_p(r):
        if r is None:
            return 0.0
        if r >= _R_HI:
            return 1.0
        if r <= _R_LO:
            return 0.0
        return float((r - _R_LO) / _R_SPAN)
    p_dup_r_arr = np.array([_r_to_p(p.get('r_min')) for p in all_pairs_list],
                           dtype=np.float64)

    # σ-outlier handling: ONLY production mode trusts it. Every other regime
    # bypasses σ-outlier and lets the regime-specific r-ramp drive the verdict.
    #   tie_panel   — fingerprint-extracted tight r-ramp is the detector;
    #                 σ-outlier would cascade on shared cable structure
    #                 (a 2 km tie panel can show 20k σ false positives that
    #                 a post-fingerprint r≥0.9 gate fails to block).
    #   all_dups    — no non-duplicate bulk to define an "outlier".
    #   short_panel — short featureless fibers give a narrow σ bulk that cascades.
    if regime in ('tie_panel', 'all_dups', 'short_panel'):
        p_dup_sigma_eff = np.zeros_like(p_dup_sigma_arr)
    else:
        p_dup_sigma_eff = p_dup_sigma_arr

    # Raw combined likelihood = max of (possibly confirmed) σ-outlier and r tiers.
    p_dup_raw_arr = np.maximum(p_dup_sigma_eff, p_dup_r_arr)

    # Physical-reality filters (mirror SOR mode):
    #
    #   1. Length-Δ filter: same fiber → same end-of-fiber length within
    #      tolerance max(10 m, length × 5e-4).
    #   2. Event-table filter: same fiber → splice events match in count,
    #      position, and loss. Required: ≥3 matched events, ≥85% match
    #      against max(n_a, n_b), mean |Δloss| ≤ 10 mdB.
    #
    # Either violation caps p_dup at 0.5 (borderline) regardless of σ/r.
    # Use the longest-wavelength λ shared by all files for length and
    # events (1550 nm is canonical when present).
    def _len_tol_m(length_m):
        if length_m is None or length_m <= 0:
            return 10.0
        return max(10.0, length_m * 5e-4)
    canonical_wl = (1550 if all(1550 in f['wl'] for f in files)
                    else sorted({wl for f in files for wl in f['wl']})[0])
    file_by_name = {f['name']: f for f in files}
    EVENT_CHECK_THRESHOLD = 0.10
    LEN_CAP = 0.5
    length_violation = np.zeros(len(all_pairs_list), dtype=bool)
    events_violation = np.zeros(len(all_pairs_list), dtype=bool)
    for i, p in enumerate(all_pairs_list):
        fa = file_by_name.get(p['a'])
        fb = file_by_name.get(p['b'])
        if fa is None or fb is None:
            continue
        wl_a = fa['wl'].get(canonical_wl) or {}
        wl_b = fb['wl'].get(canonical_wl) or {}
        len_a = wl_a.get('length_m')
        len_b = wl_b.get('length_m')
        if len_a and len_b:
            len_delta = abs(len_a - len_b)
            tol = _len_tol_m(max(len_a, len_b))
            p['length_delta_m'] = float(len_delta)
            if len_delta > tol:
                length_violation[i] = True
        if p_dup_raw_arr[i] >= EVENT_CHECK_THRESHOLD:
            # Canonical-λ event-match for the agree decision...
            n_match, n_max, n_min, mean_dloss, max_dloss = _event_match_quality(
                wl_a.get('events'), wl_b.get('events'))
            p['events_n_match'] = int(n_match)
            p['events_n_max']   = int(n_max)
            p['events_n_min']   = int(n_min)
            p['events_mean_dloss_db'] = float(mean_dloss)
            p['events_max_dloss_db']  = float(max_dloss)
            if not _events_agree(n_match, n_max, n_min, mean_dloss):
                events_violation[i] = True
            # ...plus per-λ max-Δ at matched events for the detail-table cells.
            max_dloss_per_wl = {}
            for wl in (fa.get('wl') or {}).keys():
                wa_events = (fa['wl'].get(wl) or {}).get('events')
                wb_events = (fb['wl'].get(wl) or {}).get('events')
                if wa_events is None or wb_events is None:
                    continue
                _, _, _, _, mxd = _event_match_quality(wa_events, wb_events)
                max_dloss_per_wl[wl] = float(mxd)
            p['events_max_dloss_per_wl'] = max_dloss_per_wl
    physical_violation = length_violation | events_violation
    p_dup_arr = np.where(physical_violation,
                         np.minimum(p_dup_raw_arr, LEN_CAP),
                         p_dup_raw_arr)
    for p, pd, pdr, z, lc, ec in zip(all_pairs_list, p_dup_arr, p_dup_r_arr,
                                     prob_stats['z'], length_violation,
                                     events_violation):
        p['p_dup']         = float(pd)
        p['p_dup_r']       = float(pdr)
        p['z']             = float(z)
        p['length_capped'] = bool(lc)
        p['events_capped'] = bool(ec)

    # For each file, pick the partner that gives the HIGHEST duplicate
    # likelihood (tie-broken by smallest combined disagreement). Symmetric
    # by construction: if pair (A,B) is the most-likely duplicate for both
    # files, both rows in the per-file table point at each other.
    best_partner = {}
    for f in files:
        best = None
        for g in files:
            if g['name'] == f['name']: continue
            p = pair_lookup[tuple(sorted([f['name'], g['name']]))]
            cand = {'partner': g['name'], 'sum_score': p['sum_score'],
                    'p_dup': p['p_dup'], 'pair': p}
            if (best is None
                or cand['p_dup'] > best['p_dup']
                or (cand['p_dup'] == best['p_dup'] and cand['sum_score'] < best['sum_score'])):
                best = cand
        best_partner[f['name']] = best

    return {
        'p_dup_arr':    p_dup_arr,
        'p_dup_r_arr':  p_dup_r_arr,
        'prob_stats':   prob_stats,
        'best_partner': best_partner,
        'pair_lookup':  pair_lookup,
    }


def build_report(files, all_pairs_list, truth_dups, out_path,
                 title='Duplicate Classification Report', regime='production'):
    truth_dups = truth_dups or set()
    fin = _finalize_pairs_multiwl(files, all_pairs_list, regime=regime)
    pair_lookup  = fin['pair_lookup']
    best_partner = fin['best_partner']
    p_dup_arr    = fin['p_dup_arr']

    dup_pairs = [p for p in all_pairs_list
                 if all((p['score'][wl] is not None and p['score'][wl] < _SCORE_GATE)
                        for wl in WL_ORDER)]
    used = set()
    confirmed = []
    for p in sorted(dup_pairs, key=lambda q: q['sum_score']):
        if p['a'] in used or p['b'] in used: continue
        confirmed.append(p)
        used.add(p['a']); used.add(p['b'])

    dup_names = {n for p in confirmed for n in (p['a'], p['b'])}

    truth_found = {tuple(sorted([p['a'], p['b']])) for p in confirmed}
    tp = len(truth_found & truth_dups)
    fp = len(truth_found - truth_dups)
    fn = len(truth_dups - truth_found)

    distribution_chart = chart_distribution(all_pairs_list)
    histogram_chart = chart_histogram(all_pairs_list)

    file_rows = ''
    for f in sorted(files, key=lambda x: x['name']):
        bp = best_partner[f['name']]
        partner = bp['partner']
        # Verdict reads from the best-partner pair's combined likelihood,
        # so a file flagged as DUPLICATE by its best pair stays consistent
        # with that pair's row in the confirmed-duplicate table.
        is_dup = bp['pair']['p_dup'] > 0.5
        verdict_html = (f'<td class="center"><span class="dup">DUPLICATE of {partner}</span></td>'
                        if is_dup else
                        f'<td class="center"><span class="na">unique (closest: {partner})</span></td>')
        pair = pair_lookup[tuple(sorted([f['name'], partner]))]
        wl_cells = ''
        for wl in WL_ORDER:
            sc = pair['score'][wl]
            if sc is None:
                wl_cells += '<td class="center na">---</td>'
            else:
                color = '#2d8f48' if sc < _SCORE_GATE else '#c0392b'
                wl_cells += f'<td class="center" style="color:{color};font-weight:600">{sc:.4f}</td>'
        pd_val = bp['pair']['p_dup']
        pd_color = '#2d8f48' if pd_val > 0.9 else ('#b97000' if pd_val > 0.1 else '#888')
        r_min = bp['pair'].get('r_min')
        if r_min is None:
            r_cell = '<td class="center na">—</td>'
        else:
            r_cell = f'<td class="center" style="color:{_shape_color(r_min)};font-weight:600">{r_min:.4f}</td>'
        file_rows += (f'<tr><td class="pair-cell">{f["name"]}</td>'
                      f'<td class="center">{f["test_dt"][:19]}</td>'
                      f'{wl_cells}'
                      f'<td class="center">{bp["sum_score"]:.3f}</td>'
                      f'<td class="center" style="color:{pd_color};font-weight:600">{pd_val*100:.2f}%</td>'
                      f'{r_cell}'
                      f'{verdict_html}</tr>')

    # ---- Confirmed-duplicate detail table (pairs with P(dup) > 0.5) -----
    file_by_name = {f['name']: f for f in files}
    dup_pairs_sorted = sorted(
        [p for p in all_pairs_list if p['p_dup'] > 0.5],
        key=lambda q: -q['p_dup'])
    # PDF cap — same protection as the SOR path (Zach 2026-07-21): render
    # at most PDF_DUP_ROWS_CAP rows; the Excel report carries the full list.
    dup_pairs_render, dup_overflow = _capped_rows(dup_pairs_sorted,
                                                  PDF_DUP_ROWS_CAP)
    dup_detail_rows = ''
    for p in dup_pairs_render:
        fa = file_by_name.get(p['a']); fb = file_by_name.get(p['b'])
        if fa is None or fb is None:
            continue
        # Time gap (file-level, not per-λ: one timestamp per acquisition)
        if fa.get('test_epoch') and fb.get('test_epoch'):
            gap_sec = int(abs(fa['test_epoch'] - fb['test_epoch']))
            gap_str = _fmt_time_gap(gap_sec)
        else:
            gap_str = '—'
        # Per-wavelength cells: max splice-loss Δ (mdB), span-loss Δ (mdB), shape r
        ms_cells = ''
        sl_cells = ''
        sr_cells = ''
        max_dloss_map = p.get('events_max_dloss_per_wl') or {}
        for wl in WL_ORDER:
            a_sl = fa['wl'].get(wl, {}).get('span_loss_dB')
            b_sl = fb['wl'].get(wl, {}).get('span_loss_dB')
            # For-Romeo style: max |Δloss| across MATCHED events at this λ.
            mxd = max_dloss_map.get(wl)
            if mxd is not None and mxd > 0:
                ms_cells += f'<td class="center">{mxd*1000:.0f}</td>'
            else:
                ms_cells += '<td class="center na">—</td>'
            if a_sl is not None and b_sl is not None:
                sl_cells += f'<td class="center">{abs(a_sl - b_sl)*1000:.0f}</td>'
            else:
                sl_cells += '<td class="center na">—</td>'
            r_wl = (p.get('shape_r') or {}).get(wl)
            if r_wl is None:
                sr_cells += '<td class="center na">—</td>'
            else:
                sr_cells += (f'<td class="center" style="color:{_shape_color(r_wl)};'
                             f'font-weight:600">{r_wl:.4f}</td>')
        pd_val = p['p_dup']
        pd_color = '#2d8f48' if pd_val > 0.9 else '#b97000'
        dup_detail_rows += (f'<tr><td class="pair-cell">{p["a"]} ↔ {p["b"]}</td>'
                            f'<td class="center">{gap_str}</td>'
                            f'{ms_cells}{sl_cells}{sr_cells}'
                            f'<td class="center" style="color:{pd_color};font-weight:600">{pd_val*100:.2f}%</td></tr>')
    # Boss request (2026-07-15): duplicates lead the report — section 1 on
    # page one, explicit "none" line when clean.
    if dup_detail_rows:
        ms_hdrs = ''.join(f'<th>max splice Δ @ {wl} (mdB)</th>' for wl in WL_ORDER)
        sl_hdrs = ''.join(f'<th>span loss Δ @ {wl} (mdB)</th>' for wl in WL_ORDER)
        sr_hdrs = ''.join(f'<th>similarity @ {wl}</th>' for wl in WL_ORDER)
        dup_detail_block = f'''
<div class="section-block">
<div class="dir-banner">1. Confirmed duplicate pairs (≥50% likelihood) — detail</div>
<table class="vote-table">
<tr><th style="text-align:left">Pair</th><th>Time gap</th>
  {ms_hdrs}{sl_hdrs}{sr_hdrs}<th>Duplicate likelihood</th></tr>
{dup_detail_rows}
</table>
{('<div style="padding:8px 4px;color:#b97000;font-weight:600">… and '
  f'{dup_overflow:,} more pairs at ≥50% likelihood — the complete list is '
  'in the Excel report.</div>') if dup_overflow else ''}
</div>
'''
    else:
        dup_detail_block = (
            '<div class="section-block">'
            '<div class="dir-banner">1. Confirmed duplicate pairs (\u226550% likelihood)</div>'
            '<div style="padding:10px 4px;color:#2d8f48;font-weight:600">'
            'None \u2014 no pairs at \u226550% duplicate likelihood.</div></div>')

    nonconf_sorted = sorted(
        [p for p in all_pairs_list if tuple(sorted([p['a'], p['b']])) not in truth_dups],
        key=lambda q: q['sum_score'])
    nondup_rows = ''
    for p in nonconf_sorted[:10]:
        wl_cells = ''
        for wl in WL_ORDER:
            sc = p['score'][wl]
            if sc is not None:
                color = '#2d8f48' if sc < _SCORE_GATE else '#c0392b'
                wl_cells += f'<td class="center" style="color:{color};font-weight:600">{sc:.4f}</td>'
            else:
                wl_cells += '<td class="center na">---</td>'
        pd_val = p['p_dup']
        pd_color = '#2d8f48' if pd_val > 0.9 else ('#b97000' if pd_val > 0.1 else '#888')
        r_min = p.get('r_min')
        r_cell = ('<td class="center na">—</td>' if r_min is None else
                  f'<td class="center" style="color:{_shape_color(r_min)};font-weight:600">{r_min:.4f}</td>')
        _fa2, _fb2 = file_by_name.get(p['a']), file_by_name.get(p['b'])
        _ta2 = _fa2.get('timestamp') if _fa2 else None
        _tb2 = _fb2.get('timestamp') if _fb2 else None
        _gap2 = _fmt_time_gap(abs(_ta2 - _tb2)) if _ta2 and _tb2 else '—'
        nondup_rows += (f'<tr><td class="pair-cell">{p["a"]} ↔ {p["b"]}</td>'
                        f'<td class="center">{_gap2}</td>'
                        f'{wl_cells}'
                        f'<td class="center">{p["sum_score"]:.3f}</td>'
                        f'<td class="center" style="color:{pd_color};font-weight:600">{pd_val*100:.2f}%</td>'
                        f'{r_cell}</tr>')

    n_over_50 = int((p_dup_arr > 0.5).sum())
    n_over_99 = int((p_dup_arr > 0.99).sum())
    if truth_dups:
        verdict_block = (
            '<div class="verdict-box verdict-confirm">'
            f'<b>{len(truth_dups)} / {len(truth_dups)} duplicate pairs identified. 0 false positives.</b><br>'
            'Every true duplicate sits below the decision threshold at all three wavelengths; '
            'every non-duplicate sits above it at one or more wavelengths.'
            '</div>'
        ) if tp == len(truth_dups) and fp == 0 else (
            f'<div class="verdict-box verdict-dispute">'
            f'<b>{tp}/{len(truth_dups)} TP, {fp} FP, {fn} FN.</b></div>'
        )
    else:
        verdict_block = (
            f'<div class="verdict-box verdict-confirm">'
            f'<b>{n_over_50} duplicate pair(s) identified at ≥50% likelihood; '
            f'{n_over_99} at ≥99% likelihood</b> across {len(all_pairs_list)} pairs.'
            f'</div>'
            if n_over_50 else
            f'<div class="verdict-box verdict-dispute">'
            f'<b>No duplicate pairs identified at ≥50% likelihood</b> '
            f'({len(all_pairs_list)} pairs).</div>'
        )

    generated = datetime.now().strftime('%Y-%m-%d %H:%M')
    html = f'''<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<title>{title}</title>
<style>{_BASE_CSS}</style></head><body>
{_embed_logo()}
<h1>{title}</h1>
<div class="subtitle">{len(files)} files &bull; {len(all_pairs_list)} pairs &bull; generated {generated}</div>

{verdict_block}

{dup_detail_block}

<div class="cards">
  <div class="card"><div class="card-label">Files</div><div class="card-value">{len(files)}</div></div>
  <div class="card"><div class="card-label">Pairs</div><div class="card-value">{len(all_pairs_list)}</div></div>
  <div class="card"><div class="card-label">Likelihood &gt; 99%</div>
    <div class="card-value good">{int((p_dup_arr>0.99).sum())}</div></div>
  <div class="card"><div class="card-label">Likelihood &gt; 50%</div>
    <div class="card-value">{int((p_dup_arr>0.5).sum())}</div></div>
  <div class="card"><div class="card-label">Likelihood &gt; 10%</div>
    <div class="card-value">{int((p_dup_arr>0.1).sum())}</div></div>
</div>

<div class="section-block">
<div class="dir-banner">2. Distribution</div>
<img src="data:image/png;base64,{distribution_chart}" class="chart-img" />
</div>

<div class="section-block">
<div class="dir-banner">3. Histogram — combined 3λ level of disagreement</div>
<img src="data:image/png;base64,{histogram_chart}" class="chart-img" />
</div>

<div class="section-block">
<div class="dir-banner">4. All {len(files)} files — per-file verdict</div>
<table class="vote-table">
<tr><th style="text-align:left">File</th><th>Acquisition time</th>
  <th>disagreement @ 1310</th><th>disagreement @ 1550</th><th>disagreement @ 1625</th>
  <th>combined disagreement</th><th>Duplicate likelihood</th>
  <th>similarity (min λ)</th><th>Verdict</th></tr>
{file_rows}
</table>
</div>

<div class="section-block">
<div class="dir-banner">5. Closest non-duplicate pairs</div>
<table class="vote-table">
<tr><th style="text-align:left">Pair</th><th>Time gap</th>
  <th>disagreement @ 1310</th><th>disagreement @ 1550</th><th>disagreement @ 1625</th><th>combined</th>
  <th>Duplicate likelihood</th><th>similarity (min λ)</th></tr>
{nondup_rows}
</table>
</div>

</body></html>'''

    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)
    return out_path


def _pdf_timeout_for(html_len):
    """Chrome print timeout scaled to the HTML payload.  The fixed 180 s
    budget dies on giant reports (Zach 2026-07-21: an all_dups folder put a
    62,014-row duplicate table in the HTML and Chrome blew the budget).
    Base 180 s + 60 s per MB beyond the first, capped at 480 s."""
    extra = max(0, html_len - 1_000_000)
    return int(min(480, 180 + 60 * (extra / 1_000_000)))


def _capped_rows(rows_sorted, cap):
    """(rows_to_render, overflow_count) for the PDF duplicate table.  The
    PDF is the human summary — a 62k-row table is unreadable AND breaks
    Chrome's print budget; the Excel report always carries the full list."""
    if cap is None or len(rows_sorted) <= cap:
        return rows_sorted, 0
    return rows_sorted[:cap], len(rows_sorted) - cap


PDF_DUP_ROWS_CAP = 500   # max confirmed-duplicate rows rendered in a PDF


def html_to_pdf_bytes(html_str, base_url=None):
    """Render an HTML string to PDF bytes. WeasyPrint preferred (cloud-friendly);
    Chrome used as a fallback when WeasyPrint's native libs aren't installed."""
    try:
        from weasyprint import HTML
        return HTML(string=html_str, base_url=base_url).write_pdf()
    except Exception:
        pass
    chrome = _find_chrome()
    if not chrome:
        raise RuntimeError(
            'Neither WeasyPrint nor Chrome is available. '
            'Install WeasyPrint system libs (brew install pango) '
            'or Google Chrome.')
    import tempfile
    with tempfile.NamedTemporaryFile('w', suffix='.html', delete=False,
                                     encoding='utf-8') as hf:
        hf.write(html_str)
        html_path = hf.name
    pdf_path = html_path.replace('.html', '.pdf')
    try:
        res = subprocess.run(
            [chrome, '--headless=new', '--disable-gpu', '--no-sandbox',
             '--run-all-compositor-stages-before-draw',
             '--virtual-time-budget=5000',
             f'--print-to-pdf={pdf_path}',
             '--print-to-pdf-no-header', '--no-pdf-header-footer',
             'file://' + html_path],
            capture_output=True, timeout=_pdf_timeout_for(len(html_str)),
            # no flashing console window when the windowed desktop app shells
            # out to Chrome/Edge on Windows (no-op elsewhere)
            creationflags=getattr(subprocess, 'CREATE_NO_WINDOW', 0))
        if res.returncode != 0:
            raise RuntimeError(f'Chrome failed: {res.stderr.decode(errors="ignore")[:400]}')
        with open(pdf_path, 'rb') as fh:
            return fh.read()
    finally:
        for p in (html_path, pdf_path):
            try:
                os.remove(p)
            except OSError:
                pass


def html_to_pdf(html_path, pdf_path):
    """File-to-file wrapper. Returns True on success."""
    with open(html_path, 'r', encoding='utf-8') as fh:
        html = fh.read()
    base = os.path.dirname(os.path.abspath(html_path))
    try:
        pdf_bytes = html_to_pdf_bytes(html, base_url=base)
    except Exception:
        return False
    with open(pdf_path, 'wb') as fh:
        fh.write(pdf_bytes)
    return True


def build_xlsx_multiwl(files, all_pairs_list, truth_dups, out_xlsx,
                       title='Duplicate Classification Report',
                       wl_list=None, regime='production'):
    global WL_ORDER
    """Multi-wavelength (JSON/TRC) Excel renderer. Mirrors build_xlsx_sor's
    6-sheet layout, but every per-λ metric becomes its own column.

    Sheets:
      Summary                — header counts and verdict
      Per-file verdict       — per-file row with per-λ span loss + verdict
      Confirmed duplicates   — pairs at ≥50% likelihood, per-λ detail columns
      Top 30 — lowest disagreement
      Top 30 — highest similarity
      Charts                 — distribution + histogram PNGs
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XlsxImage

    truth_dups = truth_dups or set()
    wl_list = list(wl_list) if wl_list else list(WL_ORDER)

    fin = _finalize_pairs_multiwl(files, all_pairs_list, regime=regime)
    pair_lookup  = fin['pair_lookup']
    best_partner = fin['best_partner']
    p_dup_arr    = fin['p_dup_arr']

    n_files = len(files)
    n_pairs = len(all_pairs_list)
    n99 = int((p_dup_arr > 0.99).sum())
    n50 = int((p_dup_arr > 0.5).sum())
    n10 = int((p_dup_arr > 0.1).sum())

    wb = Workbook()

    # Calibri 12 everywhere — matches SOR Excel and the rest of the UX.
    BASE      = Font(name='Calibri', size=12)
    BASE_BOLD = Font(name='Calibri', size=12, bold=True)
    TITLE_FONT = Font(name='Calibri', size=14, bold=True)
    HDR_FONT  = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    hdr_fill  = PatternFill('solid', fgColor='2C3E50')

    # ---------- Summary ----------
    ws = wb.active
    ws.title = 'Summary'
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f'Generated {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = BASE

    wl_label = ', '.join(f'{w} nm' for w in wl_list)
    rows = [
        ('Files',             n_files),
        ('Pairs',             n_pairs),
        ('Wavelengths',       wl_label),
        ('Likelihood ≥ 99%',  n99),
        ('Likelihood ≥ 50%',  n50),
        ('Likelihood ≥ 10%',  n10),
        ('Regime',            regime),
        ('Interior window (m)', f'{_INTERIOR_MIN_M:.0f}–{_INTERIOR_MAX_M:.0f}'),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        c1 = ws.cell(row=i, column=1, value=k); c1.font = BASE_BOLD
        c2 = ws.cell(row=i, column=2, value=v); c2.font = BASE
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 32

    def _write_table(ws, headers, rows_data, col_widths=None):
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = hdr_fill
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal='center')
        for r, row in enumerate(rows_data, start=2):
            for c, v in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = BASE
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = (f'A1:{get_column_letter(len(headers))}'
                              f'{1 + len(rows_data)}')
        if col_widths:
            for c, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(c)].width = w

    # ---------- Per-file verdict ----------
    # File | Acq time | Length (km) | span loss @ <wl> (one col per λ) |
    # combined disagreement | duplicate likelihood (%) | similarity (min λ) |
    # best partner | verdict
    ws = wb.create_sheet('Per-file verdict')
    span_loss_hdrs = [f'Span loss @ {wl} nm (dB)' for wl in wl_list]
    headers = (['File', 'Acquisition time', 'Length (km)']
               + span_loss_hdrs
               + ['Combined disagreement', 'Duplicate likelihood (%)',
                  'Similarity (min λ)', 'Best partner', 'Verdict'])
    rows_data = []
    for f in sorted(files, key=lambda x: x['name']):
        bp = best_partner.get(f['name'])
        # Longest λ-reported length, in km
        lengths = [(f['wl'].get(wl) or {}).get('length_m') for wl in wl_list]
        lengths = [L for L in lengths if L]
        length_km = (max(lengths) / 1000.0) if lengths else None
        span_loss_cells = []
        for wl in wl_list:
            sl = (f['wl'].get(wl) or {}).get('span_loss_dB')
            span_loss_cells.append(sl)
        if bp is None:
            rows_data.append([f['name'], f.get('test_dt', '')[:19], length_km]
                             + span_loss_cells
                             + [None, None, None, None, '—'])
            continue
        partner = bp['partner']
        pair    = bp['pair']
        verdict = (f'DUPLICATE of {partner}' if pair['p_dup'] > 0.5
                   else f'unique (closest: {partner})')
        rows_data.append([
            f['name'],
            f.get('test_dt', '')[:19],
            length_km,
        ] + span_loss_cells + [
            bp['sum_score'],
            pair['p_dup'] * 100.0,
            pair.get('r_min'),
            partner,
            verdict,
        ])
    cw = [18, 20, 12] + [16] * len(wl_list) + [22, 22, 18, 20, 32]
    _write_table(ws, headers, rows_data, col_widths=cw)

    # ---------- Confirmed duplicates (≥50% likelihood) ----------
    ws = wb.create_sheet('Confirmed duplicates')
    ms_hdrs = [f'Max splice Δ @ {wl} (mdB)' for wl in wl_list]
    sl_hdrs = [f'Span loss Δ @ {wl} (mdB)' for wl in wl_list]
    sr_hdrs = [f'Similarity @ {wl}'         for wl in wl_list]
    headers = (['Pair A', 'Pair B', 'Time gap (s)']
               + ms_hdrs + sl_hdrs + sr_hdrs
               + ['Duplicate likelihood (%)'])
    file_by_name = {f['name']: f for f in files}
    dup_sorted = sorted([p for p in all_pairs_list if p['p_dup'] > 0.5],
                        key=lambda q: -q['p_dup'])
    rows_data = []
    for p in dup_sorted:
        fa = file_by_name.get(p['a'])
        fb = file_by_name.get(p['b'])
        if fa is None or fb is None:
            continue
        ta, tb = (fa.get('test_epoch'), fb.get('test_epoch'))
        gap = int(abs(ta - tb)) if (ta and tb) else None
        max_dloss_map = p.get('events_max_dloss_per_wl') or {}
        ms_cells = []
        sl_cells = []
        sr_cells = []
        for wl in wl_list:
            mxd = max_dloss_map.get(wl)
            ms_cells.append(mxd * 1000.0 if mxd and mxd > 0 else None)
            a_sl = (fa['wl'].get(wl) or {}).get('span_loss_dB')
            b_sl = (fb['wl'].get(wl) or {}).get('span_loss_dB')
            sl_cells.append(abs(a_sl - b_sl) * 1000.0
                            if (a_sl is not None and b_sl is not None) else None)
            sr_cells.append((p.get('shape_r') or {}).get(wl))
        rows_data.append(
            [p['a'], p['b'], gap]
            + ms_cells + sl_cells + sr_cells
            + [p['p_dup'] * 100.0]
        )
    cw = ([18, 18, 13] + [22] * len(wl_list)
          + [22] * len(wl_list) + [16] * len(wl_list) + [22])
    _write_table(ws, headers, rows_data, col_widths=cw)

    # ---------- Top 30 — lowest disagreement ----------
    ws = wb.create_sheet('Top 30 lowest disagreement')
    def _gap_s(name_a, name_b):
        _fa, _fb = file_by_name.get(name_a), file_by_name.get(name_b)
        _ta = _fa.get('timestamp') if _fa else None
        _tb = _fb.get('timestamp') if _fb else None
        return abs(_ta - _tb) if _ta and _tb else None

    headers = ['Rank', 'Pair A', 'Pair B', 'Time gap (s)',
               'Combined disagreement',
               'Duplicate likelihood (%)', 'Similarity (min λ)']
    order = sorted(range(len(all_pairs_list)),
                   key=lambda i: all_pairs_list[i]['sum_score'])
    rows_data = []
    for rank, k in enumerate(order[:30], 1):
        p = all_pairs_list[k]
        rows_data.append([
            rank, p['a'], p['b'], _gap_s(p['a'], p['b']), p['sum_score'],
            p['p_dup'] * 100.0, p.get('r_min'),
        ])
    _write_table(ws, headers, rows_data,
                 col_widths=[6, 18, 18, 13, 22, 22, 18])

    # ---------- Top 30 — highest similarity ----------
    ws = wb.create_sheet('Top 30 highest similarity')
    headers = ['Rank', 'Pair A', 'Pair B', 'Time gap (s)', 'Similarity (min λ)',
               'Combined disagreement', 'Duplicate likelihood (%)']
    sim_sorted = sorted(
        [(i, p) for i, p in enumerate(all_pairs_list)
         if p.get('r_min') is not None],
        key=lambda x: -x[1]['r_min'])[:30]
    rows_data = []
    for rank, (_, p) in enumerate(sim_sorted, 1):
        rows_data.append([
            rank, p['a'], p['b'], _gap_s(p['a'], p['b']), p.get('r_min'),
            p['sum_score'], p['p_dup'] * 100.0,
        ])
    _write_table(ws, headers, rows_data,
                 col_widths=[6, 18, 18, 13, 18, 22, 22])

    # ---------- Charts ----------
    # Reuse the same multi-λ PNG renderers that feed the PDF so the Excel
    # user sees the identical visuals. WL_ORDER is consulted by the chart
    # functions for per-λ panels; temporarily override it for TRC datasets
    # that report different wavelengths.
    saved_wl = WL_ORDER
    WL_ORDER = wl_list
    try:
        chart_b64 = chart_distribution(all_pairs_list)
        hist_b64  = chart_histogram(all_pairs_list)
    finally:
        WL_ORDER = saved_wl

    try:
        ws = wb.create_sheet('Charts')
        ws['A1'] = 'Distribution charts'
        ws['A1'].font = TITLE_FONT
        # Distribution panel — full height (4 stacked subplots at 13×9)
        png_bytes = base64.b64decode(chart_b64)
        img = XlsxImage(BytesIO(png_bytes))
        orig_w, orig_h = img.width, img.height
        target_w = 1400
        img.width  = target_w
        img.height = int(target_w * orig_h / orig_w) if orig_w else target_w // 2
        ws.add_image(img, 'A3')
        # Histogram panel — placed below the distribution chart
        if hist_b64:
            png2 = base64.b64decode(hist_b64)
            img2 = XlsxImage(BytesIO(png2))
            ow, oh = img2.width, img2.height
            img2.width  = target_w
            img2.height = int(target_w * oh / ow) if ow else target_w // 2
            # leave ~60 rows below the distribution chart
            ws.add_image(img2, 'A70')
    except Exception as exc:
        print(f'  warn: skipped Charts sheet ({exc})')

    # Boss request: duplicates up front — first sheet after Summary.
    if 'Confirmed duplicates' in wb.sheetnames:
        wb.move_sheet('Confirmed duplicates',
                      offset=1 - wb.sheetnames.index('Confirmed duplicates'))
    wb.save(out_xlsx)
    print(f'XLSX: {out_xlsx}')
    return out_xlsx


def _classify_regime_multiwl(files, batch, wl_list):
    """Three-regime classifier (matches the SOR side):

        'production' — bulk pair-r low (~0.3). σ-outlier detector works.
        'tie_panel'  — bulk pair-r high (~0.95) AND bulk pair-σ ≥ 0.10 dB.
                       Many fibers sharing launch+connector signal.
        'all_dups'   — bulk pair-r high (~0.95) AND bulk pair-σ < 0.10 dB.
                       Every file is the same physical fiber.

    Uses the canonical wavelength's σ/r matrix (1550 nm preferred, else the
    median wavelength). Replaces the prior interior-event count + 20-file
    floor heuristic — that test was a band-aid that failed when small
    homogeneous-duplicate datasets (renoduplicates, newbeta) needed
    production-style detection."""
    if 1550 in wl_list:
        canonical = 1550
    else:
        canonical = sorted(wl_list)[len(wl_list) // 2]
    if canonical not in batch:
        return 'production', 0.0, 0.0
    sm = batch[canonical]['sigma_matrix']
    rm = batch[canonical]['r_matrix']
    n = sm.shape[0]
    if n < 2:
        return 'production', 0.0, 0.0
    iu = np.triu_indices(n, k=1)
    bulk_sigma = float(np.median(sm[iu])) if len(iu[0]) else 0.0
    bulk_r = float(np.median(rm[iu])) if len(iu[0]) else 0.0
    # Fraction of pairs with elevated raw r — catches long tie panels whose
    # median r is low but a large minority of pairs share cable structure
    # near r ~ 1.0 (see report_sor.py for the CLQTILA rationale).
    frac_high_r = float((rm[iu] >= 0.95).mean()) if len(iu[0]) else 0.0
    # Compute min interior length across files for the short_panel trigger.
    # Take the canonical-λ length from each file; fall back to the longest λ
    # the file reports if canonical isn't present.
    file_lengths = []
    for f in files:
        wl_rec = (f.get('wl') or {}).get(canonical) or {}
        L = wl_rec.get('length_m')
        if not L:
            # Try any λ that does report a length
            for wl_key, wlrec in (f.get('wl') or {}).items():
                if (wlrec or {}).get('length_m'):
                    L = wlrec['length_m']
                    break
        if L:
            file_lengths.append(float(L))
    min_L = min(file_lengths) if file_lengths else 0.0
    # Same four-regime taxonomy as the SOR side — see report_sor.py for
    # full rationale. Order matters: all_dups checked first so a
    # hypothetical all-duplicates short-fiber dataset doesn't get misrouted.
    if bulk_r >= 0.7 and bulk_sigma < 0.10:
        regime = 'all_dups'
    elif min_L > 0 and min_L < 200 and n >= 50:
        regime = 'short_panel'
    elif bulk_r >= 0.7 or frac_high_r >= 0.30:
        regime = 'tie_panel'
    else:
        regime = 'production'
    return regime, bulk_sigma, bulk_r


def _build_pairs_multiwl(files, wl_list, truth_dups):
    """Compute the all_pairs list using the batch metric helper. Returns
    (all_pairs, regime). `regime` is 'production' / 'tie_panel' / 'all_dups'."""
    # Two-pass: compute raw metrics, classify, then re-compute with
    # fingerprint extraction only if the dataset is a tie panel.
    batch_raw = _compute_pair_metrics_batch_multiwl(files, wl_list,
                                                    tie_panel_mode=False)
    regime, bulk_sigma, bulk_r = _classify_regime_multiwl(files, batch_raw, wl_list)
    print(f'Regime: {regime} (bulk σ={bulk_sigma:.4f} dB, bulk r={bulk_r:.4f})')
    if regime == 'tie_panel':
        batch = _compute_pair_metrics_batch_multiwl(files, wl_list,
                                                    tie_panel_mode=True)
    else:
        batch = batch_raw
    # Build a (file_index_i, file_index_j) -> per-wavelength scalar lookup.
    n = len(files)
    pairs_by_key = {}
    for wl, b in batch.items():
        K = len(b['valid_idx'])
        sm, rm, vi = b['sigma_matrix'], b['r_matrix'], b['valid_idx']
        for ki in range(K):
            i = vi[ki]
            for kj in range(ki + 1, K):
                j = vi[kj]
                key = (i, j)
                rec = pairs_by_key.setdefault(key, {'score': {}, 'shape_r': {}})
                rec['score'][wl] = float(sm[ki, kj])
                rec['shape_r'][wl] = float(rm[ki, kj])
    truth_dups = truth_dups or set()
    all_pairs = []
    for (i, j), rec in pairs_by_key.items():
        a, b = files[i], files[j]
        sc, rs = rec['score'], rec['shape_r']
        sum_sc = sum(v for v in sc.values() if v is not None)
        rs_vals = [v for v in rs.values() if v is not None]
        r_min = min(rs_vals) if rs_vals else None
        is_dup = tuple(sorted([a['name'], b['name']])) in truth_dups
        all_pairs.append({'a': a['name'], 'b': b['name'],
                          'score': sc, 'sum_score': sum_sc, 'is_dup': is_dup,
                          'shape_r': rs, 'r_min': r_min})
    return all_pairs, regime


def build_json_html(folder, title='Duplicate Classification Report', truth_dups=None):
    paths = sorted(glob.glob(os.path.join(folder, '*.json')))
    if not paths:
        raise RuntimeError(f'No JSON files found in {folder}')
    files = _load_json_files(paths)
    all_pairs, regime = _build_pairs_multiwl(files, WL_ORDER, truth_dups)
    out_html_tmp = os.path.join(folder, '_tmp_report.html')
    build_report(files, all_pairs, truth_dups or set(), out_html_tmp,
                 title=title, regime=regime)
    with open(out_html_tmp, 'r', encoding='utf-8') as fh:
        html = fh.read()
    try:
        os.remove(out_html_tmp)
    except OSError:
        pass
    return html, files, all_pairs


def run_json_bytes(folder, title='Duplicate Classification Report', truth_dups=None):
    html, files, pairs = build_json_html(folder, title=title, truth_dups=truth_dups)
    return html_to_pdf_bytes(html, base_url=folder), len(files), len(pairs)


def build_xlsx_json(folder, title, out_xlsx, truth_dups=None):
    """Load JSON files from `folder`, run the multi-λ pipeline, and write an
    Excel workbook to `out_xlsx`. Same analysis as the PDF flow."""
    paths = sorted(glob.glob(os.path.join(folder, '*.json')))
    if not paths:
        raise RuntimeError(f'No JSON files found in {folder}')
    files = _load_json_files(paths)
    all_pairs, regime = _build_pairs_multiwl(files, WL_ORDER, truth_dups)
    build_xlsx_multiwl(files, all_pairs, truth_dups or set(), out_xlsx,
                       title=title, wl_list=WL_ORDER, regime=regime)
    return out_xlsx, files, all_pairs


def run_json_xlsx_bytes(folder, title='Duplicate Classification Report', truth_dups=None):
    """Run JSON mode and return (xlsx_bytes, n_files, n_pairs). Mirrors
    run_sor_xlsx_bytes so app.py can switch between modes uniformly."""
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        tmp = os.path.join(td, 'report.xlsx')
        _, files, pairs = build_xlsx_json(folder, title, tmp, truth_dups=truth_dups)
        with open(tmp, 'rb') as fh:
            xlsx_bytes = fh.read()
    return xlsx_bytes, len(files), len(pairs)


def build_trc_html(folder, title='Duplicate Classification Report', truth_dups=None):
    """TRC-mode equivalent of build_json_html. Loads .trc files via the TRC
    parser and reuses the JSON-mode renderer (same multi-wavelength layout)."""
    global WL_ORDER
    paths = sorted(glob.glob(os.path.join(folder, '*.trc')))
    if not paths:
        raise RuntimeError(f'No TRC files found in {folder}')
    files = _load_trc_files(paths)
    # Use whichever wavelengths the TRC files actually carry — fall back to
    # the production set if everything matches it.
    common = set(files[0]['wl'].keys())
    for f in files[1:]:
        common &= set(f['wl'].keys())
    wl_list = sorted(common) or WL_ORDER
    all_pairs, regime = _build_pairs_multiwl(files, wl_list, truth_dups)
    # Override module-level WL_ORDER for rendering when TRC carries fewer/other λ
    saved = WL_ORDER
    WL_ORDER = wl_list
    out_html_tmp = os.path.join(folder, '_tmp_report.html')
    try:
        build_report(files, all_pairs, truth_dups or set(), out_html_tmp,
                     title=title, regime=regime)
        with open(out_html_tmp, 'r', encoding='utf-8') as fh:
            html = fh.read()
    finally:
        WL_ORDER = saved
        try:
            os.remove(out_html_tmp)
        except OSError:
            pass
    return html, files, all_pairs


def run_trc_bytes(folder, title='Duplicate Classification Report', truth_dups=None):
    html, files, pairs = build_trc_html(folder, title=title, truth_dups=truth_dups)
    return html_to_pdf_bytes(html, base_url=folder), len(files), len(pairs)


def build_xlsx_trc(folder, title, out_xlsx, truth_dups=None):
    """Load TRC files from `folder`, run the multi-λ pipeline, and write an
    Excel workbook to `out_xlsx`. Uses whichever wavelengths the TRCs
    actually carry (falls back to WL_ORDER if every file matches)."""
    paths = sorted(glob.glob(os.path.join(folder, '*.trc')))
    if not paths:
        raise RuntimeError(f'No TRC files found in {folder}')
    files = _load_trc_files(paths)
    common = set(files[0]['wl'].keys())
    for f in files[1:]:
        common &= set(f['wl'].keys())
    wl_list = sorted(common) or WL_ORDER
    all_pairs, regime = _build_pairs_multiwl(files, wl_list, truth_dups)
    build_xlsx_multiwl(files, all_pairs, truth_dups or set(), out_xlsx,
                       title=title, wl_list=wl_list, regime=regime)
    return out_xlsx, files, all_pairs


def run_trc_xlsx_bytes(folder, title='Duplicate Classification Report', truth_dups=None):
    """Run TRC mode and return (xlsx_bytes, n_files, n_pairs)."""
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        tmp = os.path.join(td, 'report.xlsx')
        _, files, pairs = build_xlsx_trc(folder, title, tmp, truth_dups=truth_dups)
        with open(tmp, 'rb') as fh:
            xlsx_bytes = fh.read()
    return xlsx_bytes, len(files), len(pairs)


def run_json(folder, out_pdf, title='Duplicate Classification Report', truth_dups=None):
    pdf_bytes, _, _ = run_json_bytes(folder, title=title, truth_dups=truth_dups)
    with open(out_pdf, 'wb') as fh:
        fh.write(pdf_bytes)
    return out_pdf


def main():
    TRUTH_DUPS = {
        tuple(sorted(['VERSLK001','VERSLK013'])), tuple(sorted(['VERSLK002','VERSLK014'])),
        tuple(sorted(['VERSLK003','VERSLK015'])), tuple(sorted(['VERSLK010','VERSLK016'])),
        tuple(sorted(['VERSLK011','VERSLK017'])), tuple(sorted(['VERSLK012','VERSLK018'])),
    }

    paths = sorted(glob.glob(os.path.join(JSON_FOLDER, '*.json')))
    files = _load_json_files(paths)
    print(f'Loaded {len(files)} files')

    all_pairs = []
    for a, b in combinations(files, 2):
        sc = {wl: _score(a, b, wl) for wl in WL_ORDER}
        sum_sc = sum(v for v in sc.values() if v is not None)
        is_dup = tuple(sorted([a['name'], b['name']])) in TRUTH_DUPS
        all_pairs.append({'a': a['name'], 'b': b['name'],
                          'score': sc, 'sum_score': sum_sc, 'is_dup': is_dup})

    out_html = os.path.join(HERE, 'report.html')
    build_report(files, all_pairs, TRUTH_DUPS, out_html)
    print(f'Report: {out_html}')

    pdf = out_html.replace('.html', '.pdf')
    chrome = _find_chrome()
    if chrome:
        result = subprocess.run(
            [chrome, '--headless=new', '--disable-gpu', '--no-sandbox',
             '--run-all-compositor-stages-before-draw',
             '--virtual-time-budget=5000',
             f'--print-to-pdf={os.path.abspath(pdf)}',
             '--print-to-pdf-no-header', '--no-pdf-header-footer',
             'file://' + os.path.abspath(out_html)],
            capture_output=True, timeout=180,
            creationflags=getattr(subprocess, 'CREATE_NO_WINDOW', 0))
        if result.returncode == 0:
            print(f'   PDF: {pdf}')


if __name__ == '__main__':
    main()
