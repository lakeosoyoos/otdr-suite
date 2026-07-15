#!/usr/bin/env python3
"""
splicereportmatchexfo_April28_AOrdered.py
=========================================

Splice QC report (April 28, 2026 revision).  Produces an EXFO-style
bidirectional splice deliverable from OTDR JSON or SOR files.

Core capabilities:
  • Bidirectional splice analysis via JSON-trace wide-LSA — matches
    EXFO FastReporter's internal LSA calculation for the agreed-on
    target fibers within 0.001 dB.
  • BEND detection — two-test classifier (per-fiber linear length
    model + narrow-LSA confirmation) replaces the older flat
    geometric offset rule; eliminates the fiber-length-variation
    false positives that the old rule produced.
  • In-line REFLECTIVE EVENT (ref) classification — reflective
    events with strong Fresnel reflection where the trace continues
    past the event are tagged 'ref' instead of 'BREAK', rendered in
    deep orange with a label like 'F583 ref .862 (refl -29 dB)'.
  • FIELD-GAINER detection — strict bidirectional rule: requires
    both A and B real measurements with opposite signs, bidir avg
    in [−0.7, 0] dB, mid-span position.
  • LAUNCH-ISSUE detection — fibers broken / damaged at the launch
    end appear in the dedicated ILA column.
  • Past-A-break B-fill — recovers downstream splices on fibers
    broken on the A side.

APRIL 28 GATE SET (locked — see SCRIPT_LOGIC_FLOWCHART.pdf):
  • Reburn threshold:   0.160 dB (pink A+B reburn)
  • Bend threshold:     0.090 dB (positive signed loss required)
  • Bend Test 1:        per-fiber length-model residual
                          ≤ 50 m  → SPLICE
                          ≥ 150 m → run Test 2
                          else    → ambiguous (drop)
  • Bend Test 2:        narrow-LSA at predicted km ≥ 0.030 dB
                          → confirmed BEND
                          else → drop
  • Ref classifier:     reflective + Fresnel + trace continues past
                          (≥1 real event downstream AND EOF ≥ 3 km
                          farther) → ref (deep orange).  Otherwise
                          → BREAK (red).
  • Gainer rule:        bidir avg ∈ [−0.7, 0] dB AND both A and B
                          are real event measurements (no grey-LSA)
                          AND sign(A) ≠ sign(B).

CELL LABELS:
  '325 .172'                          A+B bidirectional reburn (pink)
  '325 .340 (B)'                       B-only event
  '325 .285 (A)'                       A-only event
  '583 ref .862 (refl -29 dB)'         in-line reflective event (deep orange)
  '583 BREAK 0.862 ...'                 break (red) — trace ends near here
  '229 broke@59.2k (B-fill OK)'        broke (red) — A trace terminated
  '841 .390 (B)'                       B-fill recovery past A-break (blue)

COLORS:
  Pink         — A+B reburn (≥ 0.160 dB bidir)
  Red          — Break / Broke (physical damage)
  Deep orange  — In-line REF event (reflective + Fresnel + trace continues)
  Blue         — B-fill past A-break
  Yellow       — Bend
  Mint         — Field gainer
  Coral        — A-only HIGH
  Lt. yellow   — A-only OK
  Purple       — B-only HIGH
  Lavender     — B-only OK
  Orange       — Launch issue
  Gray         — Dead zone

USAGE
-----
    python splicereportmatchexfo.py A_DIR/ B_DIR/ --output report.xlsx

OPTIONS
    --output PATH    Output Excel file (default: splice_report_exfo.xlsx)
    --threshold dB   Reburn threshold (default 0.160)
    --site-a NAME    A-end site name
    --site-b NAME    B-end site name
    --ribbon-size N  Fibers per ribbon (default 12)

REQUIREMENTS
    pip install numpy openpyxl
    sor_reader324802a.py and json_reader.py must be in same directory.
"""

import os
import re
import sys
import argparse
from collections import defaultdict

import math
import numpy as np

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)

from sor_reader324802a import (parse_sor_full, measure_grey_loss_from_sor,
                               measure_silent_grey_from_sor,
                               _sor_ior_from_events)
# JSON-based grey-value measurement — matches EXFO's internal LSA calculation
# (see json_reader.py for the algorithm details)
from json_reader import (
    parse_otdr_json,
    measure_grey_loss_from_json,
    load_all_json,
)


# ═══════════════════════════════════════════════════════════════════════
#  DEFAULTS
# ═══════════════════════════════════════════════════════════════════════

REBURN_THRESHOLD = 0.160   # dB — flag bidirectional reburns at or above
                           #      (boss spec: flag at >= 0.16 dB)
SINGLE_DIR_THRESHOLD = 0.250  # dB — single-direction-only events (A-only,
                              #     B-only, B-fill past A-break) need a
                              #     stricter threshold because the unseen
                              #     side can't confirm.  No averaging /
                              #     halving — the raw single-direction
                              #     loss must clear this on its own.
BIDIR_CONNECTOR_LOSS = 0.500  # dB — bidir loss at a reflective (1F)
                              #     event qualifies as a 'high connector
                              #     loss' when at or above this value.
                              #     EXFO convention: in-line connectors
                              #     and mech splices normally show
                              #     0.1-0.3 dB loss; >= 0.5 dB indicates
                              #     a degraded / dirty / damaged
                              #     connector worth calling out
                              #     separately from a normal reburn.
NOMINAL_SPLICE   = 0.159   # dB expected per splice
# ── Borderline / review band around the reburn threshold ─────────────────
# Bidir reburn is a HARD >=/< call at REBURN_THRESHOLD (set in the OTDR settings
# panel, typically 0.160): a loss flags iff it is >= threshold and anything below
# is simply not flagged.  There is deliberately NO borderline / review band — per
# the boss's workflow the threshold is a hard cutoff with no near-threshold review
# tier (_is_borderline_loss is a disabled no-op, kept only so its callers and the
# manifest's n_borderline field — now always 0 — stay in place).
RIBBON_SIZE      = 12      # fibers per ribbon
POSITION_TOL     = 1.5     # km tolerance for matching A↔B events
MIN_POP_SPLICE   = 20      # absolute floor: minimum fibers to define a splice position
MIN_POP_FRACTION = 0.25    # fractional floor: minimum % of fibers that must
                           # have an event at a km bucket to call it a real
                           # closure.  Real splices show 60-75% coverage
                           # (some low-loss splices don't generate detectable
                           # events).  Phantom one-fiber bends show <10%.
                           # 25% gives plenty of margin to separate the two.
CLOSURE_CLUSTER_GAP_KM = 0.25  # km — discover_splices splits the cable-wide
                           # event stream into closures wherever consecutive
                           # event positions are farther apart than this.
                           # Replaces the old 1 km integer rounding, which split
                           # closures straddling a *.5 km boundary and merged
                           # distinct closures sharing an integer km.
END_REGION_KM    = 3.0     # last N km considered "end of fiber"
LAUNCH_FIBER_MAX = 3.0     # km — max distance for launch connector detection

# ── B-confirmation of end-region closures (HOWLAN direction-swap fix) ───────
# The END_REGION_KM phantom filter assumed no REAL splice lives in the last
# 3 km of the cable.  HOWLAN broke that: Splice 1 sits 1.8 km from Howe, so
# loading Lancaster as the A side put it inside the end region and silently
# deleted it (plus its ~57 reburn flags).  A closure near A's far end is near
# B's LAUNCH — B's cleanest region — so before dropping we now ask the B
# direction: does a discovery-strength population of B fibers see an event at
# the mirror position?  Real splice → yes (kept).  True phantom (post-EOL
# tail cluster / noise) → B's near-launch is clean there → no (dropped, as
# before).  Population reuses MIN_POP_SPLICE / MIN_POP_FRACTION so "confirm"
# means exactly "B could have discovered this closure itself".
END_REGION_B_CONFIRM_KM      = 0.5   # km — ± match window around the B-frame
                                     # mirror (wide: A far-end smear + two
                                     # independent span estimates both err)
END_REGION_B_LAUNCH_GUARD_KM = 1.0   # km — the mirror must clear B's launch
                                     # zone (same 1 km floor discovery uses),
                                     # so B's launch connector can never
                                     # "confirm" the cable-end candidate

# ── DIRTY / BAD connector recategorization (sandbox loop, milestone 3) ──────
# An already-flagged reflective in-line event that ALSO drops a real loss step
# is a dirty / failing connector (contamination, scratched endface, bad mate).
# We do NOT change its flag decision — it already flags on loss/reflectance —
# we only refine its category + label so the report tells the tech which
# connectors to clean / reseat.  Mirrors ~/Desktop/splice-tune/verify/
# connector_quality.py :: is_dirty_connector.
DIRTY_CONN_REFL_GATE_DB   = -55.0   # reflection >= this (stronger than -55 dB)
DIRTY_CONN_LOSS_GATE_DB   = 0.10    # |loss| >= this (a real loss step, dB)
DIRTY_CONN_LAUNCH_EXCL_KM = 0.05    # exclude launch connector at normalized ~0 km


def _is_dirty_connector(dist_km, reflection, loss, is_end=False):
    """True if a reflective in-line event is a dirty/bad connector: not the
    end, past the launch connector, reflectance stronger than the gate, and
    carrying a real loss step.  Callers gate this on the event already being
    reflective (is_ref / 1F), so reflectivity itself is not re-tested here."""
    if is_end:
        return False
    if dist_km is None or dist_km <= DIRTY_CONN_LAUNCH_EXCL_KM:
        return False
    if reflection is None or reflection < DIRTY_CONN_REFL_GATE_DB:
        return False
    if loss is None or abs(loss) < DIRTY_CONN_LOSS_GATE_DB:
        return False
    return True

# Wide-LSA grey-value measurement windows (matches EXFO's FastReporter
# behavior — see discussion in json_reader.py / previous experiments):
GREY_LSA_OUTER_M = 5000    # m — outer LSA window on each side of splice
GREY_LSA_INNER_M = 60      # m — inner dead zone on each side of splice

# ── BEND detection (from ZeroDBIFTHEN's Flag 3 rule / boss's method) ─────────
#
#   An OTDR event is a BEND (not a splice) if it meets BOTH:
#     1. Its bidir / single-direction loss magnitude >= BEND_THRESHOLD
#     2. Its position is more than CLOSURE_MATCH_KM from the nearest *real*
#        splice closure center (not the splice report's coarse 1 km bin,
#        but the MODE of fiber event positions inside that bin).
#
# Bends are flagged in their own category.  Analysis of downstream splices
# continues normally past a bend — unlike a break, a bend does not make the
# trace blind past it.
#
BEND_THRESHOLD        = 0.090   # dB — minimum loss to call an event a "bend"
CLOSURE_MATCH_KM      = 0.075   # km — tight window; farther → classify as bend
# "At the splice" distance for COLUMN placement (OTDR-panel editable: "Bend
# fold distance").  A bend/break cluster within this of a validated splice
# column stays IN that splice column (cells keep their bend labels); farther
# out it gets its own "Bends @ X km" column.  Platteville–Cheyenne ground
# truth: a few short-lay fibers put their splice events 107–128 m before the
# column, and the old hard-wired 75 m (CLOSURE_MATCH_KM) gate spawned six
# phantom bend columns hugging real splices ("calling bends at splices").
# 200 m folds all of those with margin (their fold-check medians reach
# 153 m); genuinely off-splice zones (PLACHE fiber-187 damage @450 m,
# Seattle's far bends @222 m–3.6 km) keep their own columns.
BEND_SPLICE_FOLD_KM   = 0.200
                                #     (75 m default — was 0.150; sensitivity
                                #     sweep across the test corpus showed the
                                #     cliff at 60-75 m where Vernon-SLC's
                                #     per-fiber drift starts spawning false
                                #     bend columns.  75 m is the tight floor;
                                #     slider lets techs go tighter per-span.)
# ── Bend asymmetry gate (April 27 revision) ───────────────────────────────
# A real macrobend at the closure is typically ASYMMETRIC in bidirectional
# OTDR — most of the loss shows up in one direction's trace and the other
# direction reads near-zero (or even slightly negative).  A position-mismatch
# artifact (fiber whose splice km drifted from its ribbon population because
# of fiber-length variation in the cable) is SYMMETRIC: both A→B and B→A
# show the splice with normal magnitudes.  Requiring asymmetry kills the
# fiber-length false positives without blinding us to real bends.
#
#   bend  ⇔  min(|a_loss|, |b_loss|) < BEND_ASYM_LOW
#       AND max(|a_loss|, |b_loss|) ≥ BEND_THRESHOLD
#
# Only applied when BOTH a_loss and b_loss are available (i.e. true
# bidirectional cases).  A-only / B-only events fall back to the geometric
# offset test alone.
BEND_ASYM_LOW         = 0.020   # dB — "near-zero" side of an asymmetric bend
                                # (currently INERT — see _is_bend_event)
# ── Per-fiber length model + trace inspection (April 28 revision) ──────────
# Replaces the pure-geometric offset gate with a two-test classifier:
#
#   TEST 1 — Per-fiber linear length model.  Each fiber has its own
#   length factor (slope ≈ 1.000 ± a few × 1e-3) due to helical lay
#   and per-tube length variation.  We fit
#       fiber_event_km = a · cable_splice_km + b
#   through every (cable_km, event_km) pair this fiber has at any
#   closure, EXCLUDING the closure being classified.  Predict where
#   the splice should sit on this fiber, then compute the residual.
#
#       |residual| ≤ BEND_RES_SPLICE_M     → SPLICE  (never a bend)
#       |residual| ≥ BEND_RES_BEND_M       → BEND candidate (run Test 2)
#       in between                          → ambiguous (treat as splice;
#                                              tech can review high-loss
#                                              cases in the report)
#
#   TEST 2 — Narrow-LSA confirmation.  When Test 1 says BEND, run a
#   narrow-window LSA (±BEND_NARROW_OUTER_M outer / ±BEND_NARROW_INNER_M
#   inner) on the raw A trace at the predicted splice km.  If a real
#   loss step is present there (≥ BEND_NARROW_LOSS_DB), the splice is
#   confirmed at predicted_km AND a separate event exists at the
#   candidate position → BEND.  If no loss at predicted_km, fall back
#   to "ambiguous" (don't flag — single event of unclear nature).
#
# When fiber_data isn't JSON (SOR-only inputs), Test 2 isn't available
# and a Test-1 BEND verdict is taken as the final answer.
BEND_RES_SPLICE_M     = 50      # m — residual ≤ this → splice (Test 1)
BEND_RES_BEND_M       = 150     # m — residual ≥ this → bend candidate
BEND_NARROW_OUTER_M   = 5000    # m — Test-2 LSA outer window (now wide-LSA)
BEND_NARROW_INNER_M   = 60      # m — Test-2 LSA dead zone

# ── Local-step re-measure gate (the PLACHE/HOWLAN phantom killer) ───────────
# The SOR event table's stored splice_loss is EXFO's LONG-window LSA (its
# per-event markers span up to 5 km).  When the detector false-fires on a
# marginal wiggle, that convention converts gentle trace curvature into a
# manufactured 0.02–0.30 dB "loss" at a position where the glass is locally
# clean — verified on PLACHE (6 phantom bend clusters, both directions:
# marker-LSA reproduces the table to ±0.002 dB, yet no step >0.008 dB exists
# within 450 m) and HOWLAN (13 phantom one-direction flags).  Before a
# STORED loss may drive a bend/single-direction flag, re-measure the trace
# with a TIGHT two-line LSA (±LOCAL_STEP_HALF_M fit windows, LOCAL_STEP_GAP_M
# dead zone — outside the ~127 m pulse smear, far inside curvature scale) and
# require at least LOCAL_STEP_CONFIRM_DB of real step.  Ground truths:
# phantoms measure ≤0.011 dB, the weakest boss-confirmed real bend (Seattle
# F426) measures 0.043 dB — 0.025 splits with margin both ways.  Reported
# NUMBERS stay EXFO's stored values (FastReporter north star); this is a
# flag GATE only.  Unmeasurable (no trace / window truncated) → keep the
# flag — never hide a possible defect because we couldn't measure.
LOCAL_STEP_CONFIRM_RATIO = 0.35  # tight read must reach this fraction of the
                                 # stored loss (real events read 40-50% at
                                 # 2500 ns; HOWLAN's phantoms read ~0%)
LOCAL_STEP_GATE_MIN_DB   = 0.15  # only police stored claims this big — the
                                 # proven phantom class stores 0.26-0.29 dB;
                                 # smaller claims' tight reads sink into noise
LOCAL_STEP_HALF_M     = 250.0
LOCAL_STEP_GAP_M      = 50.0
LOCAL_STEP_SCAN_M     = 350.0   # scan window past the event mark (EXFO marks
                                # the ONSET; the smeared step center sits up
                                # to ~300 m downstream at 2500 ns)
BEND_NARROW_LOSS_DB   = 0.030   # dB — narrow-LSA threshold for "loss present"
BEND_PERFIBER_WIN_KM  = 0.500   # km — per-fiber pair window around closure
BEND_PERFIBER_MIN_FIT = 3       # min fit points (other closures) for the model
# Severity tiers intentionally collapsed: any bend ≥ BEND_THRESHOLD is
# rendered with the same yellow fill.  The old WATCH / REVIEW / HIGH
# tiers are retained as constants only for backward-compatibility with
# any downstream code that still references them.
BEND_HIGH_LOSS        = BEND_THRESHOLD
BEND_REVIEW_LOSS      = BEND_THRESHOLD
# ── Consensus-bend cluster off-grid gate (binary-classifier path only) ────
# The cross-fiber consensus-bend sweep hard-flags an off-grid cluster only
# when its ROBUST (median) position sits more than this far from the nearest
# closure-grid column.  A *cluster* of co-located fibers needs more clearance
# than the single-event CLOSURE_MATCH_KM (75 m) attribution window, because a
# tight cluster a little past a splice is the splice's own population spread,
# not a separate bend zone.  Calibrated on the Seattle ground truth (boss-
# confirmed): the six real bends sit >=127 m from any splice, while the 100.46
# cluster sits 84 m from Splice 20 and is the splice — 100 m cleanly separates
# them.  Local to flag_consensus_bends; does NOT touch the closure clusterer.
CONSENSUS_BEND_OFFGRID_KM = 0.100   # km — min cluster-median offset from a column
# ── Helix-aware off-grid gate (June 2026) ──────────────────────────────────
# Fibers in one cable have slightly different lengths-per-cable-metre (helical
# lay), so a shared closure lands at slightly different OTDR distances across
# fibers, and that spread GROWS with distance.  On a long / high-helix span the
# tail of a closure drifts hundreds of metres past the fixed 100 m gate above,
# so flag_consensus_bends used to emit it as a phantom BEND column shadowing the
# real splice (HOWLAN→Lancaster: 0.8% helix over 117 km → ~900 m far-end spread,
# 536 splices mis-flagged as bends).  The fix scales the off-grid tolerance by
# (distance × the span's helix half-spread × HELIX_TOL_K), never below the fixed
# floor — so a cluster within the helix-explained drift of a closure is that
# closure's tail (splice), while a genuinely off-grid bend (Seattle's six, all
# >127 m off where helix only explains ~40 m) still flags.
HELIX_TOL_K            = 2.0     # multiple of the half-spread for the extreme tail
HELIX_HALFSPREAD_MAX   = 0.010   # clamp the estimate to ≤1% so noise can't over-absorb
# Per-fiber helix-model residual gate (the PRINCIPLED, cable-agnostic discriminator):
# a consensus cluster whose fibers' events sit where THEIR OWN length model predicts
# a splice at the nearest closure (median residual below this) is that closure's
# helix-drifted SPLICE tail, not a bend.  Calibrated on BOTH ground truths in-engine:
# HOWLAN's helix tails (boss: 0 bends) have residuals 6–42 m, while Seattle's six
# boss-confirmed bends have 100–3732 m — 75 m separates them with margin (and aligns
# with CLOSURE_MATCH_KM, the "same-event" distance).  One HOWLAN near-launch cluster
# (1.72 km, 2 fibers, residual 100 m) is NOT helix-explained and correctly survives
# as a lone flag — the conservative direction (never hide a possible real bend).
HELIX_RESIDUAL_BEND_M  = 75.0
# Account-then-flag gate (split_offsplice): an off-grid event folds into its
# closure column (counts as the fiber's OWN drifted splice, no separate column)
# only when it sits within this many metres of where the fiber's per-fiber length
# model predicts its splice at the nearest closure.  FIXED (not σ-scaled — that was
# falsified on Seattle, where real bends have noisy fits).  Folds HOWLAN's drifted
# mid-span splices (<50 m) while keeping every real bend (Seattle 144-670 m off
# prediction).  The far-end ref (~140 m) is inseparable from a real feature and
# correctly stays flagged.
HELIX_SPLICE_TOL_M     = 90.0
# Far-end EOF-anchored fold (LAST closure only).  The linear per-fiber model above
# under-predicts the splice shift at the cable END because helix drift goes
# NON-LINEAR there — so the 90 m gate can't separate a helix-shifted last splice from
# a real bend (both ~140 m off the linear prediction).  But the splice→end-of-fiber
# DISTANCE is preserved per fiber (measured on HOWLAN: ~8.0 km on every fiber, normal
# AND short-reading).  So for the LAST closure, anchor off the fiber's OWN EOF:
# predict its splice at (fiber_eof - consensus(eof - last_closure)) and fold if the
# event lands within HELIX_EOF_TOL_M.  Fires ONLY for fibers reading at least
# HELIX_EOF_MIN_SHORT_KM short at their EOF (the helix signature) — so it is inert on
# non-helix spans (Seattle: no short-reading fibers → branch never runs) and cannot
# fold a mid-span bend.
HELIX_EOF_TOL_M        = 300.0   # far-end splice positions scatter ~±265 m vs the EOF anchor
HELIX_EOF_MIN_SHORT_KM = 0.15    # only fibers reading >=150 m short at EOF use the EOF anchor
# Histogram bin for the mode-based closure-center refinement:
CLOSURE_MODE_BIN_M    = 25      # m — bin width for position-mode histogram
CLOSURE_MODE_WINDOW_M = 75      # m — window around mode peak for median refinement

# ── APRIL 23 revision: closure validation (drop phantom splice columns) ──
# A cluster discovered by discover_splices() is only a *real* closure if the
# fiber-event positions inside it cluster tightly.  Loose clusters are bend
# or damage zones mis-identified as splices.  Matches the tech's mental
# model: real closures look tight, bend zones are smeared.
CLOSURE_VALID_STD_MAX_M  = 150    # m — std of positions inside closure
CLOSURE_VALID_TIGHT_FRAC = 0.13   # fraction of ALL fibers whose event is
                                   #   within CLOSURE_MATCH_KM of the refined
                                   #   center.  Set intentionally low because
                                   #   broken fibers reduce the denominator:
                                   #   catastrophic breaks kill the count for
                                   #   every closure past the break zone.
                                   #   0.13 drops clear damage/bend phantoms
                                   #   (typical tight_frac ≈ 0.05–0.12) while
                                   #   keeping real splices (≥ 0.15).

# Additional closure-validity rules derived from raw-data analysis of
# the Cle Elum → Yakima 18.89 km bend zone (tech flagged bends; we flagged
# splice).  A real splice closure should look like random fusion-loss
# variation — small median loss, SOME apparent gainers from MFD mismatch.
# A bend zone looks very different: every fiber shows positive (loss) with
# a higher median.
CLOSURE_VALID_MIN_GAINER_FRAC = 0.05   # ≥ 5% of fibers in the cluster must
                                        #   show loss < 0 (apparent gain).
                                        #   Real splices ≈ 20-40% gainers.
                                        #   Bend zones ≈ 0% gainers.
CLOSURE_VALID_MEDIAN_LOSS_MAX = 0.100   # dB — median loss inside the tight
                                        #   cluster.  Real splices ≈ 0.04-
                                        #   0.08 dB median, bend zones push
                                        #   well above 0.10 dB.

# ── LAUNCH-issue detection (fibers broken/damaged at/near the launch end) ───
#
# Fibers with launch issues often silently disappear from the splice report:
# the event table ends almost immediately, so neither Pass 1 (splice analysis
# at known closures) nor Pass 2 (B-direction scan) has anything to match on.
# These fibers need to be flagged on their own so the tech knows to go look.
#
LAUNCH_HIGH_LOSS_DB          = None   # launch-event LOSS rule disabled per tech
                                      #   direction — the gate is on reflectance,
                                      #   not loss.
LAUNCH_BAD_REFL_DB           = -49.9  # launch reflectance threshold (signed,
                                      #   inclusive greater-than-or-equal).  Rule:
                                      #     refl <  -49.9 dB → good (no flag)
                                      #     refl >= -49.9 dB → bad  (flag)
                                      #   Healthy buried launch is -50 to -55 dB.
                                      #   Floats that display as "-50.0" but are
                                      #   actually -49.95 / -49.97 round in
                                      #   their favor and pass.  Anything that
                                      #   actually reaches -49.9 (or any value
                                      #   "larger" / closer to zero) is flagged
                                      #   as a damaged / dirty connector.
                                      #   -54.8 good, -50.0 good, -49.95 good,
                                      #   -49.9 BAD, -49.8 BAD, -10 BAD.
# ── FIELD-EVENT GAINER GATE ─────────────────────────────────────────────────
# Mid-span events whose signed loss falls in the [-0.7, 0] dB range get
# flagged as suspicious gainers — these are weak-gainer / near-zero events
# that the regular bend rule would either miss (if |loss| < 0.090) or
# silently fold into a normal splice column.  Excludes anything inside
# LAUNCH_FIBER_MAX km of the launch or END_REGION_KM km of the fiber end —
# those zones are evaluated separately by the launch / end logic.
FIELD_GAINER_MIN_DB          = -0.7   # most-negative loss that still flags
FIELD_GAINER_MAX_DB          = 0.0    # least-negative loss that still flags
LAUNCH_REFL_OUTLIER_DB       = 10.0   # |fiber_refl − population_median| > this → issue
LAUNCH_NO_FIRST_SPLICE_TOL_KM = 2.0   # km — must see an event within this of the
                                      #      first population closure


# ═══════════════════════════════════════════════════════════════════════
#  AUTO-DETECT & NORMALIZE UNTRIMMED TRACES
# ═══════════════════════════════════════════════════════════════════════

def _normalize_untrimmed_events(events):
    """Detect and normalize events from SOR files where start/stop was not picked.

    Untrimmed pattern (tech did NOT pick start/stop):
      #1  1F  0.000 km   — OTDR port (instrument origin)
      #2  1F  ~1.0  km   — launch connector (fiber-under-test starts here)
      ...  splice events with ~1 km offset  ...
      #N-1  1F  ~98.3 km — far-end connector (receive fiber)
      #N    xE  ~99.3 km — end-of-fiber marker

    Trimmed pattern (tech already picked start/stop):
      #1  1F  0.000 km   — launch connector (already set as origin)
      ...  splice events at correct positions  ...
      #N    xE  ~97.2 km — end-of-fiber marker

    Detection: first TWO events are both reflective (1F) non-end events,
    with the second one at a short distance (< LAUNCH_FIBER_MAX km).

    Normalization:
      1. Remove event #1 (OTDR port)
      2. Re-reference all distances from the launch connector (event #2 → dist 0)
      3. Remove the far-end connector (last 1F within 3 km of the end event)
    """
    if len(events) < 3:
        return events

    # ── Detect untrimmed ──
    e0, e1 = events[0], events[1]
    if not (e0['is_reflective'] and not e0['is_end'] and
            e0['time_of_travel'] == 0 and
            e1['is_reflective'] and not e1['is_end'] and
            0 < e1['dist_km'] < LAUNCH_FIBER_MAX):
        return events  # already trimmed — no-op

    launch_dist = e1['dist_km']
    launch_travel = e1['time_of_travel']

    # ── Find end-of-fiber event ──
    end_idx = None
    for i, e in enumerate(events):
        if e['is_end']:
            end_idx = i
            break

    # ── Find far-end connector: last 1F just before the end event ──
    far_end_idx = None
    if end_idx is not None and end_idx > 1:
        end_dist = events[end_idx]['dist_km']
        for i in range(end_idx - 1, 0, -1):
            if events[i]['is_reflective'] and not events[i]['is_end']:
                if (end_dist - events[i]['dist_km']) < LAUNCH_FIBER_MAX:
                    far_end_idx = i
                break  # only check the immediately preceding reflective event

    # ── Compute the adjusted end position ──
    # When the tech picks start/stop, the end is set at the far-end connector,
    # not at the trace noise floor beyond the receive fiber.  Mirror that by
    # moving the end event to the far-end connector position.
    far_end_norm_dist = None
    far_end_norm_travel = None
    if far_end_idx is not None:
        far_end_norm_dist = round(events[far_end_idx]['dist_km'] - launch_dist, 4)
        far_end_norm_travel = max(0, events[far_end_idx]['time_of_travel'] - launch_travel)

    # ── Build normalized event list ──
    normalized = []
    for i, e in enumerate(events):
        if i == 0:           # skip OTDR port
            continue
        if i == far_end_idx: # skip far-end connector
            continue
        new_e = dict(e)
        new_e['dist_km'] = round(e['dist_km'] - launch_dist, 4)
        new_e['time_of_travel'] = max(0, e['time_of_travel'] - launch_travel)
        # Move end event to the far-end connector position (strip receive fiber)
        if e['is_end'] and far_end_norm_dist is not None:
            new_e['dist_km'] = far_end_norm_dist
            new_e['time_of_travel'] = far_end_norm_travel
        normalized.append(new_e)

    return normalized


def _untrimmed_launch_offset_km(events):
    """Return the launch-connector offset that _normalize_untrimmed_events will
    subtract from this fiber's event distances (0.0 when already trimmed).

    Normalization re-references event distances to the launch connector
    (launch → 0), but the raw trace samples are NOT shifted: sample 0 stays the
    OTDR port.  So a normalized position P maps to trace coordinate P + offset.
    The silent-side windower indexes the raw trace, so it needs this offset to
    land its LSA windows at the right physical place (the old learned-marker
    recipe folded ~1.0 km into its window edges to absorb exactly this; the
    EXFO-exact recipe keeps the offset explicit instead)."""
    if len(events) < 3:
        return 0.0
    e0, e1 = events[0], events[1]
    if (e0['is_reflective'] and not e0['is_end'] and e0['time_of_travel'] == 0 and
            e1['is_reflective'] and not e1['is_end'] and
            0 < e1['dist_km'] < LAUNCH_FIBER_MAX):
        return float(e1['dist_km'])
    return 0.0


# ═══════════════════════════════════════════════════════════════════════
#  TRACE-BASED SPAN & BREAK DETECTION
# ═══════════════════════════════════════════════════════════════════════

# Detection thresholds
SPIKE_MIN_DB      = 3.0    # minimum dB above baseline for connector spike
NOISE_STDDEV_THR  = 0.5    # stddev threshold separating signal from noise floor
BREAK_MEAN_THR    = 58.0   # dB mean threshold for break/saturation
BREAK_STDDEV_THR  = 1.0    # stddev threshold for noise spike at break
NOISE_WINDOW      = 50     # samples for sliding window statistics


def _sample_to_km(idx, ior, pts, acq_range):
    """Convert a trace sample index to distance in km."""
    return idx * 0.02998 * 2 * acq_range / (1000.0 * ior * pts)


def _km_to_sample(km, ior, pts, acq_range):
    """Convert distance in km to a trace sample index."""
    return int(round(km * 1000.0 * ior * pts / (0.02998 * 2 * acq_range)))


def _sliding_stats(trace, window=NOISE_WINDOW):
    """Compute sliding-window mean and stddev for the trace.

    Returns (means, stds) arrays of same length as trace.
    Uses a fast cumulative-sum approach.
    """
    n = len(trace)
    means = np.empty(n)
    stds = np.empty(n)
    for i in range(n):
        lo = max(0, i - window)
        hi = min(n, i + window + 1)
        seg = trace[lo:hi]
        means[i] = seg.mean()
        stds[i] = seg.std()
    return means, stds


def _detect_launch_from_trace(trace, pts, acq_range, ior):
    """Detect the launch connector from the raw OTDR trace.

    The launch connector creates a dead-zone DIP in the backscatter trace:
    the strong Fresnel reflection saturates the detector, causing a local
    minimum at ~0.5-2.0 km.  The fiber-under-test begins after this dip.

    Returns the sample index of the launch connector.
    """
    n = len(trace)
    # Skip the OTDR port dead zone (first ~0.1km) and search 0.1-3.0 km
    skip = _km_to_sample(0.1, ior, pts, acq_range)
    search_end = min(_km_to_sample(LAUNCH_FIBER_MAX, ior, pts, acq_range), n - 1)

    if search_end <= skip + 10:
        return 0

    region = trace[skip:search_end]
    # The launch connector is at the local minimum (bottom of the dead-zone dip)
    min_rel = int(np.argmin(region))
    launch_idx = skip + min_rel

    return launch_idx


def _detect_noise_floor_from_trace(trace, launch_idx, pts, acq_range, ior):
    """Find where the trace transitions from signal to noise floor.

    The OTDR pulse width limits how far the saved trace has clean signal.
    Beyond this point, the trace goes to saturation (~64 dB) with high noise.

    This is the TRACE noise floor — NOT necessarily the fiber end (the OTDR
    events extend further due to multi-acquisition).  Used for break detection:
    if a fiber's trace goes to noise significantly earlier than the population,
    it likely has a break.

    Returns the sample index where signal transitions to noise.
    """
    n = len(trace)
    step = max(1, NOISE_WINDOW // 2)

    # Scan backward from end to find where clean signal begins
    for i in range(n - NOISE_WINDOW - 1, launch_idx, -step):
        seg = trace[i:i + NOISE_WINDOW]
        if seg.std() < NOISE_STDDEV_THR and seg.mean() < BREAK_MEAN_THR:
            # Found clean signal — noise floor starts after this
            noise_start = i + NOISE_WINDOW
            # Refine forward
            for j in range(noise_start, min(noise_start + NOISE_WINDOW * 2, n)):
                seg2 = trace[max(0, j - 10):j + 10]
                if seg2.std() > NOISE_STDDEV_THR or seg2.mean() > BREAK_MEAN_THR:
                    return j
            return noise_start

    return n - 1  # entire trace is noise (shouldn't happen)


def _detect_breaks_from_trace(trace, launch_idx, end_idx):
    """Detect mid-span breaks from the raw trace.

    A break is a sudden jump to near-saturation (>58 dB) with elevated noise,
    preceded by normal signal (<55 dB, low noise).

    Returns list of sample indices where breaks occur.
    """
    breaks = []
    n = len(trace)
    w = NOISE_WINDOW
    i = launch_idx + w

    while i < end_idx - w:
        seg_after = trace[i:i + w]
        seg_before = trace[max(launch_idx, i - w):i]

        mean_after = seg_after.mean()
        mean_before = seg_before.mean()
        std_after = seg_after.std()

        # Break: signal was present before, saturated after
        if (mean_before < 55.0 and
                mean_after > BREAK_MEAN_THR and
                std_after > BREAK_STDDEV_THR):
            # Walk backward to find the exact transition sample
            break_idx = i
            for j in range(i, max(launch_idx, i - w), -1):
                if trace[j] < 55.0:
                    break_idx = j + 1
                    break
            breaks.append(break_idx)
            # Skip past this break region
            i += w * 4
            continue

        i += w // 2

    return breaks


def _enhance_events_with_trace(fiber_result, expected_span_km, ior=None, pop_noise_floor_km=None):
    """Enhance a fiber's event list using raw trace analysis.

    Detects the launch connector from the trace (more accurate than events),
    detects breaks where the trace goes to saturation earlier than expected,
    and re-normalizes events using the trace-detected offset.

    For end-of-fiber: uses the EVENTS (not the trace) because the OTDR's
    multi-acquisition events extend further than the single-acquisition
    saved trace.

    Modifies fiber_result['events'] in place.
    """
    trace = fiber_result.get('full_trace')
    if trace is None:
        return
    pts = fiber_result['full_points']
    acq = fiber_result['acq_range']
    if ior is None:
        ior = fiber_result.get('ior', 1.4682)

    events = fiber_result['events']

    # ── Detect if this is an untrimmed file ──
    is_untrimmed = (len(events) >= 2 and
                    events[0]['is_reflective'] and not events[0]['is_end'] and
                    events[0]['time_of_travel'] == 0 and
                    events[1]['is_reflective'] and not events[1]['is_end'] and
                    0 < events[1]['dist_km'] < LAUNCH_FIBER_MAX)

    if not is_untrimmed:
        return  # already trimmed — no trace enhancement needed

    # ── Trace-based launch detection ──
    launch_idx = _detect_launch_from_trace(trace, pts, acq, ior)
    launch_km = _sample_to_km(launch_idx, ior, pts, acq)

    # ── Trace-based noise floor detection (for break detection) ──
    noise_floor_idx = _detect_noise_floor_from_trace(trace, launch_idx, pts, acq, ior)
    noise_floor_km = _sample_to_km(noise_floor_idx, ior, pts, acq)

    # ── Break detection from trace ──
    break_indices = _detect_breaks_from_trace(trace, launch_idx, noise_floor_idx)
    break_kms = [_sample_to_km(bi, ior, pts, acq) for bi in break_indices]

    # ── Compare this fiber's noise floor to the POPULATION noise floor ──
    # Normal fibers all hit noise at roughly the same distance (pulse width
    # limit).  A fiber whose trace goes to noise significantly earlier than
    # the population has a break/broke.
    trace_span = noise_floor_km - launch_km
    ref_noise_floor = pop_noise_floor_km if pop_noise_floor_km else expected_span_km
    is_trace_broke = (ref_noise_floor > 0 and
                      trace_span < ref_noise_floor - END_REGION_KM)

    # If trace indicates broke but break detector didn't find a specific break,
    # inject a break at the noise floor transition
    if is_trace_broke and not break_kms:
        break_kms.append(noise_floor_km)

    # ── End-of-fiber from EVENTS (not trace) ──
    # The events come from multi-acquisition and extend further than the saved trace.
    # Find end event and far-end connector from the event list.
    end_evt_idx = None
    for i, e in enumerate(events):
        if e['is_end']:
            end_evt_idx = i
            break

    # Find far-end connector: last 1F before end, close to end
    far_end_evt_idx = None
    if end_evt_idx is not None:
        end_dist = events[end_evt_idx]['dist_km']
        for i in range(end_evt_idx - 1, 0, -1):
            if events[i]['is_reflective'] and not events[i]['is_end']:
                if (end_dist - events[i]['dist_km']) < LAUNCH_FIBER_MAX:
                    far_end_evt_idx = i
                break

    # Fiber end = far-end connector position (or end event if no connector found)
    if far_end_evt_idx is not None:
        fiber_end_km = events[far_end_evt_idx]['dist_km']
    elif end_evt_idx is not None:
        fiber_end_km = events[end_evt_idx]['dist_km']
    else:
        fiber_end_km = noise_floor_km

    # ── Re-normalize events ──
    launch_travel = int(round(launch_idx * 2 * acq / pts))

    normalized = []
    for i, e in enumerate(events):
        if i == 0 and e['time_of_travel'] == 0:
            continue  # skip OTDR port
        if i == far_end_evt_idx:
            continue  # skip far-end connector
        new_e = dict(e)
        new_e['dist_km'] = round(e['dist_km'] - launch_km, 4)
        new_e['time_of_travel'] = max(0, e['time_of_travel'] - launch_travel)
        # Adjust end event to fiber end (far-end connector position)
        if e['is_end']:
            new_e['dist_km'] = round(fiber_end_km - launch_km, 4)
        normalized.append(new_e)

    # ── Inject synthetic break events from trace ──
    for bk_km in break_kms:
        bk_norm = round(bk_km - launch_km, 4)
        if bk_norm < 1.0:
            continue
        # Don't inject if there's already an end event before this position
        existing_end = [ne for ne in normalized if ne['is_end'] and ne['dist_km'] < bk_norm]
        if existing_end:
            continue

        # Add a break event (1F reflective with weak Fresnel)
        normalized.append({
            'number': 999,
            'time_of_travel': int(round((bk_km * 1000.0 * ior / 0.02998) * 2)),
            'dist_km': bk_norm,
            'splice_loss': 0.0,
            'reflection': -35.0,
            'slope': 0.0,
            'type': '1F9999LS',
            'is_reflective': True,
            'is_end': False,
        })
        # Remove any end events that are AFTER this break
        normalized = [ne for ne in normalized if not (ne['is_end'] and ne['dist_km'] > bk_norm)]
        # Add end event just after the break
        normalized.append({
            'number': 1000,
            'time_of_travel': int(round(((bk_km + 0.1) * 1000.0 * ior / 0.02998) * 2)),
            'dist_km': round(bk_norm + 0.1, 4),
            'splice_loss': 0.0,
            'reflection': 0.0,
            'slope': 0.0,
            'type': '0E9999LS',
            'is_reflective': False,
            'is_end': True,
        })

    # Sort by distance
    normalized.sort(key=lambda e: (e['dist_km'], 0 if not e['is_end'] else 1))

    fiber_result['events'] = normalized
    fiber_result['_trace_launch_km'] = launch_km
    fiber_result['_trace_end_km'] = fiber_end_km
    fiber_result['_trace_noise_floor_km'] = noise_floor_km
    fiber_result['_trace_breaks'] = break_kms
    fiber_result['_trace_is_broke'] = is_trace_broke


# ═══════════════════════════════════════════════════════════════════════
#  STEP 1 — Load all fibers
# ═══════════════════════════════════════════════════════════════════════

_FILENAME_ILLEGAL_CHARS = re.compile(r'[\\/:*?"<>|\x00-\x1f]+')


def _safe_filename_part(s: str, fallback: str = "site") -> str:
    """Replace characters Windows / browser download dialogs reject in
    filenames with underscores.  Used when building xlsx filenames from
    free-text GenParams strings (orig_loc / term_loc) — a tech in a
    "Seattle/Stevens Pass" route would otherwise hit a broken
    download_button save dialog or a 0-byte saved file because the slash
    is interpreted as a path separator.

    Collapses runs of whitespace + illegal chars to a single underscore,
    strips leading/trailing underscores, and falls back to a placeholder
    when the result is empty.
    """
    if not s:
        return fallback
    cleaned = _FILENAME_ILLEGAL_CHARS.sub('_', str(s))
    cleaned = re.sub(r'\s+', '_', cleaned).strip('_')
    return cleaned or fallback


def _extract_fiber_num(fn):
    """Extract fiber number from a SOR/JSON/TRC filename.

    Handles every naming pattern that has shown up on the user's disk
    after a survey of ~38k real files across 11 cable codes (see
    Project Memory note from 2026-06-13):

      ``LAGDUR0001.sor``                  -> 1     (run after a prefix)
      ``Norsea001_1550.sor``              -> 1     (strip _<wavelength>)
      ``Seattle to Spokane d.0431.sor``   -> 431   (rightmost digit run)
      ``20260520_LAGDUR0001.sor``         -> 1     (date prefix ignored)
      ``DURSAN001_1550 .json``            -> 1     (EXFO trailing-space)
      ``VERSLK001_131015501625 .json``    -> 1     (multi-λ suffix)
      ``TEST0001_155016251310.trc``       -> 1     (multi-λ TRC)
      ``CHC-HCH-LS-089.trc``              -> 89    (dashed long-shot)
      ``._STRROM0001_1550.sor``           -> None  (macOS AppleDouble)

    Rule:
      1. Strip the extension AND any leading "._" (AppleDouble metadata
         that lands next to real files in zips extracted on Mac).
      2. Right-strip whitespace (EXFO FastReporter exports JSON with a
         trailing space between the wavelength code and ``.json``).
      3. Strip one OR MORE concatenated trailing wavelength codes,
         e.g. ``_131015501625`` is three wavelengths jammed together.
      4. Take the RIGHTMOST run of digits — fiber numbers are
         conventionally last after any cable / span / date prefix.
      5. Tie-panel filenames butt a 1-digit ILA/panel suffix straight
         against a 4-digit zero-padded port with NO delimiter, e.g.
         ``PTL1PTL60145`` (ILA1→ILA6, port 0145) or ``DNW1DNW50148``
         (port 0148).  The rightmost run then reads as one number
         (60145) instead of the real port (145) and every fiber lands
         far past any real cable → the stray-fiber guard drops them all
         and aborts.  When the run ends in a 4-char zero-padded field
         (``0NNN``) with extra digits jammed in front, trust the padded
         port.  A genuine 4-digit fiber (1050, 1152) has no leading
         zero, so it is left untouched.

    Returns None when no fiber number can be extracted (which causes
    ``_load_dir`` to skip the file rather than silently overwrite a
    valid fiber).
    """
    # AppleDouble sidecars created by macOS zip-extractors mirror the
    # real filename with a "._" prefix; they're not OTDR data.  Skip
    # them BEFORE the digit walk so they never collide with the real
    # fiber that follows.
    if os.path.basename(fn).startswith("._"):
        return None
    stem, _ = os.path.splitext(fn)
    stem = stem.rstrip()
    # Strip ONE OR MORE concatenated trailing wavelength codes.  The
    # multi-λ EXFO exports we see in the field write all three
    # wavelengths jammed together: ``_131015501625``.  Without the +
    # quantifier the whole concatenation reads as one giant fiber
    # number (~131_500_000_000) and every fiber collides.
    stem = re.sub(
        r'[\s_\-.](?:850|1300|1310|1383|1490|1550|1577|1625|1650)+$', '', stem)
    matches = re.findall(r'\d+', stem)
    if not matches:
        return None
    run = matches[-1]
    # Tie-panel filenames jam a 1-digit ILA/panel suffix onto the 4-digit
    # zero-padded port (``PTL1PTL60145`` → run ``60145``).  If the run ends
    # in a zero-padded 4-char field with digits in front of it, the padded
    # field is the real port; the prefix is the ILA suffix.  (A real 4-digit
    # fiber like 1050 has no leading zero, so ``0\d{3}$`` won't match it.)
    m = re.search(r'0\d{3}$', run)
    if m and len(m.group()) < len(run):
        return int(m.group())
    return int(run)


def _dir_has_json(d):
    """True if directory contains any .json files."""
    if not d or not os.path.isdir(d):
        return False
    for fn in os.listdir(d):
        if fn.lower().endswith('.json'):
            return True
    return False


def load_all(dir_a, dir_b):
    """Load fibers from A and B directories.  Each directory can contain
    either SOR files or EXFO JSON exports — auto-detected per directory.
    When JSON is available it is preferred (it carries the same trace
    samples as SOR plus per-event LSA markers, per-section attenuation,
    and a cleaner event list for grey-value measurement)."""
    fibers_a, fibers_b = {}, {}

    def _load_dir(d, out):
        if not d or not os.path.isdir(d):
            return
        # Pick the file type by whichever has MORE valid fiber-numbered files —
        # NOT "any .json present".  A stray .json in a direction folder (an EXFO
        # export, a report, or Secret Sauce's pairs_cache.json) used to flip this
        # to JSON-only mode and skip every .sor, so the direction loaded 0 fibers
        # and the tech saw "Loaded A=0 B=N — both directions required" with the
        # .sor sitting right there.  Prefer JSON on a tie (it is the richer export).
        try:
            names = os.listdir(d)
        except OSError:
            return
        _n_json = sum(1 for f in names if not f.startswith('._')
                      and f.lower().endswith('.json') and _extract_fiber_num(f))
        _n_sor = sum(1 for f in names if not f.startswith('._')
                     and f.lower().endswith('.sor') and _extract_fiber_num(f))
        use_json = _n_json > 0 and _n_json >= _n_sor
        ext = '.json' if use_json else '.sor'
        parser = parse_otdr_json if use_json else (lambda p: parse_sor_full(p, trim=False))
        # Tally so we can WARN if the filename pattern is ambiguous
        # enough that two real files map to the same fiber number — a
        # silent overwrite used to be how multi-cable ribbon-pair zips
        # (e.g. FTHNTXAD01_AD04_001.sor + FTHNTXAD01_AD05_001.sor) lost
        # data without any error surfaced to the tech.
        collision_count = 0
        for fn in sorted(names):
            if not fn.lower().endswith(ext):
                continue
            if os.path.basename(fn).startswith("._"):
                # macOS AppleDouble metadata sidecar; not OTDR data.
                continue
            try:
                r = parser(os.path.join(d, fn))
            except Exception as exc:
                print(f"  WARN: failed to parse {fn}: {exc}")
                continue
            if not r:
                continue
            fnum = _extract_fiber_num(fn)
            if not fnum:
                print(f"  WARN: could not extract fiber number from "
                      f"'{fn}' — skipped (engine needs a numeric "
                      f"fiber id derived from the filename).")
                continue
            if fnum in out:
                collision_count += 1
                if collision_count <= 5:
                    print(f"  WARN: fiber #{fnum} already loaded from "
                          f"'{out[fnum].get('filename', '?')}'; "
                          f"'{fn}' would overwrite it — keeping the "
                          f"first.")
                continue
            r['_source'] = 'json' if use_json else 'sor'
            out[fnum] = r
        if collision_count > 5:
            print(f"  WARN: {collision_count - 5} more fiber-number "
                  f"collisions in this directory (suppressed).  If you "
                  f"intended to load multiple cables, run each cable "
                  f"as its own A-direction.")

    _load_dir(dir_a, fibers_a)
    _load_dir(dir_b, fibers_b)
    return fibers_a, fibers_b


# ═══════════════════════════════════════════════════════════════════════
#  Helper: measure grey-value splice loss from a direction's JSON trace
# ═══════════════════════════════════════════════════════════════════════

def _grey_loss(fiber_data, splice_km):
    """Return the wide-LSA splice loss at `splice_km` from this fiber's
    raw trace.  Dispatches on data source:
      • JSON  → measure_grey_loss_from_json  (uses pre-stored trace +
                resolution metadata).
      • SOR   → measure_grey_loss_from_sor   (computes resolution from
                the EXFO sampling-period in the proprietary calibration
                block and detects pre-launch offset from the first-500-
                sample minimum).

    Both implementations return loss in dB (positive = real loss) and
    have been validated to within ~5% of the event-table splice_loss
    on bidirectional reburns (SANDUR May-1 cross-check).

    Uses wide-LSA windows (±5 km outer, ±60 m inner) matching EXFO's
    approach.  Returns None when the trace isn't available or the LSA
    can't fit (saturation, near-end clipping, etc.)."""
    if fiber_data is None:
        return None
    src = fiber_data.get('_source')
    if src == 'json':
        return measure_grey_loss_from_json(
            fiber_data, splice_km,
            outer_m=GREY_LSA_OUTER_M,
            inner_m=GREY_LSA_INNER_M,
        )
    if src == 'sor':
        # SILENT-SIDE reconstruction.  _grey_loss is only called to measure the
        # OTHER direction at a matched event THIS direction didn't detect (no
        # stored markers) — for the bidirectional average.  Use the EXFO-learned
        # adaptive windower, which reproduces EXFO's silent-side value to ~0.003
        # dB (F111 0.0136 vs true 0.014).  The old fixed-window function read
        # these flat/garbage, which tipped genuine bidir splices into the
        # single-direction path (e.g. SanDur F205: EXFO .163, but flat b_grey
        # → single-dir .311).  Bend Test-2 keeps the fixed-window function via
        # _narrow_lsa_loss; only the silent-side bidir average changes here.
        return measure_silent_grey_from_sor(fiber_data, splice_km,
                                            require_clean=True)
    return None


# ═══════════════════════════════════════════════════════════════════════
#  STEP 2 — Discover splice closure positions from the A-direction population
# ═══════════════════════════════════════════════════════════════════════

def discover_splices(fibers_a):
    """Bin every fiber's mid-span splice events into 1 km buckets and
    keep buckets that have >= MIN_POP_SPLICE entries.

    Filters applied per-fiber before binning:
      • Skip events whose distance is below 1 km (launch zone).
      • Skip end-of-fiber events directly (1E / 0E types).
      • Skip events that occur AFTER the fiber's first end-of-fiber
        marker.  EXFO event detectors sometimes emit spurious "0F"
        non-end events in the post-EOL trace tail (instrument noise
        after the receive connector) — every fiber tends to have one
        at the same km, which cluster into a phantom closure right
        at the cable boundary.  This guard drops them.
    """
    # ── Collect interior splice events across the whole cable ──
    # (km, fiber) pairs, same per-fiber filters as before.
    pairs = []
    for fnum, r in fibers_a.items():
        # First end-of-fiber marker; events past this aren't real splices.
        eof_km = None
        for e in r['events']:
            if e.get('is_end'):
                eof_km = e['dist_km']
                break
        for e in r['events']:
            if e['dist_km'] < 1.0 or e['is_end']: continue
            if eof_km is not None and e['dist_km'] >= eof_km: continue
            if not e['type'].startswith('0F') and not e['type'].startswith('1F'): continue
            pairs.append((e['dist_km'], fnum))
    if not pairs:
        return []
    pairs.sort(key=lambda p: p[0])

    # Each fiber's "reach" is the position of its last non-end event
    # (i.e. how far its OTDR trace got before terminating).  The population
    # gate tests against the fibers that physically REACH a closure — not
    # the cable-wide count — so a closure past a damage zone (where most
    # fibers broke upstream) still gets discovered from the survivors.
    fiber_reach = {}
    for fnum, r in fibers_a.items():
        max_km = 0.0
        for e in r['events']:
            if e.get('is_end'):
                continue
            if e['dist_km'] > max_km:
                max_km = e['dist_km']
        fiber_reach[fnum] = max_km

    # ── Density (gap) clustering instead of 1 km integer rounding ──
    # A real closure is a tight cluster of events at ~the same km across
    # the cable.  Walk the sorted event kms and start a NEW cluster whenever
    # the gap to the previous event exceeds CLOSURE_CLUSTER_GAP_KM.  This
    # replaces `round(dist_km)`, which split any closure straddling a *.5 km
    # boundary and (via the old "merge within 1 km" step) collapsed distinct
    # closures sharing an integer km — e.g. a real splice at 99.46 had its
    # 99.5+ half rounded into bin 100 and merged with the 100.37 closure
    # into one 460 m-wide bimodal blob, losing 99.46 and contaminating
    # 100.37.  Gap clustering keeps closures <1 km apart distinct.
    clusters = [[pairs[0]]]
    for p in pairs[1:]:
        if p[0] - clusters[-1][-1][0] > CLOSURE_CLUSTER_GAP_KM:
            clusters.append([p])
        else:
            clusters[-1].append(p)

    # Population gate: a cluster must clear BOTH an absolute floor
    # (MIN_POP_SPLICE) and a fractional floor (MIN_POP_FRACTION × fibers-
    # reaching).  Real splice closures show 60-75% coverage; low-population
    # clusters (sparse off-splice bends) fall through here and are picked up
    # downstream by create_off_splice_columns.
    splices = []
    for cl in clusters:
        kms = [p[0] for p in cl]
        avg_pos = round(float(np.mean(kms)), 2)
        n_reaching = sum(1 for km in fiber_reach.values() if km >= avg_pos)
        min_count = max(MIN_POP_SPLICE,
                        int(round(n_reaching * MIN_POP_FRACTION)))
        if len(cl) < min_count:
            continue
        splices.append({
            'bin': int(round(avg_pos)), 'position_km': avg_pos,
            'count': len(cl),
            'reach_count': n_reaching,
        })

    # NB: no post-hoc "merge within 1 km" step — gap clustering already
    # keeps genuinely separate closures apart, and proximity-merging was
    # exactly what collapsed 99.46 into 100.37 before.
    return splices


# ═══════════════════════════════════════════════════════════════════════
#  STEP 2b — Refine closure centers using the MODE of fiber event positions
#            (lets us cleanly distinguish splices from bends)
# ═══════════════════════════════════════════════════════════════════════

def _classify_phantom(sp, fibers_a):
    """Classify a dropped phantom closure as 'bend' or 'damage' based on
    fiber behavior at its position.  Damage zones show many fibers that
    terminate near this km; bend zones have surviving fibers with large
    positive losses and no apparent gainers."""
    pos = sp.get('position_km_refined', sp['position_km'])
    n_broke_near = 0
    for r in fibers_a.values():
        end = [e for e in r['events'] if e.get('is_end')]
        if not end: continue
        eof = end[0]['dist_km']
        # Fiber terminates within ±500 m of this position → counts as damage
        if abs(eof - pos) < 0.5:
            n_broke_near += 1
    sp['broke_near_count'] = n_broke_near
    if n_broke_near >= 10:
        return 'damage'
    # Otherwise it's a bend zone (or near-empty)
    return 'bend'


def _b_confirms_far_closure(sp_pos_a_km, fibers_b):
    """Does the B direction confirm an end-region closure candidate?

    A closure candidate inside the A far-end region (last END_REGION_KM of
    the cable) mirrors to the B direction's NEAR-LAUNCH view — B's cleanest
    region.  A real splice there is unmistakable from B; the phantoms the
    end-region filter exists to kill (post-EOL tail clusters, launch-mirror
    grey artifacts) leave B's near-launch empty.

    Returns (confirmed, n_hits, n_b, b_mirror_km).  Confirmation requires a
    discovery-strength population — the same MIN_POP_SPLICE absolute floor
    and MIN_POP_FRACTION fractional floor discover_splices() uses — of B
    fibers with an interior (non-end, pre-EOF, past-launch) event within
    END_REGION_B_CONFIRM_KM of the mirror position.  The mirror must clear
    END_REGION_B_LAUNCH_GUARD_KM so B's own launch connector can never
    "confirm" the cable-end candidate.
    """
    if not fibers_b:
        return False, 0, 0, 0.0
    # B-frame cable span: same top-25% median-of-EOF idiom as the A side.
    b_eofs = []
    for r in fibers_b.values():
        for e in r.get('events', []):
            if e.get('is_end'):
                b_eofs.append(e['dist_km'])
                break
    if not b_eofs:
        return False, 0, len(fibers_b), 0.0
    b_eofs.sort()
    b_span_est = float(np.median(b_eofs[int(len(b_eofs) * 0.75):]))
    b_mirror = b_span_est - sp_pos_a_km
    if b_mirror < END_REGION_B_LAUNCH_GUARD_KM:
        return False, 0, len(fibers_b), b_mirror
    n_hits = 0
    for r in fibers_b.values():
        eof_km = None
        for e in r.get('events', []):
            if e.get('is_end'):
                eof_km = e['dist_km']
                break
        for e in r.get('events', []):
            if e.get('is_end'):
                continue
            d = e.get('dist_km')
            if d is None or d < 1.0:          # discovery's launch-zone skip
                continue
            if eof_km is not None and d > eof_km:   # post-EOL tail guard
                continue
            if abs(d - b_mirror) <= END_REGION_B_CONFIRM_KM:
                n_hits += 1
                break                          # one event per fiber
    n_b = len(fibers_b)
    need = max(MIN_POP_SPLICE, int(MIN_POP_FRACTION * n_b))
    return n_hits >= need, n_hits, n_b, b_mirror


def refine_closure_centers(fibers_a, splices, validate=True,
                           valid_std_max_m=None, valid_tight_frac=None,
                           valid_min_gainer_frac=None,
                           valid_median_loss_max=None,
                           return_phantoms=False, fibers_b=None):
    """Refine each splice center to the MODE of fiber events in a ±1 km
    window; optionally VALIDATE the cluster and drop phantom closures.

    Adds fields to each splice dict:
        position_km_refined : mode-based closure center (km)
        position_spread_m   : max − min of fiber event positions in window (m)
        position_std_m      : stddev of those positions (m)
        tight_frac          : fraction of fibers with an event within
                              CLOSURE_MATCH_KM of the refined center
        is_real_closure     : True iff tight enough to be a real splice

    When validate=True, clusters that fail the tightness tests are
    REMOVED from the returned list (phantom closures
    at bend/damage zones no longer get splice columns).
    """
    std_max   = CLOSURE_VALID_STD_MAX_M      if valid_std_max_m      is None else valid_std_max_m
    tight_fr  = CLOSURE_VALID_TIGHT_FRAC     if valid_tight_frac     is None else valid_tight_frac
    min_gnr   = CLOSURE_VALID_MIN_GAINER_FRAC if valid_min_gainer_frac is None else valid_min_gainer_frac
    med_max   = CLOSURE_VALID_MEDIAN_LOSS_MAX if valid_median_loss_max is None else valid_median_loss_max
    n_fibers_total = len(fibers_a) or 1

    # Cable span (top-25% median of A-side end-of-fiber distances) — used
    # to filter out phantom closures that sit within END_REGION_KM of the
    # cable end.  A "splice" at km > span − END_REGION_KM mirrors to a
    # B-frame position inside the B-launch dead zone; grey-LSA at that
    # mirrored position reads bogus values from the launch reflection
    # and inflates the bidir-grey-B reburn count.
    eof_kms = []
    for r in fibers_a.values():
        for e in r.get('events', []):
            if e.get('is_end'):
                eof_kms.append(e['dist_km'])
                break
    eof_kms.sort()
    cable_span_est = (np.median(eof_kms[int(len(eof_kms)*0.75):])
                      if eof_kms else 0)
    end_cutoff_km = (cable_span_est - END_REGION_KM
                     if cable_span_est > 0 else float('inf'))

    out = []
    dropped = []
    for sp in splices:
        # Filter near-end phantom closures
        sp_pos = sp.get('position_km_refined', sp['position_km'])
        if sp_pos > end_cutoff_km:
            # Before dropping, give the B direction a veto: a candidate near
            # A's far end sits near B's LAUNCH, where a real splice is
            # unmistakable (the HOWLAN direction-swap bug: Splice 1 at 1.8 km
            # from Howe was silently deleted whenever Lancaster was loaded as
            # A).  A discovery-strength B population at the mirror position
            # → real splice, keep it and let the normal cluster validation
            # below judge it like any other closure.  No B data / no B
            # population → phantom, dropped exactly as before.
            confirmed, n_hits, n_b, b_mirror = _b_confirms_far_closure(
                sp_pos, fibers_b)
            if confirmed:
                print(f"  Kept end-region closure at {sp_pos:.2f} km — "
                      f"B direction confirms it near its launch "
                      f"({n_hits}/{n_b} B fibers @ {b_mirror:.2f} km B-frame)")
            else:
                # End-region phantoms are dropped entirely — they don't get
                # added to the bend/damage phantom-zone list (which becomes
                # columns in the report).  Print a short note so the operator
                # can see what happened, but otherwise discard.
                print(f"  Dropped end-region phantom closure at {sp_pos:.2f} km "
                      f"(within {END_REGION_KM:.0f} km of {cable_span_est:.2f} km cable end)")
                continue
        center_guess = sp['position_km']
        # Neighbor-aware gather window: don't let the mode/refinement pool
        # reach into an adjacent closure.  Half the distance to the nearest
        # OTHER candidate (floored at 0.3 km, capped at the legacy 1.0 km),
        # so closures <1 km apart (e.g. 99.46 vs 100.37) stay distinct
        # instead of the denser neighbor capturing this one's refined center.
        others = [o['position_km'] for o in splices if o is not sp]
        nearest_other = min((abs(center_guess - o) for o in others), default=2.0)
        gather_win = min(1.0, max(0.3, nearest_other / 2.0))
        nearby = []          # positions (km) — global pool
        nearby_losses = []   # losses (signed, dB) paired with nearby[]
        per_ribbon_pos = {}  # ribbon_idx → list of positions (km) for that ribbon
        for fnum, r in fibers_a.items():
            rib_idx = (fnum - 1) // RIBBON_SIZE
            for e in r['events']:
                if e['dist_km'] < 1.0 or e['is_end']:
                    continue
                if abs(e['dist_km'] - center_guess) < gather_win:
                    nearby.append(e['dist_km'])
                    nearby_losses.append(e.get('splice_loss') or 0.0)
                    per_ribbon_pos.setdefault(rib_idx, []).append(e['dist_km'])

        if not nearby:
            sp['position_km_refined'] = center_guess
            sp['position_spread_m'] = 0.0
            sp['position_std_m'] = 0.0
            sp['tight_frac'] = 0.0
            sp['is_real_closure'] = False
            if not validate:
                out.append(sp)
            else:
                dropped.append(sp)
            continue

        arr = np.array(nearby)
        # Mode peak
        bin_km = CLOSURE_MODE_BIN_M / 1000.0
        nbins = max(5, int(np.ceil((arr.max() - arr.min()) / bin_km)))
        hist, edges = np.histogram(arr, bins=nbins, range=(arr.min(), arr.max()))
        peak_idx = int(np.argmax(hist))
        peak_center = (edges[peak_idx] + edges[peak_idx + 1]) / 2.0

        local_mask = np.abs(arr - peak_center) < (CLOSURE_MODE_WINDOW_M / 1000.0)
        if local_mask.sum() >= 5:
            refined = float(np.median(arr[local_mask]))
        else:
            refined = float(peak_center)
        sp['position_km_refined'] = refined
        sp['position_spread_m']   = float(arr.max() - arr.min()) * 1000
        sp['position_std_m']      = float(np.std(arr)) * 1000

        # ── Steven-style display position ───────────────────────────────
        # Steven's column-header convention: walk fibers in numerical
        # order, take the FIRST (lowest-numbered) fiber that has an event
        # within ±CLOSURE_MATCH_KM of the refined closure center, and use
        # that fiber's exact event distance — truncated (not rounded) to
        # 0.01 km (10 m) precision.  This is what appears across the top
        # of his hand-prepared report.
        display_km = None
        display_src_fiber = None
        for fnum in sorted(fibers_a.keys()):
            r = fibers_a[fnum]
            for e in r['events']:
                if e.get('is_end') or e['dist_km'] < 1.0:
                    continue
                if abs(e['dist_km'] - refined) < CLOSURE_MATCH_KM:
                    display_km = e['dist_km']
                    display_src_fiber = fnum
                    break
            if display_km is not None:
                break
        if display_km is None:
            display_km = refined  # fallback (shouldn't happen for real closures)
        # Truncate to 10 m (0.01 km) precision — floor toward zero
        sp['position_km_display']        = math.floor(display_km * 100) / 100.0
        sp['position_km_display_raw']    = display_km
        sp['position_km_display_fiber']  = display_src_fiber

        # ── Per-ribbon refinement ────────────────────────────────────
        # Multi-tube cables have systematic per-tube path-length
        # offsets (different helical pitches around the central
        # strength member).  Refine each ribbon's expected closure
        # position INDEPENDENTLY of the global pool, so that an event
        # in a ribbon whose tube runs ~50 m longer than the cable
        # average isn't mis-flagged as a bend just because it falls
        # outside CLOSURE_MATCH_KM of the global median.
        #
        # Algorithm per ribbon:
        #   1. Take all events in this ribbon's ±1 km window around the
        #      original splice position.
        #   2. Find the densest 50 m bin (the ribbon's local mode peak).
        #   3. Median the events within ±CLOSURE_MATCH_KM of that peak —
        #      that's the ribbon's true expected closure km.
        #   4. Need at least 3 events to trust the per-ribbon value;
        #      ribbons with fewer events fall back to the global refined
        #      position so they still get *some* center to test against.
        per_ribbon_refined = {}
        bin_km_rib = CLOSURE_MODE_BIN_M / 1000.0
        for rib_idx, positions in per_ribbon_pos.items():
            if len(positions) < 3:
                continue
            arr_r = np.array(positions)
            # Local mode peak — independent of the global refined center
            nbins_r = max(3, int(np.ceil((arr_r.max() - arr_r.min()) / bin_km_rib)))
            if arr_r.max() == arr_r.min():
                rib_peak = float(arr_r[0])
            else:
                hist_r, edges_r = np.histogram(
                    arr_r, bins=nbins_r,
                    range=(arr_r.min(), arr_r.max()))
                pk_r = int(np.argmax(hist_r))
                rib_peak = (edges_r[pk_r] + edges_r[pk_r + 1]) / 2.0
            # Median around the ribbon's own peak (NOT the global refined km)
            mask_r = np.abs(arr_r - rib_peak) < CLOSURE_MATCH_KM
            if mask_r.sum() >= 3:
                per_ribbon_refined[rib_idx] = float(np.median(arr_r[mask_r]))
            else:
                per_ribbon_refined[rib_idx] = float(np.median(arr_r))
        sp['position_km_refined_by_ribbon'] = per_ribbon_refined

        # Tightness: fraction of REACHING fibers whose event is within
        # ±CLOSURE_MATCH_KM.  We use the count of fibers that physically
        # reach this km — not the total cable population — so a closure
        # deep into a heavily damaged span isn't penalized for the fibers
        # that broke before they got here.  See discover_splices for the
        # equivalent change to the population gate.
        tight_mask = np.abs(arr - refined) < CLOSURE_MATCH_KM
        tight_count = int(tight_mask.sum())
        # De-helixed tight count: a fiber counts as "tight" when its event is
        # within the window of ITS RIBBON's refined center, not the single GLOBAL
        # center.  Per-ribbon (per-tube) centers track the helical drift, so a
        # real closure whose fibers spread hundreds of metres around the global
        # center still has most fibers tight around their own ribbon centers.
        # Without this, helix spread silently pushes tight_frac under the 0.60
        # keep-override and the whole closure (and all its splices) is dropped as
        # a phantom bend/damage column (HOWLAN 101 km: 464 splices at tight_frac
        # 0.59).  Take the MAX so this can only KEEP a closure, never drop one —
        # no regression on low-helix spans (Seattle drops 0 closures either way).
        tight_count_dehelix = 0
        for _rib_idx, _positions in per_ribbon_pos.items():
            _ctr = per_ribbon_refined.get(_rib_idx, refined)
            tight_count_dehelix += sum(1 for _p in _positions
                                       if abs(_p - _ctr) < CLOSURE_MATCH_KM)
        tight_count = max(tight_count, tight_count_dehelix)
        n_reaching_here = sum(
            1 for r in fibers_a.values()
            if any((not e.get('is_end')) and e['dist_km'] >= sp_pos
                    for e in r.get('events', []))
        ) or n_fibers_total  # fallback if computation yields zero
        sp['tight_frac'] = tight_count / float(n_reaching_here)
        sp['reach_count'] = n_reaching_here
        # Use std within the tight zone for the quality check
        tight_std_m = float(np.std(arr[tight_mask])) * 1000 if tight_count > 3 else 999.0
        sp['tight_std_m'] = tight_std_m

        # Loss-distribution rules (anti-phantom-closure):
        # collect losses of events INSIDE the tight window; compute gainer
        # fraction and median loss
        loss_arr = np.array(nearby_losses)
        tight_losses = loss_arr[tight_mask]
        if len(tight_losses) >= 5:
            sp['gainer_frac']      = float((tight_losses < 0).sum() / len(tight_losses))
            sp['median_loss_db']   = float(np.median(tight_losses))
            sp['mean_loss_db']     = float(np.mean(tight_losses))
            sp['abs_median_loss']  = float(np.median(np.abs(tight_losses)))
        else:
            sp['gainer_frac']     = 0.0
            sp['median_loss_db']  = 0.0
            sp['mean_loss_db']    = 0.0
            sp['abs_median_loss'] = 0.0

        # Verdict: a closure is REAL unless it fails one of these tests.
        # The tight-std and tight-frac tests are independent fatal signals.
        # The loss-distribution tests (no-gainers + high-median) are COMBINED
        # into a single test — both must fail together for a phantom flag,
        # because a real splice between matched fiber lots could legitimately
        # show zero gainers, and by itself a slightly-elevated median is not
        # enough to drop a closure.  Requiring BOTH to fail at the same time
        # catches bend/damage zones (which always fail both) without risking
        # a real closure where the fibers happen to produce uniform positive
        # losses.
        # Validation now relies solely on the loss-distribution physics gate
        # (zero gainers AND elevated median = bend/damage signature).  The
        # older tight_std_m and tight_frac geometry gates were removed per
        # tech direction — they were dropping legitimate closures on long
        # cables where event scatter or low-participation clusters are
        # normal.  tight_std_m / tight_frac are still computed for diagnostic
        # display only.
        fails = []
        # Apply the loss-distribution test whenever the cluster is at least
        # as big as the MIN_POP_SPLICE candidate threshold (20 fibers).  The
        # older 50-fiber guard was too strict now that this is the only gate
        # — damage zones on long cables often have < 50 fibers in the tight
        # cluster because most of the population is already broken upstream.
        if len(tight_losses) >= MIN_POP_SPLICE:
            no_gainers_fail = sp['gainer_frac'] < min_gnr
            high_median_fail = sp['median_loss_db'] > med_max
            # Tight-frac override: a cluster with >= 60% of ALL fibers
            # represented inside the 75 m tight window is unambiguously a
            # splice closure regardless of loss distribution — real bend
            # zones never reach that participation, but a uniformly-bad
            # closure (every fiber lost ~0.11 dB, no gainers) can still
            # fail the loss-distribution test and otherwise be dropped
            # (e.g. Lagrande↔Durkey 17.47 km, 67.37 km).
            tight_frac_override = sp['tight_frac'] >= 0.60
            if no_gainers_fail and high_median_fail and not tight_frac_override:
                fails.append(
                    f'loss_distribution(gainers={sp["gainer_frac"]:.2f} + '
                    f'median={sp["median_loss_db"]:+.3f}dB)'
                )
        sp['validation_fails'] = fails
        sp['is_real_closure'] = not fails

        if not validate or sp['is_real_closure']:
            out.append(sp)
        else:
            dropped.append(sp)

    if validate and dropped:
        print(f"  Dropped {len(dropped)} phantom closure(s) (bend/damage zones):")
        for sp in dropped:
            sp['phantom_type'] = _classify_phantom(sp, fibers_a)
            sp['column_kind'] = sp['phantom_type']    # 'bend' or 'damage'
            fail_str = ' + '.join(sp.get('validation_fails', [])) or 'no_data'
            print(f"    {sp['position_km']:8.2f} km  "
                  f"[{sp['phantom_type']}]  "
                  f"(tight_frac {sp['tight_frac']:.2f}, "
                  f"gainer_frac {sp['gainer_frac']:.2f}, "
                  f"median_loss {sp['median_loss_db']:+.3f} dB, "
                  f"broke_near {sp.get('broke_near_count',0)})  "
                  f"→ FAIL: {fail_str}")

    # Every kept closure is tagged as 'splice' for downstream column rendering
    for sp in out:
        sp['column_kind'] = 'splice'

    if return_phantoms:
        return out, dropped
    return out


def _closure_km_for_fiber(sp, fiber_num, ribbon_size=RIBBON_SIZE):
    """Return the closure km this fiber's bend-vs-splice geometry test
    should use.  If the splice has a per-ribbon refined position for the
    fiber's ribbon, prefer it; otherwise fall back to the global refined
    position, then to the original position_km."""
    rib_idx = (fiber_num - 1) // ribbon_size
    by_rib = sp.get('position_km_refined_by_ribbon') or {}
    if rib_idx in by_rib:
        return by_rib[rib_idx]
    return sp.get('position_km_refined', sp['position_km'])


def _bend_reference_km(sp, fnum, fiber_events,
                       ribbon_size=RIBBON_SIZE,
                       search_window_km=0.5):
    """Return the bend-detection reference km for one (splice, fiber)
    pair.  Tries per-fiber nearest-event first, then per-ribbon refined
    median, then global refined position.  This is the value bend
    geometry tests + label offsets should be measured against."""
    closure_center = _closure_km_for_fiber(sp, fnum, ribbon_size)
    per_fiber = _per_fiber_splice_km(fiber_events, closure_center,
                                      search_window_km=search_window_km)
    if per_fiber is not None:
        return per_fiber
    return closure_center


def _per_fiber_splice_km(fiber_events, closure_center_km,
                         search_window_km=0.5,
                         exclude_pos_km=None,
                         exclude_tol_m=1.0):
    """Find this specific fiber's own splice km at a given closure.

    Bend-detection only.  Walks every event in this fiber's event list
    and returns the distance of the event NEAREST to the closure
    center within ±search_window_km (default 500 m).  That event's
    distance is taken as the fiber's true splice position — the
    candidate event being classified is then measured against it,
    not against a ribbon or global aggregate.

    When ``exclude_pos_km`` is supplied, any event within
    ``exclude_tol_m`` metres of that position is skipped.  Use this in
    Pass 2a/2b so the candidate event being tested doesn't anchor the
    reference to itself (which would make every offset zero and the
    bend gate trivially pass for any event).

    Returns the fiber's nearest-event distance in km, or None if no
    event lies within the search window (caller should fall back to
    the ribbon/global center in that case).
    """
    nearest_km = None
    nearest_d = search_window_km
    excl_tol_km = exclude_tol_m / 1000.0
    for e in fiber_events:
        if e.get('is_end') or e['dist_km'] < 1.0:
            continue
        if (exclude_pos_km is not None and
                abs(e['dist_km'] - exclude_pos_km) < excl_tol_km):
            continue
        d = abs(e['dist_km'] - closure_center_km)
        if d < nearest_d:
            nearest_d = d
            nearest_km = e['dist_km']
    return nearest_km


def _perfiber_residual_m(fiber_data, all_closure_kms, candidate_event_km):
    """Test 1 — per-fiber linear length model (leave-one-out).

    For the supplied fiber, find its nearest event (within
    ±BEND_PERFIBER_WIN_KM) at every closure.  Build (cable_km,
    event_km) pairs.  Identify which closure the candidate event
    belongs to (the one whose pair contains the candidate; if the
    candidate isn't paired because it's >500 m from every closure,
    use the nearest closure to the candidate).  Fit a line through
    the *other* closures' pairs and predict event_km at the target
    closure.  Return (residual_m, predicted_km, n_fit_pts).

    Returns (None, None, 0) when there are fewer than
    BEND_PERFIBER_MIN_FIT other closures with a paired event — in
    that case the model can't be fit and the caller should fall back.
    """
    if fiber_data is None or not all_closure_kms:
        return (None, None, 0)
    events = fiber_data.get('events') or []
    real = [e for e in events
            if not e.get('is_end') and e['dist_km'] >= 1.0]
    if not real:
        return (None, None, 0)

    pairs = []   # list of (si, cable_km, event_km)
    target_si = None
    for si, ck in enumerate(all_closure_kms):
        nearest = None
        for e in real:
            d = abs(e['dist_km'] - ck)
            if d <= BEND_PERFIBER_WIN_KM and (nearest is None or d < nearest[1]):
                nearest = (e['dist_km'], d)
        if nearest is not None:
            pairs.append((si, ck, nearest[0]))
            if abs(nearest[0] - candidate_event_km) < 0.001:
                target_si = si

    if target_si is None:
        # Candidate didn't sit within ±500 m of any closure — assign it
        # to the nearest closure for prediction purposes.
        best_si, best_d = None, float('inf')
        for si, ck in enumerate(all_closure_kms):
            d = abs(ck - candidate_event_km)
            if d < best_d:
                best_si, best_d = si, d
        target_si = best_si

    if target_si is None:
        return (None, None, 0)

    fit_pts = [(c, e) for si, c, e in pairs if si != target_si]
    if len(fit_pts) < BEND_PERFIBER_MIN_FIT:
        return (None, None, len(fit_pts))

    xs = np.array([c for c, _ in fit_pts])
    ys = np.array([e for _, e in fit_pts])
    a, b = np.polyfit(xs, ys, 1)
    target_cable_km = all_closure_kms[target_si]
    predicted_km = float(a * target_cable_km + b)
    residual_m = (candidate_event_km - predicted_km) * 1000.0
    return (residual_m, predicted_km, len(fit_pts))


def _local_step_from_event(fiber_data, event,
                           half_m=None, gap_m=None):
    """Tight local two-line LSA at a stored event's position — the re-measure
    gate's instrument.  Returns the local step in dB (positive = loss) or
    None when it can't be measured (no trace, no tot, truncated windows).

    Indexing is by the event's own ``time_of_travel`` against the digitizer
    clock (sample = tot × pts / (2 × acq_range)) — the same mapping the
    marker-LSA path uses — so it is immune to IOR/frame/launch-offset issues
    that plague km-based indexing (the raw axis is ~1% off the event frame on
    some spans).  Fit windows: [pos−half, pos−gap] and [pos+gap, pos+half];
    defaults LOCAL_STEP_HALF_M / LOCAL_STEP_GAP_M sit outside the ~127 m
    pulse smear and far inside the km-scale curvature that inflates EXFO's
    long-window numbers on false-fire events.
    """
    if fiber_data is None or event is None:
        return None
    trace = fiber_data.get('trace')
    dist_km = event.get('dist_km')
    sp = fiber_data.get('exfo_sampling_period')
    if trace is None or not dist_km or not sp:
        return None
    half = LOCAL_STEP_HALF_M if half_m is None else half_m
    gap = LOCAL_STEP_GAP_M if gap_m is None else gap_m
    try:
        tr = np.asarray(trace, float)
        n = len(tr)
        if n < 100:
            return None
        # Event positions may be NORMALIZED (Pass 0); the raw trace stays in
        # the digitizer/port frame.  Recover the raw position by matching
        # time_of_travel against the saved _raw_events.  Normalization shifts
        # every tot by the SAME launch_travel constant, so recover that shift
        # from the end-of-fiber events (present in both lists) and match
        # exactly.  (Do NOT use _trace_offset_km — its rule disagrees with
        # the normalization shift on some spans: Seattle 1.004 vs 0.143 km.)
        raw_km = dist_km
        raw_evs = fiber_data.get('_raw_events')
        tot = event.get('time_of_travel')
        if raw_evs and raw_evs is not fiber_data.get('events') and tot:
            # The shift normalize applied is the LAUNCH event's tot (raw
            # event #2 in the untrimmed pattern normalize detects; 0 when the
            # list was already trimmed and normalize no-op'd).  Don't derive
            # it from the end events — normalize MOVES the end to the far-end
            # connector, so their tot delta isn't launch_travel.
            shift = 0
            if (len(raw_evs) >= 3 and
                    raw_evs[0].get('is_reflective') and not raw_evs[0].get('is_end') and
                    raw_evs[0].get('time_of_travel') == 0 and
                    raw_evs[1].get('is_reflective') and not raw_evs[1].get('is_end') and
                    raw_evs[1].get('dist_km', 99) < LAUNCH_FIBER_MAX):
                shift = raw_evs[1].get('time_of_travel', 0)
            if shift:
                twin = next((x for x in raw_evs
                             if x.get('time_of_travel') == tot + shift), None)
                if twin is not None and twin.get('dist_km'):
                    raw_km = twin['dist_km']
        # Metres/sample from the sampling period at nominal IOR — the raw
        # trace shares the digitizer clock with raw event positions (per the
        # marker-path comment), NO scale anchor.  (The bright spike ~1 km
        # past the EOF event is the RECEIVE fiber's far end, not a scale
        # error — anchoring on it mis-scales the whole axis.)
        m0 = sp * (299792458.0 / 1.468) / 2.0
        samp_per_km = 1000.0 / m0
        def step_at(idx0):
            h = (half / 1000.0) * samp_per_km
            g = (gap / 1000.0) * samp_per_km
            la, lb = int(idx0 - h), int(idx0 - g)
            ra, rb = int(idx0 + g), int(idx0 + h)
            if la < 0 or rb >= n or lb - la < 8 or rb - ra < 8:
                return None
            def fit(a, b):
                x = np.arange(a, b, dtype=float)
                y = tr[a:b]
                m = (y > 0.5) & (y < 63.5)
                if m.sum() < 8:
                    return None
                return np.polyfit(x[m], y[m], 1)
            L, R = fit(la, lb), fit(ra, rb)
            if L is None or R is None:
                return None
            # Raw SOR traces store cumulative loss (rising with distance):
            # a real loss step raises the after-line above the before-line.
            return float(np.polyval(R, idx0) - np.polyval(L, idx0))
        # EXFO marks the event at its ONSET; with a 2500 ns pulse the smeared
        # step's center sits up to ~300 m downstream.  Take the MAX two-line
        # step over [event, event + LOCAL_STEP_SCAN_M] so onset convention
        # and pulse width can't hide a real step.  (Verified: PLACHE0012's
        # real 0.371 dB splice peaks +230 m past its event mark.)
        i_start = int(raw_km * samp_per_km)
        i_end = int((raw_km + LOCAL_STEP_SCAN_M / 1000.0) * samp_per_km)
        stride = max(1, int(20.0 / m0))              # ~20 m steps
        vals = [step_at(i) for i in range(i_start, i_end + 1, stride)]
        vals = [v for v in vals if v is not None]
        if not vals:
            return None
        return max(vals)
    except Exception:
        return None


def _local_step_confirms(fiber_data, event):
    """True when the stored event's loss is CONFIRMED by the local trace
    (or can't be measured — unmeasurable keeps the flag, never hides).

    RATIO test, not an absolute floor: at 2500 ns the pulse smears any real
    step into a ~250-400 m ramp, so the tight-window estimator reads ~45-50%
    of the true loss (control: a real 0.371 splice reads 0.169; a real
    0.135 distributed loss reads 0.063).  A stored value whose tight read is
    below LOCAL_STEP_CONFIRM_RATIO of it has no support in the glass —
    HOWLAN's 13 phantom (A) entries claim 0.26-0.29 dB and read ~0.00
    (visually flat), while every verified-real event reads 40-50%."""
    stored = event.get('splice_loss')
    # Only police claims big enough for the ratio test to separate reliably:
    # the proven phantom class (HOWLAN) stores 0.26-0.29 dB over flat glass;
    # below LOCAL_STEP_GATE_MIN_DB the tight read of a REAL loss (~45% of
    # stored) sinks into trace noise and the test can't distinguish.
    if not stored or stored < LOCAL_STEP_GATE_MIN_DB:
        return True
    step = _local_step_from_event(fiber_data, event)
    if step is None:
        return True
    return step >= LOCAL_STEP_CONFIRM_RATIO * stored


def _narrow_lsa_loss(fiber_data, position_km):
    """Test 2 — LSA at a specific position on this fiber's raw trace.
    Returns the loss in dB (positive = loss) or None when the trace
    isn't available or the LSA can't fit.

    NOTE (May 2 revision): originally used very-narrow windows
    (±100 m outer, ±15 m inner) but those land inside the OTDR's
    pulse-width smear zone (~127 m at 2.5 µs pulse), producing
    unreliable readings on SOR data.  Empirical tests showed the
    wide-LSA windows (±5000 m outer, ±60 m inner) produce both:
      • a clean off-event noise floor of < 0.005 dB on SOR
      • accurate splice-step readings (within 5% of event-table)
    Neighbor clamping in the underlying LSA handles bleed-in from
    nearby events automatically, so a candidate event ~200 m off
    the predicted km doesn't pollute the predicted-km reading.

    Dispatches on data source — both JSON and SOR are now supported
    for Test 2 confirmation."""
    if fiber_data is None:
        return None
    src = fiber_data.get('_source')
    try:
        if src == 'json':
            return measure_grey_loss_from_json(
                fiber_data, position_km,
                outer_m=BEND_NARROW_OUTER_M,
                inner_m=BEND_NARROW_INNER_M,
            )
        if src == 'sor':
            return measure_grey_loss_from_sor(
                fiber_data, position_km,
                outer_m=BEND_NARROW_OUTER_M,
                inner_m=BEND_NARROW_INNER_M,
            )
    except Exception:
        return None
    return None


def _is_bend_event(event_pos_km, splice_center_km, loss,
                   fiber_events=None, a_loss=None, b_loss=None,
                   closure_kms=None, fiber_data=None):
    """Bend classifier (April 28 revision — per-fiber length model
    + narrow-LSA trace inspection).

    Pipeline:
      1. Loss must be POSITIVE (real attenuation) AND ≥ BEND_THRESHOLD.
         A bend physically removes light from the fiber, so a real
         bend always shows positive signed loss.  Negative readings
         are gainer signatures or LSA artifacts — never bends.
      2. Geometric offset gate (legacy filter, kept as a cheap first
         cut): the event must lie more than ``CLOSURE_MATCH_KM`` from
         the fiber's nearest other splice event.  Events sitting
         right where the fiber's primary splice was already detected
         are not bends — they're the splice itself.
      3. **Test 1 — per-fiber linear length model.**  When
         ``closure_kms`` and ``fiber_data`` are both supplied, fit a
         line through this fiber's (cable_km, event_km) pairs at
         every *other* closure and predict where the splice should
         sit at the candidate's closure.
            • |residual| ≤ BEND_RES_SPLICE_M  → SPLICE  (return False)
            • |residual| ≥ BEND_RES_BEND_M    → run Test 2
            • otherwise (ambiguous)             → SPLICE  (conservative)
      4. **Test 2 — narrow-LSA at predicted_km.**  A real bend means
         the splice lives at predicted_km AND a separate event lives
         at the candidate position.  Confirm by reading a narrow LSA
         on the raw A trace at the predicted km; if it shows a real
         loss step (≥ BEND_NARROW_LOSS_DB) the bend is confirmed.
         If the trace isn't available (SOR data), Test 1 alone wins.

    a_loss / b_loss / asymmetry gate parameters are accepted for
    backward compatibility but are currently inert.
    """
    if loss < BEND_THRESHOLD:
        return False  # negative or sub-threshold → not a bend

    # ── Step 2: geometric offset gate (legacy first cut) ──
    reference_km = splice_center_km
    if fiber_events is not None:
        per_fiber_km = _per_fiber_splice_km(fiber_events, splice_center_km)
        if per_fiber_km is not None:
            reference_km = per_fiber_km
    if abs(event_pos_km - reference_km) <= CLOSURE_MATCH_KM:
        return False

    _ = (a_loss, b_loss)  # asymmetry gate currently inert

    # ── Step 3: Test 1 — per-fiber linear length model ──
    if closure_kms is not None and fiber_data is not None:
        residual_m, predicted_km, _n_fit = _perfiber_residual_m(
            fiber_data, closure_kms, event_pos_km)
        if residual_m is not None:
            ares = abs(residual_m)
            if ares <= BEND_RES_SPLICE_M:
                return False        # fits the fiber's pattern → splice
            if ares < BEND_RES_BEND_M:
                return False        # ambiguous → conservative: not a bend
            # Test-1 says BEND.  Step 4 confirms via wide-LSA at the
            # predicted km.  Both JSON and SOR are supported.  We
            # require POSITIVE evidence — when the LSA fails to fit
            # (returns None) or returns a sub-threshold magnitude, we
            # drop the candidate (conservative).  Confirming via
            # absence-of-evidence inflates the bend count when the LSA
            # is inadvertently asked to fit near saturated regions or
            # at positions overlapping a known event's neighbor-clamp
            # zone.
            pred_loss = _narrow_lsa_loss(fiber_data, predicted_km)
            if pred_loss is None or abs(pred_loss) < BEND_NARROW_LOSS_DB:
                return False
            return True
        # Insufficient pairs to fit — fall through to legacy gate

    # ── Fallback: legacy geometric offset gate already passed above ──
    return True


def _trace_continues_past(fiber_events, position_km, total_span_km,
                          min_continuation_km=3.0,
                          min_distance_to_eof_km=3.0):
    """Return True iff this fiber's trace clearly continues past
    `position_km`.  Used to distinguish a real BREAK (fiber ends or
    quits within a few km past this point) from an in-line REFLECTIVE
    EVENT (a connector / mechanical splice / angled-cleave joint that
    produces a Fresnel reflection but light keeps traveling through).

    Conditions for "continues":
      • At least one non-end event exists past `position_km`, AND
      • The fiber's end-of-fiber marker (if present) is at least
        `min_distance_to_eof_km` past `position_km`.

    Falls back to "True" when EOF can't be determined and there's a
    downstream event — better to mis-call a real break as a ref event
    (visible via the strong reflection number on the cell) than to
    label a healthy connector as BREAK and panic the tech.
    """
    if not fiber_events:
        return False
    eof_km = None
    for e in fiber_events:
        if e.get('is_end'):
            ek = e['dist_km']
            if eof_km is None or ek < eof_km:
                eof_km = ek
    has_event_past = any(
        (not e.get('is_end')) and e['dist_km'] > position_km + 0.05
        for e in fiber_events
    )
    if not has_event_past:
        return False
    if eof_km is not None and eof_km - position_km < min_distance_to_eof_km:
        return False
    return True


def _is_field_gainer(event_pos_km, total_span_km, loss):
    """Field-gainer rule: an event is flagged as a field gainer if its
    signed loss falls in [FIELD_GAINER_MIN_DB, FIELD_GAINER_MAX_DB] (i.e.
    -0.7 to 0 dB by default) AND its position is mid-span — at least
    LAUNCH_FIBER_MAX km from the launch (so the launch-loss rule owns
    it) and at least END_REGION_KM km before the fiber end (so the
    end-of-fiber region doesn't pollute the flag)."""
    if not (FIELD_GAINER_MIN_DB <= loss <= FIELD_GAINER_MAX_DB):
        return False
    if event_pos_km < LAUNCH_FIBER_MAX:
        return False
    if total_span_km > 0 and event_pos_km > (total_span_km - END_REGION_KM):
        return False
    return True


def apply_field_gainer_rule(all_results, total_span_km):
    """Post-pass annotator.  Classify an event as a true field-gainer
    only when the BIDIRECTIONAL signature matches: a real gainer
    averages near zero (or slightly negative) because the two
    directions report opposite-signed losses of similar magnitude.

    Rules (April 28 revision — strict):
      • The BIDIRECTIONAL average must lie in
        [FIELD_GAINER_MIN_DB, FIELD_GAINER_MAX_DB] (i.e. [−0.7, 0] dB).
      • BOTH a_loss AND b_loss must be present AND must be REAL
        event measurements (neither side came from wide-LSA grey).
        A gainer's whole definition is bidirectional, so we refuse
        to flag based on a single direction or on a grey-LSA noise
        reading.
      • The two directional losses MUST have OPPOSITE signs — the
        canonical gainer fingerprint (A reads negative because the
        scattering coefficient rises; B reads positive because going
        the other way it falls).  Same-sign A and B is an LSA
        artifact (slope mismatch in surrounding fiber), not a gainer.
      • Position must be mid-span (existing geometry rule).

    Skips results already classified as break / broke / dead_zone —
    those classifications win unconditionally.
    """
    flagged = 0
    for key, r in all_results.items():
        if not isinstance(r, dict):
            continue
        if r.get('is_break') or r.get('is_broke') or r.get('is_dead_zone'):
            continue
        if r.get('is_ref'):
            continue  # ref cells stand on their own — never demote to gainer
        a_loss = r.get('a_loss')
        b_loss = r.get('b_loss')
        bidir  = r.get('bidir_loss')

        # Hard requirement: both directions present AND both real
        # event measurements (not wide-LSA grey).  A gainer is
        # bidirectional by definition; refuse to call it on a single
        # direction or on a grey-LSA noise reading.
        if a_loss is None or b_loss is None or bidir is None:
            continue
        if r.get('_a_is_grey') or r.get('_b_is_grey'):
            continue
        if r.get('_a_source') == 'grey_lsa' or r.get('_b_source') == 'grey_lsa':
            continue

        pos_km = r.get('bidir_dist') or r.get('a_dist') or r.get('position_km')
        if pos_km is None:
            continue

        # Bidirectional range gate (mid-span position check inside)
        if not _is_field_gainer(pos_km, total_span_km, bidir):
            continue

        # Canonical gainer signature: A and B have OPPOSITE signs.
        # Use a 1 milli-dB dead-zone so an exactly-zero side doesn't
        # ambiguously qualify.
        EPS = 0.001
        sign_a = 1 if a_loss > EPS else (-1 if a_loss < -EPS else 0)
        sign_b = 1 if b_loss > EPS else (-1 if b_loss < -EPS else 0)
        if sign_a == sign_b or sign_a == 0 or sign_b == 0:
            continue  # same-sign / near-zero → not a real gainer

        r['is_gainer'] = True
        # Gainer wins over bend in the overlap range — strip the
        # bend tag so the cell renders cleanly with the gainer fill.
        if r.get('is_bend'):
            r['is_bend'] = False
        r['is_flagged'] = True
        flagged += 1
    return flagged


def _bend_severity(loss):
    # Severity tiers collapsed — every bend ≥ BEND_THRESHOLD is simply 'BEND'.
    return 'BEND'


def _is_borderline_loss(bidir_loss, threshold):
    """DISABLED — there is no loss borderline band.  Per the boss's workflow the
    reburn threshold (REBURN_THRESHOLD, set in the OTDR settings panel, typically
    0.16) is a HARD cutoff: a bidirectional loss flags iff it is >= threshold and
    anything below is simply not flagged — there is no near-threshold "review"
    tier.  Kept as a no-op (signature + the manifest n_borderline field unchanged,
    now always 0); the BORDERLINE_*_MARGIN constants are consequently unused."""
    return False


def apply_connector_loss_rule(all_results, threshold=None):
    """Flag any reflective (1F) event whose bidir loss reaches the
    connector-loss threshold (default BIDIR_CONNECTOR_LOSS = 0.500 dB).

    Connectors and mechanical splices normally lose 0.1–0.3 dB.  A
    bidir reading at or above the connector threshold indicates a
    degraded / dirty / damaged connector worth surfacing separately
    from a normal reburn (which fires at REBURN_THRESHOLD = 0.160 dB).

    Adds:
      r['is_high_connector_loss'] = True
      label suffix '⚠ conn' appended (no duplicate if already present)

    Existing colour / classification (reburn / ref / etc.) is preserved;
    this just decorates the cell so the tech can see at a glance which
    flagged events are connector-loss issues vs splice-loss issues.

    Returns the count of events flagged."""
    if threshold is None:
        threshold = BIDIR_CONNECTOR_LOSS
    flagged = 0
    for key, r in list(all_results.items()):
        if not isinstance(r, dict):
            continue
        etype = r.get('event_type') or ''
        if not str(etype).startswith('1F'):
            continue
        # Use bidir loss if available, else fall back to single-direction
        loss = r.get('bidir_loss')
        if loss is None:
            loss = (r.get('a_loss') if r.get('a_loss') is not None
                    else r.get('b_loss'))
        if loss is None:
            continue
        if abs(loss) < threshold:
            continue
        r['is_high_connector_loss'] = True
        # Append suffix to the label (idempotent)
        lbl = r.get('label') or ''
        if '⚠ conn' not in lbl:
            r['label'] = f"{lbl} ⚠ conn".strip()
        r['is_flagged'] = True
        flagged += 1
    return flagged


def split_offsplice_events_into_own_columns(all_results, splices,
                                              splice_dist_km=None,
                                              cluster_gap_km=0.200,
                                              broke_cluster_gap_km=0.400,
                                              total_span_km=None,
                                              fibers_a=None):
    """Reassign bend / break / broke events that sit far from any
    splice into their own phantom columns.

    Bend / break / broke events that are within ``splice_dist_km``
    (default BEND_SPLICE_FOLD_KM = 150 m, the OTDR panel's "Bend fold
    distance" — read at call time so a --overrides setattr lands) of a
    real splice stay attributed to that splice column — they're "at the
    splice."  Events farther away get clustered by km position
    (``cluster_gap_km`` for bend/break, ``broke_cluster_gap_km`` for
    broke — slightly wider because broke positions vary by per-fiber
    detection threshold) and each cluster becomes a NEW pseudo-splice
    entry inserted into the splices list at the right physical position.

    Column kind by cluster contents:
      • Pure bends → ``column_kind = 'bend'`` (yellow header)
      • Any break / broke → ``column_kind = 'damage'`` (red header)

    Returns (updated_all_results, updated_splices).  Existing splice
    indices in ``all_results`` are remapped to the new sort order.
    """
    if splice_dist_km is None:
        splice_dist_km = BEND_SPLICE_FOLD_KM
    if not splices:
        return all_results, splices

    splice_kms = [sp.get('position_km_refined', sp['position_km'])
                  for sp in splices]

    # 1. Find bend / break / broke events that sit > splice_dist_km
    #    from any splice.
    # Compute launch / tailbox exclusion zones — phantom columns here
    # are almost always tailbox connectors with legitimate ~0.2 dB loss,
    # not bends.  Drop them from the off-splice clustering pass too.
    launch_zone_max = LAUNCH_FIBER_MAX
    tailbox_zone_min = (total_span_km - LAUNCH_FIBER_MAX) if total_span_km else None
    candidates = []
    # Consensus end-of-fiber (median EOF across all fibers) — the anchor the
    # account-then-flag gate uses to fold a helix-shifted LAST-closure splice on a
    # short-reading fiber.  Computed once here, passed into the gate.
    consensus_eof = None
    if fibers_a:
        _eofs = [max(float(e['dist_km'])
                     for e in r.get('events', []) if e.get('is_end'))
                 for r in fibers_a.values()
                 if any(e.get('is_end') for e in r.get('events', []))]
        if _eofs:
            consensus_eof = float(np.median(_eofs))
    for key, r in all_results.items():
        if not isinstance(r, dict):
            continue
        # Include is_ref so off-splice reflective events (e.g. F1008
        # at km 32.15, 700 m before Splice 7) get their own column at
        # their actual km position instead of being anchored to the
        # nearest splice.  Same rule for bend / break / broke.
        # Also include is_a_only / is_b_only — single-direction cells
        # that sit > CLOSURE_MATCH_KM from their assigned splice get
        # relocated into the off-splice cluster too (e.g. F859 at
        # km 60.337 with 9.315 dB single-direction loss, 240 m from
        # Splice 8 at km 60.58, should sit in the Damage @ 60.34 km
        # column with F857 — not in Splice 8).
        # is_gainer is included so a near-zero-loss field gainer that
        # sits > CLOSURE_MATCH_KM from any closure gets its own off-splice
        # column at the gainer's true km position — not silently
        # re-anchored to the nearest splice with a wrong-km label.
        # Bug #2 fix 2026-06-13.
        if not (r.get('is_bend') or r.get('is_break') or
                r.get('is_broke') or r.get('is_ref') or
                r.get('is_a_only') or r.get('is_b_only') or
                r.get('is_gainer')):
            continue
        km = r.get('bidir_dist')
        if km is None:
            continue
        # Skip phantom-column creation inside the launch / tailbox zones.
        if km < launch_zone_max:
            continue
        if tailbox_zone_min is not None and km > tailbox_zone_min:
            continue
        nearest = min(abs(km - sk) for sk in splice_kms) if splice_kms else float('inf')
        # Per-event candidacy keeps the legacy tight gate (CLOSURE_MATCH_KM):
        # events truly AT a splice never become candidates.  The panel's
        # "Bend fold distance" (splice_dist_km) is applied per-CLUSTER below,
        # on the cluster median — folding per-event instead left straggler
        # events (a cluster's far tail) marooned in their own 1-2 cell column.
        if nearest <= min(splice_dist_km, CLOSURE_MATCH_KM):
            continue
        # ── Account-then-flag gate ──────────────────────────────────────────
        # Every fiber is spliced at every closure.  Before pulling this event
        # into its own column, ask whether it is simply THIS fiber's own closure
        # splice drifted out by helix (per-fiber length model, drift-scaled
        # tolerance).  If so, leave it attributed to the nearest splice column —
        # do NOT spawn a phantom column.  Only events that are genuinely
        # ADDITIONAL (not explained by the fiber's own drift) become candidates.
        if fibers_a is not None and _event_explained_as_splice(
                r.get('fiber'), km, splice_kms, fibers_a,
                consensus_eof=consensus_eof):
            continue
        candidates.append((key, r, km))

    if not candidates:
        return all_results, splices

    # 2. Cluster candidates by km position.  Bend/break clusters use
    #    the tighter cluster_gap_km; broke clusters use the wider
    #    broke_cluster_gap_km because broke positions on adjacent
    #    fibers can drift by a few hundred metres depending on where
    #    each fiber's A trace terminates.  When a cluster mixes
    #    bend/break and broke, the wider gap is used.
    candidates.sort(key=lambda x: x[2])
    clusters = []          # list of dict {km_center, items: [(key, r)]}
    for key, r, km in candidates:
        gap = broke_cluster_gap_km if r.get('is_broke') else cluster_gap_km
        if clusters:
            # Use the wider gap if this candidate OR the prior cluster
            # contains a broke event.
            prior_has_broke = any(rr.get('is_broke') for _, rr in clusters[-1]['items'])
            effective_gap = broke_cluster_gap_km if (r.get('is_broke') or prior_has_broke) else cluster_gap_km
            if abs(km - clusters[-1]['km_center']) < effective_gap:
                cluster = clusters[-1]
                new_count = len(cluster['items']) + 1
                cluster['km_center'] = (
                    cluster['km_center'] * len(cluster['items']) + km
                ) / new_count
                cluster['items'].append((key, r))
                continue
        clusters.append({'km_center': km, 'items': [(key, r)]})

    # 3. Build new pseudo-splice entries.  Priority of column-kind
    #    when a cluster mixes event types (rare but possible):
    #       damage (break / broke)  >  ref  >  bend
    #    A cluster of only is_ref events gets column_kind='ref'
    #    (deep-orange header, matches the ref cell color).
    new_phantoms = []
    for cluster in clusters:
        # ── Bend fold distance (OTDR-panel setting) ─────────────────────────
        # A cluster whose MEDIAN position sits within splice_dist_km of a
        # validated splice column is that splice's short/long-lay tail — its
        # cells stay attributed to the splice column (bend labels intact), no
        # separate "Bends @" column.  Platteville–Cheyenne: six 2-4 fiber
        # clusters 107-165 m from their splices were phantom columns; fiber
        # 187's real damage (450/460 m off-splice) still gets its own column.
        med_km = float(np.median([r.get('bidir_dist') for _, r in cluster['items']
                                  if r.get('bidir_dist') is not None] or
                                 [cluster['km_center']]))
        if splice_kms and min(abs(med_km - sk) for sk in splice_kms) <= splice_dist_km:
            continue
        kinds = set()
        for _, r in cluster['items']:
            if r.get('is_break') or r.get('is_broke'):
                kinds.add('damage')
            elif r.get('is_ref'):
                kinds.add('ref')
            else:
                kinds.add('bend')
        if 'damage' in kinds:
            column_kind = 'damage'
        elif 'ref' in kinds and 'bend' not in kinds:
            column_kind = 'ref'
        else:
            column_kind = 'bend'
        phantom = {
            'position_km': round(cluster['km_center'], 2),
            'position_km_refined': cluster['km_center'],
            'column_kind': column_kind,
            'phantom_type': column_kind,
            'count': len(cluster['items']),
            'is_real_closure': False,
        }
        new_phantoms.append((phantom, cluster['items']))

    # 4. Remove the affected events from all_results (we'll re-add them
    #    under the new splice_idx after sort/remap).
    affected_items = []   # list of (result_dict, new_phantom)
    for phantom, items in new_phantoms:
        for old_key, r in items:
            affected_items.append((r, phantom))
            del all_results[old_key]

    # 5. Combine + sort the splices list, including the new phantoms.
    combined = list(splices) + [p for p, _ in new_phantoms]
    combined.sort(key=lambda sp: sp.get('position_km_refined',
                                          sp['position_km']))

    # 6. Build a map: identity → new index.
    id_to_idx = {id(sp): i for i, sp in enumerate(combined)}

    # 7. Remap existing splice_idx in all_results to new indices.
    #    Build identity map from old splice objects.
    old_splice_by_idx = {i: sp for i, sp in enumerate(splices)}
    new_idx_by_old_idx = {
        old_i: id_to_idx[id(sp)]
        for old_i, sp in old_splice_by_idx.items()
    }
    remapped = {}
    for old_key, r in all_results.items():
        old_si = r.get('splice_idx')
        if old_si is None or old_si not in new_idx_by_old_idx:
            remapped[old_key] = r
            continue
        new_si = new_idx_by_old_idx[old_si]
        r['splice_idx'] = new_si
        remapped[(r['fiber'], new_si)] = r
    all_results = remapped

    # 8. Re-attach affected bend/break events under their new phantom column.
    for r, phantom in affected_items:
        new_si = id_to_idx[id(phantom)]
        r['splice_idx'] = new_si
        # Strip the "(+XXXXm)" offset annotation from the cell label — the
        # new column header carries the position; the offset is now zero.
        old_label = r.get('label', '')
        if ' (' in old_label and 'm)' in old_label:
            r['label'] = old_label.rsplit(' (', 1)[0]
        r['closure_offset_m'] = 0.0
        # Reset event_source to indicate the dedicated bend / damage column.
        if r.get('is_bend'):
            r['event_source'] = 'bend_column'
        elif r.get('is_break'):
            r['event_source'] = 'break_column'
        elif r.get('is_broke'):
            r['event_source'] = 'broke_column'
        all_results[(r['fiber'], new_si)] = r

    return all_results, combined


def _format_loss(val):
    """'.172' style — drops leading 0. like Steven's report."""
    s = f"{abs(val):.3f}"
    return s[1:] if s.startswith('0.') else s


# ═══════════════════════════════════════════════════════════════════════
#  STEP 2c — Detect launch-end issues (fibers broken / damaged at launch)
#
#  These fibers would otherwise be silent in the report because their
#  event tables are truncated immediately after the launch connector.
# ═══════════════════════════════════════════════════════════════════════

def _fiber_launch_info(r):
    """Extract launch-connector event info from a fiber's events.
    Returns (first_launch_event_dict | None, end_km | None, n_events)."""
    if r is None:
        return None, None, 0
    events = r.get('events') or []
    launch_evt = None
    if events and events[0].get('is_reflective') and events[0]['dist_km'] < 0.5:
        launch_evt = events[0]
    end_events = [e for e in events if e.get('is_end')]
    end_km = end_events[0]['dist_km'] if end_events else None
    return launch_evt, end_km, len(events)


def detect_launch_issues(fibers_a, fibers_b, first_splice_km=None,
                          high_loss_db=None, bad_refl_db=None,
                          spans_have_tailbox=True,
                          **_ignored):
    """Return {fiber_num: launch_issue_dict} for every fiber that has a
    launch-end problem in either direction.

    Optional overrides (used by the Streamlit sidebar):
      high_loss_db        — launch-connector loss >= this flags HIGH_LAUNCH_LOSS
      bad_refl_db         — launch reflectance >= this flags BAD_LAUNCH_REFL
      spans_have_tailbox  — when False, the entire BAD_TAILBOX_REFL block
                            is skipped.  Use for tie-panel / jumper-only
                            spans where the cable terminates without a
                            tailbox connector and every fiber's bare-glass
                            EOL reflection would otherwise flag.
    Any other kwargs are accepted and ignored for forward-compat.

    launch_issue_dict has:
      a_tags : list[str]   — issue tags for A direction (empty if none)
      b_tags : list[str]   — issue tags for B direction
      severity : 'HIGH' | 'REVIEW' | 'WATCH'
      summary : str        — human-readable label for the cell
    """
    hi_loss = LAUNCH_HIGH_LOSS_DB if high_loss_db is None else float(high_loss_db)
    bad_refl = LAUNCH_BAD_REFL_DB if bad_refl_db is None else float(bad_refl_db)
    # Population medians
    def _gather_launch_refls(fibers):
        refls = []
        for r in fibers.values():
            le, _, _ = _fiber_launch_info(r)
            if le is not None and le.get('reflection') is not None and le['reflection'] < 0:
                refls.append(le['reflection'])
        return float(np.median(refls)) if refls else None

    a_refl_median = _gather_launch_refls(fibers_a)
    b_refl_median = _gather_launch_refls(fibers_b)

    # ── Population tailbox-refl baseline per direction ──
    # The "tailbox refl" for a fiber = refl of the last 1F event within
    # 2 km of the EOL; if no such 1F, the 1E refl itself.  Computing the
    # median lets us tell "this fiber's tailbox is uniquely bad" apart
    # from "the whole span was shot with bare-glass cable ends" (e.g.
    # SANDUR's B-direction, where the receive jumper wasn't attached on
    # any fiber).  A BAD_TAILBOX_REFL flag now requires the fiber to be
    # an OUTLIER vs the direction's population — not just over the
    # absolute threshold.
    TAILBOX_OUTLIER_DB = 10.0    # refl must be >= this much worse
                                 # (less negative) than population median
                                 # to count as a defect
    def _fiber_tailbox_refl(r):
        if r is None:
            return None
        evts = r.get('_raw_events') or r.get('events') or []
        end_evt = next((e for e in evts if e.get('is_end')), None)
        if end_evt is None:
            return None
        end_km = end_evt['dist_km']
        for e in reversed(evts):
            if e is end_evt or e.get('is_end'):
                continue
            if e['dist_km'] >= end_km:
                continue
            if (end_km - e['dist_km']) > 2.0:
                break
            if e.get('is_reflective') or str(e.get('type','')).startswith('1F'):
                return e.get('reflection')
        # No 1F tailbox — use the 1E end-event refl
        return end_evt.get('reflection')

    def _pop_tailbox_median(fibers):
        refls = []
        for r in fibers.values():
            v = _fiber_tailbox_refl(r)
            if v is not None and v < 0:
                refls.append(v)
        return float(np.median(refls)) if refls else None

    a_tb_median = _pop_tailbox_median(fibers_a)
    b_tb_median = _pop_tailbox_median(fibers_b)

    # ── FQA: acquisition-duration check per direction ──
    # A span fails FQA if traces weren't all shot for the same duration.
    # We use the "Duration" field as shown in the EXFO viewer's Test
    # Parameters / Summary section, which is the SR-4731 AveragingTime
    # field (uint16 deciseconds at +38 of FxdParams body) — exposed by
    # the SOR parser as r['duration_sec'].  Per-direction mode wins;
    # any fiber whose duration differs from its direction's mode flags
    # DURATION_MISMATCH.
    def _duration_sec(r):
        if r is None:
            return None
        return r.get('duration_sec')

    def _mode_duration(fibers):
        from collections import Counter
        durs = Counter()
        for r in fibers.values():
            d = _duration_sec(r)
            if d is not None:
                durs[d] += 1
        if not durs:
            return None
        return durs.most_common(1)[0][0]

    a_dur_mode = _mode_duration(fibers_a)
    b_dur_mode = _mode_duration(fibers_b)

    all_fibers = sorted(set(fibers_a.keys()) | set(fibers_b.keys()))
    issues = {}

    for fnum in all_fibers:
        ra = fibers_a.get(fnum)
        rb = fibers_b.get(fnum)
        a_tags, b_tags = [], []

        def _check(r, tags, pop_median_refl, dir_is_A):
            """Flag ONLY severe launch-end issues — the kind where the fiber
            silently disappears from the splice report.  We deliberately skip
            soft signals like 'NO_FIRST_SPLICE' (too noisy; many fibers have
            sub-threshold splices that don't get detected)."""
            if r is None:
                tags.append('FILE_MISSING')
                return
            launch_evt, end_km, n_events = _fiber_launch_info(r)

            # No events at all — fiber is completely silent
            if n_events == 0:
                tags.append('NO_EVENTS')
                return

            # Launch-connector reflectance check — signed comparison (strict
            # greater-than).  A healthy buried launch reflects at -50 to -55
            # dB; damaged / dirty / partially-cut connectors reflect closer
            # to 0 (less negative).  Rule: refl > -50 → flag.
            # The launch-loss rule was disabled (hi_loss is None) per tech
            # direction.
            if launch_evt is not None:
                if hi_loss is not None:
                    launch_loss_signed = launch_evt.get('splice_loss') or 0.0
                    if launch_loss_signed > hi_loss:
                        tags.append(f'LAUNCH_LOSS{launch_loss_signed:+.2f}dB')
                refl = launch_evt.get('reflection') or 0.0
                # refl < 0 precondition: refl == 0.0 means "not
                # measured / not reflective," NOT "very bad
                # reflectance."  Only flag when there is an actual
                # negative reflection number to evaluate.
                if refl < 0 and refl >= bad_refl:
                    tags.append(f'BAD_LAUNCH_REFL{refl:+.1f}dB')

            # ── Tailbox reflectance check (mirror of launch rule) ──
            # A healthy cable end has a 1F tailbox connector reflecting
            # better than -49.9 dB.  Two failure modes flag here:
            #   (a) Missing tailbox: the cable ends in a 1E event with
            #       bad reflectance and no preceding 1F connector in the
            #       last 2 km (bare glass to air, e.g. F336 at -15.6 dB).
            #   (b) Bad tailbox: the last 1F event before EOL has
            #       refl >= -49.9 dB (dirty / damaged tailbox connector).
            # A bad reflection on the 1E itself when there IS a good
            # 1F tailbox in front of it just means the receive-pigtail
            # end face is dirty — NOT a cable defect — and is ignored.
            #
            # IMPORTANT: use _raw_events (pre-normalization).  The
            # normalize step strips the 1F tailbox event and carries the
            # receive-pigtail's bare-glass EOL reflectance onto the
            # moved 1E, which would make every fiber look like a missing
            # tailbox.  Raw events preserve the original 1F/1E pair.
            # Find this fiber's tailbox-refl (same logic as the
            # population baseline).  Then apply BOTH conditions:
            #   (a) the refl is at or above the absolute -49.9 dB
            #       threshold (so we never flag a clean tailbox)
            #   (b) AND it is >= TAILBOX_OUTLIER_DB worse than the
            #       population median for this direction (so spans like
            #       SANDUR-B, where every fiber has the same bare-glass
            #       1E refl by virtue of how the shoot was done, don't
            #       light up every fiber)
            # Tie-panel / jumper-only spans set spans_have_tailbox=False
            # to skip this entire block — they don't have tailbox
            # connectors and every bare-glass EOL would otherwise flag.
            if not spans_have_tailbox:
                return
            this_tb_refl = _fiber_tailbox_refl(r)
            pop_median   = a_tb_median if dir_is_A else b_tb_median
            # Same refl < 0 precondition as the launch check:
            # this_tb_refl == 0.0 means the OTDR didn't measure a
            # reflection (event isn't reflective), not "very bad."
            if (this_tb_refl is not None
                    and this_tb_refl < 0
                    and this_tb_refl >= bad_refl
                    and pop_median is not None
                    and (this_tb_refl - pop_median) >= TAILBOX_OUTLIER_DB):
                tags.append(f'BAD_TAILBOX_REFL{this_tb_refl:+.1f}dB')

            # ── FQA: per-trace acquisition-duration check ──
            # Compare this fiber's "Duration" (seconds — the SR-4731
            # AveragingTime as shown under Test Parameters → Summary in
            # EXFO) against the majority duration for this direction.
            # Mismatched traces fail FQA: the span shouldn't have been
            # shot with mixed durations.
            dir_mode = a_dur_mode if dir_is_A else b_dur_mode
            this_dur = _duration_sec(r)
            if dir_mode is not None and this_dur is not None and this_dur != dir_mode:
                tags.append(f'DURATION_MISMATCH({this_dur:.1f}s vs {dir_mode:.1f}s)')

        _check(ra, a_tags, a_refl_median, dir_is_A=True)
        _check(rb, b_tags, b_refl_median, dir_is_A=False)

        if not a_tags and not b_tags:
            continue

        # Severity: HIGH for immediate-end / no-events / high-launch-loss,
        # REVIEW for missing-file / bad-refl, WATCH for only outlier / no-first.
        all_tags = a_tags + b_tags
        is_high = any(t.startswith(('NO_EVENTS',
                                    'HIGH_LAUNCH_LOSS', 'FILE_MISSING'))
                      for t in all_tags)
        is_review = any(t.startswith(('BAD_LAUNCH_REFL', 'BAD_TAILBOX_REFL',
                                       'DURATION_MISMATCH')) for t in all_tags)
        severity = 'HIGH' if is_high else ('REVIEW' if is_review else 'WATCH')

        # Build a compact one-line summary (first 1–2 issue tags)
        primary = a_tags[0] if a_tags else (b_tags[0] if b_tags else '')
        dir_label = 'A' if a_tags else 'B'
        summary = f"{fnum} LAUNCH({dir_label}) {primary}"

        issues[fnum] = {
            'a_tags': a_tags,
            'b_tags': b_tags,
            'severity': severity,
            'summary': summary,
        }

    return issues


# ═══════════════════════════════════════════════════════════════════════
#  STEP 3 — Pass 1: Standard splice report analysis
#           (identical logic to splice_report_generator.py, plus A-only flagging)
# ═══════════════════════════════════════════════════════════════════════

def analyze_all(fibers_a, fibers_b, splices, threshold,
                bend_threshold=None, closure_match_km=None, **_ignored):
    """
    Pass 1: For each fiber at each known splice closure position:
      - Find A event → find matching B event → compute bidir loss → flag if above threshold
      - If no B match: flag A-only if A loss >= threshold (new vs original splice report)
      - Detect broke fibers and B-fill past breaks (same as original)

    event_source field:
      'bidir'  — both A and B direction saw it (standard splice)
      'a_only' — only A direction, no B match
      'broke'  — fiber terminates mid-span
      'bfill'  — B-direction fill past a break

    Optional kwargs (wired from the Streamlit sidebar):
      bend_threshold    — overrides BEND_THRESHOLD for this call
      closure_match_km  — overrides CLOSURE_MATCH_KM for this call
    Additional unknown kwargs are accepted and ignored for forward-compat.
    """
    bend_threshold = BEND_THRESHOLD if bend_threshold is None else float(bend_threshold)
    closure_match_km = CLOSURE_MATCH_KM if closure_match_km is None else float(closure_match_km)
    results = {}

    # Closure km list (for per-fiber length model in _is_bend_event)
    closure_kms_all = [sp.get('position_km_refined', sp['position_km'])
                       for sp in splices]

    # End-of-fiber distances for broke detection
    eof_a = {}
    for fnum, r in fibers_a.items():
        end = [e for e in r['events'] if e['is_end']]
        eof_a[fnum] = end[0]['dist_km'] if end else 999

    # Auto-detect span: top 25% median of all EOL distances
    eof_a_vals = sorted(eof_a.values())
    if eof_a_vals:
        top_quarter_a = eof_a_vals[int(len(eof_a_vals) * 0.75):]
        total_span_a = np.median(top_quarter_a)
    else:
        total_span_a = 0

    eof_b = {}
    for fnum, r in fibers_b.items():
        end = [e for e in r['events'] if e['is_end']]
        eof_b[fnum] = end[0]['dist_km'] if end else 999

    eof_b_vals = sorted([v for v in eof_b.values() if v < 999])
    if eof_b_vals:
        top_quarter_b = eof_b_vals[int(len(eof_b_vals) * 0.75):]
        total_span_b = np.median(top_quarter_b)
    else:
        total_span_b = 0

    for fnum, r in fibers_a.items():
        rb = fibers_b.get(fnum)
        b_span = None
        if rb:
            b_end = [e for e in rb['events'] if e['is_end']]
            b_span = b_end[0]['dist_km'] if b_end else total_span_b

        # ── Per-fiber B-fill coverage / dead-zone pre-compute ──
        # If this fiber is A-broken and B also has a premature end/break,
        # there may be a stretch of cable that neither trace could see.
        _fiber_end_a = eof_a.get(fnum, 999)
        _fiber_is_a_broken = (0 < _fiber_end_a < total_span_a - END_REGION_KM)
        # B-fill reach = nearest-to-A-launch A-frame km that the B trace
        # can see.  B fiber sees from B-launch (A-frame = total_span_a) back
        # through b_span of fiber, so the furthest-back A-frame position it
        # reaches is (total_span_a - b_span).
        _b_fill_reach_km = None
        _dead_zone = None  # (lo_km, hi_km) in A-frame, both ends inclusive
        if _fiber_is_a_broken and b_span and total_span_a:
            _b_fill_reach_km = max(0.0, total_span_a - b_span)
            # If B's reach stops SHORT of the A-break (i.e., reach > break),
            # there's a gap between them that neither trace saw.
            if _b_fill_reach_km > _fiber_end_a + 0.2:
                _dead_zone = (_fiber_end_a, _b_fill_reach_km)

        for si, sp in enumerate(splices):
            sp_km = sp['position_km']
            # A column may be a real splice ('splice') or a bend / damage zone
            # ('bend' / 'damage').  In a phantom column, every qualifying A event
            # is treated as a bend (never a reburn) and never gets a BEND prefix
            # / offset annotation in its label — the column header already
            # tells the tech what the zone is.
            _column_kind = sp.get('column_kind', 'splice')
            _is_phantom_column = _column_kind in ('bend', 'damage')

            # ── Broke detection ──
            fiber_end = eof_a[fnum]
            # Broke detection: A trace ends mid-span (more than END_REGION_KM
            # short of the cable end).  We deliberately do NOT require the
            # B trace to also be truncated — for high-loss splices that
            # only stop A's auto-detector but not the actual signal, B
            # may legitimately see all the way through (e.g. F841 in
            # SANDUR — A quits at km 20.14, B reaches the far-end
            # reflection at km 100.56).  The earlier "a_plus_b ≈
            # total_span" guard would mis-classify such fibers as healthy
            # when they have a real high-loss event on A.
            is_mid_span_break = (fiber_end > 0 and
                                 fiber_end < total_span_a - END_REGION_KM)

            if is_mid_span_break:
                # Mark as BROKE at the nearest splice to where it terminated
                nearest_splice = min(range(len(splices)),
                                     key=lambda i: abs(splices[i]['position_km'] - fiber_end))
                nearest_dist = abs(splices[nearest_splice]['position_km'] - fiber_end)
                # No distance cap here: a break can sit several km from
                # any splice closure (e.g. Durkee↔La Grande F311–F384 at
                # km 38.43 — 3 km from the nearest closure).  We log it
                # against the NEAREST splice and let the off-splice
                # splitter relocate it into its own damage column.
                if nearest_splice == si:
                    # Enrich label with B-fill coverage / dead-zone range
                    if _dead_zone is not None:
                        _broke_label = (f"{fnum} broke@{fiber_end:.1f}k | "
                                        f"DZ {_dead_zone[0]:.1f}-"
                                        f"{_dead_zone[1]:.1f}k")
                    elif _b_fill_reach_km is not None:
                        _broke_label = f"{fnum} broke@{fiber_end:.1f}k (B-fill OK)"
                    else:
                        _broke_label = f"{fnum} broke"
                    results[(fnum, si)] = {
                        'fiber': fnum, 'splice_idx': si,
                        'bidir_loss': None, 'a_loss': None, 'b_loss': None,
                        'bidir_dist': fiber_end,
                        'is_break': False, 'is_broke': True, 'is_bend': False,
                        'is_bfill': False, 'is_dead_zone': False,
                        'is_a_only': False, 'is_b_only': False,
                        'is_flagged': True, 'event_source': 'broke',
                        'event_type': 'BROKE', 'label': _broke_label,
                        'dead_zone_km': _dead_zone,
                        'b_fill_reach_km': _b_fill_reach_km,
                    }
                # B-fill for splices past the break.  Use the cable's
                # total span as the mirror anchor — when B itself is
                # broken short (b_span < total_span_a), the simple
                # 'b_span − b_dist' mirror would map every B event to
                # the wrong A-frame position and B-fill would silently
                # miss everything.
                elif sp_km > fiber_end and rb and b_span:
                    mirror_anchor = total_span_a if total_span_a > 0 else b_span
                    b_evt = None
                    for e in rb['events']:
                        if e['dist_km'] < 1.0 or e['is_end']: continue
                        ef_from_a = mirror_anchor - e['dist_km']
                        if abs(ef_from_a - sp_km) < POSITION_TOL:
                            if b_evt is None or abs(ef_from_a - sp_km) < abs((mirror_anchor - b_evt['dist_km']) - sp_km):
                                b_evt = e
                    if b_evt is not None:
                        b_loss_val = b_evt['splice_loss']
                        # Single-direction rule: no averaging, no /2 estimate.
                        # The raw B-fill loss must clear SINGLE_DIR_THRESHOLD
                        # on its own (default 0.250 dB) — stricter than the
                        # bidir threshold because we have no opposite-side
                        # confirmation.  Gate on the SIGNED loss (positive=loss,
                        # negative=gain) so a B-side gainer can't masquerade as a
                        # single-direction loss — mirrors the A-side gate.
                        if b_loss_val >= SINGLE_DIR_THRESHOLD:
                            loss_str = f"{b_loss_val:.3f}"
                            if loss_str.startswith('0.'): loss_str = loss_str[1:]
                            label = f"{fnum} {loss_str} (B-fill)"
                            results[(fnum, si)] = {
                                'fiber': fnum, 'splice_idx': si,
                                'bidir_loss': b_loss_val, 'a_loss': None,
                                'b_loss': b_evt['splice_loss'],
                                'bidir_dist': mirror_anchor - b_evt['dist_km'],
                                'is_break': False, 'is_broke': False, 'is_bend': False,
                                'is_bfill': True, 'is_dead_zone': False,
                                'is_a_only': False, 'is_b_only': False,
                                'is_flagged': True, 'event_source': 'bfill',
                                'event_type': b_evt['type'],
                                'label': label,
                            }
                    elif (_dead_zone is not None and
                          _dead_zone[0] < sp_km < _dead_zone[1]):
                        # Column falls inside the dead zone for this fiber —
                        # neither A (past its break) nor B (past its own
                        # break) can see it.  Mark so the tech knows this
                        # splice was unobservable for this fiber.
                        results[(fnum, si)] = {
                            'fiber': fnum, 'splice_idx': si,
                            'bidir_loss': None, 'a_loss': None, 'b_loss': None,
                            'bidir_dist': sp_km,
                            'is_break': False, 'is_broke': False, 'is_bend': False,
                            'is_bfill': False, 'is_dead_zone': True,
                            'is_a_only': False, 'is_b_only': False,
                            'is_flagged': False, 'event_source': 'dead_zone',
                            'event_type': 'DEAD_ZONE',
                            'label': f"{fnum} DZ",
                            'dead_zone_km': _dead_zone,
                            'b_fill_reach_km': _b_fill_reach_km,
                        }
                continue

            # ── Find A event near this splice (neighbor-aware) ──
            # When an adjacent closure sits closer than POSITION_TOL,
            # tighten the matching window to half the gap so a single
            # A event doesn't get claimed by both closures (double
            # attribution).  Floor at 0.30 km so legitimate per-fiber
            # length drift still resolves.  Wide-spaced spans keep the
            # full POSITION_TOL.  Use index-based exclusion so a splice
            # whose raw vs refined km differ slightly doesn't see its
            # own refined position as a "neighbor."
            nearest_other_km_a = min(
                (abs(closure_kms_all[j] - sp_km)
                 for j in range(len(closure_kms_all)) if j != si),
                default=2 * POSITION_TOL)
            local_tol_a = min(POSITION_TOL,
                               max(0.30, nearest_other_km_a / 2.0))
            ea = None
            for e in r['events']:
                if abs(e['dist_km'] - sp_km) < local_tol_a and e['dist_km'] > 1.0 and not e['is_end']:
                    if ea is None or abs(e['dist_km'] - sp_km) < abs(ea['dist_km'] - sp_km):
                        ea = e

            if ea is None:
                continue

            # ── Find matching B event ──
            # Neighbor-aware matching (May 4 revision): when the
            # nearest neighbor splice is closer than POSITION_TOL, use
            # a tighter tolerance equal to half the gap to that neighbor.
            # Index-based exclusion to skip THIS splice's own refined
            # position (raw vs refined km can differ by ~50 m).
            nearest_other_km = min(
                (abs(closure_kms_all[j] - sp_km)
                 for j in range(len(closure_kms_all)) if j != si),
                default=2 * POSITION_TOL)
            local_tol = min(POSITION_TOL, max(0.30, nearest_other_km / 2.0))

            eb = None
            b_loss = None
            b_from_a = None
            if rb and b_span:
                for e in rb['events']:
                    if e['dist_km'] < 1.0 or e['is_end']: continue
                    ef_from_a = b_span - e['dist_km']
                    if abs(ef_from_a - ea['dist_km']) >= local_tol:
                        continue
                    if eb is None or abs(ef_from_a - ea['dist_km']) < abs((b_span - eb['dist_km']) - ea['dist_km']):
                        eb = e
                        b_loss = e['splice_loss']
                        b_from_a = ef_from_a

            # ── A event but no B event in table ──
            # Try to measure the B-direction loss directly from the B trace
            # using wide-LSA (EXFO's "grey value" approach).  Convert the
            # splice position to the B-frame (B_dist = B_span - sp_km).
            if b_loss is None:
                a_loss_abs = abs(ea['splice_loss'])
                b_grey = None
                if rb is not None and b_span:
                    b_frame_km = b_span - sp_km
                    b_grey = _grey_loss(rb, b_frame_km)

                if b_grey is not None:
                    # Real bidirectional average using measured B grey
                    true_bidir = round((ea['splice_loss'] + b_grey) / 2.0, 4)
                    closure_center_km = _closure_km_for_fiber(sp, fnum)
                    bend_ref_km = (_per_fiber_splice_km(r['events'], closure_center_km)
                                    or closure_center_km)
                    is_bend_offset = _is_bend_event(ea['dist_km'], bend_ref_km, true_bidir,
                                                    fiber_events=r['events'],
                                                    a_loss=ea['splice_loss'], b_loss=b_grey,
                                                    closure_kms=closure_kms_all,
                                                    fiber_data=r)
                    is_bend = is_bend_offset or _is_phantom_column

                    if abs(true_bidir) >= threshold or is_bend:
                        loss_str = _format_loss(true_bidir)
                        if is_bend and not _is_phantom_column:
                            offset_m = round((ea['dist_km'] - bend_ref_km) * 1000, 0)
                            label = f"{fnum} BEND {loss_str} bidi ({offset_m:+.0f}m)"
                        else:
                            label = f"{fnum} {loss_str}"
                        results[(fnum, si)] = {
                            'fiber': fnum, 'splice_idx': si,
                            'bidir_loss': true_bidir,
                            'a_loss': ea['splice_loss'], 'b_loss': b_grey,
                            'bidir_dist': ea['dist_km'],
                            'is_break': False, 'is_broke': False, 'is_bend': is_bend,
                            'is_bfill': False, 'is_a_only': False, 'is_b_only': False,
                            'is_flagged': True,
                            'event_source': 'bend' if is_bend else 'bidir_grey_b',
                            'bend_severity': _bend_severity(true_bidir) if is_bend else None,
                            'closure_offset_m': round((ea['dist_km'] - bend_ref_km) * 1000, 1) if is_bend else None,
                            'event_type': ea['type'],
                            'label': label,
                            '_b_is_grey': not is_bend,
                        }
                        continue

                    # Bidir average is below threshold — but that can simply
                    # mean the B-direction is FLAT here while A shows a real
                    # loss (the Ontario↔Boise / Seattle F111 miss: A reads a
                    # clear reburn, B's grey-LSA reads ~0, so the /2 average
                    # falls under REBURN_THRESHOLD and the event used to drop
                    # silently).  Recover it as a single-direction A cell when
                    # — conservatively — the PRESENT side is clearly real
                    # (raw A clears the stricter SINGLE_DIR_THRESHOLD) AND the
                    # ABSENT side is verifiably flat (grey ≈ 0, not merely
                    # unmeasurable: b_grey is not None and below the bend
                    # floor).  Anything noisier than that stays dropped — we
                    # don't want a borderline B reading masquerading as flat.
                    # Re-measure gate: the stored table loss must be CONFIRMED
                    # by a tight local LSA on this fiber's own trace — HOWLAN
                    # shipped 13 phantom (A) flags from stored 0.26-0.30 dB
                    # entries over locally flat glass.
                    if (ea['splice_loss'] >= SINGLE_DIR_THRESHOLD and
                            abs(b_grey) < BEND_THRESHOLD and
                            _local_step_confirms(r, ea)):
                        loss_str = _format_loss(a_loss_abs)
                        closure_center_km = _closure_km_for_fiber(sp, fnum)
                        bend_ref_km = (_per_fiber_splice_km(r['events'], closure_center_km)
                                        or closure_center_km)
                        is_bend_offset = _is_bend_event(ea['dist_km'], bend_ref_km, ea['splice_loss'],
                                                        fiber_events=r['events'],
                                                        closure_kms=closure_kms_all,
                                                        fiber_data=r)
                        is_bend = is_bend_offset or _is_phantom_column
                        if is_bend and not _is_phantom_column:
                            offset_m = round((ea['dist_km'] - bend_ref_km) * 1000, 0)
                            label = f"{fnum} BEND {loss_str}(A) ({offset_m:+.0f}m)"
                        elif is_bend:
                            label = f"{fnum} {loss_str}(A)"
                        else:
                            label = f"{fnum} {loss_str} (A)"
                        results[(fnum, si)] = {
                            'fiber': fnum, 'splice_idx': si,
                            'bidir_loss': None, 'a_loss': ea['splice_loss'],
                            'b_loss': b_grey,
                            'bidir_dist': ea['dist_km'],
                            'is_break': False, 'is_broke': False, 'is_bend': is_bend,
                            'is_bfill': False,
                            'is_a_only': not is_bend, 'is_b_only': False,
                            'is_flagged': True,
                            'event_source': 'bend' if is_bend else 'a_only',
                            'bend_severity': _bend_severity(ea['splice_loss']) if is_bend else None,
                            'closure_offset_m': round((ea['dist_km'] - bend_ref_km) * 1000, 1) if is_bend else None,
                            'event_type': ea['type'],
                            'label': label,
                            '_b_is_flat_grey': True,
                        }
                        continue

                    # Below threshold and B not confirmed flat — skip
                    continue

                # No JSON trace available — fall back to conservative (A alone) check:
                # A-only single-direction needs the stricter SINGLE_DIR_THRESHOLD
                # (default 0.250 dB).  No averaging.  The raw A loss alone must
                # clear it — the unseen B side can't confirm a single-direction
                # reburn.  Re-measure gate: stored loss must be locally real.
                if (ea['splice_loss'] >= SINGLE_DIR_THRESHOLD and
                        _local_step_confirms(r, ea)):
                    loss_str = _format_loss(a_loss_abs)
                    closure_center_km = _closure_km_for_fiber(sp, fnum)
                    bend_ref_km = (_per_fiber_splice_km(r['events'], closure_center_km)
                                    or closure_center_km)
                    is_bend_offset = _is_bend_event(ea['dist_km'], bend_ref_km, ea['splice_loss'],
                                                    fiber_events=r['events'],
                                                    closure_kms=closure_kms_all,
                                                    fiber_data=r)
                    is_bend = is_bend_offset or _is_phantom_column
                    if is_bend and not _is_phantom_column:
                        offset_m = round((ea['dist_km'] - bend_ref_km) * 1000, 0)
                        label = f"{fnum} BEND {loss_str}(A) ({offset_m:+.0f}m)"
                    elif is_bend:
                        label = f"{fnum} {loss_str}(A)"
                    else:
                        label = f"{fnum} {loss_str} (A)"
                    results[(fnum, si)] = {
                        'fiber': fnum, 'splice_idx': si,
                        'bidir_loss': None, 'a_loss': ea['splice_loss'], 'b_loss': None,
                        'bidir_dist': ea['dist_km'],
                        'is_break': False, 'is_broke': False, 'is_bend': is_bend,
                        'is_bfill': False,
                        'is_a_only': not is_bend, 'is_b_only': False,
                        'is_flagged': True,
                        'event_source': 'bend' if is_bend else 'a_only',
                        'bend_severity': _bend_severity(ea['splice_loss']) if is_bend else None,
                        'closure_offset_m': round((ea['dist_km'] - bend_ref_km) * 1000, 1) if is_bend else None,
                        'event_type': ea['type'],
                        'label': label,
                    }
                continue

            # ── A+B bidirectional ──
            bidir_loss = round((ea['splice_loss'] + b_loss) / 2.0, 4)
            bidir_dist = round((ea['dist_km'] + b_from_a) / 2.0, 4)

            is_reflective = ea['type'].startswith('1F')
            has_weak_fresnel = ea['reflection'] < -25.0
            mid_span = ea['dist_km'] < (total_span_a - END_REGION_KM)
            # Reflective + Fresnel + mid-span is a *candidate* for either
            # BREAK (trace terminates near this point) or REF (in-line
            # reflective event — connector / mechanical splice / angled
            # cleave with the trace continuing through).  Disambiguate by
            # asking whether the fiber's trace has real events past this
            # position and an EOF that's farther downstream.
            is_refl_event_candidate = is_reflective and has_weak_fresnel and mid_span
            trace_continues = _trace_continues_past(
                r['events'], ea['dist_km'], total_span_a)
            is_break = is_refl_event_candidate and not trace_continues
            is_ref   = is_refl_event_candidate and trace_continues

            # ── BEND check (ZeroDBIFTHEN Flag-3 rule) ──
            # If the event position is offset from the true closure center
            # by more than CLOSURE_MATCH_KM (150 m), this is a BEND not a
            # splice reburn.  Use the refined (mode-based) center, falling
            # back to the coarse position_km if refinement hasn't run.
            closure_center_km = _closure_km_for_fiber(sp, fnum)
            bend_ref_km = (_per_fiber_splice_km(r['events'], closure_center_km)
                            or closure_center_km)
            is_bend_offset = _is_bend_event(ea['dist_km'], bend_ref_km, bidir_loss,
                                            fiber_events=r['events'],
                                            a_loss=ea['splice_loss'], b_loss=b_loss,
                                            closure_kms=closure_kms_all,
                                            fiber_data=r)
            # Phantom columns always classify as bends (unless they're breaks
            # or in-line reflective events).
            is_bend = (not is_break) and (not is_ref) and (is_bend_offset or _is_phantom_column)

            is_flagged = (abs(bidir_loss) >= threshold) or is_break or is_ref or is_bend
            # Borderline band: surface a sub-threshold loss sitting on the
            # reburn knife-edge for review even though it isn't flagged.  Emit
            # when flagged OR borderline (break/ref/bend are never borderline).
            is_borderline = (not is_break and not is_ref and not is_bend
                             and _is_borderline_loss(bidir_loss, threshold))
            if not (is_flagged or is_borderline):
                continue

            if is_break:
                offset_m = round((bidir_dist - sp_km) * 1000, 1)
                uni_loss = abs(ea['splice_loss'])
                refl_db = ea['reflection']
                refl_str = f" {uni_loss:.3f} uni reflection {refl_db:.0f}"
                if refl_db > -35.0:
                    break_type = " air gap"
                else:
                    break_type = ""
                label = f"{fnum} BREAK {bidir_loss:.3f} ({abs(offset_m):.0f}m from splice){refl_str}{break_type}"
            elif is_ref:
                refl_db = ea['reflection']
                label = f"{fnum} ref {_format_loss(bidir_loss)} (refl {refl_db:.0f}dB)"
            elif is_bend:
                loss_str = _format_loss(bidir_loss)
                if _is_phantom_column:
                    # Column already says 'bends@X km' — cell just shows fiber + loss
                    label = f"{fnum} {loss_str}"
                else:
                    offset_m = round((ea['dist_km'] - bend_ref_km) * 1000, 0)
                    label = f"{fnum} BEND {loss_str} bidi ({offset_m:+.0f}m)"
            else:
                loss_str = _format_loss(bidir_loss)
                label = f"{fnum} {loss_str}"

            # Borderline / review marker (display-only): is_borderline was
            # computed at the emission gate above (so a sub-threshold knife-edge
            # cell is surfaced for review).  Break / ref / bend are their own
            # decisions, never "borderline".
            if is_borderline:
                label = f"{label}  ⚠ borderline"

            # For is_ref classifications, the defining feature is the
            # A-side reflective signature itself.  Use the A-event km as
            # the cell's km (NOT the A/B average) so the off-splice
            # column lands at the actual reflection's position, not
            # halfway between A's reflection and B's matched loss event
            # (which may be 700 m away — e.g. F1008 at km 32.15 paired
            # with the B-side loss event at A-frame 32.83).
            cell_km = ea['dist_km'] if is_ref else bidir_dist

            # ── DIRTY/BAD connector recategorization (additive) ──
            # A reflective in-line event that also drops a real loss step is a
            # dirty connector.  Refine its category + label only; the flag
            # decision (is_ref / is_flagged) is unchanged.
            _event_source = ('break' if is_break else
                             'ref' if is_ref else
                             'bend' if is_bend else 'bidir')
            if is_ref and _is_dirty_connector(ea['dist_km'], ea['reflection'],
                                               bidir_loss, is_end=ea.get('is_end')):
                _event_source = 'dirty_connector'
                label = f"{label} DIRTY CONNECTOR"

            results[(fnum, si)] = {
                'fiber': fnum, 'splice_idx': si,
                'bidir_loss': bidir_loss,
                'a_loss': ea['splice_loss'], 'b_loss': b_loss,
                'bidir_dist': cell_km,
                'is_break': is_break, 'is_broke': False, 'is_bend': is_bend,
                'is_ref': is_ref,
                'is_bfill': False, 'is_a_only': False, 'is_b_only': False,
                'is_borderline': is_borderline,
                'is_flagged': is_flagged,
                'event_source': _event_source,
                'bend_severity': _bend_severity(bidir_loss) if is_bend else None,
                'closure_offset_m': round((ea['dist_km'] - bend_ref_km) * 1000, 1) if is_bend else None,
                'event_type': ea['type'],
                'label': label,
                'fresnel': ea['reflection'] if is_reflective else None,
            }

    return results


# ═══════════════════════════════════════════════════════════════════════
#  STEP 4 — Pass 2: Scan all B-direction events not caught in Pass 1
# ═══════════════════════════════════════════════════════════════════════

def scan_b_events(fibers_a, fibers_b, splices, threshold, existing_results, total_span_a,
                  bend_threshold=None, closure_match_km=None, **_ignored):
    """
    Pass 2: For every B-direction event above threshold that was NOT already
    caught in Pass 1, find the nearest splice position (within 1.5 km) and report it.

    This is how EXFO finds events like fiber 325's 0.340 dB entry that only
    exists in the B-direction event table with no matching A-direction event.

    Returns a dict of (fnum, si) -> result — same structure as analyze_all().
    Does NOT overwrite any existing_results entries.
    """
    new_results = {}
    closure_kms_all = [sp.get('position_km_refined', sp['position_km'])
                       for sp in splices]

    for fnum, rb in fibers_b.items():
        ra = fibers_a.get(fnum)

        # B-direction span (EOL)
        b_end_events = [e for e in rb['events'] if e['is_end']]
        if not b_end_events:
            continue
        b_span = b_end_events[0]['dist_km']

        # A-direction EOL (to know if this fiber is broken)
        ra_end_km = total_span_a
        if ra:
            a_end = [e for e in ra['events'] if e['is_end']]
            if a_end:
                ra_end_km = a_end[0]['dist_km']

        for e in rb['events']:
            if e['dist_km'] < 1.0 or e['is_end']:
                continue
            # B-side tailbox region — within LAUNCH_FIBER_MAX km of the
            # B-direction EOL.  This is the launch connector on the OTHER
            # cable end as seen from B; mirrored to A-frame it falls inside
            # the A-launch zone, but POSITION_TOL (1.5 km) can pull these
            # events onto the first real splice column.  Drop them here —
            # detect_launch_issues() owns this region.
            if e['dist_km'] > (b_span - LAUNCH_FIBER_MAX):
                continue

            b_loss_signed = e['splice_loss']
            b_loss_abs = abs(b_loss_signed)
            # Gate: skip clearly-too-small B events.  Use B alone (not B/2)
            # because the real bidir depends on the A grey value we haven't
            # measured yet.  Anything with single-dir loss below threshold
            # can't possibly produce a bidir above threshold unless A grey
            # is even larger, which is unlikely.
            if b_loss_abs < threshold * 0.75:
                continue

            # Convert B-frame position to A-frame
            a_frame_km = b_span - e['dist_km']
            if a_frame_km < 0.5:
                continue  # launch artifact near B-end

            # Find nearest splice position within tolerance
            nearest_si = None
            nearest_dist = float('inf')
            for si, sp in enumerate(splices):
                d = abs(sp['position_km'] - a_frame_km)
                if d < nearest_dist:
                    nearest_dist = d
                    nearest_si = si

            if nearest_si is None or nearest_dist > POSITION_TOL:
                continue  # not near any known splice position

            # Already caught by Pass 1?
            if (fnum, nearest_si) in existing_results:
                continue

            # Already found a better match in this pass?
            if (fnum, nearest_si) in new_results:
                existing_a_frame = new_results[(fnum, nearest_si)]['bidir_dist']
                if nearest_dist >= abs(splices[nearest_si]['position_km'] - existing_a_frame):
                    continue

            # Look for A-direction event near the same A-frame position
            a_evt = None
            if ra:
                for ae in ra['events']:
                    if ae['dist_km'] < 1.0 or ae['is_end']: continue
                    if abs(ae['dist_km'] - a_frame_km) < POSITION_TOL:
                        if a_evt is None or abs(ae['dist_km'] - a_frame_km) < abs(a_evt['dist_km'] - a_frame_km):
                            a_evt = ae

            closure_center_km = _closure_km_for_fiber(splices[nearest_si], fnum)
            # Per-fiber bend reference: use this fiber's own A-direction
            # events (when available) to anchor the bend gate to the
            # fiber's actual splice km rather than the global closure.
            ra_events = ra['events'] if ra else None
            bend_ref_km = closure_center_km
            if ra_events is not None:
                pf = _per_fiber_splice_km(ra_events, closure_center_km,
                                           exclude_pos_km=a_frame_km)
                if pf is not None:
                    bend_ref_km = pf

            if a_evt is not None:
                # A event exists — compute bidirectional
                bidir = round((a_evt['splice_loss'] + b_loss_signed) / 2.0, 4)
                is_bend = _is_bend_event(a_frame_km, bend_ref_km, bidir,
                                         fiber_events=ra_events,
                                         a_loss=a_evt['splice_loss'],
                                         b_loss=b_loss_signed,
                                         closure_kms=closure_kms_all,
                                         fiber_data=ra)
                if abs(bidir) < threshold and not is_bend:
                    continue
                loss_str = _format_loss(bidir)
                if is_bend:
                    offset_m = round((a_frame_km - bend_ref_km) * 1000, 0)
                    label = f"{fnum} BEND {loss_str} bidi ({offset_m:+.0f}m)"
                else:
                    label = f"{fnum} {loss_str}"
                new_results[(fnum, nearest_si)] = {
                    'fiber': fnum, 'splice_idx': nearest_si,
                    'bidir_loss': bidir,
                    'a_loss': a_evt['splice_loss'], 'b_loss': b_loss_signed,
                    'bidir_dist': a_frame_km,
                    'is_break': False, 'is_broke': False, 'is_bend': is_bend,
                    'is_bfill': False, 'is_a_only': False, 'is_b_only': False,
                    'is_flagged': True,
                    'event_source': 'bend' if is_bend else 'bidir',
                    'bend_severity': _bend_severity(bidir) if is_bend else None,
                    'closure_offset_m': round((a_frame_km - bend_ref_km) * 1000, 1) if is_bend else None,
                    'event_type': a_evt['type'],
                    'label': label,
                }
            else:
                # B event but no A event in table — measure the A-direction
                # loss at this position from the A JSON trace (grey value).
                a_grey = _grey_loss(ra, a_frame_km) if ra is not None else None

                if a_grey is not None:
                    true_bidir = round((a_grey + b_loss_signed) / 2.0, 4)
                    is_bend = _is_bend_event(a_frame_km, bend_ref_km, true_bidir,
                                              fiber_events=ra_events,
                                              a_loss=a_grey, b_loss=b_loss_signed,
                                              closure_kms=closure_kms_all,
                                              fiber_data=ra)
                    if abs(true_bidir) < threshold and not is_bend:
                        continue
                    loss_str = _format_loss(true_bidir)
                    if is_bend:
                        offset_m = round((a_frame_km - bend_ref_km) * 1000, 0)
                        label = f"{fnum} BEND {loss_str} bidi ({offset_m:+.0f}m)"
                    else:
                        label = f"{fnum} {loss_str}"
                    new_results[(fnum, nearest_si)] = {
                        'fiber': fnum, 'splice_idx': nearest_si,
                        'bidir_loss': true_bidir,
                        'a_loss': a_grey, 'b_loss': b_loss_signed,
                        'bidir_dist': a_frame_km,
                        'is_break': False, 'is_broke': False, 'is_bend': is_bend,
                        'is_bfill': False, 'is_a_only': False, 'is_b_only': False,
                        'is_flagged': True,
                        'event_source': 'bend' if is_bend else 'bidir_grey_a',
                        'bend_severity': _bend_severity(true_bidir) if is_bend else None,
                        'closure_offset_m': round((a_frame_km - bend_ref_km) * 1000, 1) if is_bend else None,
                        'event_type': e['type'],
                        'label': label,
                        '_a_is_grey': not is_bend,
                    }
                    continue

                # No JSON trace — fall back to single-direction check.
                # B-only needs the stricter SINGLE_DIR_THRESHOLD (default
                # 0.250 dB).  No averaging — the raw B loss must clear it
                # on its own.  Gate on the SIGNED loss (positive=loss) so a
                # B-side gainer can't surface as a single-dir loss — mirrors A.
                # Re-measure gate: stored loss must be locally real.
                if (b_loss_signed >= SINGLE_DIR_THRESHOLD and
                        _local_step_confirms(rb, e)):
                    is_bend = _is_bend_event(a_frame_km, bend_ref_km, b_loss_signed,
                                              fiber_events=ra_events,
                                              closure_kms=closure_kms_all,
                                              fiber_data=ra)
                    loss_str = _format_loss(b_loss_abs)
                    if is_bend:
                        offset_m = round((a_frame_km - bend_ref_km) * 1000, 0)
                        label = f"{fnum} BEND {loss_str}(B) ({offset_m:+.0f}m)"
                    else:
                        label = f"{fnum} {loss_str} (B)"
                    new_results[(fnum, nearest_si)] = {
                        'fiber': fnum, 'splice_idx': nearest_si,
                        'bidir_loss': None,
                        'a_loss': None, 'b_loss': b_loss_signed,
                        'bidir_dist': a_frame_km,
                        'is_break': False, 'is_broke': False, 'is_bend': is_bend,
                        'is_bfill': False,
                        'is_a_only': False, 'is_b_only': not is_bend,
                        'is_flagged': True,
                        'event_source': 'bend' if is_bend else 'b_only',
                        'bend_severity': _bend_severity(b_loss_signed) if is_bend else None,
                        'closure_offset_m': round((a_frame_km - bend_ref_km) * 1000, 1) if is_bend else None,
                        'event_type': e['type'],
                        'label': label,
                    }

    return new_results


# ═══════════════════════════════════════════════════════════════════════
#  STEP 4b — APRIL 23 revision: A-first standalone-event classifier
#  (bends / breaks that are not at any VALID splice closure)
# ═══════════════════════════════════════════════════════════════════════

def scan_a_standalone_events(fibers_a, splices, existing_results, total_span_a,
                             bend_threshold=None, closure_match_km=None,
                             fibers_b=None):
    """Every A-direction non-end event that was NOT covered by Pass 1 gets
    classified as a BEND or a BREAK/BROKE.  This replaces the old behaviour
    where events inside a phantom splice column would render as splice
    reburns even though they were actually bends.

    Rule (from the tech's instructions):
      • We build the template off A direction events.
      • Any A events NOT at a splice column → bend or break.

    Returns dict (fnum, synthetic_si) → result-dict.  synthetic_si is the
    index of the NEAREST valid closure (used only for ribbon-grid placement;
    the event is displayed with a BEND/BREAK label + offset annotation)."""
    bt = BEND_THRESHOLD   if bend_threshold   is None else bend_threshold
    cm = CLOSURE_MATCH_KM if closure_match_km is None else closure_match_km

    new_results = {}
    seen_keys = set(existing_results.keys())

    # Pre-compute the refined center of every closure for fast nearest lookup.
    # We use the global refined position here for the nearest-splice search
    # (since we don't yet know which fiber we're matching).  Once a nearest
    # splice is picked, the bend-vs-splice geometry test uses the
    # per-ribbon refined position via _closure_km_for_fiber().
    closure_centers = [(si, sp.get('position_km_refined', sp['position_km']))
                       for si, sp in enumerate(splices)]

    for fnum, ra in fibers_a.items():
        # Skip fibers that are broke — broke fibers get special treatment
        events = ra.get('events') or []
        end_events = [e for e in events if e.get('is_end')]
        if not end_events:
            continue
        eof_a = end_events[0]['dist_km']
        is_broken = eof_a < (total_span_a - END_REGION_KM)

        for e in events:
            if e['is_end']:
                continue
            if e['dist_km'] < 1.0:
                continue  # launch region — handled separately
            if eof_a is not None and e['dist_km'] >= eof_a:
                continue  # post-EOL event detector noise (instrument tail)
            # Tailbox region — within LAUNCH_FIBER_MAX km of the cable end.
            # Mirror to the launch zone: events here belong to the tailbox
            # connector or the receive pigtail and are handled exclusively
            # by detect_launch_issues() (BAD_TAILBOX_REFL).  Normal tailbox
            # connectors have 0.1–0.3 dB of legitimate connector loss that
            # would otherwise leak into the splice report as false bends.
            if eof_a is not None and e['dist_km'] > (eof_a - LAUNCH_FIBER_MAX):
                continue
            loss = e.get('splice_loss') or 0.0
            # Bends are real attenuation → require POSITIVE signed loss
            # ≥ threshold.  Negative-signed events are gainer signatures
            # or LSA artifacts and are handled by the gainer pass / not
            # at all (respectively).
            if loss < bt:
                continue

            # Find nearest closure (by refined center)
            best_si, best_d = None, float('inf')
            for si, c in closure_centers:
                d = abs(e['dist_km'] - c)
                if d < best_d:
                    best_d = d
                    best_si = si
            if best_si is None:
                # No closure was discovered for this span (a tiny / all-bend /
                # mostly-broke span where discover_splices returned []).  There
                # is no column to attribute this event to — skip it rather than
                # indexing splices[None] / storing a (fnum, None) key, which was
                # a hard TypeError downstream that surfaced to the tech as
                # "report failed" with no manifest.
                continue

            # PER-FIBER bend gate: the candidate event is a "splice event"
            # for this specific fiber if it lies within CLOSURE_MATCH_KM of
            # this fiber's own nearest event to the closure (i.e. it IS the
            # fiber's primary splice event, even if that splice is offset by
            # >150 m from the global closure center due to length variation).
            # If so, Pass 1 has already classified it — skip.
            best_sp = splices[best_si] if best_si is not None else None
            best_closure_km = closure_centers[best_si][1] if best_si is not None else None
            # Exclude the candidate event itself from the per-fiber search
            # so the reference km is the fiber's PRIMARY splice event (the
            # one Pass 1 already classified), not this candidate.
            per_fiber_km = (_per_fiber_splice_km(events, best_closure_km,
                                                  exclude_pos_km=e['dist_km'])
                             if best_closure_km is not None else None)
            bend_ref_km = per_fiber_km if per_fiber_km is not None else best_closure_km
            if bend_ref_km is not None and abs(e['dist_km'] - bend_ref_km) <= cm:
                continue
            # If very far from every closure, still bucket to the nearest
            # one for display purposes.  The off-splice splitter will
            # relocate it into its own bend/damage column later.

            # Already flagged by Pass 1 / Pass 2 at this closure position?
            key = (fnum, best_si)
            if key in seen_keys or key in new_results:
                continue

            is_reflective = e.get('is_reflective') or str(e.get('type','')).startswith('1F')
            refl = e.get('reflection') or 0.0
            has_weak_fresnel = refl < -25.0

            # Reflective + weak Fresnel + mid-span — disambiguate BREAK
            # vs in-line REF (trace continues past the event).
            if is_reflective and has_weak_fresnel and e['dist_km'] < (total_span_a - END_REGION_KM):
                trace_continues = _trace_continues_past(events, e['dist_km'], total_span_a)
                if trace_continues:
                    # In-line reflective event — connector / mech splice / etc.
                    label = f"{fnum} ref {_format_loss(loss)} (refl {refl:.0f}dB)"
                    # ── DIRTY/BAD connector recategorization (additive) ──
                    # Refine category + label only; the flag decision is
                    # unchanged (this event already flags as a reflective).
                    _src = 'ref_standalone'
                    if _is_dirty_connector(e['dist_km'], refl, loss,
                                           is_end=e.get('is_end')):
                        _src = 'dirty_connector'
                        label = f"{label} DIRTY CONNECTOR"
                    new_results[key] = {
                        'fiber': fnum, 'splice_idx': best_si,
                        'bidir_loss': loss, 'a_loss': loss, 'b_loss': None,
                        'bidir_dist': e['dist_km'],
                        'is_break': False, 'is_broke': False, 'is_bend': False,
                        'is_ref': True,
                        'is_bfill': False, 'is_a_only': False, 'is_b_only': False,
                        'is_flagged': True, 'event_source': _src,
                        'event_type': e['type'],
                        'label': label,
                        'fresnel': refl,
                    }
                    continue
                # Real BREAK — fiber doesn't carry on past this point
                loss_str = _format_loss(loss)
                refl_str = f" {abs(loss):.3f} uni reflection {refl:.0f}"
                break_type = " air gap" if refl > -35.0 else ""
                label = (f"{fnum} BREAK {loss_str} @ {e['dist_km']:.3f}km"
                         f"{refl_str}{break_type}")
                new_results[key] = {
                    'fiber': fnum, 'splice_idx': best_si,
                    'bidir_loss': loss, 'a_loss': loss, 'b_loss': None,
                    'bidir_dist': e['dist_km'],
                    'is_break': True, 'is_broke': False, 'is_bend': False,
                    'is_ref': False,
                    'is_bfill': False, 'is_a_only': False, 'is_b_only': False,
                    'is_flagged': True, 'event_source': 'break_standalone',
                    'event_type': e['type'],
                    'label': label,
                    'fresnel': refl,
                }
                continue

            # BEND verification: cross-check against the B-direction data.
            # Per tech direction, the BIDIRECTIONAL AVERAGE (A_loss + B_loss)/2
            # must reach 0.090 dB for an event to qualify as a bend — that
            # way an A-only artifact gets averaged down to roughly half the
            # threshold and stops triggering false bends.  The B value can
            # come from a discrete B event in the table OR from a wide-LSA
            # grey reading on the B raw trace at the mirrored A-frame
            # position.
            b_value = None
            b_source = None
            b_event = None
            if fibers_b is not None and total_span_a > 0:
                rb = fibers_b.get(fnum)
                if rb is not None:
                    b_end_b = [be for be in rb.get('events', []) if be.get('is_end')]
                    if b_end_b:
                        # Discrete B event search (mirror via total_span_a)
                        for be in rb.get('events', []):
                            if be.get('is_end') or be['dist_km'] < 1.0:
                                continue
                            b_a_frame = total_span_a - be['dist_km']
                            d = abs(b_a_frame - e['dist_km'])
                            if d < 0.300:  # ±300 m bend matching tolerance
                                if b_event is None or d < abs((total_span_a - b_event['dist_km']) - e['dist_km']):
                                    b_event = be
                        if b_event is not None:
                            b_value = b_event.get('splice_loss')
                            b_source = 'event'
                        else:
                            # Fall back to wide-LSA grey on the B trace
                            try:
                                b_frame_km = total_span_a - e['dist_km']
                                grey = _grey_loss(rb, b_frame_km)
                            except Exception:
                                grey = None
                            if grey is not None:
                                b_value = grey
                                b_source = 'grey_lsa'

            # Bidir-average bend rule: drop the bend unless (A + B) / 2
            # reaches BEND_THRESHOLD (0.090 dB by default).  When no B
            # value can be obtained at all, skip — we don't trust an
            # A-only loss on its own for bend classification.
            if b_value is None:
                continue
            bidir_avg = (loss + b_value) / 2.0
            if bidir_avg < bt:
                continue  # bidir must be POSITIVE ≥ threshold for a bend

            # ── Per-fiber length model + narrow-LSA gate (April 28) ──
            # Test 1: predict where this fiber's splice should sit at this
            # closure using a leave-one-out linear fit through the fiber's
            # (cable_km, event_km) pairs at every other closure.  If the
            # candidate event sits within BEND_RES_SPLICE_M of the
            # prediction, it IS this fiber's splice (just offset from the
            # ribbon median by accumulated fiber-length variation) — drop.
            # If between SPLICE and BEND thresholds, treat as ambiguous
            # and drop conservatively.
            # Test 2: when residual is large enough to call BEND, confirm
            # by reading a narrow-LSA on the raw A trace at predicted_km.
            # A real bend means a separate splice exists at predicted_km
            # (loss step ≥ BEND_NARROW_LOSS_DB) AND the candidate event
            # exists at this position.  No splice signature → drop.
            closure_kms_local = [c for _, c in closure_centers]
            residual_m, predicted_km, _n_fit = _perfiber_residual_m(
                ra, closure_kms_local, e['dist_km'])
            if residual_m is not None:
                ares = abs(residual_m)
                if ares <= BEND_RES_SPLICE_M:
                    continue          # SPLICE — fits the fiber's pattern
                if ares < BEND_RES_BEND_M:
                    continue          # ambiguous — conservative drop
                # Confirm via wide-LSA at predicted_km — both JSON and
                # SOR sources supported (dispatch in _narrow_lsa_loss).
                # Conservative: require positive evidence.  Drop the
                # candidate when the LSA can't fit (None) or returns a
                # sub-threshold magnitude.
                if ra is not None:
                    pred_loss = _narrow_lsa_loss(ra, predicted_km)
                    if pred_loss is None or abs(pred_loss) < BEND_NARROW_LOSS_DB:
                        continue      # no corroborating splice — drop

            # BEND: everything else above threshold.  If the nearest column
            # is itself a phantom bend/damage zone, the column header already
            # describes the zone — keep the cell label clean.
            # Label shows the bidirectional-average loss (the value the gate
            # used) with a 'bidi' tag to make clear it's not the raw A reading.
            bidir = round((loss + b_value) / 2.0, 4)
            target_sp = splices[best_si]
            target_is_phantom = target_sp.get('column_kind') in ('bend', 'damage')
            loss_str = _format_loss(bidir)
            if target_is_phantom:
                label = f"{fnum} {loss_str} bidi"
            else:
                offset_m = round((e['dist_km'] - bend_ref_km) * 1000, 0)
                label = f"{fnum} BEND {loss_str} bidi ({offset_m:+.0f}m)"
            new_results[key] = {
                'fiber': fnum, 'splice_idx': best_si,
                'bidir_loss': bidir, 'a_loss': loss, 'b_loss': b_value,
                'bidir_dist': e['dist_km'],
                'is_break': False, 'is_broke': False, 'is_bend': True,
                'is_bfill': False, 'is_a_only': False, 'is_b_only': False,
                'is_flagged': True, 'event_source': 'bend_standalone',
                'bend_severity': _bend_severity(bidir),
                'closure_offset_m': float(offset_m) if not target_is_phantom else 0.0,
                'event_type': e['type'],
                'label': label,
                '_b_source': b_source,
            }
    return new_results


def _estimate_helix_halfspread(splices, fibers_a):
    """Per-span helix half-spread (a small fraction, e.g. 0.004 = 0.4%) from the
    cross-fiber variation in the per-fiber length factor.

    Fits, per fiber, a line through its (closure_km, this-fiber's event_km) pairs
    across the splice closures; the slope is the fiber's length factor (helical
    lay + per-tube length).  Returns half the robust (p10–p90) spread of those
    slopes, clamped to HELIX_HALFSPREAD_MAX.  A consensus cluster within
    (distance × this) of a closure is that closure's helix-drifted tail, not a
    bend.  Returns 0.0 when it can't be estimated (caller keeps the fixed gate)."""
    centers = [sp.get('position_km_refined', sp['position_km'])
               for sp in splices if sp.get('column_kind') == 'splice']
    if len(centers) < 4:
        return 0.0
    slopes = []
    for _fnum, r in fibers_a.items():
        evs = r.get('events', [])
        xs, ys = [], []
        for c in centers:
            p = _per_fiber_splice_km(evs, c, search_window_km=0.35)
            if p is not None:
                xs.append(c)
                ys.append(p)
        if len(xs) >= 5:
            try:
                slopes.append(float(np.polyfit(xs, ys, 1)[0]))
            except Exception:
                pass
    if len(slopes) < 20:
        return 0.0
    s = np.array(slopes)
    half = (float(np.percentile(s, 90)) - float(np.percentile(s, 10))) / 2.0
    return max(0.0, min(HELIX_HALFSPREAD_MAX, half))


def _cluster_helix_residuals_m(cluster, fibers_a, splice_kms, nearest_col_km):
    """Per-fiber residual (metres) between each cluster fiber's event and where
    ITS OWN length model predicts a splice at the nearest closure.

    For each fiber in the cluster, fit a line through its events at the OTHER
    splice columns (excluding the nearest), predict its splice km at
    ``nearest_col_km``, and compare to the fiber's cluster event.  SMALL
    residuals mean the cluster IS the fibers' helix-drifted splice (the fiber's
    slope already explains the offset) — a SPLICE tail, not a bend.  A genuine
    bend sits where the fiber's helix model does NOT predict a splice (large
    residual).  Cable-agnostic: each fiber uses its own slope, no global rate.
    Needs >=5 other closures per fiber to fit; fibers with fewer are skipped
    (so a span with too few closures yields no residuals → caller keeps the
    bend, the conservative direction)."""
    out = []
    for tup in cluster:
        a_km, fnum = tup[0], tup[1]
        evs = [float(ev['dist_km']) for ev in fibers_a.get(fnum, {}).get('events', [])
               if not ev.get('is_end')]
        xs, ys = [], []
        for s in splice_kms:
            if abs(s - nearest_col_km) < 0.30:
                continue                       # exclude the cluster's own closure
            cand = [k for k in evs if abs(k - s) < 1.5]
            if cand:
                xs.append(s)
                ys.append(min(cand, key=lambda k: abs(k - s)))
        if len(xs) >= 5:
            slope, intercept = np.polyfit(xs, ys, 1)
            pred = slope * nearest_col_km + intercept
            out.append(abs(a_km - pred) * 1000.0)
    return out


def _event_explained_as_splice(fnum, event_km, splice_kms, fibers_a,
                               tol_m=HELIX_SPLICE_TOL_M,
                               win_km=0.6, min_fit=3, consensus_eof=None):
    """Account-then-flag core test (the CORRECT model, per the boss's reasoning):
    every fiber is spliced at every closure, so before an off-grid event is
    flagged as a SEPARATE feature, ask whether it is simply THIS fiber's own
    closure splice, drifted out by helix.

    Build the fiber's length model from its events at the OTHER closures
    (event_km vs closure_km, leave-one-out on the nearest closure), predict where
    ITS splice falls at that nearest closure, and accept the event as that splice
    only when it lands within a FIXED tolerance of that per-fiber prediction.

    The tolerance is FIXED, not scaled by the fiber's fit spread.  Drift-scaling
    was tried and FALSIFIED on the Seattle ground truth: real bends produce noisy
    per-fiber fits (σ 450-570 m), so a σ-scaled window swallowed them (z=res/σ runs
    BACKWARDS — clean drifted splices score higher than noisy real bends).  With a
    fixed window, only a clean per-fiber prediction (event sits right where the
    fiber's helix model puts its splice) folds; a noisy fit yields a large residual
    → stays flagged (the safe, never-hide-a-real-bend direction).  Empirically
    separates HOWLAN's drifted mid-span splices (<50 m) from every real feature
    (Seattle bends 144-670 m).  NOTE: it does NOT separate the HOWLAN far-end ref
    (~140 m) from real bends (~144 m) — that one is genuinely ambiguous and
    correctly stays flagged for the tech/boss to adjudicate.

    Returns True  → the event IS the fiber's drifted splice (ACCOUNT for it,
                    no separate column);
            False → unexplained, a genuine additional event (FLAG it).
    Falls back to False (flag — the conservative direction) when the model can't
    be fit (< ``min_fit`` other closures with a paired event)."""
    if not splice_kms:
        return False
    C = min(splice_kms, key=lambda s: abs(s - event_km))

    # EOF-anchored fold for the LAST closure (helix non-linearity at the cable
    # end; see HELIX_EOF_TOL_M / HELIX_EOF_MIN_SHORT_KM).  The splice→end-of-fiber
    # distance is preserved per fiber, so a fiber reading short at its end reads
    # its last splice short by the same amount.  Fires ONLY for fibers reading
    # materially short (the helix signature), so it is inert on non-helix spans
    # and cannot fold a mid-span bend.
    if consensus_eof is not None and abs(C - max(splice_kms)) < 1e-9:
        ends = [float(e['dist_km'])
                for e in fibers_a.get(fnum, {}).get('events', []) if e.get('is_end')]
        if ends:
            fiber_eof = max(ends)
            if (consensus_eof - fiber_eof) >= HELIX_EOF_MIN_SHORT_KM:   # reads short → helix
                pred_eof = fiber_eof - (consensus_eof - C)             # preserve splice→EOF gap
                if abs(event_km - pred_eof) * 1000.0 <= HELIX_EOF_TOL_M:
                    return True

    evs = [float(e['dist_km'])
           for e in fibers_a.get(fnum, {}).get('events', [])
           if not e.get('is_end') and float(e['dist_km']) >= 0.5]
    xs, ys = [], []
    for s in splice_kms:
        if abs(s - C) < 0.30:
            continue                       # leave-one-out on the target closure
        cand = [k for k in evs if abs(k - s) < win_km]
        if cand:
            xs.append(s)
            ys.append(min(cand, key=lambda k: abs(k - s)))
    if len(xs) < min_fit:
        return False                       # can't model the fiber → flag (safe)
    slope, intercept = np.polyfit(xs, ys, 1)
    pred = slope * C + intercept
    return abs(event_km - pred) * 1000.0 <= tol_m


def flag_consensus_bends(all_results, fibers_a, fibers_b, splices, total_span_a,
                         min_fibers=2, bend_threshold=None):
    """ADDITIVE, never-demote BINARY bend classifier.

    The per-fiber length-model + narrow-LSA bend test silently DROPS real bends:
    the closure clusterer merges an off-splice bend zone into a nearby splice
    column (Seattle 84.82 km → Splice 12 @84.59), and Test-2 `_narrow_lsa_loss`
    returns None on fibers whose own splice confounds the fit (F338 @98.40).  This
    pass recovers them from the cross-fiber POPULATION: at any OFF-grid position
    where >= min_fibers fibers carry a co-located bidirectional event reaching
    BEND_THRESHOLD (bidir from each fiber's OWN b_span, not the cable-wide span),
    and whose CLUSTER (median) position is more than CONSENSUS_BEND_OFFGRID_KM
    from the nearest closure-grid column, HARD-FLAG the qualifying fibers as bends
    (is_flagged=True, is_borderline=False, event_source/category 'bend').

    Per the no-review-cells rule, bends are flagged outright, not queued for
    review.  The cluster-level off-grid gate excludes clusters the boss
    attributes to a nearby splice (Seattle 100.46 km, ~70 m off a splice).  This
    pass only ADDS uncovered cells keyed by a synthetic splice_idx; it never
    touches or demotes an existing flagged/borderline cell.  Returns new
    (fnum, synthetic_si) → result-dict entries to merge.
    """
    bt = BEND_THRESHOLD if bend_threshold is None else bend_threshold
    splice_kms = [sp.get('position_km_refined', sp['position_km'])
                  for sp in splices if sp.get('column_kind') == 'splice']
    closure_centers = [(si, sp.get('position_km_refined', sp['position_km']))
                       for si, sp in enumerate(splices)]

    def offgrid(km):
        return all(abs(km - s) > CLOSURE_MATCH_KM for s in splice_kms)

    # Per-span helix drift rate — scales the cluster-level off-grid gate below
    # so a closure's helix-drifted tail isn't mis-flagged as a separate bend.
    helix_halfspread = _estimate_helix_halfspread(splices, fibers_a)

    # 1. Off-grid co-located A+B bend candidates (own-b_span mirror).
    cands = []   # (a_km, fnum, a_event, a_idx, bidir, a_loss, b_loss)
    for fnum, ra in fibers_a.items():
        rb = fibers_b.get(fnum) if fibers_b else None
        if rb is None:
            continue
        b_ends = [be for be in rb.get('events', []) if be.get('is_end')]
        if not b_ends:
            continue
        b_span_own = b_ends[0]['dist_km']
        b_evs = [be for be in rb.get('events', [])
                 if not be.get('is_end') and be['dist_km'] >= 1.0]
        for ai, e in enumerate(ra.get('events', [])):
            if e.get('is_end') or e.get('is_reflective') or e['dist_km'] < 1.0:
                continue
            a_loss = e.get('splice_loss') or 0.0
            # Gate on the BIDIR average below, not the A side alone — bends are
            # often asymmetric (e.g. F361: A 0.09 + B 0.174 → bidir 0.132).  Only
            # require a positive A loss here (drop gainers / off-grid only).
            if a_loss <= 0.0 or not offgrid(e['dist_km']):
                continue
            best = None
            for be in b_evs:
                d = abs((b_span_own - be['dist_km']) - e['dist_km'])
                if d < 0.300 and (best is None or d < best[1]):
                    best = (be, d)
            if best is None:
                continue
            b_loss = best[0].get('splice_loss') or 0.0
            bidir = (a_loss + b_loss) / 2.0
            if bidir >= bt:
                cands.append((e['dist_km'], fnum, e, ai,
                              round(bidir, 4), round(a_loss, 4), round(b_loss, 4),
                              best[0]))
    if not cands:
        return {}

    # 2. Gap-cluster; keep clusters reaching min_fibers distinct fibers.
    cands.sort()
    clusters, cur = [], [cands[0]]
    for c in cands[1:]:
        if c[0] - cur[-1][0] <= 0.20:
            cur.append(c)
        else:
            clusters.append(cur); cur = [c]
    clusters.append(cur)

    new_results = {}
    for cl in clusters:
        if len({c[1] for c in cl}) < min_fibers:
            continue
        cluster_km = float(np.median([c[0] for c in cl]))
        nfib = len({c[1] for c in cl})
        # CLUSTER-LEVEL OFF-GRID GATE: only emit a hard bend when the cluster's
        # robust (median) position is MORE than CONSENSUS_BEND_OFFGRID_KM from
        # the nearest closure-grid column.  The per-candidate offgrid() check
        # above uses each event's own (noisy) dist_km; the cluster median is
        # the stable position.  This excludes clusters the boss attributes to a
        # nearby splice (Seattle 100.46 km, 84 m from Splice 20 → the splice),
        # while keeping the six real bends (all >=127 m from any column).
        if closure_centers:
            # Launch-closure guard: a consensus 'bend' at/just past the FIRST
            # closure (the launch closure) is treated as that closure's splice,
            # not a bend.  This is where the tech sets the first-ribbon reference
            # distance and where the OTDR launch dead zone lives, so near-launch
            # positions are unreliable AND helix is ~0 there (so the residual
            # test below can't discriminate).  Seattle-safe — its real bends are
            # all >=24 km.  (HOWLAN 1.72 km, 2 fibers, residual 100 m: the boss
            # calls it Splice 1 by assuming the first-ribbon distance.)
            if splice_kms and cluster_km <= min(splice_kms) + CLOSURE_MATCH_KM:
                continue
            nearest_col = min(abs(cc[1] - cluster_km) for cc in closure_centers)
            # Helix-aware tolerance: a cluster within (distance × the span's
            # helix half-spread × K) of a closure is that closure's helix-drifted
            # tail, not a bend.  Never below the fixed floor (preserves the
            # short-span behaviour + Seattle's six real bends, which sit far
            # past what helix explains at their distance).
            helix_tol = max(CONSENSUS_BEND_OFFGRID_KM,
                            cluster_km * helix_halfspread * HELIX_TOL_K)
            if nearest_col <= helix_tol:
                continue
            # Principled discriminator: if the cluster's fibers sit where THEIR
            # OWN helix slope predicts a splice at the nearest closure (small
            # per-fiber residual), this is that closure's helix-drifted SPLICE
            # tail, not a bend.  A real bend is NOT at the helix-predicted
            # position (large residual).  Cable-agnostic; ground-truth-separated
            # (HOWLAN tails <100 m vs Seattle real bends >400 m).
            nearest_col_km = min((cc[1] for cc in closure_centers),
                                 key=lambda s: abs(s - cluster_km))
            resids = _cluster_helix_residuals_m(cl, fibers_a, splice_kms, nearest_col_km)
            if resids and float(np.median(resids)) < HELIX_RESIDUAL_BEND_M:
                continue
        for a_km, fnum, e, ai, bidir, a_loss, b_loss, _be in cl:
            # Skip when an existing pass already surfaced this fiber near here —
            # NEVER demote or duplicate; this pass only ADDS uncovered cells.
            if any(k[0] == fnum
                   and abs((all_results[k].get('bidir_dist')
                            or all_results[k].get('dist_km') or -9) - a_km) < 0.30
                   and (all_results[k].get('is_flagged')
                        or all_results[k].get('is_borderline'))
                   for k in all_results if k[0] == fnum):
                continue
            best_si = min(range(len(closure_centers)),
                          key=lambda i: abs(closure_centers[i][1] - a_km)) \
                if closure_centers else 0
            key = (fnum, 90000 + ai)    # synthetic si, offset to avoid collision
            if key in all_results or key in new_results:
                continue
            new_results[key] = {
                'fiber': fnum, 'splice_idx': best_si,
                'bidir_loss': bidir, 'a_loss': a_loss, 'b_loss': b_loss,
                'bidir_dist': a_km,
                'is_break': False, 'is_broke': False, 'is_bend': True,
                'is_bfill': False, 'is_a_only': False, 'is_b_only': False,
                # BINARY CLASSIFIER: off-grid consensus bends are HARD flags,
                # not review queue (per the no-review-cells rule).  category
                # is 'bend' via is_bend; counted in n_flagged.
                'is_flagged': True, 'is_borderline': False,
                'event_source': 'bend',
                'bend_severity': _bend_severity(bidir),
                'closure_offset_m': 0.0,
                'event_type': e.get('type', ''),
                'label': f"{fnum} bend {_format_loss(bidir)} bidi "
                         f"({nfib} fib @{cluster_km:.2f}km)",
                '_b_source': 'consensus',
            }
    return new_results


# ═══════════════════════════════════════════════════════════════════════
#  STEP 4b' — May 12 addition: bidirectional ghost-reflection scan
#  (mid-span 1F reflective events with near-zero loss that show in BOTH
#   directions at the same mirrored km — e.g. F64 at km 40.27 on Long
#   Shots).  These slip past the loss-threshold gates of every other
#   pass but are real physical features (faint connector, mech splice,
#   ghost from a downstream reflector) that the report should call out.
# ═══════════════════════════════════════════════════════════════════════

# Mirror tolerance for the bidirectional confirmation.  Real physical
# features mirror within ±50 m on well-OTDR'd cables; 100 m gives
# breathing room for fibers with slightly different total lengths.
GHOST_REFL_MIRROR_TOL_KM = 0.100
# Maximum absolute splice loss either direction may show — anything
# above this is no longer a "no-loss" event and the other passes will
# pick it up under their own rules.
GHOST_REFL_MAX_LOSS_DB   = 0.030


def scan_bidir_ghost_reflections(fibers_a, fibers_b, splices, existing_results,
                                 total_span_a, closure_match_km=None):
    """Flag mid-span 1F reflective events that appear in BOTH A and B at
    mirror-matched positions with near-zero loss.  Filters:
      • Event must be 1F (is_reflective=True), not is_end.
      • Inside cable proper: 1 km from launch, LAUNCH_FIBER_MAX from EOL.
      • Not within CLOSURE_MATCH_KM of any known splice closure (otherwise
        it's just a reflective splice already accounted for).
      • |splice_loss| <= GHOST_REFL_MAX_LOSS_DB in BOTH directions.
      • B-direction has a matching 1F within GHOST_REFL_MIRROR_TOL_KM of
        the mirror km (b_span - a_km).
    Returns dict (fnum, splice_idx) -> result, mirroring the structure of
    the other scan_* functions.  splice_idx is the nearest known closure
    (display anchor only).
    """
    cm = CLOSURE_MATCH_KM if closure_match_km is None else closure_match_km
    new_results = {}
    closure_centers = [(si, sp.get('position_km_refined', sp['position_km']))
                       for si, sp in enumerate(splices)]

    for fnum, ra in fibers_a.items():
        rb = fibers_b.get(fnum)
        if rb is None:
            continue
        # Need a B-direction EOL to mirror against
        b_ends = [e for e in rb.get('events', []) if e.get('is_end')]
        if not b_ends:
            continue
        b_span = b_ends[0]['dist_km']
        a_ends = [e for e in ra.get('events', []) if e.get('is_end')]
        if not a_ends:
            continue
        a_eof = a_ends[0]['dist_km']

        # Pre-index B reflective non-end events for fast lookup.
        b_refl_events = [
            be for be in rb.get('events', [])
            if (be.get('is_reflective') or str(be.get('type', '')).startswith('1F'))
               and not be.get('is_end')
               and be['dist_km'] >= 1.0
               and be['dist_km'] <= (b_span - LAUNCH_FIBER_MAX)
               and abs(be.get('splice_loss') or 0.0) <= GHOST_REFL_MAX_LOSS_DB
        ]
        if not b_refl_events:
            continue

        for ae in ra.get('events', []):
            if ae.get('is_end'):
                continue
            if not (ae.get('is_reflective') or str(ae.get('type', '')).startswith('1F')):
                continue
            a_km = ae['dist_km']
            if a_km < 1.0:
                continue                  # launch zone
            if a_km > (a_eof - LAUNCH_FIBER_MAX):
                continue                  # tailbox zone
            if abs(ae.get('splice_loss') or 0.0) > GHOST_REFL_MAX_LOSS_DB:
                continue                  # has loss — other passes own it

            # Skip if at a known splice closure (already classified there)
            at_closure = any(abs(a_km - c) <= cm for _, c in closure_centers)
            if at_closure:
                continue

            # Mirror-match in B direction
            target_b_km = b_span - a_km
            be_match = None
            best_d = float('inf')
            for be in b_refl_events:
                d = abs(be['dist_km'] - target_b_km)
                if d < best_d and d <= GHOST_REFL_MIRROR_TOL_KM:
                    best_d = d
                    be_match = be
            if be_match is None:
                continue

            # Anchor to nearest known closure for display
            nearest_si = None
            nearest_d = float('inf')
            for si, c in closure_centers:
                d = abs(a_km - c)
                if d < nearest_d:
                    nearest_d = d
                    nearest_si = si
            if nearest_si is None:
                continue

            # Skip if already accounted for at this anchor closure
            if (fnum, nearest_si) in existing_results or (fnum, nearest_si) in new_results:
                continue

            a_refl = ae.get('reflection') or 0.0
            b_refl = be_match.get('reflection') or 0.0
            label = (f"{fnum} ref @ {a_km:.2f}km "
                     f"(refl {a_refl:.0f}/{b_refl:.0f}dB bidir)")
            new_results[(fnum, nearest_si)] = {
                'fiber': fnum, 'splice_idx': nearest_si,
                'bidir_loss': round(((ae.get('splice_loss') or 0.0)
                                     + (be_match.get('splice_loss') or 0.0)) / 2.0, 4),
                'a_loss': ae.get('splice_loss') or 0.0,
                'b_loss': be_match.get('splice_loss') or 0.0,
                'bidir_dist': a_km,
                'is_break': False, 'is_broke': False, 'is_bend': False,
                'is_ref': True,
                'is_bfill': False, 'is_a_only': False, 'is_b_only': False,
                'is_flagged': True,
                'event_source': 'ref_bidir_ghost',
                'event_type': ae.get('type'),
                'label': label,
                'fresnel': a_refl,
            }
    return new_results


# ═══════════════════════════════════════════════════════════════════════
#  STEP 4b'' — May 12 addition: EXFO "merged reflective" event detection
#  EXFO classifies some mid-span events as type='0F' with is_reflective=
#  False, yet still records a non-zero negative reflection value.  Its
#  viewer labels them "Merged Reflective; Non-reflective" — the event is
#  both a small splice loss AND a faint connector/mech-splice reflection.
#  Our earlier passes all skip these because they trust the 0F type
#  and the is_reflective=False flag.  F541 ELMMIL @ km 62.51 is the
#  canonical example (refl=-83 dB, loss=0.096 dB, type=0F).
# ═══════════════════════════════════════════════════════════════════════

ECHO_PARENT_TOL_KM = 0.7   # how close to cand_km/n a parent reflector must sit

# Mid-span reflectance thresholds — the OTDR settings panel's "Mid-span reflectance"
# row (overridable per customer profile via --overrides; read as module globals so
# setattr() in run_splicereport takes effect).  A mid-span reflective event is
# flagged only when its reflectance is at least MIDSPAN_REFL_WARN_DB (the floor);
# >= MIDSPAN_REFL_FAIL_DB → FAIL, between floor and fail → WARN.  Signed dB,
# less-negative = stronger reflection.
MIDSPAN_REFL_FAIL_DB = -50.0
MIDSPAN_REFL_WARN_DB = -80.0

def _is_likely_echo(cand_km, cand_refl, refl_events, tol_km=ECHO_PARENT_TOL_KM):
    """True if the reflective event at ``cand_km`` is most likely a bounce ECHO
    (ghost) of a STRONGER reflector upstream — not a real feature.

    Physics: a strong reflector at distance D bounces light off the launch and
    back, so the OTDR draws a phantom reflection at ~n·D (n=2,3,4…) that is always
    WEAKER than D and has nothing physically there.  So: if a reflective event
    sits at an integer fraction of this candidate's distance (cand_km/n) and is
    STRONGER (reflectance closer to 0), this candidate is probably its echo.

    ``refl_events`` = [(km, reflection_dB)] for the SAME fiber, own-frame km;
    reflectances are signed dB (less-negative = stronger).  Conservative — only
    fires when a clearly-stronger parent exists at the predicted echo distance, so
    a genuine isolated reflection (no upstream parent, e.g. TOPMIL0195 @30.92) is
    kept."""
    if cand_refl is None:
        return False
    for n in (2, 3, 4):
        parent_km = cand_km / n
        if parent_km < 0.5:
            continue
        for k, rf in refl_events:
            if rf is None:
                continue
            if abs(k - parent_km) <= tol_km and rf > cand_refl:
                return True
    return False


def scan_merged_reflective_events(fibers_a, fibers_b, splices,
                                   existing_results, total_span_a,
                                   closure_match_km=None):
    """Catch any mid-span REFLECTIVE event that slips past every other pass —
    both the EXFO '0F Merged Reflective; Non-reflective' case AND genuine 1F
    reflective events.  The A-direction passes only see A-frame reflections, so a
    B-only reflection (e.g. TOPMIL0195 @30.92 km, the boss's flagged event) is
    invisible to them; this pass scans BOTH directions and surfaces it.
    Criteria for either direction:
      • Mid-span (≥ LAUNCH_FIBER_MAX from launch and EOL) — so the launch and
        far-end connectors are NOT re-flagged
      • refl < 0 (a real negative reflectance measurement) — the ONLY signal we
        gate on.  Loss is NOT part of the gate:
        a reflective event is a reflective event regardless of how
        large or small the accompanying splice-loss measurement is.
      • Trace continues past (EOL at least 1 km after the event)
      • Not already flagged elsewhere (existing_results dedupe)
    Anchored to the nearest closure for ribbon-grid display.  Surfaces
    as is_ref=True (deep-orange 'ref' tier).
    """
    cm = CLOSURE_MATCH_KM if closure_match_km is None else closure_match_km
    new_results = {}
    closure_centers = [(si, sp.get('position_km_refined', sp['position_km']))
                       for si, sp in enumerate(splices)]

    def _classify_one(fibers, frame_to_a_km, dir_label):
        for fnum, r in fibers.items():
            evs = r.get('events') or []
            end_evt = next((e for e in evs if e.get('is_end')), None)
            if end_evt is None:
                continue
            eof_km = end_evt['dist_km']
            # All reflective events on this fiber (own-frame km), for the echo guard.
            refl_events = [(float(x['dist_km']), x.get('reflection')) for x in evs
                           if not x.get('is_end')
                           and (x.get('is_reflective') or str(x.get('type','')).startswith('1F'))
                           and x.get('reflection') is not None]
            for e in evs:
                if e.get('is_end'):
                    continue
                refl = e.get('reflection')
                if refl is None or refl >= 0:
                    continue            # need a real negative reflectance
                # Surface BOTH the EXFO 0F "Merged Reflective; Non-reflective"
                # case AND genuine 1F reflective events.  Skipping 1F (the old
                # behavior) lost B-only reflections like TOPMIL0195 @30.92 km that
                # the A-direction passes never see; the existing_results dedupe
                # below still prevents double-flagging A-side refs already caught.
                # NO loss filter — a reflective event is a reflective
                # event regardless of the splice-loss magnitude.
                # Mid-span only
                if e['dist_km'] < LAUNCH_FIBER_MAX:
                    continue
                if e['dist_km'] > (eof_km - LAUNCH_FIBER_MAX):
                    continue
                # Echo/ghost guard: skip if a STRONGER reflector sits at an
                # integer fraction of this distance (its 2x/3x bounce-echo).
                if _is_likely_echo(e['dist_km'], refl, refl_events):
                    continue
                # Mid-span reflectance threshold (OTDR-panel editable): flag only
                # reflections at/above the warn floor; classify FAIL vs WARN.
                if refl < MIDSPAN_REFL_WARN_DB:
                    continue
                _sev = "FAIL" if refl >= MIDSPAN_REFL_FAIL_DB else "WARN"
                # Translate to A-frame for closure matching / dedup
                a_km = frame_to_a_km(e['dist_km'], eof_km)
                # The trace must clearly continue past — simple check
                # rather than _trace_continues_past which requires
                # another non-end event past the candidate (often
                # absent on a clean span past the last splice).
                if (eof_km - e['dist_km']) < 1.0:
                    continue
                # Anchor to nearest closure for display.  Unlike the
                # bidir-ghost scan, we do NOT skip events that sit at
                # a known closure — those are precisely the ones we
                # want to surface (Pass 1 missed them because the
                # bidir loss came out too small even though one
                # direction shows a measurable reflection).  The
                # existing_results dedupe below prevents double-
                # flagging anything Pass 1 already caught.
                nearest_si, best_d = None, float('inf')
                for si, c in closure_centers:
                    d = abs(a_km - c)
                    if d < best_d:
                        best_d, nearest_si = d, si
                if nearest_si is None:
                    continue
                if ((fnum, nearest_si) in existing_results
                        or (fnum, nearest_si) in new_results):
                    continue
                loss = e.get('splice_loss') or 0.0
                _kind = "1F" if (e.get('is_reflective') or str(e.get('type','')).startswith('1F')) else "merged"
                label = (f"{fnum} ref @ {a_km:.2f}km "
                         f"(refl {refl:.0f}dB {_sev} {_kind}, {dir_label}-side)")
                new_results[(fnum, nearest_si)] = {
                    'fiber': fnum, 'splice_idx': nearest_si,
                    'bidir_loss': loss, 'a_loss': loss if dir_label=='A' else None,
                    'b_loss': loss if dir_label=='B' else None,
                    'bidir_dist': a_km,
                    'is_break': False, 'is_broke': False, 'is_bend': False,
                    'is_ref': True,
                    'is_bfill': False, 'is_a_only': False, 'is_b_only': False,
                    'is_flagged': True,
                    'event_source': 'ref_merged_exfo',
                    'event_type': e.get('type'),
                    'label': label,
                    'fresnel': refl,
                }

    # A-direction frame = A's own km
    _classify_one(fibers_a, lambda km, _eof: km, 'A')
    # B-direction frame mirrors to A-frame via b_span - b_km
    _classify_one(fibers_b, lambda km, eof: eof - km, 'B')
    return new_results


# ═══════════════════════════════════════════════════════════════════════
#  STEP 4c — APRIL 23 revision: restricted past-break B-fill scan
#  (replaces the blanket Pass-2 B scan with a targeted past-break pass)
# ═══════════════════════════════════════════════════════════════════════

def scan_b_past_breaks(fibers_a, fibers_b, splices, threshold, existing_results,
                       total_span_a):
    """For fibers that are BROKE on the A side, scan the B direction in the
    zone PAST the A-side break (so closer to the B-end than the break
    position) for events we haven't seen yet — these populate B-fill cells.

    This is the narrower B-direction usage the tech asked for: 'We use B
    direction only to look after the breaks on A sides to see if there are
    other events we are missing.'"""
    new_results = {}
    seen_keys = set(existing_results.keys())

    # Cache A break positions (km, in A-frame)
    a_break_km = {}
    for fnum, ra in fibers_a.items():
        end = [e for e in ra['events'] if e.get('is_end')]
        if not end:
            continue
        eof = end[0]['dist_km']
        if eof < total_span_a - END_REGION_KM:
            a_break_km[fnum] = eof

    if not a_break_km:
        return new_results

    for fnum, brk_km in a_break_km.items():
        rb = fibers_b.get(fnum)
        if rb is None:
            continue
        b_end_events = [e for e in rb['events'] if e.get('is_end')]
        if not b_end_events:
            continue
        b_span = b_end_events[0]['dist_km']
        # Use the CABLE's total span (not B's measured span) when
        # mirroring B-frame distances to A-frame.  When B itself is
        # broken short, b_span < total_span_a and the simple
        # b_span − b_dist formula gives wildly wrong A-frame
        # positions for every B event — so B-fill misses the very
        # cases it's meant to handle.  Anchoring on total_span_a
        # restores correct mirroring even when both directions are
        # broken.
        mirror_anchor = total_span_a if total_span_a > 0 else b_span

        # Scan B events whose A-frame position is GREATER than brk_km
        for e in rb['events']:
            if e.get('is_end'):
                continue
            if e['dist_km'] < 1.0:
                continue
            a_frame = mirror_anchor - e['dist_km']
            if a_frame <= brk_km + 0.2:   # 200m buffer past the break
                continue
            b_loss = e.get('splice_loss') or 0.0
            # Single-direction B-fill: stricter gate (0.250 default), no
            # averaging or /2 estimate.  The raw B loss alone must clear
            # SINGLE_DIR_THRESHOLD — no opposite-side confirmation possible.
            if abs(b_loss) < SINGLE_DIR_THRESHOLD:
                continue

            # Find nearest splice position (A-frame)
            nearest_si, nearest_d = None, float('inf')
            for si, sp in enumerate(splices):
                c = sp.get('position_km_refined', sp['position_km'])
                d = abs(a_frame - c)
                if d < nearest_d:
                    nearest_d = d
                    nearest_si = si
            if nearest_si is None:
                continue
            # Keep loose tolerance for B-fill (the whole point is bridging a gap)
            if nearest_d > POSITION_TOL:
                continue

            key = (fnum, nearest_si)
            if key in seen_keys or key in new_results:
                continue

            b_loss_abs = abs(b_loss)
            loss_str = _format_loss(b_loss_abs)
            new_results[key] = {
                'fiber': fnum, 'splice_idx': nearest_si,
                'bidir_loss': b_loss_abs, 'a_loss': None,
                'b_loss': b_loss, 'bidir_dist': a_frame,
                'is_break': False, 'is_broke': False, 'is_bend': False,
                'is_bfill': True,
                'is_a_only': False, 'is_b_only': False,
                'is_flagged': True, 'event_source': 'bfill',
                'event_type': e['type'],
                'label': f"{fnum} {loss_str} (B-fill)",
            }

    return new_results


# ═══════════════════════════════════════════════════════════════════════
#  STEP 4d — Symmetric B-side broke detection
# ═══════════════════════════════════════════════════════════════════════

def scan_b_side_breaks(fibers_a, fibers_b, splices, existing_results,
                        total_span_a):
    """Catch fibers that terminate mid-span on the B trace but whose A
    trace passes through (B-only breaks).  These represent a real break
    that one direction's optics can punch through but the other can't —
    e.g. Lagrande↔Durkey F1-F12 where A reaches 23.96 km but B
    terminates at 13.91 km (= 90.46 km in A-frame).

    Boss reports a DAMAGE column at the mirrored A-frame position; we
    were missing it because Pass 1's broke detection walks only the A
    direction.

    Returns dict {(fnum, nearest_si): result_dict} of new broke entries.
    """
    new_results = {}
    seen_keys = set(existing_results.keys())

    # Zero closures discovered (e.g. a span below MIN_POP_SPLICE): there is no
    # closure to attribute a B-side break to, and the nearest-closure min() over
    # range(len(splices)) below would raise "min() arg is an empty sequence" and
    # kill the whole report.  Mirror scan_a_standalone_events' zero-closure guard
    # and return nothing to add.
    if not splices:
        return new_results

    # Cache A end positions to detect A-broken fibers (skip those — Pass 1
    # already logged them; double-flagging would create duplicate columns).
    a_eof = {}
    for fnum, ra in fibers_a.items():
        end = [e for e in ra['events'] if e.get('is_end')]
        a_eof[fnum] = end[0]['dist_km'] if end else None

    # Cable total span in B's frame (median of top-quarter B-trace ends)
    eof_b = []
    for fnum, rb in fibers_b.items():
        end = [e for e in rb['events'] if e.get('is_end')]
        if end:
            eof_b.append(end[0]['dist_km'])
    if not eof_b:
        return new_results
    eof_b.sort()
    total_span_b = float(np.median(eof_b[int(len(eof_b)*0.75):]))

    for fnum, rb in fibers_b.items():
        end = [e for e in rb['events'] if e.get('is_end')]
        if not end:
            continue
        b_eof = end[0]['dist_km']

        # B-trace must terminate mid-span (≥ END_REGION_KM short of B's far end).
        if b_eof >= total_span_b - END_REGION_KM:
            continue
        if b_eof < 1.0:
            continue

        # Mirror B's eof into A-frame so we attribute to the right closure.
        a_frame_break_km = total_span_a - b_eof if total_span_a > 0 else b_eof

        # If A is also broken, the A-side broke entry already lives at the
        # A-frame position of the A-break — typically a different km than
        # the B-side break.  Only suppress if the two breaks mirror to
        # nearly the same A-frame position (within END_REGION_KM); otherwise
        # log BOTH (boss does the same — F1-F12 in Lagrande↔Durkey are
        # listed as broken at both 23.95 km on A and 90.46 km on B).
        a_end = a_eof.get(fnum)
        if (a_end is not None and 0 < a_end < total_span_a - END_REGION_KM
                and abs(a_end - a_frame_break_km) < END_REGION_KM):
            continue

        # Pick nearest closure by A-frame km (same convention as A-side).
        nearest_si = min(range(len(splices)),
                         key=lambda i: abs(splices[i]['position_km'] - a_frame_break_km))
        key = (fnum, nearest_si)
        # When Pass 1 left a dead-zone marker at this cell — or Pass 2c
        # speculatively wrote a B-fill bridge there — the B-side broke
        # supersedes it.  A concrete broke entry is the physical truth
        # ("this fiber is broken"); DZ is "neither trace saw it" and
        # B-fill is "B saw it past the A break."  Both are weaker
        # signals than a confirmed B-side termination.  Without this
        # override, F1-F12 on Lagrande↔Durkey rendered as blue B-fill
        # cells at km 90.46 instead of red BROKE — bug #3 fix 2026-06-13.
        prior = existing_results.get(key)
        if prior is not None and not (prior.get('is_dead_zone')
                                       or prior.get('is_bfill')):
            continue

        label = f"{fnum} broke@{a_frame_break_km:.1f}k (B-only)"
        new_results[key] = {
            'fiber': fnum, 'splice_idx': nearest_si,
            'bidir_loss': None, 'a_loss': None, 'b_loss': None,
            'bidir_dist': a_frame_break_km,
            'is_break': False, 'is_broke': True, 'is_bend': False,
            'is_bfill': False, 'is_dead_zone': False,
            'is_a_only': False, 'is_b_only': False,
            'is_flagged': True, 'event_source': 'broke_b',
            'event_type': 'BROKE_B', 'label': label,
        }

    return new_results


# ═══════════════════════════════════════════════════════════════════════
#  STEP 4.5 — DISTRIBUTED SECTION-LOSS DETECTOR  (additive, fully isolated)
# ═══════════════════════════════════════════════════════════════════════
#
#  The event-based report above is blind to a *stretch* of fiber carrying
#  elevated attenuation (a stressed / aged / water-affected segment with a
#  higher dB/km but NO discrete event).  This pass surfaces those stretches.
#
#  It is deliberately kept OUTSIDE the event/flag machinery:
#    • It reads the A-direction backscatter traces directly and emits its own
#      `distributed_loss` records with their OWN count (`n_distributed_loss`).
#    • It NEVER touches `all_results`, `splices`, the ribbon grid, or any
#      event-flag decision — so the historical n_flagged is untouched.
#
#  Ported verbatim from the validated standalone verify/section_loss.py
#  (SanDur 31 sections, Seattle 0).  A-DIRECTION ONLY: the census showed the
#  B direction explodes geometrically (fiber inside A's near-launch dead zone
#  reads elevated from B's far terminal — not a real degradation).

# ── tunable thresholds (mirror verify/section_loss.py) ──
DIST_SLOPE_EXCESS_DBKM = 0.05    # flag when a segment slope exceeds the per-fiber
                                 #   median by at least this much (dB/km)
DIST_MIN_RUN_KM        = 2.0     # ... sustained over a run at least this long
DIST_EVENT_GUARD_KM    = 0.2     # trim this off each interior side of a segment so
                                 #   the fit measures FIBER, not the discrete step
DIST_WINDOW_KM         = 1.0     # window length for the per-fiber median baseline
DIST_LAUNCH_GUARD_KM   = 6.0     # exclude < this from launch (dead zone + settling)
DIST_EOF_GUARD_KM      = 2.0     # exclude the last this-many km before EOF
DIST_SEVERE_LOSS_DB    = 1.0     # a loss event this big casts a recovery shadow
DIST_RECOVERY_GUARD_KM = 2.0     # a segment starting within this of a severe
                                 #   event is a recovery tail (excluded, not flagged)

# ── cross-fiber AGGREGATION thresholds (turn per-fiber sections into findings) ──
DIST_CLUSTER_GAP_KM    = 1.0     # per-fiber sections whose km-ranges overlap or sit
                                 #   within this of each other chain (transitively)
                                 #   into one candidate cable-wide region
MIN_FIBERS_FINDING     = 5       # a candidate region is reported as a FINDING only if
                                 #   it carries sections from at least this many distinct
                                 #   fibers — lone / scattered sections are dropped

_DIST_LIGHT_C = 299_792_458.0    # m/s


def _dist_trace_km_db(fiber_rec):
    """Map an A-direction fiber record's backscatter trace to (km, db).

    `km` is fiber distance per sample with the launch face at 0 km; `db` is the
    backscatter in dB (rises with distance = accumulated attenuation).  Uses the
    record's `_trace_offset_km` (set by the runner's Pass-0) so the trace frame
    matches the normalized events.  Returns (None, None) when the trace is
    unavailable.  Read-only — never mutates the record."""
    tr = fiber_rec.get('trace')
    if tr is None:
        return None, None
    tr = np.asarray(tr, dtype=float)
    if tr.size < 3:
        return None, None
    ior = _sor_ior_from_events(fiber_rec)
    res_m = _DIST_LIGHT_C * (fiber_rec.get('exfo_sampling_period') or 5e-8) / 2.0 / ior
    offset = fiber_rec.get('_trace_offset_km') or 0.0
    km = np.arange(tr.size) * res_m / 1000.0 - offset
    return km, tr


def _dist_local_slope(km, db, a_km, b_km):
    """Least-squares slope (dB/km) of db over [a_km, b_km]; None if too few pts."""
    mask = (km >= a_km) & (km <= b_km)
    x = km[mask]
    y = db[mask]
    if len(x) < 3:
        return None
    return float(np.polyfit(x, y, 1)[0])


def _dist_eof_km(events, km):
    """Distance of the end-of-fiber event, or the last sample if none flagged."""
    for e in events:
        if e.get('is_end'):
            return e['dist_km']
    return float(km[-1])


def _dist_median_slope(km, db, eof_km):
    """Median local slope across the usable interior — the per-fiber baseline.

    Sampled in DIST_WINDOW_KM windows over the launch-guarded, EOF-guarded
    interior.  The median is robust to the (few) elevated stretches we hunt and
    to splice steps, so it tracks the fiber's intrinsic dB/km."""
    lo = DIST_LAUNCH_GUARD_KM
    hi = eof_km - DIST_EOF_GUARD_KM
    slopes = []
    a = lo
    while a + DIST_WINDOW_KM <= hi:
        s = _dist_local_slope(km, db, a, a + DIST_WINDOW_KM)
        if s is not None:
            slopes.append(s)
        a += DIST_WINDOW_KM
    if not slopes:
        return None
    return float(np.median(slopes))


def _dist_severe_event_kms(events, eof_km):
    """Distances of loss events >= DIST_SEVERE_LOSS_DB inside the usable fiber."""
    out = []
    for e in events:
        if e.get('is_end'):
            continue
        loss = e.get('splice_loss')
        if loss is None:
            continue
        if loss >= DIST_SEVERE_LOSS_DB and DIST_LAUNCH_GUARD_KM <= e['dist_km'] <= eof_km:
            out.append(e['dist_km'])
    return out


def _dist_segment_bounds(events, eof_km):
    """Inter-event fiber segments [start_km, end_km] over the usable interior.

    Discrete (non-end) event distances become internal cut points (plus the
    launch / EOF guards), so each measured stretch lies BETWEEN events and a
    splice/connector step never masquerades as fiber attenuation.  The event
    guard is trimmed off each interior side."""
    lo = DIST_LAUNCH_GUARD_KM
    hi = eof_km - DIST_EOF_GUARD_KM
    cuts = sorted(e['dist_km'] for e in events
                  if not e.get('is_end') and lo < e['dist_km'] < hi)
    bounds = [lo] + cuts + [hi]
    segs = []
    for a, b in zip(bounds[:-1], bounds[1:]):
        sa = a + DIST_EVENT_GUARD_KM if a > lo else a
        sb = b - DIST_EVENT_GUARD_KM if b < hi else b
        if sb - sa >= DIST_MIN_RUN_KM:
            segs.append((sa, sb))
    return segs


def _dist_scan_fiber(km, db, events, eof_km):
    """Return (median_slope, [flagged_sections]) for one fiber.

    For every inter-event segment >= DIST_MIN_RUN_KM, fit a slope and flag it
    when it exceeds the per-fiber median by >= DIST_SLOPE_EXCESS_DBKM.  Segments
    that START within DIST_RECOVERY_GUARD_KM downstream of a severe loss event
    are recovery tails and are EXCLUDED from the flagged list."""
    med = _dist_median_slope(km, db, eof_km)
    if med is None:
        return None, []
    thresh = med + DIST_SLOPE_EXCESS_DBKM
    severe = _dist_severe_event_kms(events, eof_km)

    sections = []
    for sa, sb in _dist_segment_bounds(events, eof_km):
        slope = _dist_local_slope(km, db, sa, sb)
        if slope is None or slope < thresh:
            continue
        # recovery-tail exclusion: segment STARTS just downstream of a severe event?
        if any(0.0 <= (sa - ev) <= DIST_RECOVERY_GUARD_KM for ev in severe):
            continue
        sections.append({
            'start_km': round(sa, 2),
            'end_km':   round(sb, 2),
            'run_km':   round(sb - sa, 2),
            'slope':    round(slope, 4),
            'excess':   round(slope - med, 4),
        })
    return med, sections


def scan_distributed_loss(fibers_a):
    """A-direction post-analysis pass: surface degrading fiber SECTIONS.

    Scans each A-direction fiber's backscatter trace for inter-event runs whose
    linear attenuation slope exceeds the per-fiber median by a margin, excluding
    the launch region, the EOF tail, and severe-loss recovery tails.

    Returns a list of `distributed_loss` records (one per flagged section),
    each carrying: fiber, start_km, end_km, run_km, slope, excess_over_median,
    and a human label like "DISTRIBUTED LOSS 0.25 dB/km @74.8-81.5 km".

    Purely additive: reads the (already-loaded, Pass-0-normalized) A-direction
    records read-only and returns its own list.  Any per-fiber error is swallowed
    so a single bad trace cannot crash the report."""
    out = []
    for fnum in sorted(fibers_a.keys()):
        rec = fibers_a[fnum]
        try:
            km, db = _dist_trace_km_db(rec)
            if km is None:
                continue
            events = rec.get('events') or []
            eof_km = _dist_eof_km(events, km)
            _med, sections = _dist_scan_fiber(km, db, events, eof_km)
        except Exception as _exc:        # never let one bad fiber kill the pass
            print("splicereport: distributed-loss scan skipped fiber %s (%s)"
                  % (fnum, _exc), file=sys.stderr)
            continue
        for s in sections:
            out.append({
                'fiber':              int(fnum),
                'start_km':           s['start_km'],
                'end_km':             s['end_km'],
                'run_km':             s['run_km'],
                'slope':              s['slope'],
                'excess_over_median': s['excess'],
                'label': "DISTRIBUTED LOSS %.2f dB/km @%.1f-%.1f km"
                         % (s['slope'], s['start_km'], s['end_km']),
            })
    out.sort(key=lambda r: (r['fiber'], r['start_km']))
    return out


def aggregate_distributed_loss(sections):
    """Collapse per-fiber distributed-loss sections into CABLE-WIDE findings.

    The per-fiber `scan_distributed_loss` pass emits one record per elevated
    fiber stretch — on a cable with a genuine cable-wide degradation that is
    hundreds of near-identical rows (the same km region seen on many fibers).
    This step clusters those sections into the handful of real regions:

      • Sort sections by km and chain them TRANSITIVELY: a section joins the
        current cluster when its start lies within DIST_CLUSTER_GAP_KM of the
        running cluster end (i.e. its km-range overlaps or is within the gap
        of any section already in the cluster).
      • A cluster becomes a FINDING only when it spans >= MIN_FIBERS_FINDING
        DISTINCT fibers; scattered / lone sections below that occupancy gate
        are dropped (they remain available as the raw per-fiber count).

    Each finding carries: km_start (cluster min), km_end (cluster max),
    n_fibers (distinct), n_sections (raw rows folded in), median_excess_dbkm,
    median_slope_dbkm, example_fibers (a few fiber numbers), and a human label.

    Pure function of the section list — does not read traces or touch any
    flag/event state.  Returns the findings sorted by km_start."""
    if not sections:
        return []
    secs = sorted(sections, key=lambda s: (s['start_km'], s['end_km']))
    clusters = []
    cur = None
    cur_end = None
    for s in secs:
        if cur is None or s['start_km'] > cur_end + DIST_CLUSTER_GAP_KM:
            cur = [s]
            cur_end = s['end_km']
            clusters.append(cur)
        else:
            cur.append(s)
            if s['end_km'] > cur_end:
                cur_end = s['end_km']

    findings = []
    for c in clusters:
        fibers = sorted({s['fiber'] for s in c})
        if len(fibers) < MIN_FIBERS_FINDING:
            continue
        km_start = round(min(s['start_km'] for s in c), 2)
        km_end = round(max(s['end_km'] for s in c), 2)
        med_excess = round(float(np.median([s['excess_over_median'] for s in c])), 4)
        med_slope = round(float(np.median([s['slope'] for s in c])), 4)
        findings.append({
            'km_start':           km_start,
            'km_end':             km_end,
            'n_fibers':           len(fibers),
            'n_sections':         len(c),
            'median_excess_dbkm': med_excess,
            'median_slope_dbkm':  med_slope,
            'example_fibers':     fibers[:6],
            'label': "DISTRIBUTED LOSS region @%.1f-%.1f km — %d fibers, "
                     "median %.2f dB/km (+%.3f over baseline)"
                     % (km_start, km_end, len(fibers), med_slope, med_excess),
        })
    findings.sort(key=lambda f: f['km_start'])
    return findings


# ═══════════════════════════════════════════════════════════════════════
#  STEP 5 — Group into ribbons and build cell values
# ═══════════════════════════════════════════════════════════════════════

def build_ribbon_data(results, n_fibers, ribbon_size, n_splices, launch_issues=None):
    """Group flagged events into ribbon rows × splice columns.  If
    launch_issues is provided, each ribbon gets an extra 'launch_cell' entry
    summarising which of its fibers have launch-end issues — the write_xlsx
    function renders these into the ILA:A column."""
    n_ribbons = (n_fibers + ribbon_size - 1) // ribbon_size
    grid = {}

    for (fnum, si), res in results.items():
        ri = (fnum - 1) // ribbon_size
        key = (ri, si)
        if key not in grid:
            grid[key] = []
        grid[key].append(res)

    cells = {}
    for (ri, si), res_list in grid.items():
        res_list.sort(key=lambda x: x['fiber'])

        # Group fibers with same loss and same source type
        groups = []
        for res in res_list:
            merged = False
            for g in groups:
                if (res['bidir_loss'] is not None and g['loss'] is not None and
                        abs(res['bidir_loss'] - g['loss']) < 0.002 and
                        not res['is_break'] and not res['is_broke'] and
                        not res.get('is_bend', False) and not g.get('is_bend', False) and
                        not g['is_break'] and not g['is_broke'] and
                        res.get('event_source') == g.get('event_source')):
                    g['fibers'].append(res['fiber'])
                    merged = True
                    break
            if not merged:
                groups.append({
                    'fibers': [res['fiber']],
                    'loss': res['bidir_loss'],
                    'is_break': res['is_break'],
                    'is_broke': res['is_broke'],
                    'is_bend':  res.get('is_bend', False),
                    'is_ref':   res.get('is_ref', False),
                    'is_bfill': res.get('is_bfill', False),
                    'is_dead_zone': res.get('is_dead_zone', False),
                    'is_gainer': res.get('is_gainer', False),
                    'is_a_only': res.get('is_a_only', False),
                    'is_b_only': res.get('is_b_only', False),
                    'is_borderline': res.get('is_borderline', False),
                    'is_flagged': res.get('is_flagged', True),
                    'event_source': res.get('event_source', 'bidir'),
                    'label': res['label'],
                    'res': res,
                })

        # Build cell text — label shows source for A-only and B-only
        parts = []
        for g in groups:
            # Connector-loss decoration: appended to every cell text
            # when the underlying result is flagged is_high_connector_loss.
            # The label-based branches already include it via the
            # apply_connector_loss_rule append; the reconstructed
            # branches (A-only / B-only / B-fill / generic reburn) need
            # the suffix added explicitly here.
            conn_tag = ('  ⚠ conn'
                        if g['res'].get('is_high_connector_loss')
                        else '')
            if g.get('is_dead_zone'):
                # Collapse multi-fiber dead zones into "F1,F2,... DZ"
                fib_str = ','.join(str(f) for f in g['fibers'])
                parts.append(f"{fib_str} DZ")
            elif g['is_broke']:
                # Use the enriched broke label (includes position + DZ range)
                parts.append(g['label'])
            elif g['is_break']:
                parts.append(g['label'])
            elif g.get('is_ref'):
                parts.append(g['label'])
            elif g.get('is_bend'):
                # Use the full label (includes "BEND" marker and offset)
                parts.append(g['label'])
            elif g['is_a_only']:
                fib_str = ','.join(str(f) for f in g['fibers'])
                raw_loss = g['res']['a_loss']
                loss_abs = abs(raw_loss) if raw_loss is not None else 0
                loss_str = f"{loss_abs:.3f}"
                if loss_str.startswith('0.'): loss_str = loss_str[1:]
                # Single-direction display: raw A loss, no /2 bidir estimate.
                # Threshold (SINGLE_DIR_THRESHOLD, default 0.250) was already
                # gated upstream — anything in this branch cleared 0.250 dB
                # on its own.
                parts.append(f"{fib_str} {loss_str} (A){conn_tag}")
            elif g['is_b_only']:
                fib_str = ','.join(str(f) for f in g['fibers'])
                raw_loss = g['res']['b_loss']
                loss_abs = abs(raw_loss) if raw_loss is not None else 0
                loss_str = f"{loss_abs:.3f}"
                if loss_str.startswith('0.'): loss_str = loss_str[1:]
                parts.append(f"{fib_str} {loss_str} (B){conn_tag}")
            elif g.get('is_bfill'):
                fib_str = ','.join(str(f) for f in g['fibers'])
                loss = g['loss']
                loss_str = f"{loss:.3f}" if loss is not None else "?"
                if loss_str.startswith('0.'): loss_str = loss_str[1:]
                parts.append(f"{fib_str} {loss_str} (B-fill){conn_tag}")
            else:
                fib_str = ','.join(str(f) for f in g['fibers'])
                loss = g['loss']
                loss_str = f"{loss:.3f}" if loss is not None else "?"
                if loss_str.startswith('0.'): loss_str = loss_str[1:]
                # Additive borderline / review marker on a generic reburn cell
                # that sits on the threshold knife-edge (display-only).
                border_tag = '  ⚠ borderline' if g.get('is_borderline') else ''
                parts.append(f"{fib_str} {loss_str}{conn_tag}{border_tag}")

        cell_text = ' '.join(parts)
        is_break = any(g['is_break'] for g in groups)
        is_broke = any(g['is_broke'] for g in groups)
        is_ref   = any(g.get('is_ref', False) for g in groups)
        is_bend  = any(g.get('is_bend', False) for g in groups)
        is_bfill = any(g.get('is_bfill', False) for g in groups)

        # Has a standard bidir reburn in this cell?
        has_standard_reburn = any(
            not g['is_break'] and not g['is_broke'] and
            not g.get('is_bend') and
            not g.get('is_bfill') and not g.get('is_a_only') and
            not g.get('is_b_only')
            for g in groups
        )
        # A-only / B-only only drive color if no higher-priority event present
        is_a_only = (any(g.get('is_a_only', False) for g in groups) and
                     not is_break and not is_broke and not is_bfill and not has_standard_reburn)
        is_b_only = (any(g.get('is_b_only', False) for g in groups) and
                     not is_break and not is_broke and not is_bfill and not has_standard_reburn)

        # If estimated bidir still clears threshold, use a stronger shade.
        # Same rule applies to A-only / B-only / B-fill — all three use the
        # (single-dir / 2) ≥ threshold check to escalate to their darker tier.
        est_bidir_flagged = any(g['res'].get('est_bidir_flagged', False) for g in groups
                                if g.get('is_a_only') or g.get('is_b_only')
                                   or g.get('is_bfill'))

        max_loss = max((g['loss'] for g in groups if g['loss'] is not None), default=0)

        is_dead_zone = any(g.get('is_dead_zone', False) for g in groups)
        is_gainer    = any(g.get('is_gainer', False) for g in groups)
        is_high_connector_loss = any(
            g['res'].get('is_high_connector_loss', False) for g in groups)
        # Additive review marker — surfaces a threshold-edge reburn cell.
        # Display-only: does not drive colour / classification / counts.
        is_borderline = any(g.get('is_borderline', False) for g in groups)
        # A cell is "flagged" only if a constituent group is a real flag
        # (reburn / break / ref / bend / single-dir).  A borderline-only group
        # (sub-threshold knife-edge surfaced for review) carries is_flagged=False,
        # so a cell that is ONLY borderline stays unflagged → gets the distinct
        # borderline fill (not reburn pink) and is not counted as a reburn.
        is_flagged_cell = any(g.get('is_flagged', True) for g in groups)

        cells[(ri, si)] = {
            'text': cell_text,
            'is_break': is_break,
            'is_broke': is_broke,
            'is_ref':   is_ref,
            'is_bend':  is_bend,
            'is_bfill': is_bfill,
            'is_dead_zone': is_dead_zone,
            'is_gainer': is_gainer,
            'is_a_only': is_a_only,
            'is_b_only': is_b_only,
            'is_high_connector_loss': is_high_connector_loss,
            'is_borderline': is_borderline,
            'is_flagged': is_flagged_cell,
            'est_bidir_flagged': est_bidir_flagged,
            'max_loss': max_loss,
        }

    # ── Per-ribbon launch-issue summaries (for the ILA:A / ILA:B columns) ──
    launch_cells_a = {}   # ribbon_index → dict {text, severity}
    launch_cells_b = {}
    if launch_issues:
        per_ribbon_a = {}   # ri → list of (fnum, severity, tag)
        per_ribbon_b = {}
        for fnum, info in launch_issues.items():
            ri = (fnum - 1) // ribbon_size
            for tag in info.get('a_tags', []):
                per_ribbon_a.setdefault(ri, []).append((fnum, info['severity'], tag))
            for tag in info.get('b_tags', []):
                per_ribbon_b.setdefault(ri, []).append((fnum, info['severity'], tag))

        def _sev_order(s):
            return {'HIGH': 0, 'REVIEW': 1, 'WATCH': 2}.get(s, 3)

        for ri, items in per_ribbon_a.items():
            worst = min(items, key=lambda x: _sev_order(x[1]))[1]
            # Compact label: fiber# + abbreviated tag
            parts = [f"{f} {tag.split('@')[0].split('+')[0]}" for f, _, tag in items]
            launch_cells_a[ri] = {'text': ' '.join(parts[:6]) +
                                          (f" +{len(parts)-6} more" if len(parts) > 6 else ''),
                                   'severity': worst}
        for ri, items in per_ribbon_b.items():
            worst = min(items, key=lambda x: _sev_order(x[1]))[1]
            parts = [f"{f} {tag.split('@')[0].split('+')[0]}" for f, _, tag in items]
            launch_cells_b[ri] = {'text': ' '.join(parts[:6]) +
                                          (f" +{len(parts)-6} more" if len(parts) > 6 else ''),
                                   'severity': worst}

    return cells, launch_cells_a, launch_cells_b


# ═══════════════════════════════════════════════════════════════════════
#  STEP 6 — Generate Excel
# ═══════════════════════════════════════════════════════════════════════

def ribbon_label(ri, ribbon_size, n_fibers):
    first = ri * ribbon_size + 1
    last = min(first + ribbon_size - 1, n_fibers)
    ribbon_num = ri + 1
    tube = ''
    if ri < 48:
        tube_letter = chr(ord('A') + ri // 2)
        tube_num = (ri % 2) + 1
        tube = f" ({tube_letter}{tube_num})"
    return f"Fiber {first}-{last} ({ribbon_num}){tube}"


def write_xlsx(cells, splices, n_fibers, ribbon_size, output_path, site_a, site_b, span_km,
               launch_cells_a=None, launch_cells_b=None,
               fibers_a=None, fibers_b=None, all_results=None,
               distributed_loss=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Splice Report"

    # ── Acquisition-parameters audit sheet (added at the END of write_xlsx
    #    so we can insert it as sheet 0 once we have the workbook).  When
    #    fibers_a / fibers_b are passed, the audit checks every trace's
    #    test date / OTDR model + serial / wavelength / pulse width /
    #    averaging for consistency and reports majority-vs-outliers.
    _audit_payload = None
    if fibers_a is not None or fibers_b is not None:
        try:
            from acquisition_audit import audit_acquisition
            _audit_payload = audit_acquisition(fibers_a or {}, fibers_b or {})
        except Exception as _exc:
            print(f"  WARN: acquisition audit failed: {_exc}")
            _audit_payload = None

    n_ribbons = (n_fibers + ribbon_size - 1) // ribbon_size
    n_splices = len(splices)

    # ── Styles ──
    # Font family / size: Calibri 12 across the board per tech direction.
    FONT_NAME = "Calibri"
    FSIZE     = 12
    hdr_font    = Font(name=FONT_NAME, bold=True, size=FSIZE, color="FFFFFF")
    hdr_fill    = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    data_font   = Font(name=FONT_NAME, size=FSIZE)
    ribbon_font = Font(name=FONT_NAME, size=FSIZE)
    a_km_font   = Font(name=FONT_NAME, bold=True, size=FSIZE, color="1F4E79")
    b_km_font   = Font(name=FONT_NAME, bold=True, size=FSIZE, color="8B0000")

    # Cell fill/font for each event type
    red_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # A+B reburn
    break_fill  = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")   # break
    break_font  = Font(name=FONT_NAME, bold=True, size=FSIZE, color="FFFFFF")
    # BROKE is now rendered the same as BREAK — both are
    # physical damage where the fiber has lost transmission.  Same red fill,
    # same white bold text, same category in the tech's eye.
    broke_fill  = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")   # red
    broke_font  = Font(name=FONT_NAME, bold=True, size=FSIZE, color="FFFFFF")
    # In-line REFLECTIVE event (connector / mech splice / angled cleave)
    # — reflective + Fresnel but the trace clearly continues past it.
    # Distinct from BREAK red so the eye separates "fiber's still alive
    # and reflecting" from "fiber's gone".
    ref_fill    = PatternFill(start_color="E64A19", end_color="E64A19", fill_type="solid")   # deep orange-red
    ref_font    = Font(name=FONT_NAME, bold=True, size=FSIZE, color="FFFFFF")
    # Two-tier B-fill (parallel to A-only / B-only):
    #   bfill_fill / bfill_font  → est bidir (B/2) < threshold (light blue, conservative)
    #   bfill_fill2 / bfill_font2 → est bidir (B/2) ≥ threshold (deep blue, escalated)
    bfill_fill  = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    bfill_font  = Font(name=FONT_NAME, size=FSIZE, color="1F4E79")
    bfill_fill2 = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    bfill_font2 = Font(name=FONT_NAME, bold=True, size=FSIZE, color="FFFFFF")
    dz_fill     = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")   # dead zone (gray)
    dz_font     = Font(name=FONT_NAME, size=FSIZE, italic=True, color="3F3F3F")
    gainer_fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")   # field gainer (mint green)
    gainer_font = Font(name=FONT_NAME, bold=True, size=FSIZE, color="1B5E20")
    aonly_fill  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # A-only (light yellow, est bidir OK)
    aonly_font  = Font(name=FONT_NAME, size=FSIZE, color="7F6000")
    aonly_fill2 = PatternFill(start_color="FF7043", end_color="FF7043", fill_type="solid")   # A-only (coral, est bidir >= threshold) — deliberately non-yellow
    aonly_font2 = Font(name=FONT_NAME, bold=True, size=FSIZE, color="FFFFFF")
    bonly_fill  = PatternFill(start_color="E8D5F5", end_color="E8D5F5", fill_type="solid")   # B-only (lavender, est bidir OK)
    bonly_font  = Font(name=FONT_NAME, size=FSIZE, color="4B0082")
    bonly_fill2 = PatternFill(start_color="C084FC", end_color="C084FC", fill_type="solid")   # B-only (purple, est bidir >= threshold)
    bonly_font2 = Font(name=FONT_NAME, bold=True, size=FSIZE, color="1A0033")
    # BEND: teal / cyan — clearly distinct from splice colors so bends stand out
    # BEND cells: single yellow fill for every bend (no severity shading).
    # Matches the tech's yellow-highlight style on Cle Elum.
    bend_fill        = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
    bend_font        = Font(name=FONT_NAME, bold=True, size=FSIZE, color="5D4037")
    # Keep the old three-name aliases pointing at the single fill so any
    # downstream reference still resolves.  bend_font_high is just bend_font.
    bend_fill_watch  = bend_fill
    bend_fill_review = bend_fill
    bend_fill_high   = bend_fill
    bend_font_high   = bend_font
    # LAUNCH ISSUE: single orange fill — warns the tech a fiber had
    # launch-end trouble (broken at launch, damaged connector, truncated
    # event table, bad reflectance).  Orange was chosen to be clearly
    # distinct from the pink A+B reburn fill.  Severity tiers were
    # collapsed per tech feedback — a launch issue is a launch issue.
    launch_fill        = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    borderline_fill    = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")   # borderline / review — near-threshold knife-edge, NOT a flagged reburn
    launch_font        = Font(name=FONT_NAME, bold=True, size=FSIZE, color="5D2E00")
    # Back-compat aliases (in case anything else in the codebase still
    # references the per-severity names)
    launch_fill_high = launch_fill_review = launch_fill_watch = launch_fill
    launch_font_high = launch_font_watch = launch_font

    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # Each splice column now occupies TWO physical Excel columns: a km
    # column (left) and a ft column (right).  Headers go side-by-side,
    # data cells are merged across both so they visually span the pair.
    #   physical col = 2*si + 3 (km)  |  2*si + 4 (ft)
    def _km_col(si):
        return 2 * si + 3
    def _ft_col(si):
        return 2 * si + 4

    end_col = 2 * n_splices + 3                # ILA:B column

    # ── Row 1: B→A distances (km column + ft column, separate cells) ──
    # ── Row 2: A→B distances (km column + ft column, separate cells) ──
    # Convention swap: B→A on top, A→B on bottom — keeps the lowest-
    # numbered fiber's "near end" reading at the row directly above the
    # column header.
    ws.cell(row=1, column=2, value="B→A:").font = b_km_font
    ws.cell(row=2, column=2, value="A→B:").font = a_km_font
    for si, sp in enumerate(splices):
        # Steven's convention: the A→B header is the lowest-numbered fiber's
        # event distance, truncated to 10 m.  Falls back to the refined
        # cluster center, then the raw position, if no display value was set
        # (e.g. for phantom bend / damage columns).
        km = sp.get('position_km_display',
                    sp.get('position_km_refined', sp['position_km']))
        ft = km * 3280.84
        # B→A is the cable span minus the A-side display value, also
        # truncated to 10 m so the two headers stay self-consistent.
        b_km = math.floor((span_km - km) * 100) / 100.0
        b_ft = b_km * 3280.84
        # Row 1 = B→A
        c = ws.cell(row=1, column=_km_col(si), value=f"{b_km:.2f}km")
        c.font = b_km_font; c.alignment = Alignment(horizontal='center')
        c = ws.cell(row=1, column=_ft_col(si), value=f"{b_ft:,.0f}ft")
        c.font = b_km_font; c.alignment = Alignment(horizontal='center')
        # Row 2 = A→B
        c = ws.cell(row=2, column=_km_col(si), value=f"{km:.2f}km")
        c.font = a_km_font; c.alignment = Alignment(horizontal='center')
        c = ws.cell(row=2, column=_ft_col(si), value=f"{ft:,.0f}ft")
        c.font = a_km_font; c.alignment = Alignment(horizontal='center')
    # ILA:B end column — keep km / ft combined here (single column, no split)
    ws.cell(row=1, column=end_col, value="0.00km / 0ft").font = b_km_font
    ws.cell(row=2, column=end_col,
             value=f"{span_km:.2f}km / {span_km*3280.84:,.0f}ft").font = a_km_font

    # ── Row 3: Headers (splice label merged across km+ft pair) ──
    ws.cell(row=3, column=1, value="Ribbon").font = hdr_font
    ws.cell(row=3, column=1).fill = hdr_fill
    ws.cell(row=3, column=2, value=f"A-dir ILA: {site_a}").font = hdr_font
    ws.cell(row=3, column=2).fill = hdr_fill
    # Alternate fill colors for phantom-column headers so they stand out
    # from the blue splice headers at a glance.
    # Bend  -column header: yellow      (matches bend cells)
    # Damage-column header: red         (matches break/broke cells)
    # Ref   -column header: deep orange (matches ref cells)
    hdr_fill_bend   = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
    hdr_fill_damage = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
    hdr_fill_ref    = PatternFill(start_color="E64A19", end_color="E64A19", fill_type="solid")
    for si, sp in enumerate(splices):
        km_c, ft_c = _km_col(si), _ft_col(si)
        kind = sp.get('column_kind', 'splice')
        if kind == 'bend':
            ref_km = sp.get('position_km_refined', sp['position_km'])
            header = f"Bends @ {ref_km:.2f}km"
            cell = ws.cell(row=3, column=km_c, value=header)
            cell.fill = hdr_fill_bend
            # paint the ft side of the merged pair with the same fill so
            # the merged appearance is consistent
            ws.cell(row=3, column=ft_c).fill = hdr_fill_bend
        elif kind == 'damage':
            ref_km = sp.get('position_km_refined', sp['position_km'])
            header = f"Damage @ {ref_km:.2f}km"
            cell = ws.cell(row=3, column=km_c, value=header)
            cell.fill = hdr_fill_damage
            ws.cell(row=3, column=ft_c).fill = hdr_fill_damage
        elif kind == 'ref':
            ref_km = sp.get('position_km_refined', sp['position_km'])
            header = f"Ref @ {ref_km:.2f}km"
            cell = ws.cell(row=3, column=km_c, value=header)
            cell.fill = hdr_fill_ref
            ws.cell(row=3, column=ft_c).fill = hdr_fill_ref
        else:
            disp_n = sp.get('splice_display_num', si + 1)
            cell = ws.cell(row=3, column=km_c, value=f"Splice {disp_n}")
            cell.fill = hdr_fill
            ws.cell(row=3, column=ft_c).fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Merge the splice header across the km + ft pair
        ws.merge_cells(start_row=3, start_column=km_c,
                       end_row=3,   end_column=ft_c)
    ws.cell(row=3, column=end_col, value=f"B-dir ILA: {site_b}").font = hdr_font
    ws.cell(row=3, column=end_col).fill = hdr_fill

    # ── Data rows ──
    def _launch_fill(sev):
        # Single light-pink tier for all launch issues (severity ignored)
        return launch_fill, launch_font

    for ri in range(n_ribbons):
        row = ri + 4
        ws.cell(row=row, column=1, value=ribbon_label(ri, ribbon_size, n_fibers)).font = ribbon_font

        # ── ILA:A column (col 2) — launch-issue summary for A direction ──
        ila_a_cell = ws.cell(row=row, column=2)
        ila_a_cell.border = border
        ila_a_cell.alignment = Alignment(wrap_text=True, vertical='center')
        if launch_cells_a and ri in launch_cells_a:
            lc = launch_cells_a[ri]
            ila_a_cell.value = lc['text']
            f, fn = _launch_fill(lc['severity'])
            ila_a_cell.fill = f
            ila_a_cell.font = fn

        # ── ILA:B column (end_col) — launch-issue summary for B direction ──
        ila_b_cell = ws.cell(row=row, column=end_col)
        ila_b_cell.border = border
        ila_b_cell.alignment = Alignment(wrap_text=True, vertical='center')
        if launch_cells_b and ri in launch_cells_b:
            lc = launch_cells_b[ri]
            ila_b_cell.value = lc['text']
            f, fn = _launch_fill(lc['severity'])
            ila_b_cell.fill = f
            ila_b_cell.font = fn

        for si in range(n_splices):
            km_c, ft_c = _km_col(si), _ft_col(si)
            key = (ri, si)
            cell = ws.cell(row=row, column=km_c)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center',
                                        horizontal='center')
            # Border on the ft side so the merged appearance is consistent
            ws.cell(row=row, column=ft_c).border = border
            # Each data cell spans both km and ft columns
            ws.merge_cells(start_row=row, start_column=km_c,
                           end_row=row,   end_column=ft_c)

            if key in cells:
                cd = cells[key]
                cell.value = cd['text']
                if cd['is_break']:
                    cell.fill = break_fill
                    cell.font = break_font
                elif cd['is_broke']:
                    cell.fill = broke_fill
                    cell.font = broke_font
                elif cd.get('is_ref'):
                    cell.fill = ref_fill
                    cell.font = ref_font
                elif cd.get('is_bend'):
                    cell.fill = bend_fill
                    cell.font = bend_font
                elif cd.get('is_bfill'):
                    if cd.get('est_bidir_flagged'):
                        cell.fill = bfill_fill2
                        cell.font = bfill_font2
                    else:
                        cell.fill = bfill_fill
                        cell.font = bfill_font
                elif cd.get('is_dead_zone'):
                    cell.fill = dz_fill
                    cell.font = dz_font
                elif cd.get('is_gainer'):
                    cell.fill = gainer_fill
                    cell.font = gainer_font
                elif cd.get('is_b_only'):
                    if cd.get('est_bidir_flagged'):
                        cell.fill = bonly_fill2
                        cell.font = bonly_font2
                    else:
                        cell.fill = bonly_fill
                        cell.font = bonly_font
                elif cd.get('is_a_only'):
                    if cd.get('est_bidir_flagged'):
                        cell.fill = aonly_fill2
                        cell.font = aonly_font2
                    else:
                        cell.fill = aonly_fill
                        cell.font = aonly_font
                elif cd.get('is_borderline') and not cd.get('is_flagged', True):
                    # Sub-threshold knife-edge cell surfaced for review — give
                    # it a distinct fill, NOT the reburn pink (it is not a reburn).
                    cell.fill = borderline_fill
                    cell.font = data_font
                else:
                    cell.fill = red_fill
                    cell.font = data_font

    # ── Legend sheet ──
    ws_leg = wb.create_sheet("Legend")
    ws_leg.column_dimensions['A'].width = 14
    ws_leg.column_dimensions['B'].width = 65
    legend_items = [
        ("Pink",       "FFC7CE", "000000", "A+B — Bidirectional reburn: both directions confirmed, bidir loss >= threshold. Needs re-splice."),
        ("Red",        "FF4444", "FFFFFF", "Break — 1F reflective event (clean cut, glass-to-air Fresnel reflection). label: 'BREAK'"),
        ("Red (broke)","FF4444", "FFFFFF", "Broke — fiber trace terminates mid-span (crush / stress fracture).  Rendered with the same red fill as a break; label reads 'broke' or 'BREAK' depending on reflective vs non-reflective signature."),
        ("Deep Orange","E64A19", "FFFFFF", "REF — in-line reflective event (connector / mechanical splice / angled cleave).  Reflective + Fresnel but trace continues past it. label: 'F# ref .xxx (refl -XX dB)'"),
        ("Lt. Blue",   "BDD7EE", "1F4E79", "B-fill — B-direction loss past an A-side break (A trace is blind here). Single-direction: no averaging. Flagged only when the raw B loss alone clears the single-direction threshold (default 0.250 dB). label: 'F# .xxx (B-fill)'"),
        ("Gray",       "BFBFBF", "3F3F3F", "Dead zone — fiber broke on A side AND B trace also ends before reaching the A-break. Neither trace could see this splice for this fiber. Broke cell shows 'F# broke@XXk | DZ lo-hi k'; affected columns show 'F# DZ'."),
        ("Lt. Yellow", "FFF2CC", "7F6000", "A-only — A saw it, no B counterpart at the mirror. Single-direction: no averaging. Flagged only when the raw A loss alone clears the single-direction threshold (default 0.250 dB). label: 'F# .xxx (A)'"),
        ("Lavender",   "E8D5F5", "4B0082", "B-only — B saw it, no A counterpart at the mirror. Single-direction: no averaging. Flagged only when the raw B loss alone clears the single-direction threshold (default 0.250 dB). label: 'F# .xxx (B)'"),
        ("Yellow",     "FFEB3B", "5D4037", "BEND — event ≥ 0.090 dB at a position more than 150 m from the closure center.  Inspect conduit for pinch or tight bend."),
        ("Orange",     "FFA500", "5D2E00", "LAUNCH — fiber has a launch-end issue.  Loss rule: launch_loss >= -0.5 dB (anything weaker than a -0.5 dB gainer flags).  Reflectance rule: refl > -15 dB (damaged / dirty connector).  Plus missing file, empty event table.  Single tier — no WATCH/REVIEW/HIGH split.  Appears in ILA column.  Distinct from pink A+B reburn."),
        ("Mint Green", "A5D6A7", "1B5E20", "FIELD GAINER — mid-span event whose signed loss is in [-0.7, 0] dB (suspicious near-zero / weak-gainer event).  Excludes events within the launch zone or end-of-fiber region.  Overrides the geometric BEND tag in the [-0.7, -0.090] overlap range."),
    ]
    ws_leg.cell(row=1, column=1, value="Color").font = Font(name=FONT_NAME, bold=True, size=FSIZE)
    ws_leg.cell(row=1, column=2, value="Meaning").font = Font(name=FONT_NAME, bold=True, size=FSIZE)
    for i, (name, fc, tc, desc) in enumerate(legend_items, 2):
        c = ws_leg.cell(row=i, column=1, value=name)
        c.fill = PatternFill(start_color=fc, end_color=fc, fill_type="solid")
        c.font = Font(name=FONT_NAME, bold=True, size=FSIZE, color=tc)
        ws_leg.cell(row=i, column=2, value=desc).font = Font(name=FONT_NAME, size=FSIZE)

    # ── Distributed Loss sheet (ADDITIVE, fully separate from the grid) ──
    # Lists CABLE-WIDE distributed-loss FINDINGS produced by
    # aggregate_distributed_loss — the per-fiber sections from
    # scan_distributed_loss clustered into the handful of real km regions that
    # show elevated attenuation across many fibers.  Each row is ONE finding
    # (km region + how many fibers + the median slope), NOT one per-fiber row.
    # A different category from the discrete event flags in the Splice Report
    # sheet.  Empty (header only) when no region clears the occupancy gate.
    _dl = distributed_loss or []
    ws_dl = wb.create_sheet("Distributed Loss")
    for _col, _w in (('A', 18), ('B', 11), ('C', 12), ('D', 16),
                     ('E', 20), ('F', 18), ('G', 60)):
        ws_dl.column_dimensions[_col].width = _w
    ws_dl.cell(row=1, column=1,
               value="DISTRIBUTED SECTION LOSS — A direction").font = \
        Font(name=FONT_NAME, bold=True, size=FSIZE)
    ws_dl.cell(row=2, column=1,
               value=("Cable-wide regions of elevated attenuation (higher "
                      "dB/km) with NO discrete event, seen across many fibers.  "
                      "Per-fiber sections (slope exceeding the per-fiber median "
                      "by >= %.2f dB/km over >= %.1f km) are clustered by "
                      "km-range; a region is reported only when >= %d distinct "
                      "fibers participate.  Launch region, EOF tail, and "
                      "severe-loss recovery tails excluded.  Separate from the "
                      "event flags."
                      % (DIST_SLOPE_EXCESS_DBKM, DIST_MIN_RUN_KM,
                         MIN_FIBERS_FINDING))).font = \
        Font(name=FONT_NAME, size=FSIZE, italic=True)
    _dl_hdr = ["Region (km)", "Fibers", "Sections",
               "Median slope dB/km", "Median excess dB/km",
               "Example fibers", "Label"]
    for _ci, _h in enumerate(_dl_hdr, 1):
        _hc = ws_dl.cell(row=4, column=_ci, value=_h)
        _hc.font = hdr_font
        _hc.fill = hdr_fill
    _r = 5
    for _f in _dl:
        ws_dl.cell(row=_r, column=1,
                   value="%.2f – %.2f" % (_f['km_start'], _f['km_end']))
        ws_dl.cell(row=_r, column=2, value=_f['n_fibers'])
        ws_dl.cell(row=_r, column=3, value=_f.get('n_sections'))
        ws_dl.cell(row=_r, column=4, value=_f['median_slope_dbkm'])
        ws_dl.cell(row=_r, column=5, value=_f['median_excess_dbkm'])
        ws_dl.cell(row=_r, column=6,
                   value=", ".join(str(_x) for _x in _f.get('example_fibers', [])))
        ws_dl.cell(row=_r, column=7, value=_f['label'])
        for _ci in range(1, 8):
            ws_dl.cell(row=_r, column=_ci).font = Font(name=FONT_NAME, size=FSIZE)
        _r += 1
    if not _dl:
        ws_dl.cell(row=5, column=1,
                   value="(none — no cable-wide distributed-loss region detected)").font = \
            Font(name=FONT_NAME, size=FSIZE, italic=True)

    # ── Column widths — TRUE minimum-fit (no column wider than its content) ──
    # Calibri 12 is ~1.1–1.2 Excel-width-units/char; keep a hair of margin so
    # nothing clips.  Key subtlety: a cell that ANCHORS a multi-column merge
    # (the splice headers + the merged loss cells span the km+ft pair) must NOT
    # size a single column — its width is shared across the columns it spans
    # (handled in step 2).  Sizing a column to a merged value, plus equalizing
    # the km/ft pair, is what used to make the columns wider than necessary.
    CHAR_W   = 1.2
    PADDING  = 0.7
    MIN_W    = 3.0
    MAX_W    = 60.0
    # Total column count: col A (ribbon) + col B (ILA:A) +
    # (2 * n_splices) splice km/ft pairs + 1 ILA:B
    n_cols = 2 + 2 * n_splices + 1
    multicol_ranges = [mr for mr in ws.merged_cells.ranges
                       if mr.max_col > mr.min_col]
    multicol_anchors = {(mr.min_row, mr.min_col) for mr in multicol_ranges}

    def _needed(value):
        return max((len(line) for line in str(value).splitlines()),
                   default=0) * CHAR_W + PADDING

    # 1) base width = each column's widest OWN (non-spanning) content.
    raw_widths = {}
    for col_idx in range(1, n_cols + 1):
        widest = 0.0
        for r in range(1, ws.max_row + 1):
            if (r, col_idx) in multicol_anchors:
                continue                      # spans >1 column — sized in step 2
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            widest = max(widest, _needed(v))
        raw_widths[col_idx] = max(MIN_W, min(MAX_W, widest))

    # 2) widen a span ONLY if its merged header/value wouldn't otherwise fit,
    #    distributing just the deficit so the total stays minimal.
    for mr in multicol_ranges:
        v = ws.cell(row=mr.min_row, column=mr.min_col).value
        if v is None:
            continue
        cols = list(range(mr.min_col, mr.max_col + 1))
        deficit = _needed(v) - sum(raw_widths.get(c, MIN_W) for c in cols)
        if deficit > 0:
            add = deficit / len(cols)
            for c in cols:
                raw_widths[c] = min(MAX_W, raw_widths.get(c, MIN_W) + add)

    for col_idx, w in raw_widths.items():
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = w

    # ── Force Calibri 12 on EVERY cell ──
    # openpyxl's workbook default ("Normal" style) is Calibri 11.  Cells
    # that we don't explicitly assign a Font to (merged-cell siblings,
    # blank splice cells, blank ILA cells, ribbon names without explicit
    # font, etc.) inherit that default and end up at size 11.  Walk
    # every cell in the used range and bump it to Calibri 12 unless it
    # already has a deliberate non-default font (e.g. bold white on red
    # for break/broke).  Preserves bold / italic / color decisions while
    # standardising name + size.
    default_font_kwargs = {'name': FONT_NAME, 'size': FSIZE}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                             min_col=1, max_col=ws.max_column):
        for cell in row:
            f = cell.font
            # Preserve any explicitly-set styling (bold/italic/color);
            # only standardise the family + size.
            cell.font = Font(
                name=FONT_NAME,
                size=FSIZE,
                bold=bool(f.bold),
                italic=bool(f.italic),
                color=f.color,
            )

    # Auto-fit the legend sheet too.
    for col_idx in range(1, ws_leg.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        widest = 0
        for row in ws_leg.iter_rows(min_col=col_idx, max_col=col_idx,
                                    values_only=True):
            v = row[0]
            if v is None:
                continue
            line_len = max((len(line) for line in str(v).splitlines()), default=0)
            if line_len > widest:
                widest = line_len
        # Legend descriptions get a wider cap so they stay readable.
        cap = 90.0 if col_idx == 2 else MAX_W
        w = max(MIN_W, min(cap, widest * CHAR_W + PADDING))
        ws_leg.column_dimensions[col_letter].width = w

    # Force Calibri 12 on every legend cell too — same logic as above.
    for row in ws_leg.iter_rows(min_row=1, max_row=ws_leg.max_row,
                                 min_col=1, max_col=ws_leg.max_column):
        for cell in row:
            f = cell.font
            cell.font = Font(
                name=FONT_NAME,
                size=FSIZE,
                bold=bool(f.bold),
                italic=bool(f.italic),
                color=f.color,
            )

    ws.freeze_panes = 'C4'

    # ── Reburn summary sheet ──
    # Inserted BEFORE the acquisition audit so it lands at index 1
    # after the audit insertion below (insert_at=0 for audit, =1 for
    # reburn → ordering becomes [Acquisition, Reburn, Splice Report,
    # Legend]).  Counts ribbon × splice cells that contain at least
    # one A+B reburn fiber.
    if all_results is not None:
        try:
            from reburn_summary import compute_reburn_summary, \
                render_xlsx_sheet as _render_reburn
            _reburn = compute_reburn_summary(all_results, splices,
                                              n_fibers, ribbon_size)
            _render_reburn(wb, _reburn,
                           insert_at=0,                # before any audit
                           font_name=FONT_NAME, font_size=FSIZE)
            print(f"  Reburn summary: {_reburn['reburn_cells']} of "
                  f"{_reburn['total_cells']} cells "
                  f"({_reburn['reburn_percentage']:.2f}%)")
        except Exception as _exc:
            print(f"  WARN: failed to render reburn summary: {_exc}")

    # Insert the acquisition-parameters audit as the FIRST sheet, and set
    # it as the active sheet so the workbook opens on it.  Done last so
    # both the Splice Report sheet and the Reburn Summary above land
    # AFTER it in the final sheet order.
    if _audit_payload is not None:
        try:
            from acquisition_audit import render_xlsx_sheet
            render_xlsx_sheet(wb, _audit_payload,
                              font_name=FONT_NAME, font_size=FSIZE)
        except Exception as _exc:
            print(f"  WARN: failed to render acquisition sheet: {_exc}")

    wb.save(output_path)
    print(f"  Saved: {output_path}")


# ═══════════════════════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description='Splice QC report with EXFO-style bidirectional event matching.',
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('dir_a', help='A-direction SOR files directory')
    ap.add_argument('dir_b', nargs='?', help='B-direction SOR files directory')
    ap.add_argument('--output', '-o', default='splice_report_exfo.xlsx')
    ap.add_argument('--threshold', type=float, default=REBURN_THRESHOLD,
                    help=f'Flag threshold in dB (default {REBURN_THRESHOLD})')
    ap.add_argument('--ribbon-size', type=int, default=RIBBON_SIZE)
    ap.add_argument('--site-a', default=None,
                    help='A-end site name (auto-detected from directory names if not set)')
    ap.add_argument('--site-b', default=None,
                    help='B-end site name (auto-detected from directory names if not set)')
    ap.add_argument('--span-km', type=float, default=0,
                    help='Span distance in km (0 = auto-detect)')
    args = ap.parse_args()

    # ── Auto-detect site names from directory names ──
    # Directory names like "NEWELM 15 sec" encode the route as site_a + site_b
    # A-direction dir = A→B (e.g., NEWELM = from NEW to ELM)
    # B-direction dir = B→A (e.g., ELMNEW = from ELM to NEW)
    if args.site_a is None or args.site_b is None:
        import re
        a_base = os.path.basename(args.dir_a.rstrip('/'))
        # Extract the alphabetic prefix (e.g., "NEWELM" from "NEWELM 15 sec")
        alpha = re.match(r'([A-Za-z]+)', a_base)
        if alpha:
            route_str = alpha.group(1).upper()
            # Try to split into two 3-letter site codes (e.g., NEWELM → NEW + ELM)
            if len(route_str) >= 6:
                half = len(route_str) // 2
                if args.site_a is None:
                    args.site_a = route_str[:half]
                if args.site_b is None:
                    args.site_b = route_str[half:]
                print(f"  Auto-detected site names: {args.site_a} → {args.site_b}")
        if args.site_a is None:
            args.site_a = 'A'
        if args.site_b is None:
            args.site_b = 'B'

    print("Loading SOR files...")
    fibers_a, fibers_b = load_all(args.dir_a, args.dir_b)
    n_fibers = max(fibers_a.keys()) if fibers_a else 0
    print(f"  A: {len(fibers_a)} fibers   B: {len(fibers_b)} fibers   max fiber #{n_fibers}")

    # ── Pass 0: Normalize events for splice discovery ──
    # Save original events (needed for trace enhancement), normalize copies
    for r in list(fibers_a.values()) + list(fibers_b.values()):
        r['_raw_events'] = r['events']  # save originals
        # Offset between normalized event coords and raw trace samples, so the
        # silent-side windower can index the (unshifted) trace correctly.
        r['_trace_offset_km'] = _untrimmed_launch_offset_km(r['events'])
        r['events'] = _normalize_untrimmed_events(r['events'])

    print("Discovering splice closure positions...")
    splice_candidates = discover_splices(fibers_a)
    real_splices, phantom_zones = refine_closure_centers(
        fibers_a, splice_candidates, return_phantoms=True, fibers_b=fibers_b)
    print(f"  Found {len(real_splices)} real splice closures:")
    for i, sp in enumerate(real_splices, 1):
        ref_km = sp.get('position_km_refined', sp['position_km'])
        offset_m = (ref_km - sp['position_km']) * 1000
        print(f"    Splice {i:2d}: {sp['position_km']:8.2f} km  "
              f"(refined {ref_km:7.3f} km, {offset_m:+5.0f} m offset, "
              f"{sp['count']} fibers, spread {sp.get('position_spread_m', 0):.0f} m)")

    # Every bend / damage zone gets its own named column in the grid, sorted
    # with the real splice columns by physical position.  This mirrors the
    # tech's Cle Elum layout exactly (Splice 12, bends@11.51, Splice 11,
    # damage@17.46, bends@18.86, Splice 10, …).
    if phantom_zones:
        print(f"  Adding {len(phantom_zones)} bend/damage column(s) to the grid:")
        for sp in phantom_zones:
            ref = sp.get('position_km_refined', sp['position_km'])
            print(f"    [{sp['column_kind']:<6}] {ref:7.2f} km")
    splices = sorted(
        list(real_splices) + list(phantom_zones),
        key=lambda sp: sp.get('position_km_refined', sp['position_km']),
    )
    # Re-index splice display numbers for the real closures only
    splice_num = 0
    for sp in splices:
        if sp.get('column_kind') == 'splice':
            splice_num += 1
            sp['splice_display_num'] = splice_num

    # ── Launch-issue detection (must run BEFORE events get normalized again) ──
    first_splice_km = splices[0]['position_km'] if splices else None
    print("\nDetecting launch-end issues...")
    launch_issues = detect_launch_issues(fibers_a, fibers_b, first_splice_km)
    high_n   = sum(1 for v in launch_issues.values() if v['severity'] == 'HIGH')
    review_n = sum(1 for v in launch_issues.values() if v['severity'] == 'REVIEW')
    watch_n  = sum(1 for v in launch_issues.values() if v['severity'] == 'WATCH')
    print(f"  {len(launch_issues)} fibers with launch-end issues "
          f"(HIGH={high_n}, REVIEW={review_n}, WATCH={watch_n})")

    # Auto-detect span (preliminary, from normalized events)
    span_km = args.span_km
    if span_km == 0:
        all_ends = sorted([e['dist_km'] for r in fibers_a.values()
                           for e in r['events'] if e['is_end']])
        if all_ends:
            top_quarter = all_ends[int(len(all_ends) * 0.75):]
            span_km = round(np.median(top_quarter), 2)

    # ── Trace-based enhancement: detect breaks and refine span from raw trace ──
    # Trace enhancement re-detects launch and far-end connector from the raw
    # backscatter samples — only applicable when ``full_trace`` is present
    # (JSON files carry it; SOR sources expose ``trace`` instead, which the
    # enhancement function isn't wired to consume).  For SOR-only spans, the
    # Pass-0 ``_normalize_untrimmed_events`` already did the launch / far-end
    # trim from the events list, so we KEEP those normalized events instead
    # of unconditionally restoring raw and then no-oping the enhancement —
    # otherwise the launch 1F at ~1 km and the far-end connector 1F just
    # before EOL slide back into the event list and get classified as
    # REF / BEND on every fiber.
    n_trace_breaks = 0
    n_trace_enhanced = 0
    has_trace_data = any(r.get('full_trace') is not None for r in fibers_a.values())
    if has_trace_data:
        # Restore original events so trace enhancement can detect untrimmed
        for r in list(fibers_a.values()) + list(fibers_b.values()):
            r['events'] = r.pop('_raw_events')
    else:
        # SOR-only span — Pass-0 normalization stays.  Drop the raw-event
        # stash to free the memory.
        for r in list(fibers_a.values()) + list(fibers_b.values()):
            r.pop('_raw_events', None)
    if has_trace_data:
        print(f"\nTrace analysis: detecting breaks and span boundaries from raw trace...")

        # Phase 1: detect noise floors for all fibers to get population baseline
        all_noise_floors = []
        for r in list(fibers_a.values()) + list(fibers_b.values()):
            trace = r.get('full_trace')
            if trace is None:
                continue
            pts = r['full_points']
            acq = r['acq_range']
            ior_val = r.get('ior', 1.4682)
            launch_idx = _detect_launch_from_trace(trace, pts, acq, ior_val)
            nf_idx = _detect_noise_floor_from_trace(trace, launch_idx, pts, acq, ior_val)
            nf_km = _sample_to_km(nf_idx, ior_val, pts, acq)
            launch_km = _sample_to_km(launch_idx, ior_val, pts, acq)
            all_noise_floors.append(nf_km - launch_km)

        if all_noise_floors:
            pop_noise_floor = np.median(sorted(all_noise_floors)[int(len(all_noise_floors)*0.75):])
            print(f"  Population trace noise floor: {pop_noise_floor:.1f} km from launch")
        else:
            pop_noise_floor = span_km

        # Phase 2: enhance events using population noise floor as reference
        for fnum, r in fibers_a.items():
            _enhance_events_with_trace(r, span_km, pop_noise_floor_km=pop_noise_floor)
            if r.get('_trace_breaks'):
                n_trace_breaks += len(r['_trace_breaks'])
            if r.get('_trace_launch_km') is not None:
                n_trace_enhanced += 1
        for fnum, r in fibers_b.items():
            _enhance_events_with_trace(r, span_km, pop_noise_floor_km=pop_noise_floor)
            if r.get('_trace_breaks'):
                n_trace_breaks += len(r['_trace_breaks'])
        print(f"  Enhanced {n_trace_enhanced} A-fibers, {n_trace_breaks} breaks detected from trace")

        # Re-compute span after trace enhancement (use events, not trace)
        if args.span_km == 0:
            all_ends = sorted([e['dist_km'] for r in fibers_a.values()
                               for e in r['events'] if e['is_end']])
            if all_ends:
                top_quarter = all_ends[int(len(all_ends) * 0.75):]
                span_km = round(np.median(top_quarter), 2)

    print(f"  Span: {span_km} km ({span_km * 3280.84:,.0f} ft)")

    print(f"\nPass 1: Analyzing {len(fibers_a)} fibers at {len(splices)} splice positions "
          f"(threshold={args.threshold:.3f} dB)...")
    results = analyze_all(fibers_a, fibers_b, splices, args.threshold)
    n_p1_bidir  = sum(1 for r in results.values() if r.get('event_source') in ('bidir', 'bidir_grey_b'))
    n_p1_aonly  = sum(1 for r in results.values() if r.get('is_a_only'))
    n_p1_broke  = sum(1 for r in results.values() if r['is_broke'])
    n_p1_break  = sum(1 for r in results.values() if r['is_break'])
    n_p1_bend   = sum(1 for r in results.values() if r.get('is_bend'))
    n_p1_bfill  = sum(1 for r in results.values() if r.get('is_bfill'))
    print(f"  Pass 1 results: {len(results)} events")
    print(f"    A+B bidir:  {n_p1_bidir}")
    print(f"    A-only:     {n_p1_aonly}")
    print(f"    Breaks:     {n_p1_break}")
    print(f"    Broke:      {n_p1_broke}")
    print(f"    Bends:      {n_p1_bend}")
    print(f"    B-fill:     {n_p1_bfill}")

    # APRIL 23 revision: replace the old Pass-2 B-scan with two narrower passes
    #   Pass 2a — A-first standalone event classifier (bends / breaks
    #             at non-closure positions)
    #   Pass 2b — past-break B-fill scan (only uses B direction past
    #             an A-side break)
    print(f"\nPass 2a: Scanning A-direction standalone events (bends / breaks)...")
    a_standalone = scan_a_standalone_events(
        fibers_a, splices, results, span_km, fibers_b=fibers_b,
    )
    n_p2a_bend  = sum(1 for r in a_standalone.values() if r.get('is_bend'))
    n_p2a_break = sum(1 for r in a_standalone.values() if r.get('is_break'))
    print(f"  Pass 2a results: {len(a_standalone)} events "
          f"(bends={n_p2a_bend}, breaks={n_p2a_break})")

    print(f"\nPass 2b: Scanning bidirectional ghost reflections "
          f"(mid-span 1F with near-zero loss in BOTH directions)...")
    seen_so_far = {**results, **a_standalone}
    ghost_refl = scan_bidir_ghost_reflections(
        fibers_a, fibers_b, splices, seen_so_far, span_km,
    )
    print(f"  Pass 2b results: {len(ghost_refl)} ghost-reflection events")

    print(f"\nPass 2b': Scanning EXFO 'Merged Reflective; Non-reflective' "
          f"events (refl<0 on a 0F-typed loss event)...")
    seen_so_far = {**results, **a_standalone, **ghost_refl}
    merged_refl = scan_merged_reflective_events(
        fibers_a, fibers_b, splices, seen_so_far, span_km,
    )
    print(f"  Pass 2b' results: {len(merged_refl)} merged-reflective events")

    print(f"\nPass 2c: Scanning B-direction PAST A-side breaks (B-fill only)...")
    b_pastbreak = scan_b_past_breaks(
        fibers_a, fibers_b, splices, args.threshold, results, span_km,
    )
    print(f"  Pass 2c results: {len(b_pastbreak)} B-fill events")

    # Pass 2d — B-only mid-span breaks (A trace passes through, B trace
    # terminates).  Boss caught Lagrande↔Durkey F1-F12 with this signature
    # at km 90.46; Pass 1's A-driven broke check can't see them.
    print(f"\nPass 2d: Scanning B-only mid-span breaks "
          f"(A trace healthy, B trace terminates mid-span)...")
    pre_existing = {**results, **a_standalone, **ghost_refl, **merged_refl, **b_pastbreak}
    b_side_breaks = scan_b_side_breaks(
        fibers_a, fibers_b, splices, pre_existing, span_km,
    )
    print(f"  Pass 2d results: {len(b_side_breaks)} B-only broke fibers")

    # Merge — Pass 1 takes priority; then standalone; then ghost refl; then merged refl; then B-fill; then B-side broke
    all_results = {**results, **a_standalone, **ghost_refl, **merged_refl,
                    **b_pastbreak, **b_side_breaks}
    b_results = {**a_standalone, **ghost_refl, **merged_refl,
                  **b_pastbreak, **b_side_breaks}

    # Field-gainer annotation — flag mid-span events whose signed loss
    # falls in [-0.7, 0] dB (suspicious near-zero / weak-gainer events).
    n_field_gainers = apply_field_gainer_rule(all_results, span_km)
    print(f"  Field gainers: {n_field_gainers} (loss in "
          f"[{FIELD_GAINER_MIN_DB}, {FIELD_GAINER_MAX_DB}] dB, mid-span)")

    # High-connector-loss annotation — flag any reflective (1F) event
    # whose bidir loss reaches BIDIR_CONNECTOR_LOSS.  Decorates the
    # cell label with '⚠ conn' so the tech can spot connector-loss
    # issues separately from normal splice-loss reburns.
    n_high_conn = apply_connector_loss_rule(all_results, BIDIR_CONNECTOR_LOSS)
    print(f"  High connector loss: {n_high_conn} (1F events with "
          f"bidir >= {BIDIR_CONNECTOR_LOSS} dB)")

    # Off-splice bend / break / broke columns: any such event sitting
    # more than CLOSURE_MATCH_KM (150 m) from a real splice gets pulled
    # out of the nearest splice's column and placed in its own bend
    # (yellow) or damage (red) column at the event's actual km
    # position.  Events clustered within 200 m (bends/breaks) or 400 m
    # (broke fibers) share a column.
    # Additive review-bend sweep: surface off-grid consensus bends the
    # length-model/LSA test silently drops (display-only; never demotes).
    all_results.update(
        flag_consensus_bends(all_results, fibers_a, fibers_b, splices, span_km))
    pre_splice_ids = {id(sp) for sp in splices}
    # Account-then-flag: keep each fiber's helix-drifted OWN splice attributed to
    # its closure column (one column per closure, like the tech grid); only spin
    # off genuinely ADDITIONAL events into their own columns.
    all_results, splices = split_offsplice_events_into_own_columns(
        all_results, splices, total_span_km=span_km, fibers_a=fibers_a)
    newly_added = [sp for sp in splices if id(sp) not in pre_splice_ids]
    if newly_added:
        print(f"  Pulled off-splice events into {len(newly_added)} "
              f"new column(s):")
        for sp in newly_added:
            km = sp.get('position_km_refined', sp['position_km'])
            print(f"    [{sp['column_kind']:<6}] @ {km:.2f} km  "
                  f"({sp.get('count', 0)} events)")

    n_total   = len(all_results)
    n_bend    = sum(1 for r in all_results.values() if r.get('is_bend'))
    n_bidir   = sum(1 for r in all_results.values()
                    if r.get('event_source') in ('bidir', 'bidir_grey_a', 'bidir_grey_b'))
    n_a_only  = sum(1 for r in all_results.values() if r.get('is_a_only'))
    n_b_only  = sum(1 for r in all_results.values() if r.get('is_b_only'))
    n_breaks  = sum(1 for r in all_results.values() if r['is_break'])
    n_broke   = sum(1 for r in all_results.values() if r['is_broke'])
    n_ref     = sum(1 for r in all_results.values() if r.get('is_ref'))
    n_bfill   = sum(1 for r in all_results.values() if r.get('is_bfill'))
    n_dz      = sum(1 for r in all_results.values() if r.get('is_dead_zone'))
    n_dz_fibers = len({r['fiber'] for r in all_results.values()
                       if r['is_broke'] and r.get('dead_zone_km')})
    n_reburn  = n_bidir - n_breaks

    # Bend severity breakdown
    n_bend_high   = sum(1 for r in all_results.values()
                        if r.get('is_bend') and r.get('bend_severity') == 'HIGH')
    n_bend_review = sum(1 for r in all_results.values()
                        if r.get('is_bend') and r.get('bend_severity') == 'REVIEW')
    n_bend_watch  = sum(1 for r in all_results.values()
                        if r.get('is_bend') and r.get('bend_severity') == 'WATCH')

    print(f"\nBuilding ribbon grid...")
    cells, launch_cells_a, launch_cells_b = build_ribbon_data(
        all_results, n_fibers, args.ribbon_size, len(splices),
        launch_issues=launch_issues,
    )
    print(f"  {len(cells)} cells with flagged events, "
          f"{len(launch_cells_a)} ribbons with A-launch issues, "
          f"{len(launch_cells_b)} ribbons with B-launch issues")

    print(f"Writing Excel report...")
    write_xlsx(cells, splices, n_fibers, args.ribbon_size, args.output,
               args.site_a, args.site_b, span_km,
               launch_cells_a=launch_cells_a, launch_cells_b=launch_cells_b,
               fibers_a=fibers_a, fibers_b=fibers_b,
               all_results=all_results)

    print(f"\n{'═'*60}")
    print(f"  SPLICE REPORT (EXFO-MATCH + BENDS) COMPLETE")
    print(f"{'═'*60}")
    print(f"  Fibers:       {n_fibers}")
    print(f"  Splices:      {len(splices)}")
    print(f"  Span:         {span_km} km")
    print(f"  Threshold:    {args.threshold:.3f} dB   (bend threshold {BEND_THRESHOLD:.3f} dB, "
          f"offset > {CLOSURE_MATCH_KM*1000:.0f} m)")
    print(f"  ──────────────────────────────────")
    print(f"  A+B reburns:  {n_reburn}  (pink)   — both directions, bidir >= threshold, near closure center")
    print(f"  Breaks:       {n_breaks}  (red)    — 1F reflective event, trace ends near here")
    print(f"  Broke:        {n_broke}  (red)    — trace terminates mid-span (same red fill as break)")
    print(f"  Ref events:   {n_ref}  (deep orange) — in-line reflective event, trace continues past it")
    print(f"  B-fill:       {n_bfill}  (blue)   — B-direction past a break")
    print(f"  Dead zone:    {n_dz}  (gray)   — neither trace could see; {n_dz_fibers} broken fibers have a dead zone")
    print(f"  A-only:       {n_a_only}  (yellow) — A saw it, B did not")
    print(f"  B-only:       {n_b_only}  (purple) — B saw it, A did not  ← EXFO extra")
    print(f"  Bends:        {n_bend}  (yellow) — event >= {BEND_THRESHOLD:.3f} dB, > 150 m from closure center")
    n_gainer = sum(1 for r in all_results.values() if r.get('is_gainer'))
    print(f"  Field gainers:{n_gainer}  (mint)   — mid-span loss in [{FIELD_GAINER_MIN_DB}, {FIELD_GAINER_MAX_DB}] dB")
    print(f"  Launch:       {len(launch_issues)}  (orange) — launch-end issues (single tier)")
    print(f"  ──────────────────────────────────")
    print(f"  Total:        {n_total}")
    print(f"  Output:       {args.output}")
    print()


if __name__ == '__main__':
    main()
