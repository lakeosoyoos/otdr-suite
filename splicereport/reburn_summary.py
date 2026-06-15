"""
Reburn summary sheet
====================
Counts how many ribbon × splice cells contain at least one fiber that
the engine flagged as an A+B bidirectional reburn, divides by the total
number of (ribbon, real-splice) cells, and renders the calculation onto
a workbook sheet.

A "reburn" is the same thing the engine counts as "A+B reburns" in its
CLI summary: any flagged result whose event_source is in
('bidir', 'bidir_grey_a', 'bidir_grey_b').

A "real splice" is a column with column_kind == 'splice' — bend / damage
/ ref columns are excluded from both the numerator and the denominator
(they aren't closures the tech splices at).

Cell counting rule (per the spec):
  * One ribbon × one real-splice column = one cell in the grid.
  * The cell counts as a "reburn cell" if ANY fiber in that ribbon was
    flagged as a reburn at that splice — 1 fiber or 12 fibers, same
    contribution.
  * Total cells = n_ribbons × n_real_splices.
  * Percentage = reburn cells / total cells × 100.
"""
from __future__ import annotations

from collections import defaultdict


REBURN_SOURCES = ('bidir', 'bidir_grey_a', 'bidir_grey_b')


def compute_reburn_summary(all_results: dict,
                            splices: list,
                            n_fibers: int,
                            ribbon_size: int) -> dict:
    """Return a dict with the headline stats and the per-splice / per-
    ribbon breakdowns the sheet renders.

    Shape:
      {
        "n_ribbons":          int,
        "n_real_splices":     int,
        "total_cells":        int,    # n_ribbons * n_real_splices
        "reburn_cells":       int,
        "reburn_percentage":  float,  # 0..100, rounded to 2 dp
        "per_splice":  [{"label", "km", "reburn_cells",
                          "percent_of_ribbons"}],
        "per_ribbon":  [{"label", "reburn_cells",
                          "percent_of_splices"}],
        "real_splice_indices": [int, ...],   # diagnostic
      }
    """
    n_ribbons = max(1, (n_fibers + ribbon_size - 1) // ribbon_size)

    # Identify the splice indices we count against (column_kind=='splice').
    real_splice_indices = []
    for si, sp in enumerate(splices):
        if sp.get('column_kind', 'splice') == 'splice':
            real_splice_indices.append(si)
    n_real_splices = len(real_splice_indices)
    total_cells = n_ribbons * n_real_splices

    # For each (ribbon, splice) cell, true iff any reburn lives there.
    cell_has_reburn = defaultdict(bool)
    real_set = set(real_splice_indices)

    for (fnum, si), r in (all_results or {}).items():
        if si not in real_set:
            continue
        if not isinstance(r, dict):
            continue
        if r.get('event_source') not in REBURN_SOURCES:
            continue
        ri = (fnum - 1) // ribbon_size
        cell_has_reburn[(ri, si)] = True

    reburn_cells = sum(1 for v in cell_has_reburn.values() if v)
    pct = (reburn_cells / total_cells * 100.0) if total_cells else 0.0

    # Per-splice breakdown
    per_splice = []
    for si in real_splice_indices:
        sp = splices[si]
        km = sp.get('position_km_refined', sp.get('position_km', 0.0))
        cells_here = sum(1 for (rri, ssi) in cell_has_reburn
                         if ssi == si and cell_has_reburn[(rri, ssi)])
        per_splice.append({
            "splice_idx":     si,
            "label":          f"Splice {real_splice_indices.index(si) + 1}",
            "km":             float(km),
            "reburn_cells":   cells_here,
            "percent_of_ribbons": (cells_here / n_ribbons * 100.0
                                    if n_ribbons else 0.0),
        })

    # Per-ribbon breakdown
    per_ribbon = []
    for ri in range(n_ribbons):
        cells_here = sum(1 for (rri, ssi) in cell_has_reburn
                         if rri == ri and cell_has_reburn[(rri, ssi)])
        per_ribbon.append({
            "ribbon_idx":     ri,
            "label":          f"Ribbon {ri + 1}",
            "reburn_cells":   cells_here,
            "percent_of_splices": (cells_here / n_real_splices * 100.0
                                    if n_real_splices else 0.0),
        })

    return {
        "n_ribbons":           n_ribbons,
        "n_real_splices":      n_real_splices,
        "total_cells":         total_cells,
        "reburn_cells":        reburn_cells,
        "reburn_percentage":   round(pct, 2),
        "per_splice":          per_splice,
        "per_ribbon":          per_ribbon,
        "real_splice_indices": real_splice_indices,
    }


def render_xlsx_sheet(wb, summary: dict, *,
                       insert_at: int = 1,
                       font_name: str = "Calibri",
                       font_size: int = 12) -> None:
    """Insert the reburn summary as sheet `insert_at` in `wb` (default
    index 1 — right after Acquisition Parameters, before Splice Report).

    Layout:
        Header                           bold + dark fill
        Headline calculation             one row per term, then the final %
        Per-splice table                 (Splice # | km | cells | % of ribbons)
        Per-ribbon table                 (Ribbon # | cells | % of splices)
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ws = wb.create_sheet("Reburn Summary", insert_at)

    fnt_h1     = Font(name=font_name, size=font_size + 2, bold=True)
    fnt_h2     = Font(name=font_name, size=font_size + 1, bold=True)
    fnt_bold   = Font(name=font_name, size=font_size, bold=True)
    fnt_normal = Font(name=font_name, size=font_size)
    fnt_small  = Font(name=font_name, size=font_size - 1, italic=True,
                       color="555555")
    fnt_pct    = Font(name=font_name, size=font_size + 4, bold=True,
                       color="9C0006")

    fill_header = PatternFill("solid", fgColor="1F4E79")
    fill_subhd  = PatternFill("solid", fgColor="D9D9D9")
    fill_amber  = PatternFill("solid", fgColor="FFF2CC")
    fill_pink   = PatternFill("solid", fgColor="FFC7CE")

    fnt_header_text = Font(name=font_name, size=font_size + 1, bold=True,
                            color="FFFFFF")

    thin = Side(style="thin", color="BFBFBF")
    box  = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22

    # ── Title ──
    ws.cell(row=1, column=1, value="Reburn Summary").font = fnt_h1
    ws.cell(row=1, column=2,
            value=f"{summary['n_ribbons']} ribbons × "
                  f"{summary['n_real_splices']} splices").font = fnt_small

    # ── Headline calculation ──
    rows = [
        ("Ribbons",                          summary["n_ribbons"]),
        ("Real splice columns",              summary["n_real_splices"]),
        ("Total ribbon × splice cells",      summary["total_cells"]),
        ("Cells with at least one reburn",   summary["reburn_cells"]),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = fnt_normal
        b = ws.cell(row=r, column=2, value=value)
        b.font = fnt_normal
        b.alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=1).border = box
        b.border = box
        r += 1

    # Headline %
    ws.cell(row=r, column=1, value="Reburn percentage").font = fnt_h2
    pct_cell = ws.cell(row=r, column=2,
                       value=f"{summary['reburn_percentage']:.2f}%")
    pct_cell.font = fnt_pct
    pct_cell.alignment = Alignment(horizontal="right")
    pct_cell.fill = fill_pink
    ws.cell(row=r, column=1).border = box
    pct_cell.border = box
    r += 2

    # ── Per-splice breakdown ──
    ws.cell(row=r, column=1, value="By splice").font = fnt_h2
    r += 1
    headers = ["Splice", "km", "Cells with reburn", "% of ribbons"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = fnt_header_text
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center")
        c.border = box
    r += 1
    for entry in summary["per_splice"]:
        ws.cell(row=r, column=1, value=entry["label"]).font = fnt_normal
        ws.cell(row=r, column=2, value=f"{entry['km']:.2f} km").font = fnt_normal
        cc = ws.cell(row=r, column=3, value=entry["reburn_cells"])
        cc.font = fnt_normal
        cc.alignment = Alignment(horizontal="right")
        pc = ws.cell(row=r, column=4,
                     value=f"{entry['percent_of_ribbons']:.1f}%")
        pc.font = fnt_normal
        pc.alignment = Alignment(horizontal="right")
        if entry["reburn_cells"] > 0:
            for col in range(1, 5):
                ws.cell(row=r, column=col).fill = fill_amber
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = box
        r += 1

    r += 1
    # ── Per-ribbon breakdown ──
    ws.cell(row=r, column=1, value="By ribbon").font = fnt_h2
    r += 1
    headers = ["Ribbon", "Cells with reburn", "% of splices"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = fnt_header_text
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center")
        c.border = box
    r += 1
    for entry in summary["per_ribbon"]:
        ws.cell(row=r, column=1, value=entry["label"]).font = fnt_normal
        cc = ws.cell(row=r, column=2, value=entry["reburn_cells"])
        cc.font = fnt_normal
        cc.alignment = Alignment(horizontal="right")
        pc = ws.cell(row=r, column=3,
                     value=f"{entry['percent_of_splices']:.1f}%")
        pc.font = fnt_normal
        pc.alignment = Alignment(horizontal="right")
        if entry["reburn_cells"] > 0:
            for col in range(1, 4):
                ws.cell(row=r, column=col).fill = fill_amber
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = box
        r += 1
