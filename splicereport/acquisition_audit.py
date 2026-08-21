"""
Acquisition-parameters audit
============================
Inspects the per-file records produced by the SOR / JSON parser and
reports whether every trace in the run was shot with the same
instrument and settings.

For each of:
    * test timestamp        (calendar-day bucketed)
    * OTDR model
    * OTDR serial number
    * wavelength            (compares the SET of wavelengths)
    * pulse width           (ns; per wavelength)
    * averaging             (count for SOR/TRC, time for JSON; per wavelength)

…the helper returns either an "all match" verdict (the value is THE
SPEC) or a "majority" verdict with the count + the list of files whose
value differs.

Result shape:
    {
        "file_fields": [
            {"name": "Test date",     "result": <Verdict>},
            {"name": "OTDR model",    "result": <Verdict>},
            ...
        ],
        "per_wavelength": [
            {"wavelength_nm": 1550.0,
             "rows": [
                {"name": "Pulse width", "result": <Verdict>},
                {"name": "Averaging",   "result": <Verdict>},
             ]},
            ...
        ],
        "n_files": <int>,
        "earliest_iso": <str>,
        "latest_iso":   <str>,
    }

Verdict shape (one consistent contract — the renderer doesn't care
which field produced it):
    {
        "all_match":   bool,
        "spec":        str,                  # the displayed value
        "majority":    str | None,
        "majority_n":  int,
        "outliers":    [(filename, value_str), ...],
        "all_missing": bool,                  # field not stored in this format
    }
"""
from __future__ import annotations

import datetime as _dt
from collections import Counter, defaultdict
from typing import Any, Iterable


# ─────────────────────────────────────────────────────────────────────────────
#  The one helper, reused for every field
# ─────────────────────────────────────────────────────────────────────────────
def _is_null(v) -> bool:
    """A value that counts as 'not recorded' for consistency purposes."""
    if v is None: return True
    if isinstance(v, float) and v != v: return True   # NaN
    if isinstance(v, str)   and not v.strip(): return True
    # _averaging returns (kind, value) tuples and (None, None) when averaging
    # is absent; without this, an all-missing column rendered a green
    # "✓ All match: (missing)" instead of the "Not available" verdict.
    if isinstance(v, tuple) and v and all(_is_null(x) for x in v): return True
    return False


def consistency_check(samples: list[tuple[str, Any]],
                      display: callable = None) -> dict:
    """Given [(filename, value), ...], return a Verdict dict.

    `display` is an optional formatter applied to a single value for
    rendering (e.g. wavelength → '1550.0 nm').  Defaults to str().

    Rules:
      * A null value (None / '' / NaN) counts as differing — it shows
        as '(missing)' in the outlier list.
      * Majority = the most-common non-null value.  Ties broken by
        first-seen order (Counter.most_common is stable on ties in
        Python 3.7+).
      * all_missing = every sample is null.
    """
    display = display or (lambda v: str(v))
    nonnull_pairs = [(fn, v) for fn, v in samples if not _is_null(v)]
    if not nonnull_pairs:
        return {
            "all_match":   False,
            "spec":        "Not available (not stored in this file type)",
            "majority":    None,
            "majority_n":  0,
            "outliers":    [],
            "all_missing": True,
        }

    counts = Counter(v for _, v in nonnull_pairs)
    majority_val, majority_n = counts.most_common(1)[0]
    majority_display = display(majority_val)

    outliers = []
    for fn, v in samples:
        if _is_null(v):
            outliers.append((fn, "(missing)"))
        elif v != majority_val:
            outliers.append((fn, display(v)))

    if not outliers:
        return {
            "all_match":   True,
            "spec":        f"✓ All match: {majority_display}",
            "majority":    majority_val,
            "majority_n":  majority_n,
            "outliers":    [],
            "all_missing": False,
        }
    total = len(samples)
    return {
        "all_match":   False,
        "spec":        (f"⚠ Majority: {majority_display} "
                        f"({majority_n} of {total}) — "
                        f"{len(outliers)} differ"),
        "majority":    majority_val,
        "majority_n":  majority_n,
        "outliers":    outliers,
        "all_missing": False,
    }


# ─────────────────────────────────────────────────────────────────────────────
#  Per-field extractors — read the right key for the file's source format
# ─────────────────────────────────────────────────────────────────────────────
def _fmt_timestamp(epoch_ts: int) -> str:
    if not epoch_ts:
        return ""
    try:
        return _dt.datetime.utcfromtimestamp(int(epoch_ts)).strftime(
            "%Y-%m-%d %H:%M UTC")
    except (OSError, OverflowError, ValueError):
        return ""


def _calendar_day(epoch_ts: int) -> str:
    """Return YYYY-MM-DD for a UTC epoch, or '' if invalid/missing."""
    if not epoch_ts:
        return ""
    try:
        return _dt.datetime.utcfromtimestamp(int(epoch_ts)).strftime(
            "%Y-%m-%d")
    except (OSError, OverflowError, ValueError):
        return ""


def _pulse_ns(rec: dict) -> float | None:
    """Pulse width in nanoseconds (normalized across formats)."""
    cal = rec.get('exfo_calibration') or {}
    # JSON: the calibration block already stores NominalPulseWidth in
    # seconds (we set it to pulse_ns * 1e-9 in json_reader).  Round to
    # 1 ns to absorb float jitter so equal pulse widths compare equal.
    for k in ('NominalPulseWidth', 'CalibratedPulseWidth'):
        v = cal.get(k)
        if v is not None:
            return round(v * 1e9, 0)
    # JSON-only fallback
    if '_json_pulse_ns' in rec and rec['_json_pulse_ns'] is not None:
        return round(float(rec['_json_pulse_ns']), 0)
    return None


def _averaging(rec: dict) -> tuple[str, Any] | tuple[str, None]:
    """Returns (kind, value).  kind ∈ {'count', 'time_sec', None}.

    Format-aware:
      * JSON       — Parameters.Duration parsed to seconds  → ('time_sec', s)
      * SOR        — FxdParams Duration (the *tech-chosen* averaging time
                     shown in the EXFO viewer) when present, otherwise the
                     NumberOfAverages count from the proprietary block.
                     Duration is preferred for EXFO FTBx instruments because
                     NumberOfAverages varies adaptively per trace even when
                     every other setting is identical — using the count
                     marks a healthy run as "⚠ 852 differ" when nothing is
                     actually wrong.  Count is still the right answer for
                     instruments that let the tech set an exact average
                     count, so it's the documented fallback.
    """
    src = rec.get('_source')
    if src == 'json':
        dur = rec.get('duration_sec')
        if dur is not None:
            return ('time_sec', round(float(dur), 1))
        return (None, None)
    # SOR path
    dur = rec.get('duration_sec')
    if dur is not None and dur > 0:
        return ('time_sec', round(float(dur), 1))
    cal = rec.get('exfo_calibration') or {}
    n = cal.get('NumberOfAverages')
    if n is not None:
        return ('count', int(n))
    return (None, None)


def _wavelength_nm(rec: dict) -> float | None:
    """Test wavelength in nm.  Prefer the exact wavelength from the EXFO
    proprietary block (~1545.8 / 1625.5 for matched-lot lasers); fall
    back to the nominal FxdParams / JSON value."""
    v = rec.get('exfo_wavelength_nm')
    if v is not None:
        return round(float(v), 1)
    v = rec.get('_json_wavelength_nm')
    if v is not None:
        return round(float(v), 1)
    v = rec.get('wavelength')
    if v is not None:
        return round(float(v), 1)
    return None


# ─────────────────────────────────────────────────────────────────────────────
#  Per-direction self-consistency
#
#  A bidirectional span is DEFINED by A and B being shot from opposite ends,
#  usually on different days and often on a second unit — which is why #86
#  removed the block that compared the two.  What that block could never see,
#  because audit_acquisition concatenates A and B before checking anything, is
#  a direction that disagrees with ITSELF.  That one is a real finding: one
#  end's set was shot on two different instruments, so its losses are not one
#  calibration and cannot be compared fibre-to-fibre.
#
#  WHICH FIELDS.  Census of 43 direction folders / 29,520 traces on disk:
#
#     field                     fires on          healthy spans left clean
#     OTDR unit (model+serial)   9 of 43 (21%)    34 of 43 gain zero rows
#     Wavelength                 tracks the unit  (each unit has its own
#                                                  exact laser wavelength)
#     Test date (calendar day)  32 of 43 (74%)     8 of 43 gain zero rows
#
#  Test date is NOT checked per direction.  A 1152-fibre 64 km span takes a
#  week to shoot: Sacramento↔Suisun — the FastReporter ground-truth span, and
#  as clean as anything on disk — spans five calendar days in its A direction
#  alone.  Flagging that reproduces exactly the structural noise #86 deleted,
#  one direction at a time instead of A-against-B.  The span's date range is
#  still printed, as a fact, in the sheet header.
#
#  Model is folded into the serial rather than checked separately: on disk
#  every direction with two models also has two serials, so a separate row
#  would duplicate one finding — and "FTBx-730D-SM3 #1876272" tells the reader
#  which instrument, which two bare serial numbers do not.
# ─────────────────────────────────────────────────────────────────────────────
PER_DIRECTION_MAX_OUTLIER_FILES = 12   # above this the roster stops helping:
                                       #   a 432-of-864 split is a crew fact,
                                       #   not a list of files to go re-shoot,
                                       #   and the value×count breakdown in
                                       #   the verdict already says it all.


def _otdr_unit(rec: dict) -> tuple[str, str] | None:
    """The instrument identity: (model, serial).  None when neither stored."""
    model  = (rec.get('otdr_model')  or '').strip()
    serial = (rec.get('otdr_serial') or '').strip()
    if not model and not serial:
        return None
    return (model, serial)


def _display_unit(t: tuple[str, str]) -> str:
    model, serial = t
    if model and serial:
        return f"{model} #{serial}"
    return model or f"#{serial}"


def _mixed_verdict(verdict: dict, samples: list, display: callable) -> dict:
    """Restate a failed Verdict as a per-direction finding.

    consistency_check's own spec reads "⚠ Majority: X (N of M) — K differ",
    which is the right voice for a span-wide rollup and the wrong one here:
    the reader needs to see that ONE END was shot as two jobs, and what the
    two jobs were.  The numbers are the same numbers, re-laid-out as the
    actual split.
    """
    counts, n_missing = Counter(), 0
    for _fn, v in samples:
        if _is_null(v):
            n_missing += 1
        else:
            counts[v] += 1
    parts = [f"{display(v)} × {n}" for v, n in counts.most_common()]
    if n_missing:
        parts.append(f"(missing) × {n_missing}")
    out = dict(verdict)
    out["spec"] = ("⚠ Mixed within this direction: " + "; ".join(parts))
    if len(verdict.get("outliers") or []) > PER_DIRECTION_MAX_OUTLIER_FILES:
        out["outliers"] = []
    return out


def _direction_consistency(records: list, label: str) -> dict | None:
    """Verdicts for ONE direction checked against itself.

    Returns None when the direction agrees with itself on every field — which
    is the normal case and the reason a healthy span gains no rows at all.
    """
    if not records:
        return None
    by_file = [(r.get('filename') or '?', r) for r in records]
    checks = (
        ("OTDR unit",  [(fn, _otdr_unit(r))     for fn, r in by_file],
         _display_unit),
        ("Wavelength", [(fn, _wavelength_nm(r)) for fn, r in by_file],
         lambda v: f"{v:.1f} nm"),
    )
    rows = []
    for name, samples, disp in checks:
        v = consistency_check(samples, display=disp)
        if v["all_match"] or v["all_missing"]:
            continue
        rows.append({"name": name, "result": _mixed_verdict(v, samples, disp)})
    if not rows:
        return None
    return {"label": label, "n_files": len(records), "rows": rows}


# ─────────────────────────────────────────────────────────────────────────────
#  FastReporter "Test Settings" panel, one table per direction
#
#  WHAT THIS IS.  FastReporter shows an eight-row "Test Settings" panel for a
#  trace: the refractive index and backscatter the losses were computed
#  against, plus the four analysis thresholds that decided which events got
#  into the event table at all.  The boss reads that panel daily.  Printing it
#  twice — once per direction — answers the only question it is being asked:
#  were the A and B shots taken with the same settings, so that averaging them
#  into a bidirectional loss is legitimate?
#
#  WHY IT MATTERS CONCRETELY.  These are not decorative.  A direction shot with
#  ReflectanceThreshold -72 dB instead of -78 dB simply does not record the
#  weak reflective events the other direction did, so a bidirectional compare
#  sees "A saw it, B didn't" and blames the fiber.  A different IOR scales every
#  distance.  A different splice-loss threshold changes the event roster.
#
#  COLLAPSING MANY TRACES TO ONE VALUE.  A direction is hundreds of traces and
#  the panel has one cell per row, so the cell shows the MODE — computed by the
#  same consistency_check() the rest of this sheet uses, so "the value for this
#  direction" means the same thing everywhere in the workbook.
#
#  But the mode alone is a TRAP here, and this is not hypothetical: PLACHE LS
#  (1152 traces) is split 586 at -78.00 dB against 566 at -72.00 dB, while its
#  partner CHEPLA LS is a uniform -78.00 dB.  Take the mode of each and the two
#  directions "match" — and the report would quietly assert comparability for
#  the 566 fibers where it does not hold.  So an internally-mixed direction is
#  reported as a finding in its own right: the cell is flagged, and the split is
#  spelled out underneath.  A direction that disagrees with ITSELF cannot be
#  summarised by one number, and the sheet says so instead of picking one.
# ─────────────────────────────────────────────────────────────────────────────

#  Number formatting is FastReporter's, to the digit — IOR at 6 dp, dB at 2 dp
#  except the splice-loss / splitter / end-of-fiber thresholds at 3, percent at
#  2.  The boss reads FR daily; a table that formats differently is one he has
#  to stop and re-read before he can trust it.
def _f_ior(v):  return f"{v:.6f}"
def _f_db2(v):  return f"{v:.2f} dB"
def _f_db3(v):  return f"{v:.3f} dB"
def _f_pct2(v): return f"{v:.2f} %"


def _ts(rec: dict) -> dict:
    """The per-trace Test Settings dict; {} when the file had no EXFO block."""
    v = rec.get('test_settings')
    return v if isinstance(v, dict) else {}


def _splitter(rec: dict):
    """(enabled, threshold_dB) — FR renders these as one checkbox + value row."""
    ts = _ts(rec)
    en, thr = ts.get('SplitterDetection'), ts.get('SplitterDetectionThreshold')
    if en is None and thr is None:
        return None
    return (bool(en), thr)


def _f_splitter(t) -> str:
    _en, thr = t
    return _f_db3(thr) if thr is not None else "n/a"


def _splitter_label(t) -> str:
    """FR puts the checkbox in the label: '[ ] Splitter loss (SM Only)'.

    The box is greyed/disabled in the screenshot because SplitterDetection is
    0.  Keeping the state in the LABEL (rather than the value) is deliberate:
    if one direction ran splitter detection and the other did not, the two
    tables differ in a column the eye is already scanning.
    """
    if t is None:
        return "[ ] Splitter loss (SM Only)"
    en, _thr = t
    return f"[{'x' if en else ' '}] Splitter loss (SM Only)"


def _core_size(rec: dict):
    """The stored proxies for FR's 'Fiber core size' row.

    THE MICRON FIGURE IS NOT IN THE FILE.  A full dump of the proprietary
    block's 1,096 named fields contains no core-size or mode-field-diameter
    field, and no numeric field anywhere in the file holds 9 (or 9e-6).  What
    IS stored is the glass designation, in two independent places, and both are
    invariant across the 1,106-trace census — GenParams fiber type 652 (ITU-T
    G.652, standard single-mode) and proprietary FiberCode 0.

    G.652 does imply a ~9 µm core, but with no second value anywhere on disk
    the mapping cannot be tested, so it is not asserted.  The row prints the
    designation it can prove and says the micron figure is unavailable.  The
    comparison still runs on the stored codes, so a genuine glass mismatch
    between directions would flag even though the micron value never resolves.
    """
    ft = rec.get('fiber_type')
    fc = _ts(rec).get('FiberCode')
    if ft is None and fc is None:
        return None
    return (ft, fc)


def _f_core_size(t) -> str:
    ft, _fc = t
    if ft:
        return f"n/a (fiber type G.{ft})"
    return "n/a"


#  (FR row label, extractor, value formatter, label override)
#  Order is FastReporter's panel order, top to bottom.
_FR_ROWS = (
    ("IOR",
     lambda r: _ts(r).get('Ior') if _ts(r).get('Ior') is not None else r.get('ior'),
     _f_ior,   None),
    ("Backscatter",
     lambda r: (_ts(r).get('Rbs') if _ts(r).get('Rbs') is not None
                else r.get('backscatter_db')),
     _f_db2,   None),
    ("Helix factor",
     lambda r: _ts(r).get('HelixFactor'),           _f_pct2, None),
    ("Splice loss detection threshold",
     lambda r: _ts(r).get('SpliceLossThreshold'),   _f_db3,  None),
    ("[ ] Splitter loss (SM Only)",
     _splitter,                                     _f_splitter, _splitter_label),
    ("Reflectance detection threshold",
     lambda r: _ts(r).get('ReflectanceThreshold'),  _f_db2,  None),
    ("End-of-fiber detection threshold",
     lambda r: _ts(r).get('EndOfFiberThreshold'),   _f_db3,  None),
    ("Fiber core size",
     _core_size,                                    _f_core_size, None),
)


def _direction_test_settings(records: list, label: str) -> dict:
    """One direction's Test Settings panel: eight rows, each collapsed to its
    modal value plus an explicit note when this direction disagrees with
    itself."""
    by_file = [(r.get('filename') or '?', r) for r in records]
    rows = []
    for name, extract, fmt, label_fn in _FR_ROWS:
        samples = [(fn, extract(r)) for fn, r in by_file]
        verdict = consistency_check(samples, display=fmt)
        majority = verdict["majority"]
        missing = verdict["all_missing"]
        # Mixed = this direction's own traces disagree.  Spelled out in full
        # (every distinct value with its count) rather than summarised, because
        # "which half of my fibers used the other setting" is the actionable
        # part.
        mixed_detail = None
        if not missing and not verdict["all_match"]:
            counts, n_missing = Counter(), 0
            for _fn, v in samples:
                if _is_null(v):
                    n_missing += 1
                else:
                    counts[v] += 1
            parts = [f"{fmt(v)} × {n}" for v, n in counts.most_common()]
            if n_missing:
                parts.append(f"(missing) × {n_missing}")
            mixed_detail = "; ".join(parts)
        rows.append({
            "name":         name,
            "label":        (label_fn(majority) if label_fn and not missing
                             else name),
            "value":        ("n/a" if missing else fmt(majority)),
            "raw":          None if missing else majority,
            "missing":      missing,
            "mixed":        mixed_detail is not None,
            "mixed_detail": mixed_detail,
            "n_files":      len(samples),
        })
    return {"label": label, "n_files": len(records), "rows": rows}


def compute_test_settings(records_a: list, records_b: list,
                          label_a: str, label_b: str) -> dict | None:
    """Both directions' panels plus the row-by-row A-vs-B verdict.

    Returns None when neither direction stored any test settings at all (e.g.
    a JSON-sourced run), so the renderer emits nothing rather than a table of
    eight 'n/a's.
    """
    if not records_a and not records_b:
        return None
    a = _direction_test_settings(records_a, label_a) if records_a else None
    b = _direction_test_settings(records_b, label_b) if records_b else None
    if a is None and b is None:
        return None
    if all(r["missing"] for r in (a or b)["rows"]) and (a is None or b is None):
        return None
    if a and b and all(r["missing"] for r in a["rows"]) \
             and all(r["missing"] for r in b["rows"]):
        return None

    verdicts = []
    n_differ = 0
    for i, (name, _extract, _fmt, _lf) in enumerate(_FR_ROWS):
        ra = a["rows"][i] if a else None
        rb = b["rows"][i] if b else None
        # Comparable only when both sides actually stored the field.
        if ra is None or rb is None or ra["missing"] or rb["missing"]:
            differs = False
            comparable = False
        else:
            differs = ra["raw"] != rb["raw"]
            comparable = True
        # A direction that disagrees with itself is flagged on its own merit —
        # even when the two modes happen to be equal.  This is the PLACHE case:
        # mode-vs-mode says "match", and the match is not real.
        flag_a = bool(ra and (differs or ra["mixed"]))
        flag_b = bool(rb and (differs or rb["mixed"]))
        if differs:
            n_differ += 1
        verdicts.append({"name": name, "differs": differs,
                         "comparable": comparable,
                         "flag_a": flag_a, "flag_b": flag_b})

    n_comparable = sum(1 for v in verdicts if v["comparable"])
    mixed_notes = []
    for d in (a, b):
        if not d:
            continue
        for r in d["rows"]:
            if r["mixed"]:
                mixed_notes.append(
                    f"{d['label']} — {r['name']}: not one setting across this "
                    f"direction's own {r['n_files']} trace(s): "
                    f"{r['mixed_detail']}. The cell shows the most common "
                    f"value; the rest were shot differently.")
    if a is None or b is None:
        # One direction only.  The settings ARE stored; there is simply no
        # counterpart to check them against, and saying "not stored" here
        # would be a false negative in a table whose job is verification.
        headline = ("⚠ Only one direction loaded — the A/B comparison this "
                    "table exists for could not be made")
    elif n_differ:
        headline = (f"⚠ {n_differ} of {n_comparable} comparable parameter(s) "
                    f"DIFFER between the two directions")
    elif mixed_notes:
        headline = (f"⚠ The two directions' most-common values agree, but a "
                    f"direction disagrees with ITSELF — see below")
    elif n_comparable:
        headline = f"✓ All {n_comparable} comparable parameter(s) match"
    else:
        headline = "Test settings not stored in these files"
    # "clean" drives the green/amber on the headline, so it must mean "I
    # checked, and they agree" — not merely "I found nothing to say".  A
    # one-sided run has not checked anything and does not get a green tick.
    clean = (a is not None and b is not None
             and n_differ == 0 and not mixed_notes and n_comparable > 0)
    return {"a": a, "b": b, "verdicts": verdicts, "n_differ": n_differ,
            "n_comparable": n_comparable, "mixed_notes": mixed_notes,
            "headline": headline, "clean": clean}


# ─────────────────────────────────────────────────────────────────────────────
#  Public entry point
# ─────────────────────────────────────────────────────────────────────────────
def audit_acquisition(fibers_a: dict, fibers_b: dict,
                      label_a: str = "A-direction",
                      label_b: str = "B-direction") -> dict:
    """Run the audit across every per-file record in fibers_a + fibers_b.

    Each fibers_* entry is a dict whose value is the record returned by
    parse_sor_full() or parse_otdr_json() — i.e. carries the same keys
    we just extended in those parsers.

    `label_a` / `label_b` name the two directions in the per-direction
    section only; nothing else in the payload uses them.
    """
    records_a, records_b = [], []
    for fnum, r in sorted((fibers_a or {}).items()):
        if isinstance(r, dict):
            records_a.append(r)
    for fnum, r in sorted((fibers_b or {}).items()):
        if isinstance(r, dict):
            records_b.append(r)
    records = records_a + records_b

    # ── File-level fields ────────────────────────────────────────────
    by_file = [(r.get('filename') or '?', r) for r in records]

    # Calendar day for test timestamp
    day_samples = [(fn, _calendar_day(r.get('date_time'))) for fn, r in by_file]
    date_verdict = consistency_check(day_samples)

    # Earliest / latest full timestamp span (for context display)
    epoch_vals = [int(r.get('date_time') or 0) for _, r in by_file]
    epoch_vals = [v for v in epoch_vals if v > 0]
    earliest_iso = _fmt_timestamp(min(epoch_vals)) if epoch_vals else ""
    latest_iso   = _fmt_timestamp(max(epoch_vals)) if epoch_vals else ""

    model_samples  = [(fn, (r.get('otdr_model')  or '').strip()) for fn, r in by_file]
    serial_samples = [(fn, (r.get('otdr_serial') or '').strip()) for fn, r in by_file]
    model_verdict  = consistency_check(model_samples)
    serial_verdict = consistency_check(serial_samples)

    # Wavelength as a set per file (handles the multi-wavelength case
    # where one file holds several traces; here each record is one
    # trace, so the set is a single value — but we compare across the
    # full collection).
    wl_samples = [(fn, _wavelength_nm(r)) for fn, r in by_file]
    wl_verdict = consistency_check(
        wl_samples, display=lambda v: f"{v:.1f} nm")

    file_fields = [
        {"name": "Test date (calendar day)", "result": date_verdict},
        {"name": "OTDR model",                "result": model_verdict},
        {"name": "OTDR serial",               "result": serial_verdict},
        {"name": "Wavelength",                "result": wl_verdict},
    ]

    # ── Per-wavelength section: pulse width + averaging ─────────────
    by_wl = defaultdict(list)
    for fn, r in by_file:
        wl = _wavelength_nm(r)
        by_wl[wl].append((fn, r))

    per_wavelength = []
    for wl in sorted(by_wl.keys(),
                     key=lambda v: (v is None, v if v is not None else 0)):
        pairs = by_wl[wl]
        # Pulse width
        pw_samples = [(fn, _pulse_ns(r)) for fn, r in pairs]
        pw_verdict = consistency_check(
            pw_samples, display=lambda v: f"{int(v)} ns")
        # Averaging — keep the (kind, value) tuples so different formats
        # don't appear "equal" by coincidence.
        avg_pairs = [(fn, _averaging(r)) for fn, r in pairs]
        avg_samples = [(fn, t) for fn, t in avg_pairs]

        def _display_avg(t):
            kind, v = t
            if kind == 'count':   return f"{v} avg"
            if kind == 'time_sec': return f"{v:g} s"
            return "(missing)"
        avg_verdict = consistency_check(avg_samples, display=_display_avg)

        per_wavelength.append({
            "wavelength_nm": wl,
            "wavelength_label": (f"{wl:.1f} nm" if wl is not None
                                   else "(unknown wavelength)"),
            "n_files": len(pairs),
            "rows": [
                {"name": "Pulse width", "result": pw_verdict},
                {"name": "Averaging",   "result": avg_verdict},
            ],
        })

    # ── Per-direction self-consistency ─────────────────────────────
    # Empty list = both directions agree with themselves.  The renderer
    # emits nothing at all in that case, which is the whole point.
    per_direction = [d for d in (_direction_consistency(records_a, label_a),
                                 _direction_consistency(records_b, label_b))
                     if d is not None]

    return {
        "n_files":       len(records),
        "earliest_iso":  earliest_iso,
        "latest_iso":    latest_iso,
        "file_fields":   file_fields,
        "per_wavelength": per_wavelength,
        "per_direction": per_direction,
        # FastReporter's Test Settings panel, one table per direction.  None
        # when the sources carry no test settings at all.
        "test_settings": compute_test_settings(records_a, records_b,
                                               label_a, label_b),
    }


def engine_stamp_text() -> str:
    """One line naming the engine that produced this workbook, e.g.
    ``OTDR Suite · app build 239 (2026-08-21 09:12 PDT) · engine: update 239
    applied 2026-08-21 09:10 PDT`` — or ``OTDR Suite · dev`` in a checkout.

    Deliberately the SAME wording the hub's sidebar footer prints (app.py's
    'OTDR Suite · app {appv} · engine: {engv}'), so a boss holding two
    disagreeing workbooks can line the stamp up against what each tech's
    sidebar shows without translating between two vocabularies.

    Resolved HERE, inside the engine process, not handed down from the hub:
    the hub imported its app.py at boot while this module is loaded off disk
    when the subprocess starts, so after a cache swap the two can genuinely
    differ — and the number in the workbook came from THIS process.

    Never raises: a missing/broken error_report degrades to 'unknown' rather
    than taking the whole acquisition sheet down (its caller only warns)."""
    try:
        from error_report import version_labels
        appv, engv = version_labels()
    except Exception:
        return "OTDR Suite · unknown"
    if appv == "dev" and engv == "dev":
        return "OTDR Suite · dev"
    return f"OTDR Suite · app {appv} · engine: {engv}"


# ─────────────────────────────────────────────────────────────────────────────
#  Excel renderer — inserts as sheet 0 + sets as active
# ─────────────────────────────────────────────────────────────────────────────
def render_xlsx_sheet(wb, audit: dict, font_name: str = "Calibri",
                       font_size: int = 11,
                       per_trace_detail: bool = True,
                       per_direction_detail: bool = False,
                       test_settings_block: bool = False) -> None:
    """Insert the audit as the FIRST sheet of `wb` (an openpyxl Workbook).
    Sets it as the active sheet so the workbook opens on it.

    Layout:
        Header row              Calibri 12 bold, light grey fill
        Parameter | Result      two-column table
        Green fill              all_match rows
        Amber fill              outlier rows
        Indent (col B)          outlier files listed one per row

    `per_trace_detail` controls the trailing "Per-trace detail" block —
    the file-level Test date / OTDR model / OTDR serial / Wavelength
    checks from ``audit['file_fields']``.

    It is OFF for the BIDIRECTIONAL splice report.  Those four checks
    compare every A trace against every B trace, and a bidirectional span
    is *defined* by A and B being shot from opposite ends — usually on a
    different day, and on WSC/SUI-class spans with a second unit at a
    different wavelength.  Three of the four then report exactly half the
    traces as "differing" and list ~250 rows of filenames warning about
    normal structure.  Nothing above the block is affected: the useful
    per-wavelength Pulse width / Averaging rollups (the ones that catch a
    real reshoot, e.g. SUIWSC1100 at 19 s against a 60 s majority) are
    emitted first and are untouched.

    It stays ON for the UNIDIRECTIONAL (uni one-shot) report, whose input
    is a single direction.  There a date / model / serial disagreement is
    a genuine finding — it means one direction's set was shot with two
    units or re-shot on another day — and on a clean single-direction
    folder the four rows come back green with no filename list at all.

    `per_direction_detail` controls the "Per-direction acquisition
    consistency" block, which reports each direction checked against
    ITSELF (see the section comment above _direction_consistency).

    It is ON for the BIDIRECTIONAL splice report, which is the report that
    has two directions to check and the one whose combined view hides the
    finding.  It emits NOTHING when both directions are self-consistent —
    34 of the 43 direction folders on disk — so a clean span gains zero
    rows, which is what separates it from the block #86 deleted.

    It is OFF for the UNIDIRECTIONAL report and must stay off: uni's input
    IS one direction, so `per_trace_detail` above already reports exactly
    these facts (LAMBEY: 216 of 432 on a second OTDR unit).  Turning both
    on would print the same finding twice.

    `test_settings_block` appends FastReporter's Test Settings panel twice,
    once per direction, side by side (see _render_test_settings).

    It is ON for the BIDIRECTIONAL splice report — the report that has two
    directions to compare, which is the entire purpose of printing it twice.

    It is OFF for the UNIDIRECTIONAL report, where there is no second
    direction to compare against and a lone panel answers no question.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ws = wb.create_sheet("Acquisition Parameters", 0)
    wb.active = wb.sheetnames.index(ws.title)

    fnt_normal = Font(name=font_name, size=font_size)
    fnt_bold   = Font(name=font_name, size=font_size, bold=True)
    fnt_header = Font(name=font_name, size=font_size + 1, bold=True)
    fnt_small  = Font(name=font_name, size=font_size - 1, italic=True,
                       color="555555")

    fill_header = PatternFill("solid", fgColor="D9D9D9")
    fill_green  = PatternFill("solid", fgColor="E2EFDA")   # all-match
    fill_amber  = PatternFill("solid", fgColor="FFF2CC")   # outliers
    fill_grey   = PatternFill("solid", fgColor="F2F2F2")   # section banner

    thin = Side(style="thin", color="BFBFBF")
    box  = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 80

    # Folder coverage.  Present only when the caller supplied it AND the
    # folder was not fully covered — a clean report gains nothing at all.
    cov = audit.get("coverage") or {}
    cov_lines = audit.get("coverage_lines") or []
    cov_incomplete = bool(cov) and not cov.get("complete") and bool(cov_lines)
    fnt_alarm = Font(name=font_name, size=font_size + 1, bold=True,
                     color="FFFFFF")
    fill_alarm = PatternFill("solid", fgColor="C00000")
    fnt_alarm_detail = Font(name=font_name, size=font_size, bold=True,
                            color="C00000")
    fill_alarm_soft = PatternFill("solid", fgColor="FFE1E1")

    row = 1
    if cov_incomplete:
        for col, val in ((1, "COVERAGE"),
                         (2, audit.get("coverage_headline") or "")):
            c = ws.cell(row=row, column=col, value=val)
            c.font = fnt_alarm
            c.fill = fill_alarm
            c.border = box
            c.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1
        for line in cov_lines:
            ws.cell(row=row, column=1, value="").fill = fill_alarm_soft
            ws.cell(row=row, column=1).border = box
            c = ws.cell(row=row, column=2, value=line)
            c.font = fnt_alarm_detail
            c.fill = fill_alarm_soft
            c.border = box
            c.alignment = Alignment(vertical="top", wrap_text=True)
            row += 1
        ws.cell(row=row, column=1, value="").fill = fill_grey
        ws.cell(row=row, column=2, value="").fill = fill_grey
        row += 1

    ws.cell(row=row, column=1, value="Acquisition Parameters").font = fnt_header
    # The trace count states coverage whenever it is not 100%: the same number
    # a reviewer skims for "how big was this run".
    if cov_incomplete:
        _count = (f"{cov['n_loaded']} of {cov['n_candidates']} trace(s) in "
                  f"folder analysed — {cov['n_dropped']} NOT analysed")
        _fnt = fnt_alarm_detail
    else:
        _count = f"{audit['n_files']} trace(s)"
        _fnt = fnt_small
    ws.cell(row=row, column=2,
            value=(f"{_count}; "
                    f"first {audit['earliest_iso']} → last "
                    f"{audit['latest_iso']}")).font = _fnt
    row += 1
    ws.cell(row=row, column=1, value="").fill = fill_grey
    ws.cell(row=row, column=2, value="").fill = fill_grey
    row += 1

    # Header
    ws.cell(row=row, column=1, value="Parameter").font = fnt_bold
    ws.cell(row=row, column=2, value="Result").font = fnt_bold
    ws.cell(row=row, column=1).fill = fill_header
    ws.cell(row=row, column=2).fill = fill_header
    ws.cell(row=row, column=1).border = box
    ws.cell(row=row, column=2).border = box
    row += 1
    first_data_row = row          # freeze below the header, wherever it landed

    def _emit(name: str, verdict: dict):
        nonlocal row
        fill = fill_green if verdict["all_match"] else fill_amber
        a = ws.cell(row=row, column=1, value=name)
        b = ws.cell(row=row, column=2, value=verdict["spec"])
        a.font = fnt_normal
        b.font = fnt_normal
        a.fill = fill
        b.fill = fill
        a.border = box
        b.border = box
        a.alignment = Alignment(vertical="top")
        b.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1
        # Outlier list (indented)
        for fn, val in verdict.get("outliers", [])[:60]:
            ws.cell(row=row, column=1, value="").border = box
            c = ws.cell(row=row, column=2, value=f"      {fn} = {val}")
            c.font = fnt_small
            c.fill = fill_amber
            c.border = box
            c.alignment = Alignment(vertical="top", wrap_text=True)
            row += 1
        if len(verdict.get("outliers", [])) > 60:
            extra = len(verdict["outliers"]) - 60
            c = ws.cell(row=row, column=2,
                          value=f"      … and {extra} more")
            c.font = fnt_small
            c.fill = fill_amber
            c.border = box
            row += 1

    # Per-wavelength SUMMARY first — the green all-match rollup is the quick read,
    # so it goes at the top, above the (optional) per-trace detail.
    for w in audit["per_wavelength"]:
        a = ws.cell(row=row, column=1,
                      value=f"— {w['wavelength_label']}  "
                            f"({w['n_files']} trace(s))")
        a.font = fnt_bold
        a.fill = fill_grey
        ws.cell(row=row, column=2, value="").fill = fill_grey
        row += 1
        for entry in w["rows"]:
            _emit(entry["name"], entry["result"])
        row += 1   # blank spacer after each wavelength block

    # Per-DIRECTION consistency: each end's set checked against itself.
    # Present only when a direction actually disagrees with itself, so on a
    # healthy span nothing below this line is written at all.
    if per_direction_detail and audit.get("per_direction"):
        a = ws.cell(row=row, column=1,
                      value="Per-direction acquisition consistency")
        a.font = fnt_bold
        a.fill = fill_grey
        b = ws.cell(row=row, column=2,
                      value="Listed only when ONE direction disagrees with "
                            "ITSELF — a second OTDR unit, or a second "
                            "wavelength, inside a single end's set. Losses "
                            "from two instruments are not one calibration. A "
                            "direction shot as one job has no rows here.")
        b.font = fnt_small
        b.fill = fill_grey
        b.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1
        for d in audit["per_direction"]:
            for entry in d["rows"]:
                _emit(f"{d['label']} — {entry['name']}", entry["result"])
        row += 1

    # Per-trace DETAIL below (file-level fields + their outlier file lists).
    # Suppressed for the bidirectional splice report — see the docstring.
    if per_trace_detail:
        a = ws.cell(row=row, column=1, value="Per-trace detail")
        a.font = fnt_bold
        a.fill = fill_grey
        ws.cell(row=row, column=2, value="").fill = fill_grey
        row += 1
        for entry in audit["file_fields"]:
            _emit(entry["name"], entry["result"])

    # FastReporter Test Settings — two side-by-side panels, one per direction.
    # APPENDED, never inserted: everything above keeps the row it already had,
    # and `first_data_row` (and so freeze_panes) is decided before we start.
    if test_settings_block and audit.get("test_settings"):
        row = _render_test_settings(
            ws, audit["test_settings"], row + 1,
            font_name=font_name, font_size=font_size,
            fill_header=fill_header, fill_green=fill_green,
            fill_amber=fill_amber, fill_grey=fill_grey, box=box)

    # Freeze the header row for scrolling on long outlier lists
    ws.freeze_panes = f"A{first_data_row}"

    # Engine stamp, LAST row of the sheet.  Two workbooks that disagree about a
    # loss are otherwise indistinguishable — manifest 238 prints KANLAN F676 as
    # .168 and 239 prints .167, and nothing in either file said which engine
    # measured it.  Appended (never inserted) so every row above keeps the index
    # it already had and `first_data_row`/freeze_panes are decided before this
    # runs.  Read from THIS process — acquisition_audit.py is engine code loaded
    # out of OTDR_SUITE_HOME, so the label names the code that actually produced
    # the numbers, not what the hub UI believes it is running.
    #
    # Written at the cursor as-is, with NO extra increment: the blocks above
    # already leave `row` one past their last line, which is the one blank
    # separator this footer wants.  Every additional gap row is a row openpyxl
    # then has to materialise as empty cells in iter_rows(), so the increment
    # would cost two blank cells and buy nothing.
    a = ws.cell(row=row, column=1, value="Report engine")
    a.font = fnt_bold
    a.fill = fill_grey
    b = ws.cell(row=row, column=2, value=engine_stamp_text())
    b.font = fnt_small
    b.fill = fill_grey
    b.alignment = Alignment(vertical="top", wrap_text=True)

# ─────────────────────────────────────────────────────────────────────────────
#  Test Settings renderer
# ─────────────────────────────────────────────────────────────────────────────
#  LAYOUT.  Two tables, side by side, sharing every row index:
#
#      col A            col B          col C            col D
#      label (A-dir)    value (A-dir)  label (B-dir)    value (B-dir)
#
#  Side by side, and strictly row-aligned, is the whole point — "do these two
#  shots match" has to be answerable by running one finger down the page.  That
#  is also why a mixed direction gets NO extra row inside the table: an extra
#  row on one side would slide the two panels out of step and destroy the
#  comparison the table exists to make.  Those details go in a note block
#  underneath instead.
#
#  Columns A and B already carry widths set by the caller for the audit table
#  above (38 / 80) and are left exactly as they are; only the two NEW columns
#  are sized here.  Column D is narrower than B — the values are short and
#  left-aligned in both, so the eye still tracks label→value identically.
#
#  Colours are the sheet's existing three, unchanged: green = agrees, amber =
#  needs a look, grey = section banner.  No new colour vocabulary.
_TS_LABEL_W = 38
_TS_VALUE_W = 44


def _render_test_settings(ws, ts: dict, row: int, font_name: str,
                          font_size: int, fill_header, fill_green,
                          fill_amber, fill_grey, box) -> int:
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    fnt_normal = Font(name=font_name, size=font_size)
    fnt_bold   = Font(name=font_name, size=font_size, bold=True)
    fnt_small  = Font(name=font_name, size=font_size - 1, italic=True,
                      color="555555")
    fnt_warn   = Font(name=font_name, size=font_size, bold=True,
                      color="7F6000")

    # get_column_letter, not ws.cell(...).column_letter — reading a cell to
    # ask its letter would CREATE that cell and inflate max_row/max_column.
    for col, width in ((3, _TS_LABEL_W), (4, _TS_VALUE_W)):
        ws.column_dimensions[get_column_letter(col)].width = width

    a, b = ts.get("a"), ts.get("b")

    # ── Banner + headline verdict ──
    c = ws.cell(row=row, column=1, value="Test Settings")
    c.font = fnt_bold
    c.fill = fill_grey
    c = ws.cell(row=row, column=2, value=ts["headline"])
    c.font = fnt_bold if ts["clean"] else fnt_warn
    c.fill = fill_green if ts["clean"] else fill_amber
    c.alignment = Alignment(vertical="top", wrap_text=True)
    c.border = box
    for col in (3, 4):
        ws.cell(row=row, column=col, value="").fill = fill_grey
    row += 1

    c = ws.cell(row=row, column=2,
                value="FastReporter's Test Settings panel, read from each "
                      "trace. One table per direction so the A and B shots "
                      "can be checked against each other: same IOR, same "
                      "backscatter, same detection thresholds. Each cell is "
                      "that direction's most common value.")
    c.font = fnt_small
    c.alignment = Alignment(vertical="top", wrap_text=True)
    row += 1

    # ── Header row: the two panels' titles ──
    for col, text in ((1, "Test Settings"),
                      (2, (a or {}).get("label", "A-direction")),
                      (3, "Test Settings"),
                      (4, (b or {}).get("label", "B-direction"))):
        c = ws.cell(row=row, column=col, value=text)
        c.font = fnt_bold
        c.fill = fill_header
        c.border = box
    row += 1
    first_row = row

    # ── The eight FastReporter rows, both panels in lock-step ──
    for i, verdict in enumerate(ts["verdicts"]):
        for side, base_col, flag_key in ((a, 1, "flag_a"), (b, 3, "flag_b")):
            lab_c = ws.cell(row=row, column=base_col)
            val_c = ws.cell(row=row, column=base_col + 1)
            if side is None:
                lab_c.value = verdict["name"]
                lab_c.font = fnt_normal
                val_c.value = "n/a (direction not loaded)"
                val_c.font = fnt_normal
            else:
                r = side["rows"][i]
                lab_c.value = r["label"]
                # A mixed direction is marked in its own cell, so the reader
                # sees WHICH side is inconsistent without leaving the table.
                val_c.value = (f"{r['value']}  ⚠ mixed" if r["mixed"]
                               else r["value"])
                flagged = verdict[flag_key]
                fill = fill_amber if flagged else fill_green
                lab_c.fill = fill
                val_c.fill = fill
                lab_c.font = fnt_normal
                val_c.font = fnt_warn if flagged else fnt_normal
            lab_c.border = box
            val_c.border = box
            lab_c.alignment = Alignment(vertical="top")
            val_c.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1
    assert row - first_row == len(_FR_ROWS)   # panels stayed in lock-step

    # ── Notes: the detail that would have broken row alignment above ──
    notes = []
    for i, v in enumerate(ts["verdicts"]):
        if v["differs"]:
            ra = a["rows"][i] if a else None
            rb = b["rows"][i] if b else None
            notes.append(
                f"{v['name']}: {a['label']} = {ra['value']} but "
                f"{b['label']} = {rb['value']}. The two directions were not "
                f"shot with the same setting, so their losses are not "
                f"strictly comparable.")
    notes.extend(ts["mixed_notes"])
    notes.append(
        "Fiber core size is not stored anywhere in a .sor file — neither the "
        "Bellcore blocks nor EXFO's proprietary block carries a core size or "
        "mode-field diameter. What is stored is the glass designation (ITU-T "
        "fiber type), which is what the cell shows; FastReporter's \"9 µm\" is "
        "its own display of that designation. The directions are still "
        "compared on the stored designation.")
    if notes:
        row += 1
        for text in notes:
            ws.cell(row=row, column=1, value="").border = box
            c = ws.cell(row=row, column=2, value=text)
            c.font = fnt_small
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = box
            row += 1
    return row
